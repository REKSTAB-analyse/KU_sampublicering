import streamlit as st
import json
import os
import tempfile
import networkx as nx
from pyvis.network import Network
from matplotlib import colors as mcolors
from math import cos, sin, pi
import numpy as np
import colorsys
import pyarrow as pa
import io
import csv
from datetime import datetime
from collections import defaultdict, OrderedDict
import plotly.graph_objects as go
from networkx.algorithms.community.quality import modularity as modq
from networkx.algorithms.community import greedy_modularity_communities
import ast
import math
import re
from pathlib import Path
import hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import random
import paramiko
import base64
import subprocess

# ---------------------------------------------------------------------------
# CONFIG - all paths in one place
# ---------------------------------------------------------------------------

_ERDA = st.secrets["erda"]
DATA_PATH = _ERDA["data_path"]

@st.cache_resource
def get_sftp():
    transport = paramiko.Transport((_ERDA["host"], 22))
    transport.connect(username=_ERDA["username"], password=_ERDA["password"])
    return paramiko.SFTPClient.from_transport(transport)

def read_file(filename: str) -> bytes:
    sftp = get_sftp()
    with sftp.open(f"{DATA_PATH}/{filename}", "rb") as f:
        return f.read()

# ---------------------------------------------------------------------------
# CONSTANTS
# ---------------------------------------------------------------------------

HIERARKI = {
    "Særlig stilling": -1,
    "Øvrige VIP (DVIP)": 0,
    "Ph.d.": 1,
    "Stillinger u. adjunktniveau": 2,
    "Postdoc": 3,
    "Adjunkt": 4,
    "Lektor": 5,
    "Professor": 6,
}

CPR = {
    "m": "Mænd",
    "k": "Kvinder"
}

GROUP_ORDER = sorted(HIERARKI.keys(), key=lambda g: HIERARKI[g])
LVL_MIN, LVL_MAX = min(HIERARKI.values()), max(HIERARKI.values())

#FAC_ORDER = ["SAMF", "JUR", "TEO", "SUND", "HUM", "SCIENCE"]
FAC_ORDER = ["SAMF", "SCIENCE", "TEO", "SUND", "HUM", "JUR"]
FAC_ABBRS = {
    "Det Teologiske Fakultet": "TEO",
    "Det Juridiske Fakultet": "JUR",
    "Det Humanistiske Fakultet": "HUM",
    "Det Natur- og Biovidenskabelige Fakultet": "SCIENCE",
    "Det Samfundsvidenskabelige Fakultet": "SAMF",
    "Det Sundhedsvidenskabelige Fakultet": "SUND",
}

def make_abbr(name: str, existing: set = None) -> str:
    """Lav forkortelse af et navn ved at tage forbogstaver af ord, minus stopord."""
    _stop = {"for", "og", "i", "af", "til", "det", "den", "de", "en", "et",
             "med", "på", "ved", "om", "fra", "under", "over"}
    words = [w for w in name.split() if w.lower() not in _stop]
    abbr = "".join(w[0].upper() for w in words if w)
    if not abbr:
        abbr = name[:4].upper()
    # Hvis forkortelsen kolliderer, tilføj ekstra bogstav fra første ord
    if existing is not None:
        original = abbr
        i = 1
        while abbr in existing:
            abbr = original + (words[0][i].upper() if i < len(words[0]) else str(i))
            i += 1
    return abbr


# ---------------------------------------------------------------------------
# DATA LOADING
# ---------------------------------------------------------------------------

@st.cache_data(show_spinner="Indlæser netværksdata...")
def load_network_data() -> dict:
    raw = json.loads(read_file("vip_transformed.json"))
    return {int(k): v for k, v in raw.items()}

@st.cache_data
def load_ku_colors() -> dict:
    return json.loads(read_file("ku-farver02.json"))

@st.cache_data
def load_stilling_map() -> dict:
    result = {}
    reader = csv.DictReader(
        read_file("KU_stillingstyper.csv").decode("utf-8").splitlines(),
        delimiter=";"
    )
    for row in reader:
        raw = row["Stillingstype"].strip()
        med = row["medtages?"].strip()
        grp = row["Stillingsgruppe"].strip()
        if med == "1" and grp:
            result[raw] = grp
    return result

@st.cache_data(show_spinner="Indlæser forfatterdata...")
def load_forfatterantal() -> dict:
    return json.loads(read_file("forfatterantal.json"))

@st.cache_data(show_spinner="Indlæser forfatterdata...")
def load_forfatterantal_dist() -> dict:
    return json.loads(read_file("forfatterantal_dist.json"))

@st.cache_data
def load_publikationstyper() -> dict:
    return json.loads(read_file("publikationstyper.json"))

@st.cache_data
def load_inst_filter() -> tuple[set, dict]:
    inst_ok = set()
    inst_to_fac = {}
    reader = csv.DictReader(
        read_file("Fakulteter_institutter.csv").decode("utf-8-sig").splitlines(),
        delimiter=";"
    )
    for row in reader:
        if row["medtages?"].strip() == "1":
            fac  = row["Fakultet"].strip()
            inst = row["Institut"].strip()
            alt  = row["Alternativ"].strip()
            if inst:
                inst_ok.add(inst)
                inst_to_fac[inst] = fac
            if alt:
                inst_ok.add(alt)
                inst_to_fac[alt] = fac
    return inst_ok, inst_to_fac

@st.cache_data
def load_pubtype_map() -> dict:
    mapping = OrderedDict()
    lines = read_file("Publikationstyper_mod.csv").decode("utf-8").splitlines()
    reader = csv.DictReader(lines, delimiter=";")
    field_map = {h.strip(): h for h in reader.fieldnames or []}
    col_raw = next((field_map[k] for k in field_map if k.lower() == "publikationstype"), None)
    col_collapsed = next((field_map[k] for k in field_map if k.lower() == "kollapset"), None)
        
    if not col_raw or not col_collapsed:
        raise ValueError()

    for row in reader:
        raw = (row.get(col_raw) or "").strip()
        col = (row.get(col_collapsed) or "").strip()
        if not raw:
            continue
        mapping[raw] = col or raw
    
    return dict(mapping)

@st.cache_data(show_spinner="Indlæser forfatterdata...")
def load_forfatterpositioner() -> dict:
    return json.loads(read_file("forfatterpositioner.json"))

@st.cache_data
def load_ku_totals() -> dict:
    raw = json.loads(read_file("ku_totals.json"))
    return {int(k): v for k, v in raw.items()}

@st.cache_data  
def load_logo() -> bytes:
    return read_file("KU-logo.png")

@st.cache_data
def load_svg(filename: str) -> str | None:
    try:
        return read_file(filename).decode("utf-8")
    except Exception:
        return None

# ---------------------------------------------------------------------------
# GitHub
# ---------------------------------------------------------------------------

def _get_last_deploy_date() -> str:
    try:
        repo_dir = os.path.dirname(os.path.abspath(__file__))
        ts = subprocess.check_output(
            ["git", "log", "-1", "--format=%ci"],
            cwd=repo_dir,
            stderr=subprocess.DEVNULL,
        ).decode().strip()
        dt = datetime.fromisoformat(ts)
        return f"{dt.day}. {dt.strftime('%B').lower()} {dt.year}"
    except Exception:
        d = datetime.today()
        return f"{d.day}. {d.strftime('%B').lower()} {d.year}"

_DEPLOY_DATE = _get_last_deploy_date()

# ---------------------------------------------------------------------------
# COLOR HELPERS
# ---------------------------------------------------------------------------
_KU_PALETTE_RAW = [
    # Mørke - høj kontrast, bruges først
    "#122947",  # Blå mørk
    "#901a1E",  # Rød mørk
    "#39641c",  # Grøn mørk
    "#0a5963",  # Petroleum mørk
    "#3d3d3d",  # Grå mørk
    "#7d5402",  # Brun mørk (JUR)
    # Mellem - god læsbarhed
    "#ffbd38",  # Gul (adskiller)
    "#4b8325",  # Grøn mellem
    "#c73028",  # Rød mellem
    "#197f8e",  # Petroleum mellem
    "#425570",  # Blå mellem
    "#666666",  # Grå mellem
    # Lyse - bruges sidst, kun ved mange kategorier
    "#bac7d9",  # Blå lys
    "#dB3B0A",  # Rød-orange lys
    "#becaa8",  # Grøn lys
    "#b7d7de",  # Petroleum lys
    "#e1dfdf",  # Grå lys
]

def ku_color_sequence(n: int, seed: int = 26) -> list[str]:
    if n <= len(_KU_PALETTE_RAW):
        return _KU_PALETTE_RAW[:n]
    # Fallback ved flere kategorier end paletten rækker
    plotly_defaults = [
        "#3A1A5F",  # Lilla mørk (TEO)
        "#7d5402",  # Brun mellem
        "#c45c5f",  # Rosa-rød
        "#5C1012",  # Rød meget mørk
        "#6B84A0",  # Blå-grå
        "#fefaf2",  # Champagne
        "#7A131A",  # Bordeaux
        "#aaaaaa",  # Grå neutral
        "#ffbd38",  # Gul (gentaget)
        "#becaa8",  # Grøn lys (gentaget)
    ]
    extras = plotly_defaults * ((n - len(_KU_PALETTE_RAW)) // len(plotly_defaults) + 1)
    return _KU_PALETTE_RAW + extras[:n - len(_KU_PALETTE_RAW)]

def build_faculty_colors(ku_farver: dict) -> dict:
    return {
        "TEO": "#3A1A5F",
        "JUR": "#7d5402",
        "HUM": ku_farver["Blaa"]["Moerk"],
        "SCIENCE": ku_farver["Groen"]["Moerk"],
        "SAMF": ku_farver["Petroleum"]["Moerk"],
        "SUND": "#7A131A",
    }

def stillingsgruppe_colors(ku_farver: dict) -> dict:
    return {
        "Særlig stilling": "#D4D4D4",
        "Øvrige VIP (DVIP)": "#BAC7D9",
        "Ph.d.": "#6B84A0",
        "Stillinger u. adjunktniveau":"#425570",
        "Postdoc": "#AAAAAA",
        "Adjunkt": "#C45C5F",
        "Lektor": "#901A1E",
        "Professor": "#5C1012",
    }

def adjust_color(hex_color: str, lightness_factor: float = 1.0, saturation_factor: float = 1.0) -> str:
    r, g, b = mcolors.to_rgb(hex_color)
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    l = max(0.0, min(1.0, l * lightness_factor))
    s = max(0.0, min(1.0, s * saturation_factor))
    r2, g2, b2 = colorsys.hls_to_rgb(h, l, s)
    return mcolors.to_hex((r2, g2, b2))

def add_alpha(hex_color: str, alpha: float) -> str:
    r, g, b = mcolors.to_rgb(hex_color)
    return f"rgba({int(r*255)}, {int(g*255)}, {int(b*255)}, {alpha})"

# ---------------------------------------------------------------------------
# TRANSLATE NATIONALITIES
# ---------------------------------------------------------------------------

_COUNTRY_NAMES_DA = {
    "DK":  "Danmark",
    "D":   "Tyskland",
    "CN":  "Kina",
    "I":   "Italien",
    "GB":  "Storbritannien",
    "E":   "Spanien",
    "USA": "USA",
    "S":   "Sverige",
    "NL":  "Holland",
    "IND": "Indien",
    "F":   "Frankrig",
    "GR":  "Grækenland",
    "N":   "Norge",
    "PL":  "Polen",
    "IR":  "Iran",
    "AUS": "Australien",
    "CDN": "Canada",
    "P":   "Portugal",
    "BR":  "Brasilien",
    "B":   "Belgien",
    "RUS": "Rusland",
    "SF":  "Finland",
    "A":   "Østrig",
    "IRL": "Irland",
    "CH":  "Schweiz",
    "MEX": "Mexico",
    "J":   "Japan",
    "TR":  "Tyrkiet",
    "PAK": "Pakistan",
    "ROK": "Sydkorea",
    "R":   "Rumænien",
    "LTU": "Litauen",
    "IS":  "Island",
    "H":   "Ungarn",
    "ETH": "Etiopien",
    "RCH": "Chile",
    "CZE": "Tjekkiet",
    "CO":  "Colombia",
    "HRV": "Kroatien",
    "BG":  "Bulgarien",
    "IL":  "Israel",
    "UKR": "Ukraine",
    "NEP": "Nepal",
    "LVA": "Letland",
    "SVN": "Slovenien",
    "SVK": "Slovakiet",
    "EST": "Estland",
    "SRB": "Serbien",
    "VN":  "Vietnam",
    "PE":  "Peru",
    "RI":  "Indonesien",
    "ZA":  "Sydafrika",
    "ET":  "Egypten",
    "T":   "Thailand",
    "AR":  "Argentina",
    "NZ":  "New Zealand",
    "PI":  "Filippinerne",
    "ZW":  "Zimbabwe",
    "EAK": "Kenya",
    "RC":  "Taiwan",
    "ARM": "Armenien",
    "RL":  "Libanon",
    "MAL": "Malaysia",
    "BD":  "Bangladesh",
    "GH":  "Ghana",
    "SGP": "Singapore",
    "HKJ": "Jordan",
    "GDA": "Ukendt",
    "BHU": "Bhutan",
    "MOZ": "Mozambique",
    "CL":  "Sri Lanka",
    "L":   "Luxembourg",
    "UZB": "Usbekistan",
    "EAT": "Tanzania",
    "BH":  "Bahrain",
    "EC":  "Ecuador",
    "DY":  "Benin",
    "MDA": "Moldova",
    "RWA": "Rwanda",
    "EAU": "Uganda",
    "YV":  "Venezuela",
    "MS":  "Mauritius",
    "BLR": "Belarus",
    "AL":  "Albanien",
    "BIH": "Bosnien-Hercegovina",
    "SN":  "Senegal",
    "YMN": "Yemen",
    "WAN": "Nigeria",
    "KAZ": "Kasakhstan",
    "SU":  "Sovjetunionen",
    "MAK": "Nordmakedonien",
    "MDG": "Madagaskar",
    "SWA": "Namibia",
    "CY":  "Cypern",
    "BOL": "Bolivia",
    "DZ":  "Algeriet",
    "SYR": "Syrien",
    "KWT": "Kuwait",
    "GEO": "Georgien",
    "TN":  "Tunesien",
    "DOM": "Dominikanske Republik",
    "CAM": "Cameroun",
    "NIC": "Nicaragua",
    "FL":  "Liechtenstein",
    "MA":  "Marokko",
    "OMN": "Oman",
    "Ukendt": "Ukendt",
}

def country_name(code: str) -> str:
    """Returnér det danske landenavn for en bil-kendingskode. Falder tilbage til koden selv hvis ukendt."""
    if not code:
        return "Ukendt"
    return _COUNTRY_NAMES_DA.get(code, _COUNTRY_NAMES_DA.get(code.upper(), code))


# ---------------------------------------------------------------------------
# SIZE SCALING
# ---------------------------------------------------------------------------

def scale_size_log(val: float, max_auth: float, px_min: float = 5, px_max: float = 60) -> float:
    if val <= 0 or max_auth <= 0:
        return px_min
    return px_min + (math.log1p(val) / math.log1p(max_auth)) * (px_max - px_min)

# ---------------------------------------------------------------------------
# PYARROW HELPERS
# ---------------------------------------------------------------------------

def build_table(rows: list, schema_fields: list) -> pa.Table:
    cols = {name: [r.get(name) for r in rows] for name, _ in schema_fields}
    arrays = [pa.array(cols[name], type=dtype) for name, dtype in schema_fields]
    return pa.Table.from_arrays(arrays, names=[name for name, _ in schema_fields])

def rows_to_csv_bytes(rows: list, field_order: list) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=field_order, extrasaction="ignore")
    writer.writeheader()
    for r in rows:
        writer.writerow({k: r.get(k) for k in field_order})
    return buf.getvalue().encode("utf-8")

def rows_to_excel_bytes(rows: list, field_order:list, sheet_name: str = "Data") -> bytes:
    """Return a styled .xlsx file as bytes with KU red header styling."""

    KU_RED   = "901a1E"   # KU primary red (no #)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_font    = Font(name="Arial", bold=False, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", start_color=KU_RED, end_color=KU_RED)
    header_align   = Alignment(horizontal="left", vertical="center", wrap_text=False)
    thin_side      = Side(style="thin", color="DDDDDD")
    cell_border    = Border(bottom=thin_side)
    data_font      = Font(name="Arial", size=10)
    data_align     = Alignment(horizontal="left", vertical="center")

    # Header row
    for col_idx, col_name in enumerate(field_order, start = 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    
    # Data rows
    for row_idx, row in enumerate(rows, start = 2):
        for col_idx, col_name in enumerate(field_order, start=1):
            val = row.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = cell_border    
            if isinstance(val, float):
                cell.number_format = '0.00'    
            elif isinstance(val, int):
                cell.number_format = '0'       

    for col_idx, col_name in enumerate(field_order, start=1):
        max_len = max(
            len(str(col_name)),
            *(len(str(row.get(col_name, "") or "")) for row in rows),
        ) if rows else len(col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)
    
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def fmt_dk(value, decimals: int = 1) -> str:
    """Formater tal med dansk notation: punktum som tusindtalsseparator, komma som decimal."""
    if value is None:
        return ""
    formatted = f"{value:,.{decimals}f}"          # engelsk: 1,234.56
    return formatted.replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_ui(value, decimals=1):
    """Formater tal til UI: punktum som decimal, ingen tusindseparator."""
    if value is None:
        return ""
    fmt = f"{{:.{decimals}f}}"
    return fmt.format(value)


# ---------------------------------------------------------------------------
# NODE-LABEL HELPER
# ---------------------------------------------------------------------------

def make_node_label(m: dict) -> str:
    parts = [m[k] for k in ("fac", "inst", "grp") if m.get(k)]
    return " | ".join(parts)

# ---------------------------------------------------------------------------
# GRAPH MERGE FUNCTIONS
# ---------------------------------------------------------------------------

def merge_grp_to_facgrp(nodes: dict, edges: list):
    merged_meta = {}
    old_to_new = {}
    size_acc = {}

    for nid, m in nodes.items():
        if m.get("type") != "grp":
            continue
        fac, grp = m["fac"], m["grp"]
        new_id = f"{fac}|{grp}"
        old_to_new[nid] = new_id
        if new_id not in merged_meta:
            merged_meta[new_id] = {"type": "grp", "fac": fac, "inst": "", "grp": grp, "size": 0, "children": []}
            size_acc[(fac, grp)] = 0
        size_acc[(fac, grp)] += m.get("size", 0)
        merged_meta[new_id]["children"].append(nid)

    for nid, m in nodes.items():
        if m.get("type") == "fac":
            merged_meta[nid] = m

    for (fac, grp), s in size_acc.items():
        merged_meta[f"{fac}|{grp}"]["size"] = s

    edge_acc: dict[tuple, float] = {}
    for edge in edges:
        u, v, w = edge[0], edge[1], edge[2]
        sex_combo = edge[3] if len(edge) > 3 else None
        if u not in old_to_new or v not in old_to_new:
            continue
        u2, v2 = old_to_new[u], old_to_new[v]
        if u2 == v2:
            continue
        key = (min(u2, v2), max(u2, v2), sex_combo)
        edge_acc[key] = edge_acc.get(key, 0) + w

    return merged_meta, [(u, v, w, sc) for (u, v, sc), w in edge_acc.items()]

def merge_grp_to_inst(nodes: dict, edges: list):
    merged_meta = {}
    old_to_new = {}

    for nid, m in nodes.items():
        if m.get("type") != "grp":
            continue
        fac, inst = m["fac"], m["inst"]
        new_id = f"INST:{fac}|{inst}"
        old_to_new[nid] = new_id
        if new_id not in merged_meta:
            merged_meta[new_id] = {"type": "inst", "fac": fac, "inst": inst, "grp": "", "size": 0, "children": []}
        merged_meta[new_id]["size"] += int(m.get("size", 0))
        merged_meta[new_id]["children"].append(nid)

    for nid, m in nodes.items():
        t = m.get("type")
        if t == "inst":
            if nid not in merged_meta:
                merged_meta[nid] = dict(m)
                merged_meta[nid].setdefault("children", [])
            old_to_new[nid] = nid
        elif t == "fac":
            merged_meta[nid] = m
            old_to_new[nid] = nid

    edge_acc = {}
    for edge in edges:
        u, v, w = edge[0], edge[1], edge[2]
        if u in old_to_new and v in old_to_new:
            u2, v2 = old_to_new[u], old_to_new[v]
            if u2 != v2:
                key = tuple(sorted((u2, v2)))
                edge_acc[key] = edge_acc.get(key, 0) + w

    return merged_meta, [(u, v, w) for (u, v), w in edge_acc.items()]

def merge_all_to_fac(nodes: dict, edges: list):
    merged_meta = {}
    old_to_new = {}

    for nid, m in nodes.items():
        fac = m["fac"]
        new_id = f"FAC:{fac}"
        old_to_new[nid] = new_id
        if new_id not in merged_meta:
            merged_meta[new_id] = {"type": "fac", "fac": fac, "inst": "", "grp": "", "size": 0, "children": []}
        merged_meta[new_id]["size"] += m.get("size", 0)
        merged_meta[new_id]["children"].append(nid)

    edge_acc = {}
    for edge in edges:
        u, v, w = edge[0], edge[1], edge[2]
        if u not in old_to_new or v not in old_to_new:
            continue
        u2, v2 = old_to_new[u], old_to_new[v]
        if u2 == v2:
            continue
        key = tuple(sorted((u2, v2)))
        edge_acc[key] = edge_acc.get(key, 0) + w

    return merged_meta, [(u, v, w) for (u, v), w in edge_acc.items()]

def merge_fac_by_sex(nodes: dict, edges: list):
    """Split each faculty into two nodes - one per sex - giving mode FS."""
    merged_meta = {}
    old_to_new  = {}

    for nid, m in nodes.items():
        if m.get("type") != "grp":
            continue
        fac = m["fac"]
        sex = m.get("sex", "m")
        new_id = f"FAC:{fac}|{sex}"
        old_to_new[nid] = new_id

        if new_id not in merged_meta:
            merged_meta[new_id] = {
                "type":     "fac_sex",
                "fac":      fac,
                "sex":      sex,
                "inst":     "",
                "grp":      "",
                "size":     0,
                "children": [],
            }
        merged_meta[new_id]["size"]     += m.get("size", 0)
        merged_meta[new_id]["children"].append(nid)

    edge_acc = {}
    for edge in edges:
        u_raw, v_raw, w = edge[0], edge[1], edge[2]
        sex_combo = edge[3] if len(edge) > 3 else None

        if u_raw not in old_to_new or v_raw not in old_to_new:
            continue
        u2, v2 = old_to_new[u_raw], old_to_new[v_raw]
        if u2 == v2:
            continue

        key = tuple(sorted((u2, v2)))
        if key not in edge_acc:
            edge_acc[key] = {"weight": 0, "sex_combo": sex_combo}
        edge_acc[key]["weight"] += w

    merged_edges = [
        (u, v, acc["weight"], acc["sex_combo"])
        for (u, v), acc in edge_acc.items()
    ]
    return merged_meta, merged_edges

def merge_inst_by_sex(nodes: dict, edges: list):
    """Split each institute into two nodes - one per sex."""
    merged_meta = {}
    old_to_new  = {}

    for nid, m in nodes.items():
        if m.get("type") != "grp":
            continue
        fac  = m["fac"]
        inst = m["inst"]
        sex  = m.get("sex", "m")
        new_id = f"INST:{fac}|{inst}|{sex}"
        old_to_new[nid] = new_id

        if new_id not in merged_meta:
            merged_meta[new_id] = {
                "type":     "inst_sex",
                "fac":      fac,
                "inst":     inst,
                "sex":      sex,
                "grp":      "",
                "size":     0,
                "children": [],
            }
        merged_meta[new_id]["size"]     += m.get("size", 0)
        merged_meta[new_id]["children"].append(nid)

    edge_acc = {}
    for edge in edges:
        u_raw, v_raw, w = edge[0], edge[1], edge[2]
        sex_combo = edge[3] if len(edge) > 3 else None

        if u_raw not in old_to_new or v_raw not in old_to_new:
            continue
        u2, v2 = old_to_new[u_raw], old_to_new[v_raw]
        if u2 == v2:
            continue

        key = tuple(sorted((u2, v2)))
        if key not in edge_acc:
            edge_acc[key] = {"weight": 0, "sex_combo": sex_combo}
        edge_acc[key]["weight"] += w

    merged_edges = [
        (u, v, acc["weight"], acc["sex_combo"])
        for (u, v), acc in edge_acc.items()
    ]
    return merged_meta, merged_edges

def merge_grp_by_sex(nodes: dict, edges: list):
    """Split each position group into two nodes - one per sex."""
    merged_meta = {}
    old_to_new  = {}

    for nid, m in nodes.items():
        if m.get("type") != "grp":
            continue
        grp = m["grp"]
        sex = m.get("sex", "m")
        new_id = f"GRP:{grp}|{sex}"
        old_to_new[nid] = new_id

        if new_id not in merged_meta:
            merged_meta[new_id] = {
                "type":     "grp_sex",
                "fac":      m.get("fac", ""),
                "inst":     m.get("inst", ""),
                "grp":      grp,
                "sex":      sex,
                "size":     0,
                "children": [],
            }
        merged_meta[new_id]["size"]     += m.get("size", 0)
        merged_meta[new_id]["children"].append(nid)

    edge_acc = {}
    for edge in edges:
        u_raw, v_raw, w = edge[0], edge[1], edge[2]
        sex_combo = edge[3] if len(edge) > 3 else None

        if u_raw not in old_to_new or v_raw not in old_to_new:
            continue
        u2, v2 = old_to_new[u_raw], old_to_new[v_raw]
        if u2 == v2:
            continue

        key = tuple(sorted((u2, v2)))
        if key not in edge_acc:
            edge_acc[key] = {"weight": 0, "sex_combo": sex_combo}
        edge_acc[key]["weight"] += w

    merged_edges = [
        (u, v, acc["weight"], acc["sex_combo"])
        for (u, v), acc in edge_acc.items()
    ]
    return merged_meta, merged_edges

def merge_fac_by_nat(nodes: dict, edges: list):
    """Split each faculty into one node per nationality."""
    merged_meta = {}
    old_to_new  = {}

    for nid, m in nodes.items():
        if m.get("type") != "grp":
            continue
        fac = m["fac"]
        nat = m.get("statsborgerskab", "Ukendt") or "Ukendt"
        new_id = f"FAC:{fac}|{nat}"
        old_to_new[nid] = new_id

        if new_id not in merged_meta:
            merged_meta[new_id] = {
                "type":             "fac_nat",
                "fac":              fac,
                "statsborgerskab":  nat,
                "inst":             "",
                "grp":              "",
                "size":             0,
                "children":         [],
            }
        merged_meta[new_id]["size"]     += m.get("size", 0)
        merged_meta[new_id]["children"].append(nid)

    edge_acc = {}
    for edge in edges:
        u_raw, v_raw, w = edge[0], edge[1], edge[2]
        if u_raw not in old_to_new or v_raw not in old_to_new:
            continue
        u2, v2 = old_to_new[u_raw], old_to_new[v_raw]
        if u2 == v2:
            continue
        key = tuple(sorted((u2, v2)))
        edge_acc[key] = edge_acc.get(key, 0.0) + w

    merged_edges = [(u, v, w) for (u, v), w in edge_acc.items()]
    return merged_meta, merged_edges



# ---------------------------------------------------------------------------
# MODE HELPER FUNCTIONS
# ---------------------------------------------------------------------------

def sex_in_mode(mode: str) -> bool:
    return "S" in mode

def nat_in_mode(mode: str) -> bool:
    return "N" in mode

def fac_in_mode(mode: str) -> bool:
    return "F" in mode

def inst_in_mode(mode: str) -> bool:
    return "I" in mode

def grp_in_mode(mode: str) -> bool:
    return "G" in mode

def base_mode(mode: str) -> str:
    """Returnerer mode uden S-suffikset, fx 'FIG' fra 'FIGS'."""
    return mode.replace("S", "").replace("N", "")

def network_mode(mode: str) -> str:

    return mode


# ---------------------------------------------------------------------------
# APPLY MODE MERGE
# ---------------------------------------------------------------------------
def _split_nodes_sex(nm: dict, es: list, raw_nodes: dict = None) -> tuple:
    """Split grp-noder på køn og opdater kanter tilsvarende."""
    nm_sex = {}
    old_to_new = {}  # original_id -> {sex -> new_id}

    for nid, m in nm.items():
        if m.get("type") == "grp":
            old_to_new[nid] = {}
            # Beregn størrelse per køn fra children
            size_by_sex: dict[str, int] = {"k": 0, "m": 0}
            for child_id in m.get("children", []):
                parts = child_id.split("|")
                child_sex = parts[3] if len(parts) > 3 else None
                if child_sex in size_by_sex and raw_nodes and child_id in raw_nodes:
                    size_by_sex[child_sex] += raw_nodes[child_id].get("size", 0)
            for sx in ["k", "m"]:
                new_id = nid + f"|{sx}"
                nm_sex[new_id] = {**m, "sex": sx, "type": "grp_sex", "size": size_by_sex[sx]}
                old_to_new[nid][sx] = new_id
        else:
            nm_sex[nid] = m

    es_sex = []
    for edge in es:
        u, v, w = edge[0], edge[1], edge[2]
        sex_combo = edge[3] if len(edge) > 3 and edge[3] else None
        if sex_combo and "-" in sex_combo:
            sx_u, sx_v = sex_combo.split("-", 1)
        else:
            sx_u, sx_v = None, None

        u_new = old_to_new.get(u, {}).get(sx_u, u) if u in old_to_new else u
        v_new = old_to_new.get(v, {}).get(sx_v, v) if v in old_to_new else v

        if u_new != v_new:
            es_sex.append((u_new, v_new, w, sex_combo))

    return nm_sex, es_sex

def apply_mode_merge(mode: str, raw_nodes: dict, raw_edges: list):
    """Return (node_meta, edge_source) for the current display mode."""

    _base = base_mode(mode)
    _sex  = sex_in_mode(mode)

    if _base == "F":
        if nat_in_mode(mode):
            return {nid: m for nid, m in raw_nodes.items() if m.get("type") == "grp"}, list(raw_edges)
        nm, es = merge_fac_by_sex(raw_nodes, raw_edges) if _sex else merge_all_to_fac(raw_nodes, raw_edges)
        return nm, es
    elif _base == "I":
        if nat_in_mode(mode):
            return {nid: m for nid, m in raw_nodes.items() if m.get("type") == "grp"}, list(raw_edges)
        if _sex:
            return merge_inst_by_sex(raw_nodes, raw_edges)
        raw_grp = {nid: m for nid, m in raw_nodes.items() if m.get("type") == "grp"}
        nm, es = merge_grp_to_inst(raw_grp, raw_edges)
        return {nid: m for nid, m in nm.items() if m.get("type") == "inst"}, es
    elif _base == "G":
        if nat_in_mode(mode):
            # Bevar rå grp-noder med nationalitet - _needs_aggregation splitter dem korrekt
            return {nid: m for nid, m in raw_nodes.items() if m.get("type") == "grp"}, list(raw_edges)
        if _sex:
            return merge_grp_by_sex(raw_nodes, raw_edges)
        else:
            # Merge alle rå grp-noder til én node per stillingsgruppe
            merged = {}
            sizes = {}
            for nid, m in raw_nodes.items():
                if m.get("type") != "grp":
                    continue
                grp = m["grp"]
                if grp not in merged:
                    merged[grp] = {"type": "grp", "grp": grp, "fac": "", "inst": "", "size": 0}
                    sizes[grp] = 0
                sizes[grp] += m.get("size", 0)
            for g, s in sizes.items():
                merged[g]["size"] = s
            edge_acc: dict[tuple, float] = {}
            for edge in raw_edges:
                u, v, w = edge[0], edge[1], edge[2]
                sc = edge[3] if len(edge) > 3 else None
                if u not in raw_nodes or v not in raw_nodes:
                    continue
                grp_u = raw_nodes[u]["grp"]
                grp_v = raw_nodes[v]["grp"]
                if grp_u == grp_v:
                    continue
                key = (min(grp_u, grp_v), max(grp_u, grp_v), sc)
                edge_acc[key] = edge_acc.get(key, 0) + w
            return merged, [(u, v, w, sc) for (u, v, sc), w in edge_acc.items()]
    elif _base == "FI":
        if nat_in_mode(mode):
            return {nid: m for nid, m in raw_nodes.items() if m.get("type") == "grp"}, list(raw_edges)
        if _sex:
            nm, es = merge_inst_by_sex(raw_nodes, raw_edges)
        else:
            nm, es = merge_grp_to_inst(raw_nodes, raw_edges)
        return nm, es
    elif _base == "FIG":
        if nat_in_mode(mode):
            return {nid: m for nid, m in raw_nodes.items() if m.get("type") == "grp"}, list(raw_edges)
        nm, es = merge_grp_variants(raw_nodes, raw_edges)
        nm = {nid: m for nid, m in nm.items() if m.get("type") != "inst"}
        if _sex:
            nm, es = _split_nodes_sex(nm, es, raw_nodes)
        return nm, es
    elif _base == "FG":
        if nat_in_mode(mode):
            return {nid: m for nid, m in raw_nodes.items() if m.get("type") == "grp"}, list(raw_edges)
        nm, es = merge_grp_to_facgrp(raw_nodes, raw_edges)
        if _sex:
            nm, es = _split_nodes_sex(nm, es, raw_nodes)
        return nm, es
    elif _base == "IG":
        if nat_in_mode(mode):
            return {nid: m for nid, m in raw_nodes.items() if m.get("type") == "grp"}, list(raw_edges)
        nm, merged_edges = merge_grp_variants(raw_nodes, raw_edges)
        node_meta = {}
        for nid, m in nm.items():
            if m.get("type") == "grp":
                node_meta[nid] = m
                fac, inst = m.get("fac", ""), m.get("inst", "")
                inst_id = f"INST:{fac}|{inst}"
                if inst_id not in node_meta:
                    node_meta[inst_id] = {"type": "inst", "fac": fac, "inst": inst, "grp": "", "size": 0}
        if _sex:
            node_meta, merged_edges = _split_nodes_sex(node_meta, merged_edges, raw_nodes)
        return node_meta, merged_edges

    return raw_nodes, raw_edges

def merge_grp_variants(nodes: dict, edges: list):
    """Collapse grp nodes that differ only in sex/citizenship into one node."""
    old_to_new = {}
    merged = {}

    for nid, m in nodes.items():
        if m.get("type") != "grp":
            merged[nid] = m
            continue
        new_id = f"{m['fac']}|{m['inst']}|{m['grp']}"
        old_to_new[nid] = new_id
        if new_id not in merged:
            merged[new_id] = {"type": "grp", "fac": m["fac"], "inst": m["inst"],
                               "grp": m["grp"], "size": 0, "children": [],
                               "statsborgerskab": m.get("statsborgerskab", "")}
        merged[new_id]["size"] += m.get("size", 0)
        merged[new_id]["children"].append(nid)

    edge_acc: dict[tuple, float] = {}
    for edge in edges:
        u, v, w = edge[0], edge[1], edge[2]
        sex_combo = edge[3] if len(edge) > 3 else None
        u2 = old_to_new.get(u, u)
        v2 = old_to_new.get(v, v)
        if u2 == v2:
            continue
        key = (min(u2, v2), max(u2, v2), sex_combo)
        edge_acc[key] = edge_acc.get(key, 0) + w

    return merged, [(u, v, w, sc) for (u, v, sc), w in edge_acc.items()]

# ---------------------------------------------------------------------------
# NODE / EDGE FILTER HELPERS
# ---------------------------------------------------------------------------

def passes_category_filters(m: dict, mode: str, show_fac: bool, show_inst: bool,
                              show_grp: bool, selected_facs: list,
                              selected_insts: list, selected_grps: list) -> bool:
    fac  = m.get("fac")
    inst = m.get("inst")
    grp  = m.get("grp")
    t    = m.get("type")

    if show_fac and selected_facs and fac not in selected_facs:
        return False
    if show_inst and selected_insts and inst not in selected_insts:
        return False
    if base_mode(mode) in ("FI", "IG", "I"):
        return True
    if t in ("grp", "grp_sex") and show_grp and selected_grps and grp not in selected_grps:
        return False
    return True

def size_relevant_in_mode(m: dict, mode: str) -> bool:
    t     = m.get("type", "grp")
    _base = base_mode(mode)
    _sex  = sex_in_mode(mode)
    if t == "fac_sex":
        return _sex and _base == "F"
    if t == "fac":
        return not _sex and _base == "F"
    if t == "inst_sex":
        return _sex and _base == "I"
    if t == "inst":
        return _base in ("FI", "IG", "I")
    if t == "grp_sex":
        return _sex and _base in ("G", "FG", "FIG", "IG", "FI")
    if t == "grp":
        return not _sex and _base in ("G", "FG", "FIG")
    return False

def node_passes_size(m: dict, mode: str, dyn_author_min: int, dyn_author_max: int) -> bool:
    t = m.get("type")
    # Nodes whose size is tracked by the slider - apply the filter
    if size_relevant_in_mode(m, mode):
        return dyn_author_min <= m.get("size", 0) <= dyn_author_max
    # All other node types (e.g. inst placeholders in FI mode, fac nodes
    # that aren't the primary level) pass through unconditionally
    return True

def node_passes_filters(nid: str, m: dict, mode: str, show_fac: bool, show_inst: bool,
                         show_grp: bool, selected_facs: list,
                         selected_insts: list, selected_grps: list) -> bool:
    fac = m.get("fac")
    inst = m.get("inst")
    grp = m.get("grp")
    t   = m.get("type")

    if mode == "IG":
        return True
    if show_fac and selected_facs and fac not in selected_facs:
        return False
    if show_inst and selected_insts and inst not in selected_insts:
        return False
    if t in ("grp", "grp_sex") and show_grp and selected_grps and grp not in selected_grps:
        return False
    if t == "inst" and show_grp and selected_grps and base_mode(mode) not in ("FI", "IG"):
        if grp and grp not in selected_grps:
            return False
    return True

def pre_nodes_for_mode(node_meta: dict, mode: str) -> set:
    _base = base_mode(mode)
    _sex  = sex_in_mode(mode)
    # N-modes tjekkes FØRST da de overskriver basis-opførslen for alle _base-værdier
    if nat_in_mode(mode):
        return {nid for nid, m in node_meta.items() if m.get("type") == "grp"}
    if _base == "FIG":
        if _sex:
            return {nid for nid, m in node_meta.items() if m.get("type") in ("fac", "grp_sex")}
        return {nid for nid, m in node_meta.items() if m.get("type") != "inst"}
    elif _base == "FI":
        if _sex:
            return {nid for nid, m in node_meta.items() if m.get("type") == "inst_sex"}
        return set(node_meta.keys())
    elif _base == "IG":
        if _sex:
            return {nid for nid, m in node_meta.items() if m.get("type") in ("inst", "grp_sex")}
        return {nid for nid, m in node_meta.items() if m.get("type") in ("grp", "inst")}
    elif _base == "FG":
        if _sex:
            return {nid for nid, m in node_meta.items() if m.get("type") in ("fac", "grp_sex")}
        return set(node_meta.keys())
    elif _sex and _base == "F":
        return {nid for nid, m in node_meta.items() if m.get("type") == "fac_sex"}
    elif _sex and _base == "I":
        return {nid for nid, m in node_meta.items() if m.get("type") == "inst_sex"}
    elif _sex and _base == "G":
        return {nid for nid, m in node_meta.items() if m.get("type") == "grp_sex"}
    return set(node_meta.keys())

def edge_type(u: str, v: str, node_meta: dict, mode: str) -> str:
    _base = base_mode(mode)
    if _base == "G":
        return "group"
    if _base in ("I", "IG"):
        gu = node_meta[u].get("inst", "")
        gv = node_meta[v].get("inst", "")
    else:
        gu = node_meta[u].get("fac", "")
        gv = node_meta[v].get("fac", "")
    return "intra" if (gu and gu == gv) else "inter"

def edge_type_grp(u: str, v: str, node_meta: dict) -> str:
    """Position-group-level intra/inter classification."""
    gu = node_meta[u].get("grp", "")
    gv = node_meta[v].get("grp", "")
    return "intra" if (gu and gu == gv) else "inter"

def edge_type_fac(u: str, v: str, node_meta: dict) -> str:
    """Faculty-level intra/inter classification (always faculty, regardless of mode)."""
    fu = node_meta[u].get("fac", "")
    fv = node_meta[v].get("fac", "")
    return "intra" if (fu and fu == fv) else "inter"

def edge_type_inst(u: str, v: str, node_meta: dict) -> str:
    """Institute-level intra/inter classification."""
    iu = node_meta[u].get("inst", "")
    iv = node_meta[v].get("inst", "")
    return "intra" if (iu and iu == iv) else "inter"

def author_passes_filters(fac_list, inst_list, grp, selected_facs, selected_insts, selected_grps) -> bool:
    if selected_facs and not any(f in selected_facs for f in fac_list):
        return False
    if selected_insts and not any(i in selected_insts for i in inst_list):
        return False
    if selected_grps and grp not in selected_grps:
        return False
    return True

# ---------------------------------------------------------------------------
# LAYOUT
# ---------------------------------------------------------------------------

def compute_layout(nodes_keep: set, node_meta: dict, mode: str, network_scale: int = 1200,
                   n_selected_nats: int = None) -> dict:

    def _r(n: int, min_dist: float = 250, floor: float = 1200) -> float:
        """Radius så n noder har mindst min_dist pixels mellem sig."""
        if n <= 1:
            return 0
        return max(min_dist * n / (2 * pi), floor)

    # Forstærk network_scale eksponentielt så slideren har større effekt
    network_scale = int(network_scale ** 1.5 / 35)

    R_G = 100

    pos = {}

    # --- Faculty centres ---
    fac_centers = {}
    faculties = sorted({m["fac"] for m in node_meta.values() if "fac" in m})
    faculties = ["HUM", "SCIENCE", "SAMF", "JUR", "SUND", "TEO"]
    k = max(1, len(faculties))

    if base_mode(mode) == "FIG":
        _insts_per_fac = {}
        for m in node_meta.values():
            if m.get("type") in ("grp", "grp_sex", "grp_nat") and m.get("fac"):
                _insts_per_fac.setdefault(m["fac"], set()).add(m.get("inst", ""))
        _max_insts = max((len(v) for v in _insts_per_fac.values()), default=1)
        _n_nats = max(1, len({m.get("statsborgerskab","") for m in node_meta.values() if m.get("statsborgerskab")})) if nat_in_mode(mode) else 1
        _R_NAT_est = max(50, _n_nats * network_scale // 30)
        if nat_in_mode(mode):
            R_FAC = _r(k, network_scale + _max_insts * (network_scale // 10 + _R_NAT_est * 3 // 2) + 100, floor=network_scale * 2)
        else:
            R_FAC = _r(k, network_scale + _max_insts * (network_scale // 20 + _R_NAT_est * 1) + 100, floor=network_scale * 3 // 2)

    elif base_mode(mode) == "IG":     
        _insts_per_fac = {}
        for m in node_meta.values():
            if m.get("fac") and m.get("inst"):
                _insts_per_fac.setdefault(m["fac"], set()).add(m["inst"])
        _max_insts = max((len(v) for v in _insts_per_fac.values()), default=1)
        _n_nats = max(1, len({m.get("statsborgerskab","") for m in node_meta.values() if m.get("statsborgerskab")})) if nat_in_mode(mode) else 1
        _R_NAT_est = max(50, _n_nats * network_scale // 30)
        R_FAC = _r(k, network_scale + _max_insts * (network_scale // 10 + _R_NAT_est * 1) + 100, floor=network_scale * 2)

    elif base_mode(mode) == "I":         
        _insts_per_fac = {}
        for m in node_meta.values():
            if m.get("fac") and m.get("inst"):
                _insts_per_fac.setdefault(m["fac"], set()).add(m["inst"])
        _max_insts = max((len(v) for v in _insts_per_fac.values()), default=1)
        _n_nats = max(1, len({m.get("statsborgerskab","") for m in node_meta.values() if m.get("statsborgerskab")})) if nat_in_mode(mode) else 1
        _R_NAT_est = max(50, _n_nats * network_scale // 30)
        R_FAC = _r(k, network_scale + _max_insts * (network_scale // 10 + _R_NAT_est * 1) + 100, floor=network_scale * 2)

    elif nat_in_mode(mode) and base_mode(mode) == "F":
        # NF-mode: fakulteter har nationalitets-clustre rundt om sig
        _n_nats = n_selected_nats if n_selected_nats is not None else max(1, len({m.get("statsborgerskab","") for m in node_meta.values() if m.get("statsborgerskab")}))
        _R_NAT_est = max(30, _n_nats * network_scale // 80)
        R_FAC = _r(k, network_scale + _R_NAT_est * 1, floor=network_scale + _R_NAT_est // 2)

    elif nat_in_mode(mode) and base_mode(mode) == "FI":
        # NFI-mode: hvert institut har en nationalitets-cluster
        _max_insts = 1
        _insts_per_fac = {}
        for m in node_meta.values():
            if m.get("fac") and m.get("inst"):
                _insts_per_fac.setdefault(m["fac"], set()).add(m["inst"])
        _max_insts = max((len(v) for v in _insts_per_fac.values()), default=1)
        _n_nats = max(1, len({m.get("statsborgerskab","") for m in node_meta.values() if m.get("statsborgerskab")}))
        _R_NAT_est = max(50, _n_nats * network_scale // 30)
        R_FAC = _r(k, network_scale + _max_insts * (network_scale // 20 + _R_NAT_est * 1 // 2) + 100, floor=network_scale * 3 // 2)

    else:
        R_FAC = _r(k, network_scale, floor=network_scale)

    for i, fac in enumerate(faculties):
        theta = 2 * pi * i / k
        fac_centers[fac] = (R_FAC * cos(theta), R_FAC * sin(theta))
  
    # --- Institute centres ---
    inst_centers = {}

    if base_mode(mode) in ("FI", "IS"):
        inst_by_fac = {}
        for m in node_meta.values():
            if m.get("type") in ("inst", "inst_sex", "inst_nat"):
                inst_by_fac.setdefault(m["fac"], []).append(m["inst"])
        _n_nats = max(1, len({m.get("statsborgerskab","") for m in node_meta.values() if m.get("statsborgerskab")})) if nat_in_mode(mode) else 1
        _R_NAT_est = max(50, _n_nats * network_scale // 30)
        for fac, insts in inst_by_fac.items():
            cx, cy = fac_centers.get(fac, (0, 0))
            unique_insts = sorted(set(insts))
            k = max(1, len(unique_insts))
            if nat_in_mode(mode):
                R_INST = _r(k, network_scale // 3 + _R_NAT_est * 2, floor=network_scale // 3 + _R_NAT_est * 2)
            else:
                R_INST = _r(k, 200, floor = network_scale // 3)
            for j, inst in enumerate(unique_insts):
                theta = 2 * pi * j / k
                inst_centers[(fac, inst)] = (cx + R_INST * cos(theta), cy + R_INST * sin(theta))

    elif base_mode(mode) == "I":
        inst_by_fac = {}
        for m in node_meta.values():
            if m.get("fac") and m.get("inst"):
                inst_by_fac.setdefault(m["fac"], set()).add(m["inst"])
        _n_nats = max(1, len({m.get("statsborgerskab","") for m in node_meta.values() if m.get("statsborgerskab")})) if nat_in_mode(mode) else 1
        _R_NAT_est = max(50, _n_nats * network_scale // 30)
        for fac, insts in inst_by_fac.items():
            cx, cy = fac_centers.get(fac, (0, 0))
            unique_insts = sorted(insts)
            k = max(1, len(unique_insts))
            R_INST = _r(k, network_scale // 3 + _R_NAT_est * 2, floor=network_scale // 3 + _R_NAT_est * 2)
            for j, inst in enumerate(unique_insts):
                theta = 2 * pi * j / k
                inst_centers[(fac, inst)] = (cx + R_INST * cos(theta), cy + R_INST * sin(theta))

    elif base_mode(mode) == "FIG":
        insts_by_fac = {}
        for nid, m in node_meta.items():
            if m.get("type") in ("grp", "grp_sex", "grp_nat"):
                insts_by_fac.setdefault(m["fac"], []).append(m["inst"])
        _n_nats = max(1, len({m.get("statsborgerskab","") for m in node_meta.values() if m.get("statsborgerskab")})) if nat_in_mode(mode) else 1
        _R_NAT_est = max(50, _n_nats * network_scale // 30)
        for fac, insts in insts_by_fac.items():
            cx, cy = fac_centers.get(fac, (0, 0))
            insts = sorted(set(insts))
            k = max(1, len(insts))
            if nat_in_mode(mode):
                R_INST = _r(k, network_scale // 2 + _R_NAT_est * 2, floor=network_scale // 2 + _R_NAT_est * 2)
            elif sex_in_mode(mode):
                R_INST = _r(k, network_scale // 2, floor=network_scale // 3)
            else:
                R_INST = _r(k, network_scale // 3, floor=network_scale // 4)
            for j, inst in enumerate(insts):
                theta = 2 * pi * j / k
                inst_centers[(fac, inst)] = (cx + R_INST * cos(theta), cy + R_INST * sin(theta))

    elif base_mode(mode) == "IG":
        inst_by_fac = {}
        for m in node_meta.values():
            if m.get("fac") and m.get("inst"):
                inst_by_fac.setdefault(m["fac"], set()).add(m["inst"])
        _n_nats = max(1, len({m.get("statsborgerskab","") for m in node_meta.values() if m.get("statsborgerskab")})) if nat_in_mode(mode) else 1
        _R_NAT_est = max(50, _n_nats * network_scale // 30)
        for fac, insts in inst_by_fac.items():
            cx, cy = fac_centers.get(fac, (0, 0))
            unique_insts = sorted(insts)
            k = max(1, len(unique_insts))
            R_INST = _r(k, network_scale // 2 + _R_NAT_est * 2, floor=network_scale // 2 + _R_NAT_est * 2)
            for j, inst in enumerate(unique_insts):
                theta = 2 * pi * j / k
                inst_centers[(fac, inst)] = (cx + R_INST * cos(theta), cy + R_INST * sin(theta))

    elif base_mode(mode) == "G" and not nat_in_mode(mode):
        if sex_in_mode(mode):
            # SG-mode: én ring, stillingsgrupper sorteret efter hierarki,
            # k til venstre og m til højre inden for hver gruppe
            unique_grps = sorted(
                {node_meta[nid2].get("grp", "") for nid2 in nodes_keep if node_meta[nid2].get("grp")},
                key=lambda g: HIERARKI.get(g, 999)
            )
            n_grps = max(1, len(unique_grps))
            R_G = _r(n_grps, network_scale // 4, floor=network_scale // 2)
            pair_offset = network_scale // 8
            for j, grp in enumerate(unique_grps):
                theta = 2 * pi * j / n_grps
                cx_grp = R_G * cos(theta)
                cy_grp = R_G * sin(theta)
                pair = sorted(
                    [nid2 for nid2 in nodes_keep if node_meta[nid2].get("grp") == grp],
                    key=lambda n: node_meta[n].get("sex", "m")  # k før m (k < m alfabetisk)
                )
                n_pair = len(pair)
                angle_gap = 0.38  # radianer mellem k og m - juster efter smag
                for p_idx, nid2 in enumerate(pair):
                    offset_angle = angle_gap * (p_idx - (n_pair - 1) / 2)
                    t = theta + offset_angle
                    pos[nid2] = (R_G * cos(t), R_G * sin(t))
        else:
            # G-mode: fakultetsklynger
            grps_by_fac = {}
            for nid2 in nodes_keep:
                fac = node_meta[nid2].get("fac", "")
                grps_by_fac.setdefault(fac, []).append(nid2)
            for fac, fac_nodes in grps_by_fac.items():
                cx, cy = fac_centers.get(fac, (0, 0))
                fac_nodes_sorted = sorted(fac_nodes, key=lambda n: HIERARKI.get(node_meta[n].get("grp", ""), 999))
                k = max(1, len(fac_nodes_sorted))
                R_G = _r(k, network_scale // 6, floor=network_scale // 5)
                for j, nid2 in enumerate(fac_nodes_sorted):
                    theta = 2 * pi * j / k
                    pos[nid2] = (cx + R_G * cos(theta), cy + R_G * sin(theta))
        return pos

    # --- Place every node ---
    for nid in nodes_keep:
        if nid in pos:
            continue
        m = node_meta[nid]
        t = m.get("type")

        if t == "fac_sex":
            cx, cy = fac_centers.get(m["fac"], (0, 0))
            R_SEX = network_scale // 3
            pos[nid] = (cx + (R_SEX if m.get("sex") == "m" else -R_SEX), cy)

        elif t == "fac_nat":
            fac = m.get("fac", "")
            nat = m.get("statsborgerskab", "")
            cx, cy = fac_centers.get(fac, (0, 0))
            nats_for_fac = sorted({
                _m.get("statsborgerskab", "")
                for _m in node_meta.values()
                if _m.get("fac") == fac and _m.get("statsborgerskab")
            })
            k   = max(1, len(nats_for_fac))
            j   = nats_for_fac.index(nat) if nat in nats_for_fac else 0
            R_NAT = _r(k, network_scale // 10, floor=max(80, k * network_scale // 60))
            theta = 2 * pi * j / k
            pos[nid] = (cx + R_NAT * cos(theta), cy + R_NAT * sin(theta))
        
        elif t == "inst_nat":
            fac  = m.get("fac", "")
            inst = m.get("inst", "")
            nat  = m.get("statsborgerskab", "")
            cx, cy = inst_centers.get((fac, inst), fac_centers.get(fac, (0, 0)))  # ← fac som fallback
            nats_for_inst = sorted({
                _m.get("statsborgerskab", "")
                for _m in node_meta.values()
                if _m.get("inst") == inst and _m.get("statsborgerskab")
            })
            k     = max(1, len(nats_for_inst))
            j     = nats_for_inst.index(nat) if nat in nats_for_inst else 0
            R_NAT = _r(k, network_scale // 10, floor=max(80, k * network_scale // 60))
            theta = 2 * pi * j / k
            pos[nid] = (cx + R_NAT * cos(theta), cy + R_NAT * sin(theta))

        elif t == "grp_nat":
            fac  = m.get("fac", "")
            inst = m.get("inst", "")
            nat  = m.get("statsborgerskab", "")
            grp  = m.get("grp", "")
            _bm  = base_mode(mode)
            
            # Vælg cluster-center afhængig af mode
            if _bm in ("FIG", "IG", "FI", "I"):
                # Brug institut-center når institut er aktivt niveau
                cx, cy = inst_centers.get((fac, inst), fac_centers.get(fac, (0, 0)))
                # Stillingsgrupper for dette institut
                grps_for_unit = sorted(
                    {_m.get("grp","") for _m in node_meta.values() 
                    if _m.get("grp") and _m.get("fac") == fac and _m.get("inst") == inst},
                    key=lambda g: HIERARKI.get(g, 999)
                )
            elif _bm in ("FG", "F"):
                # Brug fakultets-center når kun fakultet er aktivt
                cx, cy = fac_centers.get(fac, (0, 0))
                grps_for_unit = sorted(
                    {_m.get("grp","") for _m in node_meta.values() 
                    if _m.get("grp") and _m.get("fac") == fac},
                    key=lambda g: HIERARKI.get(g, 999)
                )
            else:
                # NG-mode: ring af alle stillingsgrupper på KU-niveau (uændret)
                grps_for_unit = sorted({_m.get("grp","") for _m in node_meta.values() if _m.get("grp")}, 
                                    key=lambda g: HIERARKI.get(g, 999))
                k_grp = max(1, len(grps_for_unit))
                R_GRP = _r(k_grp, network_scale // 2, floor=network_scale)
                g_idx = grps_for_unit.index(grp) if grp in grps_for_unit else 0
                theta_grp = 2 * pi * g_idx / k_grp
                cx = R_GRP * cos(theta_grp)
                cy = R_GRP * sin(theta_grp)
            
            # For institut/fakultet-modes: placér stillingsgruppe-cluster rundt om enheds-center
            if _bm in ("FIG", "IG", "FI", "I", "FG", "F"):
                n_grps = max(1, len(grps_for_unit))
                if _bm == "FIG":
                    R_GRP_local = _r(n_grps, network_scale // 4, floor=network_scale // 3)
                else:
                    R_GRP_local = _r(n_grps, network_scale // 6, floor=network_scale // 5)
                g_idx = grps_for_unit.index(grp) if grp in grps_for_unit else 0
                theta_grp = 2 * pi * g_idx / n_grps
                cx_grp = cx + R_GRP_local * cos(theta_grp)
                cy_grp = cy + R_GRP_local * sin(theta_grp)
            else:
                # NG-mode: cx/cy er allerede stillingsgruppe-positionen
                cx_grp = cx
                cy_grp = cy
            
            # Nationaliteter for denne stillingsgruppe i den specifikke enhed
            nats_for_grp = sorted({
                _m.get("statsborgerskab","")
                for _m in node_meta.values()
                if _m.get("grp") == grp and _m.get("statsborgerskab")
                and (not fac or _m.get("fac") == fac)
                and (not inst or _bm not in ("FIG", "IG", "FI", "I") or _m.get("inst") == inst)
            })
            k_nat     = max(1, len(nats_for_grp))
            R_NAT     = max(50, k_nat * network_scale // 40)
            n_idx     = nats_for_grp.index(nat) if nat in nats_for_grp else 0
            theta_nat = 2 * pi * n_idx / k_nat
            pos[nid]  = (cx_grp + R_NAT * cos(theta_nat), cy_grp + R_NAT * sin(theta_nat))

        elif t == "inst_sex":
            fac = m.get("fac")
            sex = m.get("sex", "m")
            inst = m.get("inst")
            cx, cy = fac_centers.get(fac, (0, 0))

            # Alle institutter for dette fakultet, sorteret
            fac_insts = sorted({
                node_meta[nid2].get("inst")
                for nid2 in nodes_keep
                if node_meta[nid2].get("type") == "inst_sex"
                and node_meta[nid2].get("fac") == fac
                and node_meta[nid2].get("inst")
            })
            n_insts = max(1, len(fac_insts))
            inst_idx = fac_insts.index(inst) if inst in fac_insts else 0

            # Køn bestemmer halvkreds: k = venstre (pi/2 → 3pi/2), m = højre (3pi/2 → 5pi/2)
            if n_insts == 1:
                # Kun ét institut: placer k og m fast til venstre/højre
                R_INST = network_scale // 3
                pos[nid] = (cx + (R_INST if sex == "m" else -R_INST), cy)
            else:
                if sex == "k":
                    theta = pi / 2 + inst_idx * pi / n_insts
                else:
                    theta = 3 * pi / 2 + inst_idx * pi / n_insts
                R_INST = _r(n_insts, network_scale // 3, floor=network_scale // 3)
                pos[nid] = (cx + R_INST * cos(theta), cy + R_INST * sin(theta))

        elif t == "fac":
            pos[nid] = fac_centers.get(m["fac"], (0, 0))

        elif t == "inst":
            center = inst_centers.get((m["fac"], m["inst"]))
            if center is None:
                continue
            pos[nid] = center

        elif t in ("grp", "grp_sex"):
            fac  = m.get("fac")
            inst = m.get("inst")
            sex  = m.get("sex", "")
            _bm  = base_mode(mode)

            # SF/SFI-modes uden stillingsgrupper: placér grp_sex som "institut-stand-in"
            # k til venstre, m til højre omkring fakultets-center
            if sex_in_mode(mode) and sex and "F" in _bm and "G" not in _bm:
                cx, cy = fac_centers.get(fac, (0, 0))
                fac_insts = sorted({
                    node_meta[nid2].get("inst", "")
                    for nid2 in nodes_keep
                    if node_meta[nid2].get("fac") == fac
                    and node_meta[nid2].get("inst")
                })
                n_insts = max(1, len(fac_insts))
                inst_idx = fac_insts.index(inst) if inst in fac_insts else 0
                R_INST = _r(n_insts, network_scale // 3, floor=network_scale // 3)
                if n_insts == 1:
                    pos[nid] = (cx + (R_INST if sex == "m" else -R_INST), cy)
                else:
                    if sex == "k":
                        theta = pi / 2 + inst_idx * pi / n_insts
                    else:
                        theta = 3 * pi / 2 + inst_idx * pi / n_insts
                    pos[nid] = (cx + R_INST * cos(theta), cy + R_INST * sin(theta))

            elif _bm == "FIG":
                # SFIG: institut-center, cluster består af alle grp/grp_sex for det institut
                center = inst_centers.get((fac, inst))
                if center is None:
                    continue
                cx, cy = center
                cluster = sorted(
                    nid2 for nid2 in nodes_keep
                    if node_meta[nid2].get("type") in ("grp", "grp_sex")
                    and node_meta[nid2].get("fac") == fac
                    and node_meta[nid2].get("inst") == inst
                )
            elif _bm == "FG":
                # SFG: fakultets-center, cluster består af alle grp/grp_sex for det fakultet
                cx, cy = fac_centers.get(fac, (0, 0))
                cluster = sorted(
                    nid2 for nid2 in nodes_keep
                    if node_meta[nid2].get("type") in ("grp", "grp_sex")
                    and node_meta[nid2].get("fac") == fac
                )
            elif _bm in ("FI", "IG", "I"):
                center = inst_centers.get((fac, inst))
                if center is None:
                    continue
                cx, cy = center
                cluster = sorted(
                    nid2 for nid2 in nodes_keep
                    if node_meta[nid2].get("type") in ("grp", "grp_sex")
                    and node_meta[nid2].get("fac") == fac
                    and node_meta[nid2].get("inst") == inst
                )
            else:
                cx, cy = fac_centers.get(fac, (0, 0))
                cluster = [nid]

            if nid in pos:
                continue

            grp_pairs = {}
            for nid2 in cluster:
                g = node_meta[nid2].get("grp", nid2)
                grp_pairs.setdefault(g, []).append(nid2)
            unique_grps = sorted(grp_pairs.keys(), key=lambda g: HIERARKI.get(g, 999))
            n_grps = max(1, len(unique_grps))
            if base_mode(mode) == "FIG":
                R_GRP      = _r(n_grps, network_scale // 10, floor=network_scale // 8)
                pair_offset = network_scale // 4
            else:
                R_GRP      = _r(n_grps, network_scale // 6, floor=network_scale // 5)
                pair_offset = network_scale // 2

            grp_idx = unique_grps.index(node_meta[nid].get("grp", nid)) if node_meta[nid].get("grp") in unique_grps else 0
            theta = 2 * pi * grp_idx / n_grps
            cx_grp = cx + R_GRP * cos(theta)
            cy_grp = cy + R_GRP * sin(theta)

            # Placer k og m ved siden af hinanden
            pair = grp_pairs.get(node_meta[nid].get("grp", nid), [nid])
            p_idx = pair.index(nid) if nid in pair else 0
            n_pair = len(pair)
            offset_x = pair_offset * (p_idx - (n_pair - 1) / 2)
            pos[nid] = (cx_grp + offset_x, cy_grp)

    return pos

# ---------------------------------------------------------------------------
# CENTRALITY HELPERS
# ---------------------------------------------------------------------------

def aggregate_centrality_by(meta_key: str, nodes_list: list, node_meta: dict,
                             weighted_deg: dict, bet_cent: dict):
    acc_wd, acc_bc = {}, {}
    for n in nodes_list:
        meta = node_meta.get(n, {})
        key = meta.get(meta_key, "")
        if not key:
            continue
        acc_wd[key] = acc_wd.get(key, 0.0) + float(weighted_deg.get(n, 0.0))
        acc_bc[key] = acc_bc.get(key, 0.0) + float(bet_cent.get(n, 0.0))
    return (
        sorted(acc_wd.items(), key=lambda x: -x[1]),
        sorted(acc_bc.items(), key=lambda x: -x[1]),
    )


def build_grp_table_by_mode(weighted_deg, bet_cent, node_meta: dict, mode: str):
    wd_map = {grp: float(val) for grp, val in weighted_deg} if weighted_deg else {}
    bc_map = {grp: float(val) for grp, val in bet_cent}     if bet_cent     else {}

    rows = []
    base_fields = None

    for nid, meta in node_meta.items():
        if meta.get("type") not in ("grp", "grp_sex"):
            continue
        grp  = meta.get("grp", "")
        fac  = meta.get("fac", "")
        inst = meta.get("inst", "")
        wd   = wd_map.get(grp, 0.0)
        bc   = bc_map.get(grp, 0.0)

        if base_mode(mode) in ("FIG", "G"):
            row = {"Stillingsgruppe": grp, "Fakultet": fac, "Institut": inst,
                   "Samlet samarbejdsomfang": wd, "Brobyggerrolle": bc, "mode": mode}
            base_fields = ["Stillingsgruppe", "Fakultet", "Institut",
                           "Samlet samarbejdsomfang", "Brobyggerrolle", "mode"]
        elif mode == "IG":
            row = {"Stillingsgruppe": grp, "Institut": inst,
                   "Samlet samarbejdsomfang": wd, "Brobyggerrolle": bc, "mode": mode}
            base_fields = ["Stillingsgruppe", "Institut",
                           "Samlet samarbejdsomfang", "Brobyggerrolle", "mode"]
        elif mode == "FG":
            row = {"Stillingsgruppe": grp, "Fakultet": fac,
                   "Samlet samarbejdsomfang": wd, "Brobyggerrolle": bc, "mode": mode}
            base_fields = ["Stillingsgruppe", "Fakultet",
                           "Samlet samarbejdsomfang", "Brobyggerrolle", "mode"]
        elif base_mode(mode) == "G":
            row = {"Stillingsgruppe": grp,
                   "Samlet samarbejdsomfang": wd, "Brobyggerrolle": bc, "mode": mode}
            base_fields = ["Stillingsgruppe", "Samlet samarbejdsomfang", "Brobyggerrolle", "mode"]
        else:
            continue

        rows.append(row)

    if base_fields is None:
        return [], []

    type_map = {
        "Stillingsgruppe": pa.string(), "Fakultet": pa.string(),
        "Institut": pa.string(), "Samlet samarbejdsomfang": pa.float64(),
        "Brobyggerrolle": pa.float64(), "mode": pa.string(),
    }
    schema = [(key, type_map[key]) for key in base_fields]
    return rows, schema


# ===========================================================================
# MAIN APP
# ===========================================================================

def intra_inter_labels(mode: str) -> tuple:
    """Return (intra_label, inter_label) suited to the current mode's grouping level."""
    if base_mode(mode) in ("I", "IG"):
        return "intra-institut", "inter-institut"
    if base_mode(mode) == "G":
        return "intra-gruppe", "inter-gruppe"
    return "intra-fakultet", "inter-fakultet"

def filter_status_caption(mode: str, show_intra: bool, show_inter: bool,
                          show_intra_inst: bool, show_inter_inst: bool,
                          show_intra_grp: bool, show_inter_grp: bool,
                          show_fac: bool, show_inst: bool, show_grp: bool) -> str:
    parts = []

    _intra_lbl, _inter_lbl = intra_inter_labels(mode)

    # Hovedniveau
    if not show_intra and not show_inter:
        return "**Ingen par-typer er valgt** - juster filtrene i sidebaren under *Organisation*."
    elif not show_intra:
        parts.append(f"{_intra_lbl} par")
    elif not show_inter:
        parts.append(f"{_inter_lbl} par")

    # Institut-niveau
    if show_fac and show_inst:
        if not show_intra_inst and not show_inter_inst:
            parts.append("nogen institut-par")
        elif not show_intra_inst:
            parts.append("intra-institut par")
        elif not show_inter_inst:
            parts.append("inter-institut par")

    # Stillingsgruppe-niveau
    if show_grp:
        if not show_intra_grp and not show_inter_grp:
            parts.append("nogen stillingsgruppe-par")
        elif not show_intra_grp:
            parts.append("intra-stillingsgruppe par")
        elif not show_inter_grp:
            parts.append("inter-stillingsgruppe par")

    if not parts:
        return ""
    return f"Figuren inkluderer ikke {', '.join(parts)}."

def _compute_mod_pre(G2, nodes_keep, node_meta, comm_key):
    try:
        _connected = {n for n in G2.nodes() if G2.degree(n) > 0}
        G_conn = G2.subgraph(_connected).copy()
        comms = {}
        for nid in nodes_keep:
            gl = node_meta.get(nid, {}).get(comm_key, "")
            if gl:
                comms.setdefault(gl, []).append(nid)
        filtered = [[n for n in c if n in _connected] for c in comms.values()]
        filtered = [c for c in filtered if c]
        # Tilføj uncovered connected noder som singletons - modq kræver komplet partition
        _covered = {n for c in filtered for n in c}
        filtered += [[n] for n in _connected if n not in _covered]
        if len(filtered) < 2 or G_conn.number_of_edges() == 0:
            return float("nan")
        val = modq(G_conn, filtered, weight="weight")
        return val
    
    except Exception:
        return float("nan")

def compute_year_snapshot(content: dict, mode: str,
                           year: int, 
                           raw_nodes_pre: dict,
                           selected_genders: list,
                           selected_gender_edges: list,
                           selected_citizenships: list,
                           selected_facs: list,
                           selected_insts: list,
                           selected_grps: list,
                           show_fac: bool, show_inst: bool, show_grp: bool,
                           sampub_count_raw: int = 0,
                           ku_totals: dict = None,
                           collab_pairs: dict = None,
                           show_intra: bool = True,
                           show_inter: bool = True) -> dict:
    """Recompute aggregated stats for one year, same filters as main view."""
    raw_nodes = {}
    for nid, meta in content["nodes"]:
        m = dict(meta)
        m.setdefault("type", "grp")
        parts = nid.split("|")
        if "fac" not in m and len(parts) >= 1:
            m["fac"] = parts[0]
        if "inst" not in m and len(parts) >= 2: 
            m["inst"] = parts[1]
        if "grp" not in m and len(parts) >= 3: 
            m["grp"] = parts[2]
        if "sex" not in m and len(parts) >= 4: 
            m["sex"] = parts[3]
        if "statsborgerskab" not in m and len(parts) >= 5:
            m["statsborgerskab"] = parts[4]
        m.setdefault("size", 0)
        raw_nodes[nid] = m

    raw_nodes_pre_cs = dict(raw_nodes)

    if selected_citizenships:
        raw_nodes = {
            nid: m for nid, m in raw_nodes.items()
            if m.get("type") != "grp"
            or m.get("statsborgerskab", "") in selected_citizenships
        }

    for nid, m in list(raw_nodes.items()):
        fac  = m.get("fac", "")
        inst = m.get("inst", "")
        if not fac or not inst:
            continue
        inst_id = f"INST:{fac}|{inst}"
        if inst_id not in raw_nodes:
            raw_nodes[inst_id] = {"type": "inst", "fac": fac, "inst": inst, "grp": "", "size": 0}

    raw_edges = [
        (e[0], e[1], e[2], e[3] if len(e) > 3 else None)
        for e in content["edges"]
        if e[0] in raw_nodes and e[1] in raw_nodes
    ]

    node_meta, edge_source = apply_mode_merge(mode, raw_nodes, raw_edges)

    node_meta = {
        nid: m for nid, m in node_meta.items()
        if passes_category_filters(m, mode, show_fac, show_inst, show_grp,
                                   selected_facs, selected_insts, selected_grps)
    }
    if selected_genders and sex_in_mode(mode):
        node_meta = {nid: m for nid, m in node_meta.items() if m.get("sex") in selected_genders}

    edge_source = [e for e in edge_source if e[0] in node_meta and e[1] in node_meta]
    nodes_keep  = pre_nodes_for_mode(node_meta, mode)
    total_authors = sum(node_meta[nid].get("size", 0) for nid in nodes_keep)

    edges_keep = []
    edges_keep = []
    for edge in edge_source:
        u, v, w = edge[0], edge[1], edge[2]
        sex_combo = edge[3] if len(edge) > 3 else None
        if u not in nodes_keep or v not in nodes_keep:
            continue
        if selected_gender_edges and sex_combo and sex_combo not in selected_gender_edges:
            continue
        et = edge_type(u, v, node_meta, mode)
        if et == "intra" and not show_intra:
            continue
        if et == "inter" and not show_inter:
            continue
        edges_keep.append((u, v, w, sex_combo))

    if not sex_in_mode(mode):
        connected  = {n for u, v, *_ in edges_keep for n in (u, v)}
        nodes_keep = nodes_keep & connected
    #isolated_nodes = all_nodes_pre_isolation - nodes_keep
    total_authors = sum(node_meta[nid].get("size", 0) for nid in nodes_keep)

    total_pubs = sum(w for _, _, w, *_ in edges_keep)
    intra_pubs = sum(w for u, v, w, *_ in edges_keep
                     if base_mode(mode) != "G" and edge_type(u, v, node_meta, mode) == "intra")
    inter_pubs = sum(w for u, v, w, *_ in edges_keep
                     if base_mode(mode) != "G" and edge_type(u, v, node_meta, mode) == "inter")
    intra_inst_pubs = sum(w for u, v, w, *_ in edges_keep
                          if edge_type_inst(u, v, node_meta) == "intra")
    inter_inst_pubs = sum(w for u, v, w, *_ in edges_keep
                          if edge_type_inst(u, v, node_meta) == "inter")
    intra_grp_pubs = sum(w for u, v, w, *_ in edges_keep
                         if edge_type_grp(u, v, node_meta) == "intra")
    inter_grp_pubs = sum(w for u, v, w, *_ in edges_keep
                         if edge_type_grp(u, v, node_meta) == "inter")
    
    # Sande intra/inter inkl. intra-node par fra collab_pairs
    _cp = collab_pairs or {}
    # Filtrér collab_pairs med samme filtre som edges_keep
    _fac_intra_ew_all: dict[str, float] = {}
    _fac_inter_ew_all: dict[str, float] = {}
    _inst_intra_ew_all: dict[str, float] = {}
    _inst_inter_ew_all: dict[str, float] = {}
    _grp_intra_ew_all: dict[str, float] = {}
    _grp_inter_ew_all: dict[str, float] = {}
    _intra_pubs_all = 0
    _inter_pubs_all = 0
    _intra_inst_pubs_all = 0
    _inter_inst_pubs_all = 0
    _intra_grp_pubs_all = 0
    _inter_grp_pubs_all = 0

    for raw_key, counts in _cp.items():
        parts = raw_key.split("|")
        if len(parts) != 3:
            continue
        fac, inst, grp = parts
        if selected_facs  and fac  not in selected_facs:  continue
        if selected_insts and inst not in selected_insts: continue
        if selected_grps  and grp  not in selected_grps:  continue

        fi = counts.get("fac_intra",  0) / 2
        fx = counts.get("fac_inter",  0) / 2
        ii = counts.get("inst_intra", 0) / 2
        ix = counts.get("inst_inter", 0) / 2
        gi = counts.get("grp_intra",  0) / 2
        gx = counts.get("grp_inter",  0) / 2

        _fac_intra_ew_all[fac]  = _fac_intra_ew_all.get(fac,  0.0) + fi
        _fac_inter_ew_all[fac]  = _fac_inter_ew_all.get(fac,  0.0) + fx
        _inst_intra_ew_all[inst] = _inst_intra_ew_all.get(inst, 0.0) + ii
        _inst_inter_ew_all[inst] = _inst_inter_ew_all.get(inst, 0.0) + ix
        _grp_intra_ew_all[grp]  = _grp_intra_ew_all.get(grp,  0.0) + gi
        _grp_inter_ew_all[grp]  = _grp_inter_ew_all.get(grp,  0.0) + gx

        _intra_pubs_all      += fi
        _inter_pubs_all      += fx
        _intra_inst_pubs_all += ii
        _inter_inst_pubs_all += ix
        _intra_grp_pubs_all  += gi
        _inter_grp_pubs_all  += gx

    fac_tot, inst_tot, grp_tot = {}, {}, {}
    fac_grp_tot:  dict[str, dict[str, int]] = {}
    inst_grp_tot: dict[str, dict[str, int]] = {}
    for nid in pre_nodes_for_mode(node_meta, mode):
        m    = node_meta.get(nid, {})
        size = m.get("size", 0)
        f, i, g = m.get("fac"), m.get("inst"), m.get("grp")
        if f:   fac_tot[f]   = fac_tot.get(f,   0) + size
        if i:   inst_tot[i]  = inst_tot.get(i,  0) + size
        if g:   grp_tot[g]   = grp_tot.get(g,   0) + size
        if f and g:
            fac_grp_tot.setdefault(f, {})[g] = fac_grp_tot.setdefault(f, {}).get(g, 0) + size
        if i and g:
            inst_grp_tot.setdefault(i, {})[g] = inst_grp_tot.setdefault(i, {}).get(g, 0) + size

    fac_grp_ew:  dict[str, dict[str, float]] = {}
    inst_grp_ew: dict[str, dict[str, float]] = {}
    for u, v, w, *_ in edges_keep:
        for n in (u, v):
            m = node_meta.get(n, {})
            f, i, g = m.get("fac"), m.get("inst"), m.get("grp")
            if f and g:
                fac_grp_ew.setdefault(f, {})[g] = fac_grp_ew.setdefault(f, {}).get(g, 0.0) + w
            if i and g:
                inst_grp_ew.setdefault(i, {})[g] = inst_grp_ew.setdefault(i, {}).get(g, 0.0) + w

    G2 = nx.Graph()
    for u in pre_nodes_for_mode(node_meta, mode): G2.add_node(u)
    for u, v, w, *_ in edges_keep: G2.add_edge(u, v, weight=w)
    density = nx.density(G2)

    if mode in ("F", "I", "G"):
        modularity_pre_snap    = float("nan")
        modularity_greedy_snap = float("nan")
    else:
        #comm_key_snap = "fac" if mode in ("F","FS","FI","FG","FIG") else ("inst" if mode in ("I","IG","IS") else "grp")
        comm_key_snap = (
            "fac"  if base_mode(mode) in ("F", "FI", "FG", "FIG") else
            "inst" if base_mode(mode) in ("I", "IG") else
            "grp"
        )
        
        if comm_key_snap is None:
            modularity_pre_snap = float("nan")
            modularity_greedy_snap = float("nan")
        else:
            communities_snap, seen_snap = [], set()
            for nid in nodes_keep:
                gl = node_meta.get(nid, {}).get(comm_key_snap, "")
                if not gl or gl in seen_snap: continue
                members = [n for n in nodes_keep if node_meta.get(n, {}).get(comm_key_snap) == gl]
                if members:
                    communities_snap.append(members)
                    seen_snap.add(gl)
            try:
                _connected_snap = {n for n in G2.nodes() if G2.degree(n) > 0}
                _G2_conn_snap = G2.subgraph(_connected_snap).copy()
                _communities_snap_filtered = [
                    [n for n in c if n in _connected_snap] for c in communities_snap
                ]
                _communities_snap_filtered = [c for c in _communities_snap_filtered if c]
                if _communities_snap_filtered and _G2_conn_snap.number_of_edges() > 0 and len(_communities_snap_filtered) >= 2:
                    _mod_val = modq(_G2_conn_snap, _communities_snap_filtered, weight="weight")
                    modularity_pre_snap = _mod_val if _mod_val >= 0 else float("nan")
                else:
                    modularity_pre_snap = float("nan")
            except Exception:
                modularity_pre_snap = float("nan")
            try:
                _G2_conn_g = G2.subgraph({n for n in G2.nodes() if G2.degree(n) > 0}).copy()
                if _G2_conn_g.number_of_edges() > 0 and _G2_conn_g.number_of_nodes() > 1:
                    _gc_full          = list(greedy_modularity_communities(_G2_conn_g, weight="weight"))
                    _mod_gr_snap      = modq(_G2_conn_g, _gc_full, weight="weight")
                    modularity_greedy_snap = _mod_gr_snap if _mod_gr_snap >= 0 else float("nan")
                else:
                    modularity_greedy_snap = float("nan")
            except Exception:
                modularity_greedy_snap = float("nan")

    fac_ew, inst_ew, grp_ew = {}, {}, {}
    for u, v, w, *_ in edges_keep:
        for n in (u, v):
            m = node_meta.get(n, {})
            if m.get("fac"):  fac_ew[m["fac"]]  = fac_ew.get(m["fac"],  0.0) + w
            if m.get("inst"): inst_ew[m["inst"]] = inst_ew.get(m["inst"], 0.0) + w
            if m.get("grp"):  grp_ew[m["grp"]]  = grp_ew.get(m["grp"],  0.0) + w
    
    # Betweenness centralitet aggregeret per org-enhed
    try:
        if G2.number_of_edges() > 0 and G2.number_of_nodes() > 1:
            _bet = nx.betweenness_centrality(G2, weight="weight", normalized=True)
        else:
            _bet = {}
    except Exception:
        _bet = {}

    fac_bs, inst_bs, grp_bs = {}, {}, {}
    for nid, bval in _bet.items():
        m = node_meta.get(nid, {})
        if m.get("fac"):  fac_bs[m["fac"]]  = fac_bs.get(m["fac"],  0.0) + bval
        if m.get("inst"): inst_bs[m["inst"]] = inst_bs.get(m["inst"], 0.0) + bval
        if m.get("grp"):  grp_bs[m["grp"]]  = grp_bs.get(m["grp"],  0.0) + bval

    # Sex breakdowns
    sex_bidrag: dict[str, int] = {}
    for nid in nodes_keep:
        m = node_meta.get(nid, {})
        sex = m.get("sex", "")
        if sex:
            sex_bidrag[sex] = sex_bidrag.get(sex, 0) + m.get("size", 0)
    
    # Nationality breakdown
    nat_bidrag: dict[str, int] = {}
    nat_ew: dict[str, float] = {}
    for nid, m in raw_nodes.items():
        if m.get("type") != "grp":
            continue
        cs = m.get("statsborgerskab", "")
        if cs:
            nat_bidrag[cs] = nat_bidrag.get(cs, 0) + m.get("size", 0)
    for u, v, w, *_ in raw_edges:
        cs_u = u.split("|")[4] if len(u.split("|")) > 4 else ""
        cs_v = v.split("|")[4] if len(v.split("|")) > 4 else ""
        if not cs_u or not cs_v:
            continue
        is_dk_u = cs_u == "DK"
        is_dk_v = cs_v == "DK"
        if is_dk_u and is_dk_v:
            nat_ew["dk_dk"] = nat_ew.get("dk_dk", 0.0) + w
        elif not is_dk_u and not is_dk_v:
            nat_ew["int_int"] = nat_ew.get("int_int", 0.0) + w
        else:
            nat_ew["dk_int"] = nat_ew.get("dk_int", 0.0) + w
    
    nat_ew_all: dict[str, float] = {"dk_dk": 0.0, "dk_int": 0.0, "int_int": 0.0}
    for raw_key, counts in content.get("cs_pairs", {}).items():
        parts = raw_key.split("|")
        if len(parts) != 4: continue
        fac, inst, grp, focal_cs = parts
        if selected_facs  and fac  not in selected_facs:  continue
        if selected_insts and inst not in selected_insts: continue
        if selected_grps  and grp  not in selected_grps:  continue
        focal_dk = focal_cs == "DK"
        same  = counts.get("same",  0) / 2
        cross = counts.get("cross", 0) / 2
        if focal_dk:
            nat_ew_all["dk_dk"]  += same
            nat_ew_all["dk_int"] += cross
        else:
            nat_ew_all["int_int"] += same
            nat_ew_all["dk_int"]  += cross

    combo_pubs: dict[str, int] = {}
    for u, v, w, sex_combo in edges_keep:
        if sex_combo:
            combo_pubs[sex_combo] = combo_pubs.get(sex_combo, 0) + int(round(w))
    
    # Homofili per køn - KU-niveau fra sex_pairs
    _sp_year = content.get("sex_pairs", {})
    _hom_ku_year = {"k": {"same": 0, "cross": 0}, "m": {"same": 0, "cross": 0}}
    for raw_key, counts in _sp_year.items():
        parts = raw_key.split("|")
        if len(parts) != 4:
            continue
        focal_sex = parts[3]
        if focal_sex in _hom_ku_year:
            _hom_ku_year[focal_sex]["same"]  += counts.get("same", 0)
            _hom_ku_year[focal_sex]["cross"] += counts.get("cross", 0)

    def _hom_index(sp_data, sex_bidrag_dict):
        result = {}
        total_bidrag = sum(sex_bidrag_dict.values()) or 1
        for sex in ("k", "m"):
            baseline = sex_bidrag_dict.get(sex, 0) / total_bidrag
            same  = sp_data[sex]["same"]
            cross = sp_data[sex]["cross"]
            pairs = same + cross or 1
            rate  = same / pairs
            result[sex] = round(rate / baseline, 4) if baseline > 0 else None
        return result

    hom_index_ku = _hom_index(_hom_ku_year, sex_bidrag)

    # Homofili per nationalitet
    _cs_pairs_year = content.get("cs_pairs", {})
    _ku_cs_sp_y = {"DK": {"same": 0, "cross": 0}, "intl": {"same": 0, "cross": 0}}
    for raw_key, counts in _cs_pairs_year.items():
        parts = raw_key.split("|")
        if len(parts) != 4:
            continue
        focal_cs_dk = "DK" if parts[3] == "DK" else "intl"
        _ku_cs_sp_y[focal_cs_dk]["same"]  += counts.get("same", 0)
        _ku_cs_sp_y[focal_cs_dk]["cross"] += counts.get("cross", 0)

    _nat_bidrag_full: dict[str, int] = {}
    for _nid, _m in raw_nodes_pre_cs.items():
        if _m.get("type") != "grp":
            continue
        _cs = _m.get("statsborgerskab", "")
        if _cs:
            _nat_bidrag_full[_cs] = _nat_bidrag_full.get(_cs, 0) + _m.get("size", 0)
    _nat_tot_y = sum(_nat_bidrag_full.values()) or 1
    _baseline_dk_y   = _nat_bidrag_full.get("DK", 0) / _nat_tot_y
    _baseline_intl_y = 1 - _baseline_dk_y
    cs_hom_index_ku = {}
    for cs_grp, label, baseline in [("DK", "DK", _baseline_dk_y), ("intl", "International", _baseline_intl_y)]:
        d = _ku_cs_sp_y[cs_grp]
        pairs = d["same"] + d["cross"] or 1
        rate  = d["same"] / pairs
        cs_hom_index_ku[label] = round(rate / baseline, 4) if baseline > 0 else None

    # Alle forfatterpar inkl. intra-node (fra sex_pairs)
    combo_pubs_all: dict[str, int] = {}
    for raw_key, counts in _sp_year.items():
        parts = raw_key.split("|")
        if len(parts) != 4:
            continue
        focal_sex = parts[3]
        same  = counts.get("same", 0)
        cross = counts.get("cross", 0)
        # same-par: focal er k -> k-k, focal er m -> m-m
        same_combo = f"{focal_sex}-{focal_sex}"
        cross_combo = "k-m"
        combo_pubs_all[same_combo] = combo_pubs_all.get(same_combo, 0) + same
        combo_pubs_all[cross_combo] = combo_pubs_all.get(cross_combo, 0) + cross
    # Halvér da hvert par er talt to gange (én gang per endpoint)
    combo_pubs_all = {k: v // 2 for k, v in combo_pubs_all.items()}

    # Top co-publication pairs
    top_pairs: dict[str, float] = {}
    for u, v, w, *_ in edges_keep:
        mu = node_meta.get(u, {})
        mv = node_meta.get(v, {})
        label_u = " | ".join(p for p in (mu.get("fac",""), mu.get("inst",""), mu.get("grp","")) if p)
        label_v = " | ".join(p for p in (mv.get("fac",""), mv.get("inst",""), mv.get("grp","")) if p)
        pair_key = " ↔ ".join(sorted([label_u, label_v]))
        top_pairs[pair_key] = top_pairs.get(pair_key, 0.0) + w

    # Isolated units (nodes with no edges)
    all_unit_nodes = pre_nodes_for_mode(node_meta, mode)
    connected_nodes = {n for u, v, *_ in edges_keep for n in (u, v)}
    isolated_units: list[str] = []
    for nid in all_unit_nodes - connected_nodes:
        m = node_meta.get(nid, {})
        parts = [p for p in (m.get("fac",""), m.get("inst",""), m.get("grp","")) if p]
        if parts:
            isolated_units.append(" | ".join(parts))

    _ku = (ku_totals or {}).get(year, {})
    if mode == "F" and selected_facs:
        ku_pubs    = sum(_ku.get("by_fac", {}).get(f, {}).get("pubs",    0) for f in selected_facs)
        ku_authors = sum(_ku.get("by_fac", {}).get(f, {}).get("authors", 0) for f in selected_facs)
    elif mode == "I" and selected_insts:
        ku_pubs    = sum(_ku.get("by_inst", {}).get(i, {}).get("pubs",    0) for i in selected_insts)
        ku_authors = sum(_ku.get("by_inst", {}).get(i, {}).get("authors", 0) for i in selected_insts)
    elif mode in ("G", "IG") and selected_grps:
        ku_pubs    = sum(_ku.get("by_grp", {}).get(g, {}).get("pubs",    0) for g in selected_grps)
        ku_authors = sum(_ku.get("by_grp", {}).get(g, {}).get("authors", 0) for g in selected_grps)
    else:  # S, N, eller ingen filter valgt
        ku_pubs    = _ku.get("total_pubs",    0)
        ku_authors = _ku.get("total_authors", 0)
    
    # Per-enhed intra/inter (Mulighed A: forfatterpar hvor mindst én node tilhører enheden)
    fac_intra_ew: dict[str, float] = {}
    fac_inter_ew: dict[str, float] = {}
    for u, v, w, *_ in edges_keep:
        fu = node_meta.get(u, {}).get("fac", "")
        fv = node_meta.get(v, {}).get("fac", "")
        et = "intra" if (fu and fu == fv) else "inter"
        for f in {fu, fv}:
            if not f: continue
            if et == "intra":
                fac_intra_ew[f] = fac_intra_ew.get(f, 0.0) + w
            else:
                fac_inter_ew[f] = fac_inter_ew.get(f, 0.0) + w

    inst_intra_ew: dict[str, float] = {}
    inst_inter_ew: dict[str, float] = {}
    for u, v, w, *_ in edges_keep:
        iu = node_meta.get(u, {}).get("inst", "")
        iv = node_meta.get(v, {}).get("inst", "")
        et = "intra" if (iu and iu == iv) else "inter"
        for i in {iu, iv}:
            if not i: continue
            if et == "intra":
                inst_intra_ew[i] = inst_intra_ew.get(i, 0.0) + w
            else:
                inst_inter_ew[i] = inst_inter_ew.get(i, 0.0) + w

    grp_intra_ew: dict[str, float] = {}
    grp_inter_ew: dict[str, float] = {}
    # Krydstabuleret: fac → inst → vægt
    fac_inst_intra_ew: dict[str, dict[str, float]] = {}
    fac_inst_inter_ew: dict[str, dict[str, float]] = {}
    # Krydstabuleret: fac → grp → vægt, inst → grp → vægt
    fac_grp_intra_ew: dict[str, dict[str, float]] = {}
    # Krydstabuleret: fac → grp → vægt, inst → grp → vægt
    fac_grp_intra_ew: dict[str, dict[str, float]] = {}
    fac_grp_inter_ew: dict[str, dict[str, float]] = {}
    inst_grp_intra_ew: dict[str, dict[str, float]] = {}
    inst_grp_inter_ew: dict[str, dict[str, float]] = {}
    for u, v, w, *_ in edges_keep:
        gu = node_meta.get(u, {}).get("grp", "")
        gv = node_meta.get(v, {}).get("grp", "")
        fu = node_meta.get(u, {}).get("fac", "")
        fv = node_meta.get(v, {}).get("fac", "")
        iu = node_meta.get(u, {}).get("inst", "")
        iv = node_meta.get(v, {}).get("inst", "")
        # fac → inst krydstabulering
        et_inst = "intra" if (iu and iu == iv) else "inter"
        for inst, fac in [(iu, fu), (iv, fv)]:
            if inst and fac:
                d_fi = fac_inst_intra_ew if et_inst == "intra" else fac_inst_inter_ew
                d_fi.setdefault(fac, {})[inst] = d_fi.setdefault(fac, {}).get(inst, 0.0) + w
        et_grp = "intra" if (gu and gu == gv) else "inter"
        for g, f, i in [(gu, fu, iu), (gv, fv, iv)]:
            if not g: continue
            d_grp = grp_intra_ew if et_grp == "intra" else grp_inter_ew
            d_grp[g] = d_grp.get(g, 0.0) + w 
            if f:
                d_fg = fac_grp_intra_ew if et_grp == "intra" else fac_grp_inter_ew
                d_fg.setdefault(f, {})[g] = d_fg.setdefault(f, {}).get(g, 0.0) + w
            if i:
                d_ig = inst_grp_intra_ew if et_grp == "intra" else inst_grp_inter_ew
                d_ig.setdefault(i, {})[g] = d_ig.setdefault(i, {}).get(g, 0.0) + w
    _fac_reach_partners: dict[str, set] = {}
    for u, v, w, *_ in edges_keep:
        fu = node_meta.get(u, {}).get("fac", "")
        fv = node_meta.get(v, {}).get("fac", "")
        if fu and fv and fu != fv:
            _fac_reach_partners.setdefault(fu, set()).add(fv)
            _fac_reach_partners.setdefault(fv, set()).add(fu)
    _fac_reach = {f: len(_fac_reach_partners.get(f, set())) for f in fac_tot}

    _inst_reach_partners: dict[str, set] = {}
    for u, v, w, *_ in edges_keep:
        iu = node_meta.get(u, {}).get("inst", "")
        iv = node_meta.get(v, {}).get("inst", "")
        if iu and iv and iu != iv:
            _inst_reach_partners.setdefault(iu, set()).add(iv)
            _inst_reach_partners.setdefault(iv, set()).add(iu)
    _inst_reach = {i: len(_inst_reach_partners.get(i, set())) for i in inst_tot}
    
    _inst_fac_reach_partners: dict[str, set] = {}
    for u, v, w, *_ in edges_keep:
        iu = node_meta.get(u, {}).get("inst", "")
        iv = node_meta.get(v, {}).get("inst", "")
        fu = node_meta.get(u, {}).get("fac", "")
        fv = node_meta.get(v, {}).get("fac", "")
        if iu and fv and fu != fv:
            _inst_fac_reach_partners.setdefault(iu, set()).add(fv)
        if iv and fu and fu != fv:
            _inst_fac_reach_partners.setdefault(iv, set()).add(fu)
    _inst_fac_reach = {i: len(_inst_fac_reach_partners.get(i, set())) for i in inst_tot}

    _grp_reach_partners: dict[str, set] = {}
    for u, v, w, *_ in edges_keep:
        gu = node_meta.get(u, {}).get("grp", "")
        gv = node_meta.get(v, {}).get("grp", "")
        if gu and gv and gu != gv:
            _grp_reach_partners.setdefault(gu, set()).add(gv)
            _grp_reach_partners.setdefault(gv, set()).add(gu)
    _grp_reach = {g: len(_grp_reach_partners.get(g, set())) for g in grp_tot}

    _node_ew: dict[str, float] = {}
    for u, v, w, *_ in edges_keep:
        for n in (u, v):
            m = node_meta.get(n, {})
            parts = [p for p in (m.get("grp",""), m.get("inst",""), m.get("fac","")) if p]
            lbl = " | ".join(parts) if parts else "ukendt"
            _node_ew[lbl] = _node_ew.get(lbl, 0.0) + w
    


    return {
        "total_pubs":        total_pubs,
        "total_authors":     total_authors,
        "intra_pubs":        intra_pubs,
        "inter_pubs":        inter_pubs,
        "intra_inst_pubs":   intra_inst_pubs,
        "inter_inst_pubs":   inter_inst_pubs,
        "intra_grp_pubs":    intra_grp_pubs,
        "inter_grp_pubs":    inter_grp_pubs,
        "intra_pubs_all":      _intra_pubs_all,
        "inter_pubs_all":      _inter_pubs_all,
        "intra_inst_pubs_all": _intra_inst_pubs_all,
        "inter_inst_pubs_all": _inter_inst_pubs_all,
        "intra_grp_pubs_all":  _intra_grp_pubs_all,
        "inter_grp_pubs_all":  _inter_grp_pubs_all,
        "fac_intra_ew_all":    _fac_intra_ew_all,
        "fac_inter_ew_all":    _fac_inter_ew_all,
        "inst_intra_ew_all":   _inst_intra_ew_all,
        "inst_inter_ew_all":   _inst_inter_ew_all,
        "grp_intra_ew_all":    _grp_intra_ew_all,
        "grp_inter_ew_all":    _grp_inter_ew_all,
        "fac_grp_tot":       fac_grp_tot,
        "inst_grp_tot":      inst_grp_tot,
        "fac_grp_ew":        fac_grp_ew,
        "inst_grp_ew":       inst_grp_ew,
        "fac_tot":           fac_tot,
        "inst_tot":          inst_tot,
        "grp_tot":           grp_tot,
        "fac_ew":            fac_ew,
        "inst_ew":           inst_ew,
        "grp_ew":            grp_ew,
        "fac_reach":         _fac_reach,
        "inst_reach":        _inst_reach,
        "inst_fac_reach": _inst_fac_reach,
        "grp_reach":         _grp_reach,
        "density":           density,
        "modularity_pre":    modularity_pre_snap,
        "modularity_pre_fac":  _compute_mod_pre(G2, nodes_keep, node_meta, "fac"),
        "modularity_pre_inst": _compute_mod_pre(G2, nodes_keep, node_meta, "inst"),
        "modularity_pre_grp":  _compute_mod_pre(G2, nodes_keep, node_meta, "grp"),
        "modularity_greedy": modularity_greedy_snap,
        "sex_bidrag":        sex_bidrag,
        "combo_pubs":        combo_pubs,
        "combo_pubs_all": combo_pubs_all,
        "hom_index_ku": hom_index_ku,
        "top_pairs":         top_pairs,
        "isolated_units":    isolated_units,
        "ku_total_pubs":    ku_pubs,
        "ku_total_authors": ku_authors,
        "nat_bidrag": nat_bidrag,
        "nat_ew":     nat_ew,
        "nat_ew_all": nat_ew_all,
        "cs_hom_index_ku": cs_hom_index_ku,
        "sampub_count_raw": sampub_count_raw,
        "sampub_rate": round(total_pubs / total_authors, 4) if total_authors else 0.0,
        "sampub_rate":     round(total_pubs / total_authors, 4) if total_authors else 0.0,
        "total_pubs_all":  _intra_pubs_all + _inter_pubs_all,
        "sampub_rate_all": round((_intra_pubs_all + _inter_pubs_all) / total_authors, 4) if total_authors else 0.0,
        "fac_intra_ew":      fac_intra_ew,
        "fac_inter_ew":      fac_inter_ew,
        "inst_intra_ew":     inst_intra_ew,
        "inst_inter_ew":     inst_inter_ew,
        "grp_intra_ew":      grp_intra_ew,
        "grp_inter_ew":      grp_inter_ew,
        "fac_inst_intra_ew": fac_inst_intra_ew,
        "fac_inst_inter_ew": fac_inst_inter_ew,
        "fac_grp_intra_ew":  fac_grp_intra_ew,
        "fac_grp_inter_ew":  fac_grp_inter_ew,
        "inst_grp_intra_ew": inst_grp_intra_ew,
        "inst_grp_inter_ew": inst_grp_inter_ew,
        "grp_node_ew": _node_ew,
        "fac_bs":  fac_bs,
        "inst_bs": inst_bs,
        "grp_bs":  grp_bs,
        "fac_ew_all": {
            f: _fac_intra_ew_all.get(f, 0.0) + _fac_inter_ew_all.get(f, 0.0)
            for f in set(_fac_intra_ew_all) | set(_fac_inter_ew_all)
        },
        "inst_ew_all": {
            i: _inst_intra_ew_all.get(i, 0.0) + _inst_inter_ew_all.get(i, 0.0)
            for i in set(_inst_intra_ew_all) | set(_inst_inter_ew_all)
        },
        "grp_ew_all": {
            g: _grp_intra_ew_all.get(g, 0.0) + _grp_inter_ew_all.get(g, 0.0)
            for g in set(_grp_intra_ew_all) | set(_grp_inter_ew_all)
        },
        "sund_lektor_size": sum(
            m.get("size", 0) for nid, m in node_meta.items()
            if m.get("fac") == "SUND" and m.get("grp") == "Lektor"
        ),
    }



def main():
    # -----------------------------------------------------------------------
    # Load data 
    # -----------------------------------------------------------------------
    data_by_year_CURIS = load_network_data()

    data_by_year_OpenAlex = {
        2021: {
            "nodes": [
                # --- TEO ---
                ("TEO|Bibel|Professor|m|DK",      {"fac": "TEO",     "inst": "Bibel",     "grp": "Professor", "size": 55, "sex": "m", "statsborgerskab": "DK"}),
                ("TEO|Bibel|Professor|k|DK",      {"fac": "TEO",     "inst": "Bibel",     "grp": "Professor", "size": 63, "sex": "k", "statsborgerskab": "DK"}),
                ("TEO|Bibel|Lektor|k|DK",         {"fac": "TEO",     "inst": "Bibel",     "grp": "Lektor",    "size": 30, "sex": "k", "statsborgerskab": "DK"}),
                ("TEO|Bibel|Adjunkt|k|D",         {"fac": "TEO",     "inst": "Bibel",     "grp": "Adjunkt",   "size": 20, "sex": "k", "statsborgerskab": "D"}),
                ("TEO|Bibel|Ph.d.|m|DK",          {"fac": "TEO",     "inst": "Bibel",     "grp": "Ph.d.",     "size": 15, "sex": "m", "statsborgerskab": "DK"}),
                ("TEO|Oldgræsk|Professor|m|DK",   {"fac": "TEO",     "inst": "Oldgræsk",  "grp": "Professor", "size": 38, "sex": "m", "statsborgerskab": "DK"}),
                ("TEO|Oldgræsk|Lektor|k|UK",      {"fac": "TEO",     "inst": "Oldgræsk",  "grp": "Lektor",    "size": 29, "sex": "k", "statsborgerskab": "UK"}),
                ("TEO|Oldgræsk|Adjunkt|m|DK",     {"fac": "TEO",     "inst": "Oldgræsk",  "grp": "Adjunkt",   "size": 24, "sex": "m", "statsborgerskab": "DK"}),
                ("TEO|Oldgræsk|Ph.d.|k|D",        {"fac": "TEO",     "inst": "Oldgræsk",  "grp": "Ph.d.",     "size": 16, "sex": "k", "statsborgerskab": "D"}),
                # --- HUM ---
                ("HUM|Engelsk|Professor|k|UK",    {"fac": "HUM",     "inst": "Engelsk",   "grp": "Professor", "size": 44, "sex": "k", "statsborgerskab": "UK"}),
                ("HUM|Engelsk|Lektor|m|DK",       {"fac": "HUM",     "inst": "Engelsk",   "grp": "Lektor",    "size": 25, "sex": "m", "statsborgerskab": "DK"}),
                ("HUM|Engelsk|Ph.d.|k|DK",        {"fac": "HUM",     "inst": "Engelsk",   "grp": "Ph.d.",     "size": 19, "sex": "k", "statsborgerskab": "DK"}),
                ("HUM|Engelsk|Adjunkt|m|UK",      {"fac": "HUM",     "inst": "Engelsk",   "grp": "Adjunkt",   "size": 13, "sex": "m", "statsborgerskab": "UK"}),
                ("HUM|Historie|Professor|m|DK",   {"fac": "HUM",     "inst": "Historie",  "grp": "Professor", "size": 50, "sex": "m", "statsborgerskab": "DK"}),
                ("HUM|Historie|Lektor|k|DK",      {"fac": "HUM",     "inst": "Historie",  "grp": "Lektor",    "size": 27, "sex": "k", "statsborgerskab": "DK"}),
                ("HUM|Historie|Ph.d.|m|D",        {"fac": "HUM",     "inst": "Historie",  "grp": "Ph.d.",     "size": 17, "sex": "m", "statsborgerskab": "D"}),
                ("HUM|Historie|Adjunkt|k|DK",     {"fac": "HUM",     "inst": "Historie",  "grp": "Adjunkt",   "size": 14, "sex": "k", "statsborgerskab": "DK"}),
                # --- SCIENCE ---
                ("SCIENCE|Matematik|Professor|m|DK", {"fac": "SCIENCE", "inst": "Matematik", "grp": "Professor", "size": 60, "sex": "m", "statsborgerskab": "DK"}),
                ("SCIENCE|Matematik|Lektor|m|DK",    {"fac": "SCIENCE", "inst": "Matematik", "grp": "Lektor",    "size": 45, "sex": "m", "statsborgerskab": "DK"}),
                ("SCIENCE|Matematik|Postdoc|k|D",    {"fac": "SCIENCE", "inst": "Matematik", "grp": "Postdoc",   "size": 28, "sex": "k", "statsborgerskab": "D"}),
                ("SCIENCE|Matematik|Ph.d.|k|UK",     {"fac": "SCIENCE", "inst": "Matematik", "grp": "Ph.d.",     "size": 22, "sex": "k", "statsborgerskab": "UK"}),
                ("SCIENCE|Kemi|Professor|k|DK",      {"fac": "SCIENCE", "inst": "Kemi",      "grp": "Professor", "size": 67, "sex": "k", "statsborgerskab": "DK"}),
                ("SCIENCE|Kemi|Lektor|m|D",          {"fac": "SCIENCE", "inst": "Kemi",      "grp": "Lektor",    "size": 42, "sex": "m", "statsborgerskab": "D"}),
                ("SCIENCE|Kemi|Postdoc|k|DK",        {"fac": "SCIENCE", "inst": "Kemi",      "grp": "Postdoc",   "size": 24, "sex": "k", "statsborgerskab": "DK"}),
                ("SCIENCE|Kemi|Ph.d.|m|UK",          {"fac": "SCIENCE", "inst": "Kemi",      "grp": "Ph.d.",     "size": 20, "sex": "m", "statsborgerskab": "UK"}),
            ],
            "edges": [
                ("TEO|Oldgræsk|Lektor|k|UK",         "TEO|Bibel|Adjunkt|k|D",             3, "k-k"),
                ("TEO|Oldgræsk|Professor|m|DK",      "TEO|Oldgræsk|Lektor|k|UK",          5, "k-m"),
                ("TEO|Oldgræsk|Lektor|k|UK",         "TEO|Oldgræsk|Ph.d.|k|D",            2, "k-k"),
                ("HUM|Engelsk|Professor|k|UK",       "HUM|Engelsk|Lektor|m|DK",           5, "k-m"),
                ("HUM|Engelsk|Lektor|m|DK",          "HUM|Engelsk|Ph.d.|k|DK",            2, "k-m"),
                ("HUM|Engelsk|Professor|k|UK",       "HUM|Historie|Lektor|k|DK",          3, "k-k"),
                ("HUM|Historie|Ph.d.|m|D",           "HUM|Historie|Adjunkt|k|DK",         3, "k-m"),
                ("SCIENCE|Matematik|Professor|m|DK", "SCIENCE|Matematik|Lektor|m|DK",     7, "m-m"),
                ("SCIENCE|Matematik|Ph.d.|k|UK",     "SCIENCE|Matematik|Postdoc|k|D",     3, "k-k"),
                ("SCIENCE|Matematik|Professor|m|DK", "SCIENCE|Kemi|Lektor|m|D",           6, "m-m"),
                ("SCIENCE|Kemi|Postdoc|k|DK",        "SCIENCE|Kemi|Ph.d.|m|UK",           3, "k-m"),
                ("TEO|Bibel|Professor|k|DK",         "HUM|Engelsk|Ph.d.|k|DK",            1, "k-k"),
                ("TEO|Oldgræsk|Adjunkt|m|DK",        "HUM|Historie|Ph.d.|m|D",            2, "m-m"),
                ("HUM|Engelsk|Lektor|m|DK",          "SCIENCE|Matematik|Postdoc|k|D",     3, "k-m"),
                ("HUM|Historie|Professor|m|DK",      "SCIENCE|Kemi|Professor|k|DK",       5, "k-m"),
                ("TEO|Oldgræsk|Professor|m|DK",      "SCIENCE|Matematik|Professor|m|DK",  2, "m-m"),
            ],
        },
        2022: {
            "nodes": [
                ("TEO|Bibel|Professor|m|DK",         {"fac": "TEO",     "inst": "Bibel",     "grp": "Professor", "size": 50, "sex": "m", "statsborgerskab": "DK", "fwci": 1.3}),
                ("TEO|Bibel|Adjunkt|k|D",            {"fac": "TEO",     "inst": "Bibel",     "grp": "Adjunkt",   "size": 19, "sex": "k", "statsborgerskab": "D",  "fwci": 1.0}),
                ("TEO|Oldgræsk|Lektor|k|UK",         {"fac": "TEO",     "inst": "Oldgræsk",  "grp": "Lektor",    "size": 30, "sex": "k", "statsborgerskab": "UK", "fwci": 1.8}),
                ("HUM|Engelsk|Ph.d.|k|DK",           {"fac": "HUM",     "inst": "Engelsk",   "grp": "Ph.d.",     "size": 31, "sex": "k", "statsborgerskab": "DK", "fwci": 1.0}),
                ("HUM|Historie|Lektor|k|DK",         {"fac": "HUM",     "inst": "Historie",  "grp": "Lektor",    "size": 28, "sex": "k", "statsborgerskab": "DK", "fwci": 1.2}),
                ("SCIENCE|Matematik|Professor|m|DK", {"fac": "SCIENCE", "inst": "Matematik", "grp": "Professor", "size": 65, "sex": "m", "statsborgerskab": "DK", "fwci": 2.6}),
                ("SCIENCE|Kemi|Postdoc|k|DK",        {"fac": "SCIENCE", "inst": "Kemi",      "grp": "Postdoc",   "size": 36, "sex": "k", "statsborgerskab": "DK", "fwci": 2.7}),
            ],
            "edges": [
                ("TEO|Bibel|Professor|m|DK",         "HUM|Engelsk|Ph.d.|k|DK",           2, "k-m"),
                ("TEO|Oldgræsk|Lektor|k|UK",         "SCIENCE|Matematik|Professor|m|DK", 2, "k-m"),
                ("HUM|Historie|Lektor|k|DK",         "SCIENCE|Kemi|Postdoc|k|DK",        3, "k-k"),
                ("TEO|Bibel|Adjunkt|k|D",            "SCIENCE|Kemi|Postdoc|k|DK",        2, "k-k"),
                ("HUM|Engelsk|Ph.d.|k|DK",           "HUM|Historie|Lektor|k|DK",         2, "k-k"),
                ("SCIENCE|Matematik|Professor|m|DK", "SCIENCE|Kemi|Postdoc|k|DK",        4, "k-m"),
            ],
        },
    }

    def get_data_by_year(source: str) -> dict:

        if source == "CURIS":
            return data_by_year_CURIS

        if source == "OpenAlex":
            return data_by_year_OpenAlex

        # --- Merge both sources ---
        merged = {}
        all_years = sorted(set(data_by_year_CURIS) | set(data_by_year_OpenAlex))

        for year in all_years:
            curis  = data_by_year_CURIS.get(year,    {"nodes": [], "edges": []})
            openalex = data_by_year_OpenAlex.get(year, {"nodes": [], "edges": []})

            # Nodes: sum sizes for matching IDs; keep all unique IDs
            node_map = {}
            for nid, meta in curis["nodes"]:
                node_map[nid] = dict(meta)
                node_map[nid]["_sources"] = {"CURIS"}

            for nid, meta in openalex["nodes"]:
                if nid in node_map:
                    node_map[nid]["size"] += meta.get("size", 0)
                    node_map[nid]["_sources"].add("OpenAlex")
                else:
                    node_map[nid] = dict(meta)
                    node_map[nid]["_sources"] = {"OpenAlex"}

            # Edges: sum weights for matching pairs
            edge_map = {}
            for u, v, w, *rest in curis["edges"]:
                key = tuple(sorted((u, v)))
                sex_combo = rest[0] if rest else None
                edge_map[key] = {"w": w, "sex_combo": sex_combo}

            for u, v, w, *rest in openalex["edges"]:
                key = tuple(sorted((u, v)))
                if key in edge_map:
                    edge_map[key]["w"] += w
                else:
                    sex_combo = rest[0] if rest else None
                    edge_map[key] = {"w": w, "sex_combo": sex_combo}

            merged[year] = {
                "nodes": [(nid, meta) for nid, meta in node_map.items()],
                "edges": [(u, v, d["w"], d["sex_combo"]) for (u, v), d in edge_map.items()],
                }

        return merged
    
    with st.spinner("Indlæser publikationsdata..."):
        ku_farver = load_ku_colors()
        faculty_base_colors = build_faculty_colors(ku_farver)
        grp_colors = stillingsgruppe_colors(ku_farver)

        SVG_CENTRALITET  = load_svg("fig_centralitet_v4.svg")
        SVG_FORFATTERPAR = load_svg("fig_forfatterpar.svg")

        forfatterpositioner = load_forfatterpositioner()
        ku_totals = load_ku_totals()
        _, inst_to_fac = load_inst_filter()
        forfatterantal_data    = load_forfatterantal()
        forfatterantal_dist = load_forfatterantal_dist()
        publikationstyper_data = load_publikationstyper()

    # -----------------------------------------------------------------------
    # Page config & header
    # -----------------------------------------------------------------------

    st.set_page_config(
        page_title="REKSTAB Analyse",
        page_icon=load_logo(),
        layout="wide",
    )

    col_logo, col_title = st.columns([1, 4])
    with col_logo:
        st.markdown(
            f'<img src="data:image/png;base64,{base64.b64encode(load_logo()).decode()}" '
            f'style="max-width:180px; width:100%;">',
            unsafe_allow_html=True,
        )
    with col_title:
        st.title("Sampublicering på Københavns Universitet")

    # -----------------------------------------------------------------------
    # Sidebar filters
    # -----------------------------------------------------------------------
    
    # RET!
    #for _key in ("cb_fac", "cb_inst", "cb_grp"):
        #if _key not in st.session_state:
            #st.session_state[_key] = True

    for _key, _default in [("cb_fac", True), ("cb_inst", False), ("cb_grp", False)]:
        if _key not in st.session_state:
            st.session_state[_key] = _default

    with st.sidebar:
        st.sidebar.header("Filtre og visning")

        st.sidebar.caption("Analysen kortlægger sampubliceringsmønstre blandt KU's videnskabelige personale. \n\n Brug filtrene til at zoome ind på et bestemt fakultet, karrieretrin, år eller tilføje en diversitetsdimension.")
        
        with st.expander("**Datakilde**"):

            data_source = st.radio(
                "Vælg datakilde",
                options = ["CURIS", "OpenAlex", "Begge"],
                index = 0,
                key = "data_source_radio",
                help = "Vælg datakilden, som skal ligge til grund for analyserne"
            )
        
            year = st.selectbox("Vælg år", list(get_data_by_year(data_source).keys()), index=0, key="year_selectbox")
        
        data_by_year = get_data_by_year(data_source)
    
    institut_fakultets_map = {}
    for year_content in data_by_year.values():
        for nid, meta in year_content["nodes"]:
            fac  = meta.get("fac")
            inst = meta.get("inst")
            if fac and inst:
                institut_fakultets_map[inst] = fac

    years = list(data_by_year.keys())

    all_faculties = sorted({
        meta["fac"]
        for content in data_by_year.values()
        for _, meta in content["nodes"]
    })
    all_groups = sorted({
        meta["grp"]
        for content in data_by_year.values()
        for _, meta in content["nodes"]
    }, key=lambda g: HIERARKI.get(g, 999))

    # All citizenship codes present across all years - DK sorted first.
    _all_cs_raw = sorted({
        meta.get("statsborgerskab", "")
        for content in data_by_year.values()
        for _, meta in content["nodes"]
        if meta.get("statsborgerskab")
    })
    all_citizenships = (
        (["DK"] if "DK" in _all_cs_raw else []) +
        [c for c in _all_cs_raw if c != "DK"]
    )

    global_sizes      = [meta["size"] for yc in data_by_year.values() for _, meta in yc["nodes"]]
    global_min_auth   = min(global_sizes) if global_sizes else 0
    global_max_auth   = max(global_sizes) if global_sizes else 0

    all_edge_counts   = [int(round(e[2])) for yc in data_by_year.values() for e in yc["edges"]]
    global_min_w      = min(all_edge_counts) if all_edge_counts else 0
    global_max_w      = max(all_edge_counts) if all_edge_counts else 1
    
    def _on_diversity_change():        
        new_value = st.session_state.get("rb_diversitet", "Ingen")
        
        if new_value in ("Køn", "Statsborgerskab"):
            # Gem nuværende org-tilstand (kun hvis ikke allerede gemt)
            if "_org_state_pre_diversity" not in st.session_state:
                st.session_state["_org_state_pre_diversity"] = {
                    "cb_fac":  st.session_state.get("cb_fac",  True),
                    "cb_inst": st.session_state.get("cb_inst", True),
                    "cb_grp":  st.session_state.get("cb_grp",  True),
                }
            # Find det mest aggregerede niveau der er valgt nu
            if st.session_state.get("cb_fac", False):
                st.session_state["cb_fac"]  = True
                st.session_state["cb_inst"] = False
                st.session_state["cb_grp"]  = False
            elif st.session_state.get("cb_inst", False):
                st.session_state["cb_fac"]  = False
                st.session_state["cb_inst"] = True
                st.session_state["cb_grp"]  = False
            elif st.session_state.get("cb_grp", False):
                # Allerede kun G valgt - lad det være
                pass
            else:
                # Ingen niveauer valgt (kantcase) - default til F
                st.session_state["cb_fac"]  = True
                st.session_state["cb_inst"] = False
                st.session_state["cb_grp"]  = False
        else:
            # Brugeren slog diversitet fra: genopret tidligere org-tilstand
            prev = st.session_state.pop("_org_state_pre_diversity", None)
            if prev:
                st.session_state["cb_fac"]  = prev["cb_fac"]
                st.session_state["cb_inst"] = prev["cb_inst"]
                st.session_state["cb_grp"]  = prev["cb_grp"]

    with st.sidebar:
        # --- DIVERSITET expander (no organisation info here) ---
        with st.expander("**Diversitet**"):
            st.caption(
            "Aktivér for at tilføje køns- og nationalitetsfaner."
            "\n\nFiltrene på køn og statsborgerskab begrænser, hvilke forfattere der indgår i netværket og analyserne.")
            
            diversitetsvalg = st.radio(
                "**Diversitetsdimension**",
                options=["Ingen", "Køn", "Statsborgerskab"],
                index=0,
                key="rb_diversitet",
                on_change=_on_diversity_change,  # NY linje
            )
            
            analyse_køn = (diversitetsvalg == "Køn")
            analyse_nat = (diversitetsvalg == "Statsborgerskab")

            if analyse_køn:
                selected_genders = st.multiselect(
                    "Vælg køn (tom = begge)",
                    options = ["k", "m"],
                    default = ["k", "m"],
                    format_func = lambda x: {"k": "Kvinder", "m": "Mænd"}.get(x, x)
                )

                selected_gender_edges = st.multiselect(
                    "Forfatterpar efter kønskombination",
                    options = ["k-k", "k-m", "m-m"],
                    default = [],
                    format_func=lambda x: {"k-k": "Kvinde-Kvinde", "k-m": "Kvinde-Mand", "m-m": "Mand-Mand"}.get(x, x),
                )
            else:
                selected_genders = []
                selected_gender_edges = []
            
            if analyse_nat:
                _cs_totals = {}
                for yc in data_by_year.values():
                    for nid, meta in yc["nodes"]:
                        cs = meta.get("statsborgerskab", "")
                        if cs:
                            _cs_totals[cs] = _cs_totals.get(cs, 0) + meta.get("size", 0)
                _cs_sorted = sorted(_cs_totals, key=lambda c: -_cs_totals[c])

                _nat_top_n = st.number_input(
                    "Antal nationaliteter i netværket",
                    min_value=1,
                    max_value=len(_cs_sorted),
                    value=min(5, len(_cs_sorted)),
                    step=1,
                    key="nat_top_n_sidebar",
                    help="Netværket viser kun forfattere fra de *X* mest repræsenterede nationaliteter.",
                )
                selected_citizenships = _cs_sorted[:_nat_top_n]
                st.caption(f"Inkluderet i netværk og analyser: {', '.join(country_name(c) for c in selected_citizenships)}")
                #st.caption(f"Inkluderet i netværk og analyser: {', '.join(c for c in selected_citizenships)}")
            else:
                selected_citizenships = []

        sex_active = bool(selected_genders or selected_gender_edges)
        diversity_active = sex_active or bool(selected_citizenships)

        # --- ORGANISATION expander ---
        _org_opts = ["Fakulteter", "Institutter", "Stillingsgrupper"]

        with st.expander("**Organisation**"):
            st.caption(
"""Fakulteter og institutter er administrative enheder.

Stillingsgrupper opdeler efter karrieretrin."""
            )

            if diversitetsvalg in ("Køn", "Statsborgerskab"):
                _active_levels = sum([
                    st.session_state.get("cb_fac",  False),
                    st.session_state.get("cb_inst", False),
                    st.session_state.get("cb_grp",  False),
                ])
                if _active_levels > 1:
                    st.caption(
                        "**Bemærk:** Diversitetsdimensioner fungerer bedst med ét organisationsniveau "
                        "valgt. Med flere niveauer kan netværket blive svært at læse."
                    )

            show_fac  = st.checkbox("**Fakulteter**",       key="cb_fac")
            show_inst = st.checkbox("**Institutter**",      key="cb_inst")
            show_grp  = st.checkbox("**Stillingsgrupper**", key="cb_grp")

            if show_grp:
                st.caption("""Med stillingsgrupper aktiveret bliver netværksvisningen så kompleks,
                at linjetykkelserne ikke direkte kan aflæses som forfatterpar.
                """)

            _all_insts_by_fac: dict[str, list] = {}
            for yc in data_by_year.values():
                for _, meta in yc["nodes"]:
                    f, i = meta.get("fac", ""), meta.get("inst", "")
                    if f and i and i not in _all_insts_by_fac.setdefault(f, []):
                        _all_insts_by_fac[f].append(i)

            selected_facs = (
                st.multiselect("Vælg fakulteter (tom = alle)", all_faculties, default=[])
                if show_fac else []
            )

            if show_inst:
                _inst_opts_sidebar = sorted({
                    i
                    for f, insts in _all_insts_by_fac.items()
                    for i in insts
                    if not selected_facs or f in selected_facs
                })
                selected_insts = st.multiselect(
                    "Vælg institutter (tom = alle)",
                    options=_inst_opts_sidebar,
                    default=[v for v in st.session_state.get("selected_insts_prev", []) if v in _inst_opts_sidebar],
                    key="selected_insts_ms",
                )
                st.session_state["selected_insts_prev"] = selected_insts
            else:
                selected_insts = []

            selected_grps = (
                st.multiselect("Vælg stillingsgrupper (tom = alle)", all_groups, default=[])
                if show_grp else []
            )

            _org_levels_pre = [l for l, a in [("F", show_fac), ("I", show_inst), ("G", show_grp)] if a]
            _sex_active_pre = bool(selected_genders or selected_gender_edges)
            if _sex_active_pre and len(_org_levels_pre) == 1:
                _mode_pre = _org_levels_pre[0] + "S"
            else:
                _mode_pre = ("F" if show_fac else "") + ("I" if show_inst else "") + ("G" if show_grp else "")

            st.caption("Filtrene bestemmer, hvilke forfatterpar der indgår i analyserne. Intra-par er forfatterpar inden for samme enhed; inter-par er forfatterpar på tværs af enheder.")

            _intra_label, _inter_label = intra_inter_labels(_mode_pre)
            show_intra = st.checkbox(f"Vis {_intra_label} forfatterpar", True, key="chk_intra")
            show_inter = st.checkbox(f"Vis {_inter_label} forfatterpar", True, key="chk_inter")

            _inst_mode = show_fac and show_inst
            _grp_mode  = show_grp
            if _inst_mode:
                show_intra_inst = st.checkbox("Vis intra-instituttets forfatterpar", True, key="chk_intra_inst")
                show_inter_inst = st.checkbox("Vis inter-instituttets forfatterpar", True, key="chk_inter_inst")
            else:
                show_intra_inst = True
                show_inter_inst = True

            if show_grp:
                show_intra_grp = st.checkbox("Vis intra-stillingsgruppe forfatterpar", True, key="chk_intra_grp")
                show_inter_grp = st.checkbox("Vis inter-stillingsgruppe forfatterpar", True, key="chk_inter_grp")
            else:
                show_intra_grp = True
                show_inter_grp = True

# -----------------------------------------------------------------------
# Main controls
# -----------------------------------------------------------------------
# Derive a preliminary mode just for checkbox labels (full mode comes after)
    _org_levels_pre = [l for l, a in [("F", show_fac), ("I", show_inst), ("G", show_grp)] if a]
    _sex_active_pre = bool(selected_genders or selected_gender_edges)
    if _sex_active_pre and len(_org_levels_pre) == 1:
        _mode_pre = _org_levels_pre[0] + "S"
    else:
        _mode_pre = ("F" if show_fac else "") + ("I" if show_inst else "") + ("G" if show_grp else "")

    _inst_mode = show_fac and show_inst   # modes where both fac + inst are visible (FI, FIG)

    with st.sidebar:
        with st.expander("**Forskningsprofil**"):
            st.caption(
                "Aktivér for at tilføje faner med bibliometriske mål fra OpenAlex"
            )
            analyse_internationalt = st.checkbox(
                "**Internationalt samarbejde**",
                value=False,
                key="cb_intl",
                help="Samarbejde med forfattere uden for KU baseret på affiliering",
            )
            analyse_fwci = st.checkbox(
                "**FWCI (citationsimpact)**",
                value=False,
                key="cb_fwci",
                help="Field-Weighted Citation Impact fra OpenAlex",
            )
            analyse_output = st.checkbox(
                "**Forskningsoutput**",
                value=False,
                key="cb_output",
                help="Publikationsvolumen, open access-andel m.m. fra OpenAlex",
            )

        netvaerk_expander = st.expander("**Netværksvisning**")


    cols = st.columns(2)
    
    if not show_fac and not show_inst and not show_grp:
        st.error("Vælg mindst ét filter (fakultet, institut eller stillingsgruppe).")
        st.stop()

    # -----------------------------------------------------------------------
    # Determine mode
    # -----------------------------------------------------------------------
    # Sex can now be combined with any single organisation level.
    # Priority when multiple org levels are checked: F > I > G
    # Sex suffix "S" is appended when sex filter is active AND exactly one
    # org level is selected (so the split is meaningful).

    org_levels = [l for l, active in [("F", show_fac), ("I", show_inst), ("G", show_grp)] if active]

    mode = (
        ("F" if show_fac  else "") +
        ("I" if show_inst else "") +
        ("G" if show_grp  else "")
    )
    if sex_active:
        mode = mode + "S"
    if selected_citizenships:
        mode = mode + "N"

    _base_mode_labels = {
        "F":   "fakulteter",
        "FI":  "fakulteter og institutter",
        "FIG": "fakulteter, institutter og stillingsgrupper",
        "FG":  "fakulteter og stillingsgrupper",
        "I":   "institutter",
        "IG":  "institutter og stillingsgrupper",
        "G":   "stillingsgrupper",
    }
    _current_mode_label = _base_mode_labels.get(base_mode(mode), base_mode(mode))

    # -----------------------------------------------------------------------
    # Build raw node metadata for selected year
    # -----------------------------------------------------------------------

    content = data_by_year[year]

    raw_nodes = {}
    for nid, meta in content["nodes"]:
        m = dict(meta)
        m.setdefault("type", "grp")
        parts = nid.split("|")
        if "fac"  not in m and len(parts) >= 1: 
            m["fac"]  = parts[0]
        if "inst" not in m and len(parts) >= 2: 
            m["inst"] = parts[1]
        if "grp"  not in m and len(parts) >= 3: 
            m["grp"]  = parts[2]
        if "sex"  not in m and len(parts) >= 4: 
            m["sex"]  = parts[3]
        if "statsborgerskab" not in m and len(parts) >= 5: 
            m["statsborgerskab"] = parts[4]
        m.setdefault("size", 0)
        raw_nodes[nid] = m

    # ---------------------------------------------------------------------------
    # Citizenship filter: applied at raw grp-node level BEFORE any merge.
    # Empty selection = no filter (show all). OR logic for multiple codes.
    # ---------------------------------------------------------------------------
    raw_nodes_pre_cs = dict(raw_nodes)

    if selected_citizenships:
        raw_nodes = {
            nid: m for nid, m in raw_nodes.items()
            if m.get("type") != "grp"
            or m.get("statsborgerskab", "") in selected_citizenships
        }
        # Remove edges that reference nodes now absent
        raw_edges_prefilter = [
            e for e in content["edges"]
            if e[0] in raw_nodes and e[1] in raw_nodes
        ]
    else:
        raw_edges_prefilter = list(content["edges"])


    # Add institute placeholder nodes
    for nid, m in list(raw_nodes.items()):
        fac  = m.get("fac", "")
        inst = m.get("inst", "")
        if not fac or not inst:
            continue
        inst_id = f"INST:{fac}|{inst}"
        if inst_id not in raw_nodes:
            raw_nodes[inst_id] = {"type": "inst", "fac": fac, "inst": inst, "grp": "", "size": 0}

    # Normalise edges to 4-tuples
    raw_edges = [
        (e[0], e[1], e[2], e[3] if len(e) > 3 else None)
        for e in raw_edges_prefilter
    ]
    raw_nodes_unfiltered = dict(raw_nodes)   # snapshot after inst placeholders, before mode merge
    raw_edges_unfiltered = list(raw_edges)

    # -----------------------------------------------------------------------
    # Apply mode merge
    # -----------------------------------------------------------------------

    node_meta, edge_source = apply_mode_merge(mode, raw_nodes, raw_edges)

    all_inst = sorted({m.get("inst", "UKENDT") for m in node_meta.values()})

    # -----------------------------------------------------------------------
    # Size slider
    # -----------------------------------------------------------------------

    pre_size_nodes = {
        nid: m for nid, m in node_meta.items()
        if passes_category_filters(m, mode, show_fac, show_inst, show_grp,
                                   selected_facs, selected_insts, selected_grps)
    }

    if not pre_size_nodes:
        st.error("Ingen noder matcher valgte filtre.")
        st.stop()

    candidate_items  = [(nid, int(m.get("size", 0))) for nid, m in pre_size_nodes.items() if size_relevant_in_mode(m, mode)]
    candidate_sizes  = [s for _, s in candidate_items]
    vis_min          = min(candidate_sizes) if candidate_sizes else 0
    vis_max          = max(candidate_sizes) if candidate_sizes else 1
    if vis_min == vis_max:
        vis_max = vis_min + 1 if vis_min > 0 else 1

    author_ver = (mode, tuple(sorted(candidate_items)))
    if st.session_state.get("author_version") != author_ver:
        st.session_state["author_range"]   = (vis_min, vis_max)
        st.session_state["author_version"] = author_ver

    # -----------------------------------------------------------------------
    # Edge slider
    # -----------------------------------------------------------------------

    pre_nodes_keep = pre_nodes_for_mode(node_meta, mode)
    edge_candidates = [
        (e[0], e[1], int(round(e[2])))
        for e in edge_source
        if e[0] in node_meta and e[1] in node_meta and e[0] in pre_nodes_keep and e[1] in pre_nodes_keep
    ]
    cur_min_w = min(w for _, _, w, *_ in edge_candidates) if edge_candidates else 0
    cur_max_w = max(w for _, _, w, *_ in edge_candidates) if edge_candidates else 1

    edge_ver = (mode, year, len(edge_candidates), cur_min_w, cur_max_w)
    if st.session_state.get("edge_version") != edge_ver:
        st.session_state["edge_range"]   = (cur_min_w, cur_max_w)
        st.session_state["edge_version"] = edge_ver

    with st.sidebar:
        with netvaerk_expander:
            st.caption("Silderne afgrænser udsnittet baseret på størrelse: hvor mange forfattere "
            "en enhed mindst skal have, og hvor mange forfatterpar der mindst skal være for at "
            "kanten medtages. Vægt og netværksstørrelse styrer kun den visuelle fremstilling.")
            author_min, author_max = st.slider(
                "Filter: antal unikke forfattere (dynamisk)",
                min_value=vis_min, max_value=vis_max,
                value=st.session_state.get("author_range", (vis_min, vis_max)),
                key="author_minmax_slider",
                help="Noder med færre forfattere end minimumsværdien skjules. Nyttig til at fjerne meget små enheder fra netværket.",
            )
            st.session_state["author_range"] = (author_min, author_max)

            edge_min, edge_max = st.slider(
                "Filter: antal forfatterpar",
                min_value=int(cur_min_w), max_value=int(cur_max_w),
                value=st.session_state.get("edge_range", (int(cur_min_w), int(cur_max_w))),
                step=1,
                key=f"edge_slider_{year}_{mode}",
                help="Færre forfatterpar end minimumsværdien skjules. Sæt minimumsværdien højt for kun at se de stærkeste samarbejder.",
            )
            edge_scale = st.slider(
                "Vægt:", 
                min_value=1.0, 
                max_value=50.0, 
                value=6.0, 
                step=0.1, 
                key="edge_scale_slider",
                help="Skalerer linjetykkelsen i netværket. Påvirker kun visualiseringen, ikke de underliggende tal."
            )
            _default_scale = {"I": 400, "F": 600, "G": 400}.get(base_mode(mode), 1200)
            _scale_key = f"network_scale_default_{base_mode(mode)}"
            network_scale = st.slider(
                "Netværksstørrelse",
                min_value=100, max_value=5000,
                value=st.session_state.get(_scale_key, _default_scale),
                key="network_scale_slider",
                help="Justerer størrelsen af hele netværksplottet. Øg værdien, hvis noder overlapper."
            )
            st.session_state[_scale_key] = network_scale

    st.session_state["edge_range"] = (edge_min, edge_max)

    # -----------------------------------------------------------------------
    # Apply remaining node filters
    # -----------------------------------------------------------------------

    node_meta = {
        nid: m for nid, m in pre_size_nodes.items()
        if node_passes_size(m, mode, author_min, author_max)
    }
    if not node_meta:
        st.error("Ingen noder matcher de valgte filtre (kategori + størrelse).")
        st.stop()

    node_meta = {
        nid: m for nid, m in node_meta.items()
        if node_passes_filters(nid, m, mode, show_fac, show_inst, show_grp,
                               selected_facs, selected_insts, selected_grps)
    }
    if not node_meta:
        st.error("Ingen noder matcher de valgte filtre.")
        st.stop()

    # Sex filter on nodes - active in FS, IS, GS modes
    if selected_genders and sex_in_mode(mode):
        node_meta = {
            nid: m for nid, m in node_meta.items()
            if m.get("sex") in selected_genders
            or m.get("type") in ("fac", "inst", "fac_sex", "inst_sex")
        }
        if not node_meta:
            st.error("Ingen noder matcher det valgte kønsfilter.")
            st.stop()

    # -----------------------------------------------------------------------
    # Build edges_keep
    # -----------------------------------------------------------------------

    edge_source = [e for e in edge_source if e[0] in node_meta and e[1] in node_meta]

    nodes_keep = pre_nodes_for_mode(node_meta, mode)
    total_authors = sum(node_meta[nid].get("size", 0) for nid in nodes_keep)

    edges_keep = []
    for edge in edge_source:
        u, v, w = edge[0], edge[1], edge[2]
        sex_combo = edge[3] if len(edge) > 3 else None
        if u not in node_meta or v not in node_meta:
            continue
        if u not in nodes_keep or v not in nodes_keep:
            continue
        et = edge_type(u, v, node_meta, mode)
        if et == "intra" and not show_intra:
            continue
        if et == "inter" and not show_inter:
            continue
        # Institute-level filter (only active in FI / FIG modes)
        if _inst_mode:
            et_inst = edge_type_inst(u, v, node_meta)
            if et_inst == "intra" and not show_intra_inst:
                continue
            if et_inst == "inter" and not show_inter_inst:
                continue
        # Stillingsgruppe-level filter
        if _grp_mode:
            et_grp = edge_type_grp(u, v, node_meta)
            if et_grp == "intra" and not show_intra_grp:
                continue
            if et_grp == "inter" and not show_inter_grp:
                continue
        
        if not (edge_min <= int(round(w)) <= edge_max):
            continue
        # Filter by sex_combo if edges are filtered
        if selected_gender_edges and sex_combo and sex_combo not in selected_gender_edges:
            continue
        edges_keep.append((u, v, w, sex_combo))

    # Remove isolated nodes (skip in sex-split modes so all nodes stay visible)
    all_nodes_pre_isolation = set(nodes_keep)
    connected = {n for u, v, *_ in edges_keep for n in (u, v)}
    isolated_nodes = all_nodes_pre_isolation - connected
    nodes_keep = nodes_keep - isolated_nodes
    edges_keep = [e for e in edges_keep if e[0] in nodes_keep and e[1] in nodes_keep]

    # Byg base-noder/-kanter til netværk, Samarbejdsmønstre og Nøgleaktører
    # (fjerner køns-/nationalitetsdimensionen fra noder og kanter)
    _net_mode = network_mode(mode)
    _needs_aggregation = (_net_mode != mode) or nat_in_mode(mode)

    if _needs_aggregation:
        _base_node_meta: dict = {}
        _base_node_map: dict  = {}

        _keep_fac  = "F" in _net_mode
        _keep_inst = "I" in _net_mode
        _keep_grp  = "G" in _net_mode
        _keep_sex  = sex_in_mode(_net_mode)
        _keep_nat  = nat_in_mode(mode)

        for nid, m in node_meta.items():
            parts = []
            if _keep_fac  and m.get("fac"):              parts.append(m["fac"])
            # Ved inst-niveau: inkludér altid fac som disambiguator selv hvis F ikke er i _net_mode
            elif _keep_inst and m.get("fac"):            parts.append(m["fac"])
            if _keep_inst and m.get("inst"):             parts.append(m["inst"])
            if _keep_grp  and m.get("grp"):              parts.append(m["grp"])
            if _keep_sex  and m.get("sex"):              parts.append(m["sex"])
            if _keep_nat  and m.get("statsborgerskab"):  parts.append(m["statsborgerskab"])
            base_nid = "|".join(parts) if parts else nid
            _base_node_map[nid] = base_nid

            if base_nid not in _base_node_meta:
                _base_node_meta[base_nid] = {k: v for k, v in m.items()}
                _base_node_meta[base_nid]["size"] = 0
                # Ryd dimensioner der ikke skal vises
                if not _keep_grp:  _base_node_meta[base_nid]["grp"] = ""
                if not _keep_inst: _base_node_meta[base_nid]["inst"] = ""
                if not _keep_sex:  _base_node_meta[base_nid].pop("sex", None)
                # Ret type
                _type_map = {
                    "fac_sex": "fac_sex" if _keep_sex else ("fac_nat" if _keep_nat else "fac"),
                    "inst_sex": ("inst_sex" if (_keep_sex and _keep_inst) else
                                 "fac_sex"  if _keep_sex else
                                 "fac_nat"  if (_keep_nat and not _keep_inst) else "inst"),
                    "grp_sex": ("inst_sex" if (_keep_inst and _keep_sex) else
                                "fac_sex"  if _keep_sex else
                                "fac_nat"  if (_keep_nat and not _keep_inst) else
                                "inst"     if _keep_inst else "fac"),
                    "grp":     ("grp_nat"  if (_keep_nat and _keep_grp) else
                                "inst_nat" if (_keep_nat and _keep_inst and not _keep_grp) else
                                "fac_nat"  if (_keep_nat and not _keep_inst and not _keep_grp) else
                                "inst"     if (_keep_inst and not _keep_grp) else "fac"),
                }
                _base_node_meta[base_nid]["type"] = _type_map.get(m.get("type",""), m.get("type",""))
                if _keep_nat:
                    _base_node_meta[base_nid]["statsborgerskab"] = m.get("statsborgerskab", "")
            _base_node_meta[base_nid]["size"] += m.get("size", 0)

        _base_nodes_keep = {_base_node_map[n] for n in nodes_keep    if n in _base_node_map}
        _base_isolated   = {_base_node_map[n] for n in isolated_nodes if n in _base_node_map} - _base_nodes_keep
        _base_edge_acc: dict = {}
        for u, v, w, *_ in edges_keep:
            bu, bv = _base_node_map.get(u, u), _base_node_map.get(v, v)
            if bu == bv:
                continue
            key = tuple(sorted((bu, bv)))
            _base_edge_acc[key] = _base_edge_acc.get(key, 0.0) + w
        _base_edges_keep = [(bu, bv, w) for (bu, bv), w in _base_edge_acc.items()]
        _base_mode = _net_mode
    else:
        _base_node_meta  = node_meta
        _base_node_map   = {n: n for n in node_meta}
        _base_nodes_keep = nodes_keep
        _base_isolated   = isolated_nodes
        _base_edges_keep = [(u, v, w) for u, v, w, *_ in edges_keep]
        _base_mode       = mode


    # -----------------------------------------------------------------------
    # Layout
    # -----------------------------------------------------------------------

    pos = compute_layout(
        _base_nodes_keep | _base_isolated, _base_node_meta, _base_mode,
        network_scale=st.session_state.get("network_scale_slider", 1200),
        n_selected_nats=len([c for c in (selected_citizenships or [])
                             if any(m.get("statsborgerskab") == c
                                    for m in _base_node_meta.values())]) or None,)
    
    nodes_keep = {nid for nid in nodes_keep if _base_node_map.get(nid, nid) in pos}
    edges_keep = [e for e in edges_keep if _base_node_map.get(e[0], e[0]) in pos and _base_node_map.get(e[1], e[1]) in pos]
    # Til PyVis: brug base-noder
    def _is_ghost(nid):
        m = _base_node_meta.get(nid, {})
        t = m.get("type", "")
        if m.get("size", 0) == 0:
            return True
        if t == "fac" and base_mode(_base_mode) in ("FIG", "FG", "IG", "G"):
            return True
        return False

    _pyvis_nodes    = {nid for nid in _base_nodes_keep if nid in pos and not _is_ghost(nid)}
    _pyvis_isolated = {nid for nid in _base_isolated   if nid in pos and not _is_ghost(nid)}
    # fac-noder er layoutankre, ikke visuelle noder, i alle modes uden F som primært niveau
    if base_mode(_base_mode) in ("FIG", "FG", "IG", "G"):
        _pyvis_nodes = {nid for nid in _pyvis_nodes
                        if _base_node_meta.get(nid, {}).get("type") != "fac"}


    # -----------------------------------------------------------------------
    # Build PyVis network
    # -----------------------------------------------------------------------

    net = Network(height="700px", width="100%", directed=False)
    net.toggle_physics(False)

    _node_display: dict = {}

    for nid in _pyvis_nodes:
        meta = _base_node_meta[nid]
        fac    = meta.get("fac", "")
        inst   = meta.get("inst", "")
        grp    = meta.get("grp", "")
        sex    = meta.get("sex", "")
        size   = meta.get("size", "NA")
        t      = meta.get("type", "grp")

        nat = meta.get("statsborgerskab", "")
        _bm = base_mode(_base_mode)
        _nx = nat_in_mode(_base_mode)
        parts = []
        if fac    and ("F" in _bm):           parts.append(fac)
        if inst   and ("I" in _bm):           parts.append(inst)
        if grp    and ("G" in _bm):           parts.append(grp)
        if nat    and _nx:                    parts.append(nat)
        title_text = " | ".join(parts) + f"\n Unikke forfattere: {size}"

        _bm = base_mode(_base_mode)
        _sx = sex_in_mode(_base_mode)
        _nx = nat_in_mode(_base_mode)
        if _bm == "G" and _nx:
            nats_all = sorted({_m.get("statsborgerskab","") for _m in _base_node_meta.values() if _m.get("statsborgerskab")})
            k     = max(1, len(nats_all))
            rank  = nats_all.index(nat) if nat in nats_all else 0
            lf    = 0.5 + 1.8 * (rank / max(1, k - 1))
            sf    = 1.3 - 0.7 * (rank / max(1, k - 1))
            color = adjust_color(grp_colors.get(grp, "#888888"), lf, sf)
            label = " | ".join(p for p in [grp, nat] if p)
        elif _bm == "G":
            color = grp_colors.get(grp, "#888888")
            label = grp + (f"\n({sex})" if _sx and sex else "")
        elif _bm in ("I", "FI") and _nx:
            insts_for_fac = sorted({_m["inst"] for _m in _base_node_meta.values() if _m.get("fac") == fac and _m.get("inst")})
            nats_for_inst = sorted({_m.get("statsborgerskab","") for _m in _base_node_meta.values() if _m.get("inst") == inst and _m.get("statsborgerskab")})
            rank_i = insts_for_fac.index(inst) if inst in insts_for_fac else 0
            rank_n = nats_for_inst.index(nat)   if nat  in nats_for_inst  else 0
            factor = 1.7 - 1.2 * (rank_i / max(1, len(insts_for_fac) - 1))
            lf     = 0.5 + 1.8 * (rank_n / max(1, len(nats_for_inst) - 1))
            sf     = 1.3 - 0.7 * (rank_n / max(1, len(nats_for_inst) - 1))
            base_c = adjust_color(faculty_base_colors.get(fac, "black"), factor)
            color  = adjust_color(base_c, lf, sf)
            label  = " | ".join(p for p in [inst, nat] if p)
        elif _bm in ("I", "FI") and _sx:
            label = inst + f"\n({sex})"
            insts_for_fac = sorted({m["inst"] for m in node_meta.values() if m.get("fac") == fac})
            rank   = insts_for_fac.index(inst) if inst in insts_for_fac else 0
            k      = len(insts_for_fac)
            factor = 1.7 - 1.2 * (rank / max(1, k - 1))
            base   = adjust_color(faculty_base_colors.get(fac, "black"), factor)
            color  = adjust_color(base, 1.3, 0.7) if sex == "m" else adjust_color(base, 0.75, 1.1)
        elif _bm in ("I", "FI"):
            label = inst
            insts_for_fac = sorted({m["inst"] for m in node_meta.values() if m["fac"] == fac})
            rank   = insts_for_fac.index(inst) if inst in insts_for_fac else 0
            k      = len(insts_for_fac)
            factor = 1.7 - 1.2 * (rank / max(1, k - 1))
            color  = adjust_color(faculty_base_colors.get(fac, "black"), factor)
        elif _bm == "F" and _nx:
            nat = meta.get("statsborgerskab", "")
            nats_for_fac = sorted({
                _m.get("statsborgerskab", "")
                for _m in _base_node_meta.values()
                if _m.get("fac") == fac and _m.get("statsborgerskab")
            })
            k      = max(1, len(nats_for_fac))
            rank   = nats_for_fac.index(nat) if nat in nats_for_fac else 0
            t_val  = rank / max(1, k - 1)
            lf     = 0.5 + 1.8 * t_val
            sf     = 1.3 - 0.7 * t_val
            base_c = faculty_base_colors.get(fac, "black")
            color  = adjust_color(base_c, lf, sf)
            label  = f"{fac} | {nat}"
        elif _bm == "F" and _nx:
            nats_for_fac = sorted({
                _m.get("statsborgerskab", "")
                for _m in _base_node_meta.values()
                if _m.get("fac") == fac and _m.get("statsborgerskab")
            })
            k      = max(1, len(nats_for_fac))
            rank   = nats_for_fac.index(nat) if nat in nats_for_fac else 0
            t_val  = rank / max(1, k - 1)
            lf     = 0.5 + 1.8 * t_val
            sf     = 1.3 - 0.7 * t_val
            base_c = faculty_base_colors.get(fac, "black")
            color  = adjust_color(base_c, lf, sf)
            label  = f"{fac} | {nat}"
        elif _bm == "F" and _sx:
            base  = faculty_base_colors.get(fac, "black")
            color = adjust_color(base, 1.3, 0.7) if sex == "m" else adjust_color(base, 0.75, 1.1)
            label = f"{fac}\n({'m' if sex == 'm' else 'k'})"
        elif _bm == "F":
            label = fac
            color = faculty_base_colors.get(fac, "black")
        else:
            base  = faculty_base_colors.get(fac, "black")
            lvl   = HIERARKI.get(grp, LVL_MIN)
            lf    = 1.2 + 1.2 * (1 - (lvl / max(HIERARKI.values())))
            sf    = 0.6 + 0.4 * (1 - (lvl / max(HIERARKI.values())))
            color = adjust_color(base, lf, sf)
            if _sx and inst and grp:
                label = f"{inst} | {grp}\n({sex})"
            elif _sx and inst:
                label = f"{inst}\n({sex})"
            elif _sx and grp:
                label = grp + f"\n({sex})"
            elif inst and grp:
                label = f"{inst}\n{grp}"
            else:
                label = inst or grp

        if nid not in pos:
            continue

        x, y    = pos[nid]
        size_px = scale_size_log(meta.get("size", 1), global_max_auth)
        net.add_node(nid, label=label, x=x, y=y, size=size_px,
                     color=color, title=title_text, physics=False, font="30px")
        _node_display[nid] = {"label": label, "color": color, "size_px": size_px, "title": title_text}
    
    for nid in _pyvis_isolated:
        if nid not in _base_node_meta or nid not in pos:
            continue
        meta  = _base_node_meta[nid]
        fac   = meta.get("fac", "")
        inst  = meta.get("inst", "")
        grp   = meta.get("grp", "")
        sex   = meta.get("sex", "")
        nat   = meta.get("statsborgerskab", "")
        size  = meta.get("size", "NA")

        # Recompute label + color the same way as the main loop
        _bm = base_mode(_base_mode)
        _sx = sex_in_mode(_base_mode)
        _nx = nat_in_mode(_base_mode)
        if _bm == "G" and _nx:
            color = grp_colors.get(grp, "#888888")
            label = " | ".join(p for p in [grp, nat] if p)
        elif _bm == "G":
            color = grp_colors.get(grp, "#888888")
            label = grp + (f"\n({sex})" if _sx and sex else "")
        elif _bm in ("I", "FI") and _nx:
            insts_for_fac = sorted({_m["inst"] for _m in _base_node_meta.values() if _m.get("fac") == fac and _m.get("inst")})
            rank_i = insts_for_fac.index(inst) if inst in insts_for_fac else 0
            factor = 1.7 - 1.2 * (rank_i / max(1, len(insts_for_fac) - 1))
            base_c = adjust_color(faculty_base_colors.get(fac, "black"), factor)
            color  = base_c
            label  = " | ".join(p for p in [inst, nat] if p)
        elif _bm in ("I", "FI") and _sx:
            label = inst + f"\n({sex})"
            insts_for_fac = sorted({m["inst"] for m in node_meta.values() if m.get("fac") == fac})
            rank   = insts_for_fac.index(inst) if inst in insts_for_fac else 0
            k      = len(insts_for_fac)
            factor = 1.7 - 1.2 * (rank / max(1, k - 1))
            base   = adjust_color(faculty_base_colors.get(fac, "black"), factor)
            color  = adjust_color(base, 1.3, 0.7) if sex == "m" else adjust_color(base, 0.75, 1.1)
        elif _bm in ("I", "FI"):
            label = inst
            insts_for_fac = sorted({m["inst"] for m in node_meta.values() if m.get("fac") == fac})
            rank   = insts_for_fac.index(inst) if inst in insts_for_fac else 0
            k      = len(insts_for_fac)
            factor = 1.7 - 1.2 * (rank / max(1, k - 1))
            color  = adjust_color(faculty_base_colors.get(fac, "black"), factor)
        elif _bm == "F" and _nx:
            nats_for_fac = sorted({
                _m.get("statsborgerskab", "")
                for _m in _base_node_meta.values()
                if _m.get("fac") == fac and _m.get("statsborgerskab")
            })
            k      = max(1, len(nats_for_fac))
            rank   = nats_for_fac.index(nat) if nat in nats_for_fac else 0
            t_val  = rank / max(1, k - 1)
            lf     = 0.5 + 1.8 * t_val
            sf     = 1.3 - 0.7 * t_val
            color  = adjust_color(faculty_base_colors.get(fac, "black"), lf, sf)
            label  = f"{fac} | {nat}"
        elif _bm == "F" and _sx:
            base  = faculty_base_colors.get(fac, "black")
            color = adjust_color(base, 1.3, 0.7) if sex == "m" else adjust_color(base, 0.75, 1.1)
            label = f"{fac}\n({'m' if sex == 'm' else 'k'})"
        elif _bm == "F":
            label = fac
            color = faculty_base_colors.get(fac, "black")
        else:
            base  = faculty_base_colors.get(fac, "black")
            lvl   = HIERARKI.get(grp, LVL_MIN)
            lf    = 1.2 + 1.2 * (1 - (lvl / max(HIERARKI.values())))
            sf    = 0.6 + 0.4 * (1 - (lvl / max(HIERARKI.values())))
            color = adjust_color(base, lf, sf)
            if _sx and inst and grp:
                label = f"{inst} | {grp}\n({sex})"
            elif _sx and inst:
                label = f"{inst}\n({sex})"
            elif _sx and grp:
                label = grp + f"\n({sex})"
            elif inst and grp:
                label = f"{inst}\n{grp}"
            else:
                label = inst or grp

        parts = []
        if fac_in_mode(mode) or base_mode(mode) in ("FI", "FIG", "IG", "I") and fac:
            parts.append(fac)
        if base_mode(mode) in ("FI", "FIG", "IG", "I") and inst:
            parts.append(inst)
        if grp_in_mode(mode) or base_mode(mode) in ("FG", "FIG", "IG") and grp:
            parts.append(grp)
        if sex and sex_in_mode(mode):
            parts.append(f"({sex})")
        title_text = " | ".join(parts) + f"\n Unikke forfattere: {size}\n Ingen forfatterpar i udsnittet"

        x, y    = pos[nid]
        size_px = scale_size_log(meta.get("size", 1), global_max_auth)
        net.add_node(
            nid,
            label=label,
            x=x, y=y,
            size=size_px * 0.7,
            color={
                "background": add_alpha(color, 0.3),
            },
            borderWidth=0,
            borderWidthSelected=3,
            title=title_text,
            physics=False,
            font="24px",
        )

    # Store per-node display info for reuse in ego-netværk tab
    #_node_display: dict = {}   # {nid: {"label": ..., "color": ..., "size_px": ...}}

    # -----------------------------------------------------------------------
    # Add edges to PyVis
    # -----------------------------------------------------------------------

    _pyvis_edges = [(u, v, w) for u, v, w in _base_edges_keep
                    if u in _pyvis_nodes and v in _pyvis_nodes]
    max_w_cur = max((w for u, v, w in _pyvis_edges), default=1.0)

    for u, v, w in _pyvis_edges:
        width = 6 * edge_scale * (w / max_w_cur)
        if base_mode(mode) == "G":
            col = "gray"
            et  = "group"
        else:
            fu = _base_node_meta.get(u, {}).get("fac", "")
            fv = _base_node_meta.get(v, {}).get("fac", "")
            et  = "intra" if fu == fv else "inter"
            col = "black" if et == "inter" else adjust_color(faculty_base_colors.get(fu, "#888888"), 0.25)

        net.add_edge(
            u, v, width=width,
            color={"color": add_alpha(col, 0.25), "highlight": add_alpha(col, 0.85), "hover": col},
            title=f"Forfatterpar: {int(w)} ({et})")

    # -----------------------------------------------------------------------
    # NetworkX graph for centrality
    # -----------------------------------------------------------------------

    G = nx.Graph()
    for nid in nodes_keep:
        G.add_node(nid)
    for u, v, w, *_sc in edges_keep:
        G.add_edge(u, v, weight=w)

    weighted_deg = dict(G.degree(weight="weight"))
    bet_cent     = nx.betweenness_centrality(G, weight="weight", normalized=True)

    grp_nodes  = [n for n, m in node_meta.items() if m.get("type") in ("grp", "grp_sex")]
    inst_nodes = [n for n, m in node_meta.items() if m.get("type") in ("inst", "inst_sex")]
    fac_nodes  = [n for n, m in node_meta.items() if m.get("type") in ("fac", "fac_sex")]

    # Beregn forfatterpar per org-enhed: hver par tælles én gang hos hvert endepunkt
    # (total sum = 2 × antal unikke par; andel beregnes mod total_pubs)
    fac_ew_cent: dict[str, float] = {}
    inst_ew_cent: dict[str, float] = {}
    grp_ew_cent: dict[str, float] = {}
    for u, v, w, *_ in edges_keep:
        for n in (u, v):
            m = node_meta.get(n, {})
            if m.get("fac"):  fac_ew_cent[m["fac"]]  = fac_ew_cent.get(m["fac"],  0.0) + w
            if m.get("inst"): inst_ew_cent[m["inst"]] = inst_ew_cent.get(m["inst"], 0.0) + w
            if m.get("grp"):  grp_ew_cent[m["grp"]]  = grp_ew_cent.get(m["grp"],  0.0) + w

    source_for_fac = fac_nodes or inst_nodes or grp_nodes
    faculty_wd_sorted = sorted(fac_ew_cent.items(), key=lambda x: -x[1])
    _, faculty_bs_sorted = aggregate_centrality_by(
        "fac", source_for_fac, node_meta, weighted_deg, bet_cent)

    nodes_for_inst = grp_nodes if mode in ("IG", "FIG") else (inst_nodes or grp_nodes)
    inst_wd_sorted = sorted(inst_ew_cent.items(), key=lambda x: -x[1])
    _, inst_bs_sorted = aggregate_centrality_by(
        "inst", nodes_for_inst, node_meta, weighted_deg, bet_cent)

    grp_wd_sorted = sorted(grp_ew_cent.items(), key=lambda x: -x[1])
    _, grp_bs_sorted = aggregate_centrality_by(
        "grp", grp_nodes, node_meta, weighted_deg, bet_cent)

    # Node-level compound labels for multi-level modes (FIG, FI, FG, IG …).
    # Label: "Stillingsgruppe @ Institut @ Fakultet" (parts omitted if empty).
    def _compound_label(m: dict) -> str:
        parts = [p for p in (m.get("grp",""), m.get("inst",""), m.get("fac","")) if p]
        return " | ".join(parts) if parts else "ukendt"

    if base_mode(mode) != "G" and grp_nodes:
        _nlabels  = {n: _compound_label(node_meta[n]) for n in grp_nodes}
        _wd_nodes = sorted(
            ((lbl, float(weighted_deg.get(n, 0)) / 2) for n, lbl in _nlabels.items()),
            key=lambda x: -x[1],
        )
        _bs_nodes = sorted(
            ((lbl, float(bet_cent.get(n, 0))) for n, lbl in _nlabels.items()),
            key=lambda x: -x[1],
        )
        grp_node_wd_sorted = _wd_nodes
        grp_node_bs_sorted = _bs_nodes
    else:
        grp_node_wd_sorted = []
        grp_node_bs_sorted = []


    # -----------------------------------------------------------------------
    # Size aggregates for tabs
    # -----------------------------------------------------------------------

    fac_sizes, inst_sizes, grp_sizes = {}, {}, {}
    _all_nodes_for_size = pre_nodes_for_mode(node_meta, mode)
    for nid in _all_nodes_for_size:
        m    = node_meta[nid]
        size = m.get("size", 0)
        fac  = m.get("fac")
        inst = m.get("inst")
        grp  = m.get("grp")
        t    = m.get("type")
        if t in ("grp", "grp_sex"):
            grp_sizes.setdefault(grp, []).append(size)
        if inst:
            inst_sizes.setdefault(inst, []).append(size)
        if fac:
            fac_sizes.setdefault(fac, []).append(size)

    fac_avg_size  = {f: sum(v) / len(v) for f, v in fac_sizes.items()}
    fac_tot_size  = {f: sum(v)           for f, v in fac_sizes.items()}
    inst_avg_size = {i: sum(v) / len(v) for i, v in inst_sizes.items()}
    inst_tot_size = {i: sum(v)           for i, v in inst_sizes.items()}
    grp_avg_size  = {g: sum(v) / len(v) for g, v in grp_sizes.items()}
    grp_tot_size  = {g: sum(v)           for g, v in grp_sizes.items()}

    # Edge-weight share per org-unit (each edge split evenly between endpoints)
    fac_ew: dict[str, float] = {}
    inst_ew: dict[str, float] = {}
    grp_ew: dict[str, float] = {}
    for u, v, w, *_ in edges_keep:
        for n in (u, v):
            m = node_meta.get(n, {})
            if m.get("fac"):  fac_ew[m["fac"]]  = fac_ew.get(m["fac"],  0.0) + w
            if m.get("inst"): inst_ew[m["inst"]] = inst_ew.get(m["inst"], 0.0) + w
            if m.get("grp"):  grp_ew[m["grp"]]  = grp_ew.get(m["grp"],  0.0) + w

    # -----------------------------------------------------------------------
    # Render PyVis
    # -----------------------------------------------------------------------
    def build_click_panel_html(nodes_keep: set, node_meta: dict, edges_keep: list,
                                node_display: dict = None) -> str:
        """Load the panel template and inject node data."""
        _node_info = {}
        for nid in nodes_keep:
            m = node_meta.get(nid, {})
            neighbours_acc = {}
            for u, v, w, *_ in edges_keep:
                if u == nid or v == nid:
                    partner = v if u == nid else u
                    pm = node_meta.get(partner, {})
                    et = "Intra" if node_meta.get(nid, {}).get("fac") == pm.get("fac") else "Inter"
                    partner_label = (node_display.get(partner, {}).get("label", "").replace("\n", " ")
                                     if node_display else "")
                    if not partner_label:
                        parts = [p for p in (pm.get("fac",""), pm.get("inst",""), pm.get("grp","")) if p]
                        partner_label = " | ".join(parts) if parts else partner
                    if partner_label in neighbours_acc:
                        neighbours_acc[partner_label]["pubs"] += int(w)
                    else:
                        neighbours_acc[partner_label] = {"label": partner_label, "pubs": int(w), "type": et}
            neighbours_data = sorted(neighbours_acc.values(), key=lambda x: -x["pubs"])
            node_label = (node_display.get(nid, {}).get("label", "").replace("\n", " ")
                          if node_display else "")
            if not node_label:
                parts = [p for p in (m.get("fac",""), m.get("inst",""), m.get("grp","")) if p]
                node_label = " | ".join(parts) if parts else nid
            _node_info[nid] = {
                "label":      node_label,
                "size":       m.get("size", 0),
                "neighbours": neighbours_data,
            }

        panel_path = Path(__file__).parent / "network_panel.html"
        template   = panel_path.read_text(encoding="utf-8")
        return template.replace(
            "__NODE_INFO_JSON__",
            json.dumps(_node_info, ensure_ascii=False),
        )

    temp_path = os.path.join(tempfile.gettempdir(), f"pyvis_{year}.html")
    net.write_html(temp_path)
    with open(temp_path, "r", encoding="utf-8") as f:
        html = f.read()

    click_panel = build_click_panel_html(_base_nodes_keep, _base_node_meta, _base_edges_keep, _node_display)
    html = html.replace("</body>", click_panel + "\n</body>")

    with st.expander("Sådan læser du netværket", expanded = False):
        st.markdown(
f"""
**Hvad viser cirklerne?** 

Hver cirkel repræsenterer en organisatorisk enhed på det valgte niveau. 
Cirklens størrelse afspejler antallet af unikke forfattere: jo større cirkel, jo flere unikke forfattere. 
Farverne følger fakultetstilhørsforholdet.

**Hvad viser linjerne?**

En linje mellem to cirkler betyder, at de har publiceret sammen. Linjetykkelsen afspejler antallet af
**forfatterpar** - jo tykkere linje, jo mere sampublicering. 
Sampublicering måles i **forfatterpar**, ikke i antal publikationer. Hermed vil en publikation med fire forfattere 
kunne resultere i seks forfatterpar.
""")
        if SVG_FORFATTERPAR:
                st.markdown(f'<div style="max-width:300px;">{SVG_FORFATTERPAR}</div>', unsafe_allow_html=True)
                st.caption("Illustration af et eksempel, hvor en publikation med fire forfattere (A, B, C, D) genererer seks forfatterpar.")

        st.markdown(f"""
**Hvad kan det her værktøj?**

Appen kortlægger sampubliceringsmønstre blandt KU's VIP-forforfattere - hvem der publicerer, hvor intensivt, 
og hvordan det udvikler sig over tid. Netværkvisningen herunder er udgangspunktet for analyserne;
de øvrige faner går i dybden med fakulteter, institutter, stillingsgrupper, køn og nationaliteter. 

**Hvordan læses netværket?**

Netværksvisningen er velegnet til at identificere mønstre i sampublicering - ikke til at vurdere enkeltpersoners eller
publikationers kvalitet. 

Når du læser figuren, kan det være brugbart at spørge: 

- Hvilke enheder fungere som **knudepunkter** med mange forbindelser?
- Er samarbejdet koncentreret omkring få enheder - eller relativt bredt fordelt? 
- Domineres netværket af samarbejde **inden for** eller **på tværs af** organisatoriske enheder? 

Netværket viser **relationer**, ikke årsager: tætte forbindelser kan afspejle både strategisk samarbejde, faglig afhængighed
eller strukturelle forhold.

**Tip:** Klik på en cirkel for at se dens detaljer og nærmeste samarbejdspartnere i panelet i højre hjørne.

**Hvordan tilpasser jeg netværket?**

Brug **sidepanelet til venstre** til at jusetere, hvad netværket viser:

- **Organisation** bestemmer detaljeniveauet - fakulteter giver overblik, institutter og stillingsgrupper går i dybden.
Filtrene afgrænser analyserne til specifikke enheder. 
- **Diversitet** tilføjer køns- og nationalitetsdimensioner. 
- **Netværksvisning** justerer størrelsesfiltre og den visuelle fremstilling. 

De øvrige faner **nedenunder netværket** uddyber analyserne med figurer og tabeller for hvert organisatoriske niveau.
""")

    st.components.v1.html(html, height=800, scrolling=True)

    # Show active-filter badges so the user knows the view is subsetted
    _cpr = {"m": "Mænd", "k": "Kvinder"}
    _combo_display = {"k-k": "Kvinde-Kvinde", "k-m": "Kvinde-Mand", "m-m": "Mand-Mand"}

    active_filters = []
    if selected_citizenships:
        active_filters.append(f"Statsborgerskab: {', '.join(selected_citizenships)}")
    if selected_genders:
        active_filters.append(f"Køn: {', '.join(_cpr.get(g, g) for g in selected_genders)}")
    if selected_gender_edges:
        active_filters.append(f"Kønskombination: {', '.join(_combo_display.get(c, c) for c in selected_gender_edges)}")


    # -----------------------------------------------------------------------
    # Summary stats
    # -----------------------------------------------------------------------
    _base_edges_keep = [(u, v, w) for u, v, w in _base_edges_keep
                        if u in _base_node_meta and v in _base_node_meta]

    _stat_edges = [(u, v, w) for u, v, w in _base_edges_keep
                   if u in _base_node_meta and v in _base_node_meta]
    total_pubs     = sum(w for _, _, w in _stat_edges)
    intra_pubs     = sum(w for u, v, w in _stat_edges if _base_mode != "G" and edge_type(u, v, _base_node_meta, _base_mode) == "intra")
    inter_pubs     = sum(w for u, v, w in _stat_edges if _base_mode != "G" and edge_type(u, v, _base_node_meta, _base_mode) == "inter")
    intra_grp_pubs = sum(w for u, v, w in _stat_edges if edge_type_grp(u, v, _base_node_meta) == "intra")
    inter_grp_pubs = sum(w for u, v, w in _stat_edges if edge_type_grp(u, v, _base_node_meta) == "inter")

    # -----------------------------------------------------------------------
    # Modularity
    # -----------------------------------------------------------------------

    comm_key = "fac" if base_mode(mode) in ("F", "FI", "FG", "FIG") else ("inst" if base_mode(mode) in ("I", "IG") else "grp")
    communities = []
    communities_dict = {}
    for nid in nodes_keep:
        group_label = node_meta[nid].get(comm_key, "")
        if not group_label:
            continue
        if group_label not in communities_dict:
            communities_dict[group_label] = []
        communities_dict[group_label].append(nid)

    communities = list(communities_dict.values())
    seen_groups = set(communities_dict.keys())

    G2 = nx.Graph()
    for u in _base_nodes_keep | _base_isolated:
        G2.add_node(u)
    for u, v, w in _base_edges_keep:
        G2.add_edge(u, v, weight=w)

    density = nx.density(G2)

    covered = {n for comm in communities for n in comm}
    for nid in G2.nodes():
        if nid not in covered:
            communities.append([nid])
    _connected = {n for n in G2.nodes() if G2.degree(n) > 0}
    G2_connected = G2.subgraph(_connected).copy()
    communities_filtered = [
        [n for n in c if n in _connected] for c in communities
    ]
    communities_filtered = [c for c in communities_filtered if c]
    _cf_covered = {n for c in communities_filtered for n in c}
    communities_filtered += [[n] for n in _connected if n not in _cf_covered]
    _cf_multi = [c for c in communities_filtered if len(c) >= 2]
    modularity_pre = (
        modq(G2_connected, communities_filtered, weight="weight")
        if len(_cf_multi) >= 2 and G2_connected.number_of_edges() > 0 else float("nan")
    )
    _G2_conn = G2.subgraph({n for n in G2.nodes() if G2.degree(n) > 0}).copy()
    if _G2_conn.number_of_edges() > 0 and _G2_conn.number_of_nodes() > 1:
        greedy_comms      = list(greedy_modularity_communities(_G2_conn, weight="weight"))
        _mod_gr           = modq(_G2_conn, greedy_comms, weight="weight")
        modularity_greedy = _mod_gr if _mod_gr >= 0 else float("nan")
        n_comms           = len(greedy_comms)
        greedy_comms_labeled = [
            sorted([
                " | ".join(p for p in (node_meta.get(nid, {}).get("fac", ""),
                                        node_meta.get(nid, {}).get("inst", ""),
                                        node_meta.get(nid, {}).get("grp", "")) if p)
                for nid in comm
            ])
            for comm in greedy_comms
        ]
    else:
        greedy_comms, modularity_greedy, n_comms = [], float("nan"), 0



    # -----------------------------------------------------------------------
    # Year comparison snapshots (lightweight - no network rendering)
    # -----------------------------------------------------------------------

    all_years_data = {}
    for _yr, _content in data_by_year.items():
        try:
            all_years_data[_yr] = compute_year_snapshot(
            _content, mode, _yr, raw_nodes,
            selected_genders, selected_gender_edges, selected_citizenships,
            selected_facs, selected_insts, selected_grps,
            show_fac, show_inst, show_grp,
            sampub_count_raw=_content.get("sampub_count", 0),
            ku_totals=ku_totals,
            collab_pairs=_content.get("collab_pairs", {}),
            show_intra=show_intra,
            show_inter=show_inter,
        )
        except Exception as e:
            st.warning(f"compute_year_snapshot fejlede for år {_yr}: {e}")
    
    _base_mode_for_oversigt = base_mode(mode)
    all_years_data_base = {}
    for _yr, _content in data_by_year.items():
        try:
            all_years_data_base[_yr] = compute_year_snapshot(
                _content, _base_mode_for_oversigt, _yr, raw_nodes,
                [], [], [],  # ingen køns-, kants- eller statsborgerfiltre
                selected_facs, selected_insts, selected_grps,
                show_fac, show_inst, show_grp,
                sampub_count_raw=_content.get("sampub_count", 0),
                ku_totals=ku_totals,
                collab_pairs=_content.get("collab_pairs", {}),
                show_intra=show_intra,
                show_inter=show_inter,
            )
        except Exception as e:
            st.warning(f"compute_year_snapshot (base) fejlede for år {_yr}: {e}")
    
            
    # -----------------------------------------------------------------------
    # TABS
    # -----------------------------------------------------------------------

    # Build tab list dynamically so diversity tabs appear whenever active,
    # regardless of the current org mode.
    #_base_tabs = {
        #"FS":   ["Oversigt", "Fakulteter", "Nøgleaktører", "Samarbejdsmønstre", "Datagrundlag"],
        #"IS":   ["Oversigt", "Institutter", "Nøgleaktører", "Samarbejdsmønstre", "Datagrundlag"],
        #"GS":   ["Oversigt", "Stillingsgrupper", "Nøgleaktører", "Samarbejdsmønstre", "Datagrundlag"],
        #"FIS":  ["Oversigt", "Fakulteter", "Institutter", "Nøgleaktører", "Samarbejdsmønstre", "Datagrundlag"],
        #"FIGS": ["Oversigt", "Fakulteter", "Institutter", "Stillingsgrupper", "Nøgleaktører", "Samarbejdsmønstre", "Datagrundlag"],
        #"FGS":  ["Oversigt", "Fakulteter", "Stillingsgrupper", "Nøgleaktører", "Samarbejdsmønstre", "Datagrundlag"],
        #"IGS":  ["Oversigt", "Institutter", "Stillingsgrupper", "Nøgleaktører", "Samarbejdsmønstre", "Datagrundlag"],
        #"F":   ["Oversigt", "Fakulteter", "Nøgleaktører", "Samarbejdsmønstre", "Datagrundlag"],
        #"FI":  ["Oversigt", "Fakulteter", "Institutter", "Nøgleaktører", "Samarbejdsmønstre", "Datagrundlag"],
        #"FIG": ["Oversigt", "Fakulteter", "Institutter", "Stillingsgrupper", "Nøgleaktører", "Samarbejdsmønstre", "Datagrundlag"],
        #"FG":  ["Oversigt", "Fakulteter", "Stillingsgrupper", "Nøgleaktører", "Samarbejdsmønstre", "Datagrundlag"],
        #"IG":  ["Oversigt", "Institutter", "Stillingsgrupper", "Nøgleaktører", "Samarbejdsmønstre", "Datagrundlag"],
        #"I":   ["Oversigt", "Institutter", "Nøgleaktører", "Samarbejdsmønstre", "Datagrundlag"],
        #"G":   ["Oversigt", "Stillingsgrupper", "Nøgleaktører", "Samarbejdsmønstre", "Datagrundlag"],
    #}
    _base_tabs = {
        "FS":   ["Oversigt", "Datagrundlag"],
        "IS":   ["Oversigt", "Datagrundlag"],
        "GS":   ["Oversigt", "Datagrundlag"],
        "FIS":  ["Oversigt", "Datagrundlag"],
        "FIGS": ["Oversigt", "Datagrundlag"],
        "FGS":  ["Oversigt", "Datagrundlag"],
        "IGS":  ["Oversigt", "Datagrundlag"],
        "FN":   ["Oversigt", "Datagrundlag"],
        "IN":   ["Oversigt", "Datagrundlag"],
        "GN":   ["Oversigt", "Datagrundlag"],
        "FIN":  ["Oversigt", "Datagrundlag"],
        "FIGN": ["Oversigt", "Datagrundlag"],
        "FGN":  ["Oversigt", "Datagrundlag"],
        "IGN":  ["Oversigt", "Datagrundlag"],
        "F":   ["Oversigt", "Fakulteter", "Nøgleaktører", "Samarbejdsmønstre", "Netværksudvikling", "Datagrundlag"],
        "FI":  ["Oversigt", "Fakulteter", "Institutter", "Nøgleaktører", "Samarbejdsmønstre", "Netværksudvikling", "Datagrundlag"],
        "FIG": ["Oversigt", "Fakulteter", "Institutter", "Stillingsgrupper", "Nøgleaktører", "Samarbejdsmønstre", "Netværksudvikling", "Datagrundlag"],
        "FG":  ["Oversigt", "Fakulteter", "Stillingsgrupper", "Nøgleaktører", "Samarbejdsmønstre", "Netværksudvikling", "Datagrundlag"],
        "IG":  ["Oversigt", "Institutter", "Stillingsgrupper", "Nøgleaktører", "Samarbejdsmønstre", "Netværksudvikling", "Datagrundlag"],
        "I":   ["Oversigt", "Institutter", "Nøgleaktører", "Samarbejdsmønstre", "Netværksudvikling", "Datagrundlag"],
        "G":   ["Oversigt", "Stillingsgrupper", "Nøgleaktører", "Samarbejdsmønstre", "Netværksudvikling", "Datagrundlag"],
    }
    tabs_by_mode = {}
    for _m, _tl in _base_tabs.items():
        _extra_div    = []
        _extra_openalex = []
        if analyse_køn:
            _extra_div.append("Køn")
        if analyse_nat:
            _extra_div.append("Nationaliteter")
        if analyse_internationalt:
            _extra_openalex.append("Internationalt samarbejde")
        if analyse_fwci:
            _extra_openalex.append("FWCI")
        if analyse_output:
            _extra_openalex.append("Forskningsoutput")
        # Diversitetsfaner indsættes før Nøgleaktører
        _ins = _tl.index("Datagrundlag")
        # OpenAlex-faner indsættes før Datagrundlag
        _ins_data = _tl.index("Datagrundlag")
        _tl2 = _tl[:_ins] + _extra_div + _tl[_ins:_ins_data] + _extra_openalex + _tl[_ins_data:]
        tabs_by_mode[_m] = _tl2
    tabs_to_show = tabs_by_mode.get(mode) or tabs_by_mode.get(base_mode(mode), ["Basisstatistik"])
    tabs         = st.tabs(tabs_to_show)
    tabs_dict    = {name: tab for name, tab in zip(tabs_to_show, tabs)}
    
    _filter_caption = filter_status_caption(
        mode, show_intra, show_inter,
        show_intra_inst, show_inter_inst,
        show_intra_grp, show_inter_grp,
        show_fac, show_inst, show_grp,
    )

    with tabs_dict["Oversigt"]:
        _snap_base = all_years_data_base.get(year, {})
        render_tab_oversigt(year, mode, _base_edges_keep,
                    _snap_base.get("total_pubs", total_pubs),
                    _snap_base.get("intra_pubs", intra_pubs),
                    _snap_base.get("inter_pubs", inter_pubs),
                    _base_node_meta, *intra_inter_labels(_base_mode),
                    all_years_data=all_years_data_base,
                    isolated_nodes=_base_isolated,
                    total_authors=_snap_base.get("total_authors", sum(_base_node_meta[n].get("size", 0) for n in _base_nodes_keep)),
                    sampub_count=content.get("sampub_count", 0),
                    intra_grp_pubs=_snap_base.get("intra_grp_pubs", intra_grp_pubs),
                    inter_grp_pubs=_snap_base.get("inter_grp_pubs", inter_grp_pubs),
                    isolated_units_base=_snap_base.get("isolated_units"),
                    intra_inst_pubs=_snap_base.get("intra_inst_pubs"),
                    inter_inst_pubs=_snap_base.get("inter_inst_pubs"),
                    filter_caption=_filter_caption
                    )

    if "Fakulteter" in tabs_dict:
        with tabs_dict["Fakulteter"]:
            render_tab_fakulteter(year, mode, fac_tot_size, fac_avg_size, edges_keep, node_meta, all_years_data=all_years_data, fac_ew=fac_ew, faculty_base_colors=faculty_base_colors, total_pubs=total_pubs, filter_caption=_filter_caption, show_intra=show_intra)

    if "Institutter" in tabs_dict:
        with tabs_dict["Institutter"]:
            render_tab_institutter(year, mode, inst_tot_size, inst_avg_size, institut_fakultets_map, edges_keep, node_meta, all_years_data=all_years_data, inst_ew=inst_ew, faculty_base_colors=faculty_base_colors, total_pubs=total_pubs, filter_caption=_filter_caption, show_intra=show_intra)

    if "Stillingsgrupper" in tabs_dict:
        with tabs_dict["Stillingsgrupper"]:
            render_tab_stillingsgrupper(year, mode, all_groups, grp_tot_size, grp_avg_size,
                            selected_facs, selected_insts, selected_grps, 
                            selected_genders, selected_citizenships,
                            all_years_data=all_years_data, grp_ew=grp_ew,
                            edges_keep=edges_keep, node_meta=node_meta, 
                            forfatterpositioner=forfatterpositioner,
                            inst_to_fac=inst_to_fac, total_pubs=total_pubs,
                            filter_caption=_filter_caption, show_intra=show_intra, show_inter=show_inter)

    if "Nøgleaktører" in tabs_dict:
        with tabs_dict["Nøgleaktører"]:
            render_tab_centralitet(year, _base_mode, faculty_wd_sorted, faculty_bs_sorted,
                       inst_wd_sorted, inst_bs_sorted,
                       grp_wd_sorted, grp_bs_sorted, _base_node_meta,
                       grp_node_wd_sorted, grp_node_bs_sorted,
                       faculty_base_colors=faculty_base_colors,
                       grp_colors=grp_colors,
                       all_years_data=all_years_data,
                       svg_centralitet=SVG_CENTRALITET, 
                       filter_caption=_filter_caption)

    if "Samarbejdsmønstre" in tabs_dict:
        with tabs_dict["Samarbejdsmønstre"]:
            render_tab_netvaerksstruktur(year, _base_mode, density, modularity_pre, modularity_greedy,
                             n_comms, communities_dict, greedy_comms, comm_key,
                             _base_edges_keep, _base_node_meta, all_years_data=all_years_data, 
                       filter_caption=_filter_caption)

    if "Køn" in tabs_dict:
        with tabs_dict["Køn"]:
            render_tab_køn(year, mode, raw_nodes, raw_edges, node_meta,
                           selected_facs, selected_insts, selected_grps, all_years_data=all_years_data,
                           edges_keep=edges_keep, forfatterpositioner=forfatterpositioner,
                           sex_pairs=content.get("sex_pairs", {}),
                           filter_caption=_filter_caption, show_intra=show_intra)

    if "Nationaliteter" in tabs_dict:
        with tabs_dict["Nationaliteter"]:
            render_tab_nationaliteter(year, mode, raw_nodes, raw_edges, node_meta,
                                      selected_facs, selected_insts, selected_grps,
                                      raw_nodes_unfiltered, raw_edges_unfiltered,
                                      all_years_data,
                                      cs_pairs=content.get("cs_pairs", {}),
                                      cs_nat_pairs=content.get("cs_nat_pairs", {}),
                                      raw_nodes_pre_cs=raw_nodes_pre_cs,
                                      selected_citizenships=selected_citizenships,
                                      filter_caption=_filter_caption, show_intra=show_intra, edges_keep=edges_keep)

    if "Internationalt samarbejde" in tabs_dict:
        with tabs_dict["Internationalt samarbejde"]:
            st.subheader("Internationalt samarbejde")
            st.error(
"""**Under opbygning.** Denne fane vil vise KU-forskernes samarbejde med forfattere fra institutioner uden for KU, 
baseret på affilieringsdata fra CURIS og OpenAlex (afhænger af valgte datakilde). 
Det adskiller sig fra Nationaliteter-fanen, som viser statsborgerskabet hos KU-ansatte - her handler det om, 
*hvem* KU samarbejder med eksternt.""",
                icon="🔨",
            )

    if "FWCI" in tabs_dict:
        with tabs_dict["FWCI"]:
            st.subheader("FWCI - citationsimpact")
            st.error(
"""**Under opbygning.** Field-Weighted Citation Impact (FWCI) måler, hvor meget en publikation citeres relativt til det 
globale gennemsnit inden for samme fagområde, publikationstype og år. En FWCI over 1,0 indikerer, at publikationen 
citeres mere end gennemsnittet. Data hentes fra OpenAlex.""",
                icon="🔨",
            )

    if "Forskningsoutput" in tabs_dict:
        with tabs_dict["Forskningsoutput"]:
            st.subheader("Forskningsoutput")
            st.error(
"""**Under opbygning.** Denne fane vil undersøge, hvordan publikationsvolumenen ved brug af LLM kan kvantificeres i 
forskningsområdet. Data hentes i første omgang fra OpenAlex - og suppleres senere med CURIS.""",
                icon="🔨",
            )
    
    if "Netværksudvikling" in tabs_dict:
        with tabs_dict["Netværksudvikling"]:
            render_tab_netværksudvikling(year, _base_mode, all_years_data=all_years_data, filter_caption=_filter_caption)

    if "Datagrundlag" in tabs_dict:
        with tabs_dict["Datagrundlag"]:
            render_tab_datagrundlag(year, mode, all_groups, selected_facs, selected_insts, selected_grps,
                                    forfatterantal=forfatterantal_data,
                                    publikationstyper=publikationstyper_data,
                                    faculty_base_colors=faculty_base_colors,
                                    years_sorted=sorted(all_years_data.keys()),
                                    pubtype_map=load_pubtype_map(),
                                    all_years_data=all_years_data, forfatterantal_dist=forfatterantal_dist)



    st.markdown(f"""
<hr style="margin-top: 50px;">
<div style="text-align:center; color:#666; font-size: 0.9em;">
  REKSTAB Analyse · Amanda Schramm Petersen · <a href="mailto:ascp@adm.ku.dk">ascp@adm.ku.dk</a>
  · opdateret {_DEPLOY_DATE}
</div>
""", unsafe_allow_html=True)



def _render_year_comparison(all_years_data: dict, series: list, title: str,
                             yaxis_label: str = "Forfatterantal",
                             key_suffix: str = "",
                             colors = None,
                             description: str = None,
                             show_table: bool = True,
                             height: int = 380):
    """Render a year-comparison line chart inside a collapsible expander.

    Parameters
    ----------
    all_years_data : dict  {year: snapshot_dict}
    series : list of (label, data_key, sub_key_or_None)
        Each entry describes one line:
        - label: legend name
        - data_key: top-level key in snapshot ("fac_tot", "grp_tot", etc.)
        - sub_key: the specific org-unit name (str) or None for scalar keys
                   like "total_pubs"
    title : str  - expander title
    """
    if not all_years_data or len(all_years_data) < 2:
        return

    years_sorted = sorted(all_years_data.keys())
    if description:
        st.markdown(f"""{description}""")
    
    ku_colors = ku_color_sequence(len(series))

    fig = go.Figure()
    for i, (label, data_key, sub_key) in enumerate(series):
        y_vals = []
        for yr in years_sorted:
            snap = all_years_data.get(yr, {})
            if sub_key is None:
                y_vals.append(snap.get(data_key, 0))
            else:
                y_vals.append(snap.get(data_key, {}).get(sub_key, 0))
        _color = (colors.get(label) if colors else None) or ku_colors[i]

        fig.add_trace(go.Scatter(
            x=years_sorted, y=y_vals, name=label,
            mode="lines+markers",
            line=dict(width=2, **({"color": _color} if _color else {})),
            marker=dict(size=7, **( {"color": _color} if _color else {})),
        ))
    fig.update_layout(
        yaxis_title=yaxis_label,
        xaxis=dict(tickmode="array", tickvals=years_sorted, dtick=1),
        height=height,
        legend_title="",
        title=dict(text=title) if title else None,
        legend=dict(orientation="v", yanchor="top", y=1, xanchor="left", x=1.02),
        margin=dict(t=40 if title else 20, b=20, r=200),
    )
    st.plotly_chart(fig, width='stretch', key=f"yc_{key_suffix}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
    )
    

    # ── Downloadable table ────────────────────────────────────────────────────
    _tbl_rows = []
    for label, data_key, sub_key in series:
        row = {"Enhed": label}
        for yr in years_sorted:
            snap = all_years_data.get(yr, {})
            val = snap.get(data_key, {}).get(sub_key, 0) if sub_key else snap.get(data_key, 0)
            row[str(yr)] = round(val, 1)
        _tbl_rows.append(row)

    _tbl_schema = (
        [("Enhed", pa.string())] +
        [(str(yr), pa.float64()) for yr in years_sorted]
    )
    if show_table: 
        with st.expander("Se tabel"):
            st.dataframe(build_table(_tbl_rows, _tbl_schema), hide_index=True, width="stretch")
            _key = hashlib.md5(f"{title}{key_suffix}{','.join(str(s) for s in series)}".encode()).hexdigest()[:10]
            st.download_button(
                "Download (.xlsx)",
                data=rows_to_excel_bytes(_tbl_rows, [n for n, _ in _tbl_schema]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_yrcmp_{_key}"
            )

# ===========================================================================
# TAB RENDERERS  (unchanged except minor mode checks for new sex modes)
# ===========================================================================

def _render_org_bar(edges_keep: list, node_meta: dict, org_key: str, title_label: str,
                    color_map: dict = None, size_map: dict = None):

    _totals: dict = {}
    for u, v, w, *_ in edges_keep:
        for n in (u, v):
            k = node_meta.get(n, {}).get(org_key, "")
            if k:
                _totals[k] = _totals.get(k, 0) + w

    if not _totals:
        return

    _sorted = sorted(_totals.items(), key=lambda x: -x[1])
    _grand  = sum(v for _, v in _sorted) or 1
    _colors = [color_map.get(k, "#122947") if color_map else "#122947" for k, _ in _sorted]

    _tab_abs, _tab_ratio = st.tabs(["Forfatterpar", "Forfatterpar per forfatter"])

    with _tab_abs:
        _fig = go.Figure(go.Bar(
            y=[x[0] for x in _sorted],
            x=[x[1] for x in _sorted],
            orientation="h",
            marker_color=_colors,
            text=[f"{fmt_ui(x[1])}  ({fmt_ui(100*x[1]/_grand)}%)" for x in _sorted],
            textposition="inside",
        ))
        _fig.update_layout(
            xaxis_title="Forfatterpar (fordelt)",
            yaxis_title=title_label,
            height=max(350, 35 * len(_sorted)),
            margin=dict(l=160, t=50, r=80),
            yaxis=dict(autorange="reversed"),
        )
        st.plotly_chart(_fig, width='stretch',
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
    )

    with _tab_ratio:
        if not size_map:
            st.caption("Forfatterantal ikke tilgængeligt for dette niveau.")
        else:
            _ratio_sorted = sorted(
                [(k, v / size_map[k]) for k, v in _totals.items() if size_map.get(k)],
                key=lambda x: -x[1],
            )
            if not _ratio_sorted:
                st.caption("Ingen data.")
            else:
                _ratio_colors = [color_map.get(k, "#122947") if color_map else "#122947" for k, _ in _ratio_sorted]
                _fig_r = go.Figure(go.Bar(
                    y=[x[0] for x in _ratio_sorted],
                    x=[x[1] for x in _ratio_sorted],
                    orientation="h",
                    marker_color=_ratio_colors,
                    text=[f"{fmt_ui(x[1],3)}" for x in _ratio_sorted],
                    textposition="inside",
                ))
                _fig_r.update_layout(
                    xaxis_title="Forfatterpar per forfatter",
                    yaxis_title=title_label,
                    height=max(350, 35 * len(_ratio_sorted)),
                    margin=dict(l=160, t=50, r=80),
                    yaxis=dict(autorange="reversed"),
                )
                st.plotly_chart(_fig_r, width='stretch',
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
        )

    # ── Combined table for both tabs ──────────────────────────────────────────
    _ratio_map = (
        {k: v / size_map[k] for k, v in _totals.items() if size_map and size_map.get(k)}
        if size_map else {}
    )
    _combined_rows = [
        {
            title_label:                       k,
            "Forfatterpar (fordelt)":              round(v, 1),
            "Andel (%)":                       round(100 * v / _grand, 1),
            "Forfatterantal":                 size_map.get(k, 0) if size_map else None,
            "Forfatterpar per forfatter":    round(_ratio_map[k], 4) if k in _ratio_map else None,
        }
        for k, v in _sorted
    ]
    _combined_schema = [
        (title_label,                       pa.string()),
        ("Forfatterpar (fordelt)",              pa.float64()),
        ("Andel (%)",                       pa.float64()),
    ]
    if size_map:
        _combined_schema += [
            ("Forfatterantal",                 pa.int64()),
            ("Forfatterpar per forfatter",    pa.float64()),
        ]
    else:
        # Drop the None columns if no size_map
        _combined_rows = [{k: v for k, v in r.items()
                           if k not in ("Forfatterantal", "Forfatterpar per forfatter")}
                          for r in _combined_rows]

    _key = hashlib.md5(f"{title_label}_{org_key}".encode()).hexdigest()[:10]
    with st.expander("Se tabel"):
        st.dataframe(build_table(_combined_rows, _combined_schema), hide_index=True, width="stretch")
        st.download_button(
            "Download (.xlsx)",
            data=rows_to_excel_bytes(_combined_rows, [n for n, _ in _combined_schema]),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"org_comb"
        )

        

def render_tab_oversigt(year, mode, edges_keep, total_pubs, intra_pubs, inter_pubs,
                        node_meta, intra_label="intra-fakultet", inter_label="inter-fakultet",
                        all_years_data=None, isolated_nodes=None, total_authors=0,
                        sampub_count=0, intra_grp_pubs=0, inter_grp_pubs=0,
                        isolated_units_base=None,
                        intra_inst_pubs=None, inter_inst_pubs=None, filter_caption=None, show_intra=True
                        ):
    _mode_labels = {
        "F":   "fakulteter",
        "FI":  "fakulteter og institutter",
        "FIG": "fakulteter, institutter og stillingsgrupper",
        "FG":  "fakulteter og stillingsgrupper",
        "I":   "institutter",
        "IG":  "institutter og stillingsgrupper",
        "G":   "stillingsgrupper",
        "FS":  "fakulteter opdelt på køn",
        "IS":  "institutter opdelt på køn",
        "GS":  "stillingsgrupper opdelt på køn",
    }

    _grp_modes = {"FG", "FIG", "IG", "G", "GS"}
    _pct_intra     = round(100 * intra_pubs     / total_pubs, 1) if total_pubs else 0.0
    _pct_inter     = round(100 * inter_pubs     / total_pubs, 1) if total_pubs else 0.0
    _pct_intra_grp = round(100 * intra_grp_pubs / total_pubs, 1) if total_pubs else 0.0
    _filter_key    = f"{int(intra_pubs)}_{int(inter_pubs)}_{int(total_pubs)}"
    _has_grp_donut = base_mode(mode) in {"FG", "FIG", "IG", "G"} and total_pubs > 0
    _ys            = sorted(all_years_data.keys()) if all_years_data else []

    if base_mode(mode) in ("FI", "FIG"):
        _intra_inst_pubs = intra_inst_pubs if intra_inst_pubs is not None else sum(
            w for u, v, w, *_ in edges_keep if edge_type_inst(u, v, node_meta) == "intra")
        _inter_inst_pubs = inter_inst_pubs if inter_inst_pubs is not None else sum(
            w for u, v, w, *_ in edges_keep if edge_type_inst(u, v, node_meta) == "inter")
    else:
        _intra_inst_pubs = _inter_inst_pubs = 0

    st.subheader("Overblik over sampubliceringsaktivitet")
    
    if "S" in mode or "N" in mode: 
        st.error("Analyserne nedenfor er uden nogen diversitetsdimensioner.")
    
    st.markdown(
f"""
Fanen giver et samlet overblik over KU's sampubliceringsaktivitet i **{year}**, opgjort på 
**{_mode_labels.get(mode, mode)}**. 

Oversigten beskriver **omfang og fordeling** af sampublicering på tværs af
organisatoriske enheder, herunder hvor mange forfattere der indgår i samarbejde,
hvor intensive samarbejderne er, og hvordan sampublicering fordeler sig mellem
enheder.

Fanen er bevidst deskriptiv; den viser, **hvor meget** og **hvordan** der
sampubliceres, men **ikke hvordan samarbejdet er organiseret strukturelt**.
Oversigten fungerer derfor som et analytisk indgangspunkt til de øvrige faner,
hvor netværksstruktur, roller og dynamikker analyseres mere detaljeret.

Opgørelserne i denne fane bygger på tre centrale begreber:

- **Sampublikationer** er publikationer med mindst to KU‑VIP‑forfattere.
- **Forfatterantal** er antallet af unikke KU‑VIP‑forskere knyttet til enheden.
  Hver forsker tælles én gang, uanset antal publikationer og uanset om der også
  indgår soloartikler.
- **Forfatterpar** tæller hver unik kombination af to KU‑VIP-forfattere på én publikation.
  En publikation med fire forfattere genererer derfor seks forfatterpar.
""")

    with st.expander("Læsning af procenter"):
        st.markdown("""
Når forfatterpar opgøres per enhed, tælles hvert forfatterpar én gang hos
hvert af de to endepunkter. Et samarbejde mellem to forskellige enheder
indgår derfor i begge enheders opgørelse.

Det betyder, at procentandele på tværs af enheder **kan summere til mere end 100 %**.
Det er et bevidst valg, fordi det gør det muligt at aflæse:
*“enhed X indgår i Y % af alle forfatterpar i KU‑netværket”* direkte fra figurerne,
uden halve værdier eller korrektioner.
""")

    st.markdown(
"""Heraf beregnes **sampubliceringsraten**, som er antallet af forfatterpar divideret
med antallet af forfattere. Raten siger noget om **samarbejdsintensitet** relativt
til enhedens størrelse.
""")


    # ── 1. Nøgletal ──────────────────────────────────────────────────────────
    col1, col2, col3, col4, col5, col6 = st.columns(6)
    with col1:
        st.metric("Sampublikationer (ufiltreret)", int(sampub_count),
                  help="Antal publikationer med mindst to VIP-forfattere")
    with col2:
        st.metric("Unikke forfattere", int(total_authors),
                  help="Antal unikke KU-VIP-forskere i udsnittet (hver forsker tælles én gang, uanset antal publikationer).")
    with col3:
        _rate = round(total_pubs / total_authors, 2) if total_authors else 0.0
        st.metric("Sampubliceringsrate", _rate,
                  help="Forfatterpar divideret med unikke forfattere")
    
    with col4:
        st.metric("Forfatterpar (netværk)", int(total_pubs),
                  help="Forfatterpar mellem forskellige organisatoriske enheder - som vist i netværket")
    with col5:
        _all_pubs = int(all_years_data[year].get("intra_pubs_all", 0) + all_years_data[year].get("inter_pubs_all", 0)) \
                    if all_years_data and year in all_years_data else "-"
        st.metric("Forfatterpar (inkl. intra-enhed)", _all_pubs,
                  help="Alle forfatterpar inkl. par inden for samme organisatoriske enhed")

    if isolated_nodes is not None:
        n_iso = len(isolated_units_base) if isolated_units_base is not None else len(isolated_nodes)
        with col6:
            st.metric("Isolerede noder", n_iso,
                      help="Noder uden forfatterpar i det valgte udsnit")
        if n_iso > 0:
            with st.expander("Se isolerede noder"):
                _iso_rows = []
                if isolated_units_base is not None:
                    # Byg lookup fra label → size via node_meta
                    _label_to_size = {}
                    for _nid, _m in node_meta.items():
                        _parts = [p for p in (_m.get("fac",""), _m.get("inst",""), _m.get("grp","")) if p]
                        _lbl = " | ".join(_parts)
                        if _lbl:
                            _label_to_size[_lbl] = _label_to_size.get(_lbl, 0) + _m.get("size", 0)
                    for label in sorted(isolated_units_base):
                        parts = [p.strip() for p in label.split("|")]
                        _iso_rows.append({
                            "Fakultet":        parts[0] if len(parts) > 0 else "",
                            "Institut":        parts[1] if len(parts) > 1 else "",
                            "Stillingsgruppe": parts[2] if len(parts) > 2 else "",
                            "Forfattere": _label_to_size.get(label, 0),
                        })
                else:
                    for nid in sorted(isolated_nodes):
                        m = node_meta.get(nid, {})
                        _iso_rows.append({
                            "Fakultet":        m.get("fac", ""),
                            "Institut":        m.get("inst", ""),
                            "Stillingsgruppe": m.get("grp", ""),
                            "Forfattere": m.get("size", 0),
                        })
                _iso_rows.sort(key=lambda r: (-r["Forfattere"], r["Fakultet"]))
                _iso_schema = [
                    ("Fakultet",        pa.string()),
                    ("Institut",        pa.string()),
                    ("Stillingsgruppe", pa.string()),
                    ("Forfattere", pa.int64()),
                ]
                st.dataframe(build_table(_iso_rows, _iso_schema), hide_index=True, width="stretch")
                st.download_button(
                    "Download (.xlsx)",
                    data=rows_to_excel_bytes(_iso_rows, [n for n, _ in _iso_schema]),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"iso_rows_{year}_{mode}"
                )
    
    # ── 2. Donuts ─────────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown(
f"""#### Hvor foregår samarbejdet i {year}?
Figurerne nedenfor opdeler sampubliceringen i **intra‑** og **inter‑samarbejde**,
afhængigt af det valgte organisatoriske niveau.

- *Intra* betegner samarbejde inden for samme organisatoriske enhed.
- *Inter* betegner samarbejde på tværs af enheder.

Som metodisk udgangspunkt vises netværket kun med forfatterpar **på tværs af**
organisatoriske enheder. Forfatterpar inden for samme enhed kan ikke vises i netværket, 
da en enhed ikke kan have en forbindelseslinje til sig selv.

Disse interne forfatterpar kan dog indgå i de samlede opgørelser og kan tilføjes
via togglen nedenfor for at vise den fulde sampubliceringsaktivitet.
""")

    _use_all_donut = st.toggle(
        "Inkludér intra-enhed forfatterpar (f.eks. to professorer på samme institut)",
        value=False,
        key=f"donut_all_{year}_{mode}",
        help="Til: alle forfatterpar inkl. par inden for samme enhed (f.eks. to professorer på samme institut). Fra: kun netværkskanter.",
        disabled=not show_intra
    )

    if _use_all_donut and all_years_data and year in all_years_data:
        _snap = all_years_data[year]
        _intra_pubs_plot = _snap.get("intra_pubs_all",      intra_pubs)
        _inter_pubs_plot = _snap.get("inter_pubs_all",      inter_pubs)
        _intra_inst_plot = _snap.get("intra_inst_pubs_all", _intra_inst_pubs)
        _inter_inst_plot = _snap.get("inter_inst_pubs_all", _inter_inst_pubs)
        _intra_grp_plot  = _snap.get("intra_grp_pubs_all",  intra_grp_pubs)
        _inter_grp_plot  = _snap.get("inter_grp_pubs_all",  inter_grp_pubs)
        st.caption("Alle forfatterpar vises - inkl. par inden for samme enhed. Disse kan ikke ses i netværksvisningen.")
    else:
        _intra_pubs_plot = intra_pubs
        _inter_pubs_plot = inter_pubs
        _intra_inst_plot = _intra_inst_pubs
        _inter_inst_plot = _inter_inst_pubs
        _intra_grp_plot  = intra_grp_pubs
        _inter_grp_plot  = inter_grp_pubs
        st.caption("Kun forfatterpar *mellem* forskellige organisatoriske enheder - som vist i netværket.")

    _total_plot         = (_intra_pubs_plot + _inter_pubs_plot) or 1
    _pct_intra_plot     = round(100 * _intra_pubs_plot / _total_plot, 1)
    _pct_inter_plot     = round(100 * _inter_pubs_plot / _total_plot, 1)
    _pct_intra_grp_plot = round(100 * _intra_grp_plot  / _total_plot, 1)
    _inst_tot_plot      = (_intra_inst_plot + _inter_inst_plot) or 1
    _pct_intra_inst_plot = round(100 * _intra_inst_plot / _inst_tot_plot, 1)


    _filter_key = f"{int(_intra_pubs_plot)}_{int(_inter_pubs_plot)}"

    if total_pubs > 0 and base_mode(mode) != "G":
        _show_inst_donut = base_mode(mode) in ("FI", "FIG")
        _show_grp_donut  = _has_grp_donut
        _n_donuts = 1 + int(_show_inst_donut) + int(_show_grp_donut)

        if _n_donuts == 3:
            _col_donut, _col_donut2, _col_donut_grp = st.columns(3)
        elif _n_donuts == 2:
            _col_donut, _col_donut_grp = st.columns(2)
            _col_donut2 = None
        else:
            _col_donut = st.columns([1])[0]
            _col_donut2 = _col_donut_grp = None

        _fig_donut = go.Figure(go.Pie(
            labels=[intra_label.capitalize(), inter_label.capitalize()],
            values=[_intra_pubs_plot, _inter_pubs_plot],
            hole=0.55, marker_colors=["#122947", "#bac7d9"],
            textinfo="percent", rotation=45, sort=False,
            hoverinfo="label+value+percent",
        ))
        _fig_donut.update_layout(
            title="Fakulteter",
            height=360, margin=dict(t=60, b=60, l=60, r=60),
            showlegend=True,
            legend=dict(orientation="v", yanchor="bottom", y=-0.2, xanchor="center", x=0.5),
        )
        if filter_caption:
            st.caption(filter_caption)
        with _col_donut:
            st.plotly_chart(_fig_donut, width='stretch', key=f"donut_{year}_{mode}_{_filter_key}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}})

        if _show_inst_donut and _col_donut2 is not None:
            _fig_donut2 = go.Figure(go.Pie(
                labels=["Intra-institut", "Inter-institut"],
                values=[_intra_inst_plot, _inter_inst_plot],
                hole=0.55, marker_colors=["#122947", "#bac7d9"],
                textinfo="percent", rotation=45, sort=False,
                hoverinfo="label+value+percent",
            ))
            _fig_donut2.update_layout(
                title="Institutter",
                height=360, margin=dict(t=60, b=60, l=60, r=60),
                showlegend=True,
                legend=dict(orientation="v", yanchor="bottom", y=-0.2, xanchor="center", x=0.5),
            )
            with _col_donut2:
                st.plotly_chart(_fig_donut2, width='stretch', key=f"donut2_{year}_{mode}_{_filter_key}",
                    config={"toImageButtonOptions": {"format": "png", "scale": 3}})

        if _show_grp_donut and _col_donut_grp is not None:
            _fig_donut_grp = go.Figure(go.Pie(
                labels=["Intra-stillingsgruppe", "Inter-stillingsgruppe"],
                values=[_intra_grp_plot, _inter_grp_plot],
                hole=0.55, marker_colors=["#39641c", "#becaa8"],
                textinfo="percent", rotation=45, sort=False,
                hoverinfo="label+value+percent",
            ))
            _fig_donut_grp.update_layout(
                title="Stillingsgrupper",
                height=360, margin=dict(t=60, b=60, l=60, r=60),
                showlegend=True,
                legend=dict(orientation="v", yanchor="bottom", y=-0.2, xanchor="center", x=0.5),
            )
            with _col_donut_grp:
                st.plotly_chart(_fig_donut_grp, width='stretch', key=f"donut_grp_{year}_{mode}_{_filter_key}",
                    config={"toImageButtonOptions": {"format": "png", "scale": 3}})

    elif _has_grp_donut and base_mode(mode) == "G":
        _fig_donut_grp = go.Figure(go.Pie(
            labels=["Intra-stillingsgruppe", "Inter-stillingsgruppe"],
            values=[_intra_grp_plot, _inter_grp_plot],
            hole=0.55, marker_colors=["#39641c", "#becaa8"],
            textinfo="percent", rotation=45, sort=False,
            hoverinfo="label+value+percent",
        ))
        _fig_donut_grp.update_layout(
            height=360, margin=dict(t=60, b=60, l=60, r=60),
            showlegend=True,
            legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5),
        )
        st.plotly_chart(_fig_donut_grp, width='stretch', key=f"donut_grp_{year}_{mode}_{_filter_key}",
            config={"toImageButtonOptions": {"format": "png", "scale": 3}})

    # ── 3. Tidsserier ─────────────────────────────────────────────────────────
    st.markdown("---")

    st.markdown(
f"""
#### Udvikling over tid

Afsnittet nedenfor viser udviklingen i sampublicering over tid.
Figurerne kombinerer netværksbaserede opgørelser med ufiltrerede KU‑totaler
for at sætte udviklingen i perspektiv.

Brug fanerne til at skifte mellem:
- **Forfatterpar** (samlet samarbejdsaktivitet),
- **Forfattere** (hvor mange der indgår i samarbejde),
- **Sampubliceringsrate** (intensitet relativt til størrelse).

""")

    _tab_fp, _tab_fb, _tab_rate = st.tabs([
        "Forfatterpar", "Forfattere", "Sampubliceringsrate"
    ])

    with _tab_fp:
        st.markdown(
"""
Forfatterpar i det filtrerede netværk vises sammen med det ufiltrerede antal
sampublikationer samt KU’s samlede publikationstal.

Forskelle mellem kurverne afspejler både ændringer i samarbejdsmønstre og
ændringer i publikationsvolumen. En stigning i forfatterpar betyder øget
sampublicering, men ikke nødvendigvis flere publikationer.
""")
        if _ys:
            _render_year_comparison(
                all_years_data,
                series=[
                    ("Forfatterpar (netværk)",           "total_pubs",       None),
                    ("Forfatterpar (inkl. intra-enhed)",  "total_pubs_all",   None),
                    ("Sampublikationer (ufiltreret)",     "sampub_count_raw", None),
                    ("Alle KU-publikationer",             "ku_total_pubs",    None),
                ],
                title=f"Forfatterpar over tid, {_ys[0]}–{_ys[-1]}",
                yaxis_label="Antal forfatterpar / publikationer",
                key_suffix=f"fp_{year}_{mode}",
            )

    with _tab_fb:
        st.markdown(
"""
Figuren viser antallet af unikke KU‑VIP‑forfattere, der indgår i sampublicering,
sammenlignet med det samlede antal KU‑VIP‑forfattere.

Forskellen mellem de to kurver viser, hvor stor en del af forskerne der
ikke indgår i nogen sampublikationer i det valgte udsnit.
""")
        if _ys:
            _render_year_comparison(
                all_years_data,
                series=[
                    ("Forfattere i sampublikationer (filtreret)", "total_authors",    None),
                    ("Alle unikke KU-forfattere",                      "ku_total_authors", None),
                ],
                title=f"Forfattere over tid, {_ys[0]}–{_ys[-1]}",
                yaxis_label="Antal unikke forfattere",
                key_suffix=f"fb_{year}_{mode}",
            )

    with _tab_rate:
        st.markdown(
"""
Sampubliceringsraten viser det gennemsnitlige antal forfatterpar per forfatter.

En stigende rate indikerer øget sampubliceringsintensitet – ikke nødvendigvis
flere forfattere. Raten inklusive intra‑enhed er altid højere, da den
tæller alle forfatterpar - også dem inden for samme organisatoriske enhed.

**Eksempel:** KU har et år 100 unikke forfattere og indgår i 300 forfatterpar. Sampubliceringsraten 
er da 300 / 100 = 3,0, hvilket betyder, at den gennemsnitlige forfatter indgår i tre sampubliceringsrelationer (forfatterpar).
Et andet år er der 50 forfattere og 200 forfatter - og dermed en højere rate på 4,0, selvom det
samlede samarbejdsomfang er lavere. 
""")
        if _ys:
            _render_year_comparison(
                all_years_data,
                series=[
                    ("Rate - netværk (forfatterpar / forfattere)",           "sampub_rate",     None),
                    ("Rate - inkl. intra-enhed (forfatterpar / forfattere)", "sampub_rate_all", None),
                ],
                title=f"Sampubliceringsrate over tid, {_ys[0]}–{_ys[-1]}",
                yaxis_label="Rate",
                key_suffix=f"rate_{year}_{mode}",
            )

    
    # ── 4. De største samarbejder ─────────────────────────────────────────────
    if all_years_data and len(all_years_data) >= 2:
        years_sorted = sorted(all_years_data.keys())

        _all_pair_totals: dict[str, float] = {}
        for snap in all_years_data.values():
            for pair, w in snap.get("top_pairs", {}).items():
                _all_pair_totals[pair] = _all_pair_totals.get(pair, 0.0) + w
        st.markdown("---")
        st.markdown(f"#### De største samarbejder, {years_sorted[0]}–{years_sorted[-1]}")
        st.markdown(
"""
Afsnittet nedenfor fremhæver de mest intensive samarbejdsrelationer,
opgjort som det samlede antal forfatterpar på tværs af år. Vær opmærksom på, hvilke filtre der er valgte i sidepanelet. 
""")

        top_n_pairs = st.number_input(
            "**Antal par at vise**",
            min_value=1, max_value=min(20, max(len(_all_pair_totals), 1)),
            value=min(4, max(len(_all_pair_totals), 1)),
            step=1, key=f"top_pairs_n_{mode}",
        )
        _top_pairs = sorted(_all_pair_totals.items(), key=lambda x: -x[1])[:top_n_pairs]
        _top_pair_labels = [p for p, _ in _top_pairs]
        colors = ku_color_sequence(len(_top_pair_labels))

        _fig_pairs = go.Figure()
        for i, pair in enumerate(_top_pair_labels):
            _y = [all_years_data[yr].get("top_pairs", {}).get(pair, 0) for yr in years_sorted]
            _fig_pairs.add_trace(go.Scatter(
                x=years_sorted, y=_y, name=pair,
                mode="lines+markers",
                line=dict(width=2, color=colors[i]),
                marker=dict(size=7, color=colors[i]),
            ))
        _fig_pairs.update_layout(
            yaxis_title="Antal forfatterpar",
            xaxis=dict(tickmode="array", tickvals=years_sorted, dtick=1),
            height=420,
            title=f"De {top_n_pairs} mest sampublicerende par, {years_sorted[0]}–{years_sorted[-1]}",
            legend=dict(orientation="h", yanchor="top", y=-0.25, xanchor="left", x=0),
            margin=dict(t=50, b=120),
        )
        st.plotly_chart(_fig_pairs, width='stretch',
            config={"toImageButtonOptions": {"format": "png", "scale": 3}})
        if filter_caption:
            st.caption(filter_caption)   

        _pair_tbl_rows = [
            {"Par": pair,
             **{str(yr): round(all_years_data[yr].get("top_pairs", {}).get(pair, 0), 1)
                for yr in years_sorted},
             "Total": round(_all_pair_totals.get(pair, 0), 1)}
            for pair in _top_pair_labels
        ]
        _pair_tbl_schema = (
            [("Par", pa.string())] +
            [(str(yr), pa.float64()) for yr in years_sorted] +
            [("Total", pa.float64())]
        )
        with st.expander("Se tabel"):
            st.dataframe(build_table(_pair_tbl_rows, _pair_tbl_schema), hide_index=True, width="stretch")
            st.download_button(
                "Download (.xlsx)",
                data=rows_to_excel_bytes(_pair_tbl_rows, [n for n, _ in _pair_tbl_schema]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_top_pairs_{year}_{mode}"
            )


        # ── Hvilke samarbejder vokser? ────────────────────────────────────────
        st.markdown(f"##### Hvilke samarbejder vokser?")
        st.markdown(
f"""Figuren viser de par, der har haft den **største stigning** i antal forfatterpar
fra {years_sorted[0]} til {years_sorted[-1]}. Positive værdier indikerer voksende
samarbejde; negative indikerer aftagende.

Den **absolutte** ændring viser den rå forskel i forfatterpar. Den **relative** ændring
viser procentvis vækst - nyttig for at sammenligne par med meget forskellige udgangspunkter,
men udelader par med 0 forfatterpar i første år.""")

        _yr_first, _yr_last = years_sorted[0], years_sorted[-1]
        _growth_abs: dict[str, float] = {}
        _growth_rel: dict[str, float] = {}
        _all_pairs_union = set()
        for snap in all_years_data.values():
            _all_pairs_union |= set(snap.get("top_pairs", {}).keys())
        for pair in _all_pairs_union:
            _w_first = all_years_data[_yr_first].get("top_pairs", {}).get(pair, 0)
            _w_last  = all_years_data[_yr_last].get("top_pairs", {}).get(pair, 0)
            _growth_abs[pair] = _w_last - _w_first
            if _w_first > 0:
                _growth_rel[pair] = round(100 * (_w_last - _w_first) / _w_first, 1)

        _growth_n = st.number_input(
            "**Antal par at vise (top voksende + top aftagende)**",
            min_value=1, max_value=min(20, max(len(_growth_abs), 1)),
            value=min(4, max(len(_growth_abs), 1)),
            step=1, key=f"growth_pairs_n_{mode}",
        )

        _tab_abs_growth, _tab_rel_growth = st.tabs(["Absolut ændring", "Relativ ændring (%)"])

        def _make_growth_fig(growth_dict, xlabel, title, fmt_fn):
            _sorted = sorted(growth_dict.items(), key=lambda x: -x[1])
            _top    = _sorted[:_growth_n]
            _bot    = _sorted[-_growth_n:][::-1]
            _plot   = _top + [p for p in _bot if p not in _top]
            _plot_sorted = sorted(_plot, key=lambda x: x[1])
            _colors = ["#122947" if v >= 0 else "#901a1E" for _, v in _plot_sorted]
            fig = go.Figure(go.Bar(
                y=[p for p, _ in _plot_sorted],
                x=[v for _, v in _plot_sorted],
                orientation="h",
                marker_color=_colors,
                text=[fmt_fn(v) for _, v in _plot_sorted],
                textposition="inside",
                hovertemplate="<b>%{y}</b><br>" + xlabel + ": %{x}<extra></extra>",
            ))
            fig.add_vline(x=0, line_width=1, line_color="#666666")
            fig.update_layout(
                xaxis_title=f"{xlabel} ({_yr_first}→{_yr_last})",
                height=max(350, 28 * len(_plot_sorted)),
                margin=dict(l=200, t=50, r=80),
                yaxis=dict(autorange="reversed"),
                title=title,
            )
            tbl_rows = [
                {"Par": p, xlabel: round(v, 1),
                 str(_yr_first): round(all_years_data[_yr_first].get("top_pairs", {}).get(p, 0), 1),
                 str(_yr_last):  round(all_years_data[_yr_last].get("top_pairs",  {}).get(p, 0), 1)}
                for p, v in sorted(growth_dict.items(), key=lambda x: -x[1])
            ]
            tbl_schema = [
                ("Par", pa.string()),
                (str(_yr_first), pa.float64()),
                (str(_yr_last),  pa.float64()),
                (xlabel,         pa.float64()),
            ]
            return fig, tbl_rows, tbl_schema

        with _tab_abs_growth:
            _fig, _tbl_rows, _tbl_schema = _make_growth_fig(
                _growth_abs,
                "Ændring i forfatterpar",
                f"Absolut vækst i forfatterpar, {_yr_first}–{_yr_last}",
                lambda v: f"{v:+.1f}",
            )
            if filter_caption:
                st.caption(filter_caption)
            st.plotly_chart(_fig, width='stretch',
                config={"toImageButtonOptions": {"format": "png", "scale": 3}})
            with st.expander("Se tabel"):
                st.dataframe(build_table(_tbl_rows, _tbl_schema), hide_index=True, width="stretch")
                st.download_button("Download (.xlsx)",
                    data=rows_to_excel_bytes(_tbl_rows, [n for n, _ in _tbl_schema]),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_growth_abs_{year}_{mode}")

        with _tab_rel_growth:
            if not _growth_rel:
                st.error("Ingen par har forfatterpar i første år - relativ ændring kan ikke beregnes.")
            else:
                _fig, _tbl_rows, _tbl_schema = _make_growth_fig(
                    _growth_rel,
                    "Relativ ændring (%)",
                    f"Relativ vækst i forfatterpar, {_yr_first}–{_yr_last}",
                    lambda v: f"{v:+.1f}%",
                )
                if filter_caption:
                    st.caption(filter_caption)
                st.plotly_chart(_fig, width='stretch',
                    config={"toImageButtonOptions": {"format": "png", "scale": 3}})
                with st.expander("Se tabel"):
                    st.dataframe(build_table(_tbl_rows, _tbl_schema), hide_index=True, width="stretch")
                    st.download_button("Download (.xlsx)",
                        data=rows_to_excel_bytes(_tbl_rows, [n for n, _ in _tbl_schema]),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_growth_rel_{year}_{mode}")



def render_tab_fakulteter(year, mode, fac_tot_size, fac_avg_size, edges_keep, node_meta, all_years_data=None, fac_ew=None, faculty_base_colors=None, size_map=None, total_pubs=None, filter_caption=None, show_intra=True):
    if not fac_in_mode(mode):
        return
    if not fac_tot_size or not fac_avg_size:
        st.error("Ingen fakultetsdata for det valgte filter.")
        return

    st.markdown(
f"""
### Fakulteternes bidrag til sampublicering

Fanen kortlægger, hvordan de seks fakulteter bidrager til KU's sampubliceringsaktivitet i 
**{year}**. Fakultetsniveauet giver et **overordnet strukturelt billede** af forskningssamarbejde,
hvor interne forskelle mellem institutter og forskningsgrupper er udjævnet.
Dermed bliver brede forskelle i samarbejdsomfang, sampubliceringsintensitet og
publikationspraksis tydelige på tværs af KU’s hovedområder.

Fanen fokuserer på **bidrag og relative forskelle** mellem fakulteter – ikke på
detaljer i det interne samarbejde, som analyseres nærmere på institut‑ og
stillingsgruppeniveau i de øvrige faner.

Opgørelserne i fanen bygger på tre centrale mål:

- **Forfatterantal** viser antallet af unikke KU‑VIP‑forfattere, der er knyttet til
  hvert fakultet, uanset hvor mange publikationer de har bidraget til.
- **Forfatterpar** viser, hvor mange forfatterpar et fakultet indgår i. Hvert
  forfatterpar tælles én gang hos hvert fakultet, som indgår i samarbejdet.
- **Sampubliceringsrate** sætter antallet af forfatterpar i forhold til
  forfatterantal og gør det muligt at sammenligne samarbejdsintensitet mellem
  fakulteter af meget forskellig størrelse.
""")

    if size_map is None:
        size_map = fac_tot_size

    _grand_tot     = sum(fac_tot_size.values()) or 1
    _top_fac, _top_val = max(fac_tot_size.items(), key=lambda x: x[1])
    _bot_fac, _bot_val = min(fac_tot_size.items(), key=lambda x: x[1])
    _top_share     = 100 * _top_val / _grand_tot
    _avg_top, _avg_top_val = max(fac_avg_size.items(), key=lambda x: x[1])
    _avg_bot, _avg_bot_val = min(fac_avg_size.items(), key=lambda x: x[1])
    _facs_ord      = [f for f in FAC_ORDER if f in fac_tot_size]
    _tots_ord      = [int(fac_tot_size[f]) for f in _facs_ord]
    _avgs_ord      = [round(fac_avg_size.get(f, 0), 1) for f in _facs_ord]
    _ews_ord       = [round(fac_ew.get(f, 0), 1) if fac_ew else 0 for f in _facs_ord]
    _fac_colors    = [faculty_base_colors.get(f, "#122947") if faculty_base_colors else "#122947" for f in _facs_ord]
    _grand_ew      = total_pubs if total_pubs else 0

    
    # ── Forfatterbidrag i [år] ────────────────────────────────────────────────
    # Tekst-summary
    _ew_top_txt = ""
    if fac_ew and _grand_ew:
        _ew_top, _ew_top_val = max(fac_ew.items(), key=lambda x: x[1])
        _ew_share = 100 * _ew_top_val / _grand_ew
        _ratio_sorted = sorted(
            [(k, fac_ew[k] / fac_tot_size[k]) for k in fac_tot_size if fac_tot_size.get(k) and fac_ew.get(k)],
            key=lambda x: -x[1],
        )
        if not _ratio_sorted:
            _rat_top = _rat_bot = ""
            _rat_top_val = _rat_bot_val = 0.0
        else:
            _rat_top, _rat_top_val = _ratio_sorted[0]
            _rat_bot, _rat_bot_val = _ratio_sorted[-1]
        _ew_top_txt = (
            f" Målt på forfatterpar har **{_ew_top}** flest med {fmt_ui(_ew_top_val)} "
            f"({fmt_ui(_ew_share)}% af samtlige forfatterpar). "
            f"Relativt til størrelse - forfatterpar per forfatter - er **{_rat_top}** "
            f"mest sampublicerende ({fmt_ui(_rat_top_val,3)}), mens **{_rat_bot}** er mindst ({fmt_ui(_rat_bot_val,3)})."
        )

    st.markdown("---")

    st.markdown(
f""" 
#### Fakulteternes forfattere i {year}

Figuren viser fordelingen af **KU‑VIP‑forfattere** på tværs af fakulteterne.
Store fakulteter vil naturligt bidrage med flere forfattere, men fordelingen giver
alligevel et vigtigt udgangspunkt for at forstå forskelle i sampublicering og
samarbejdsintensitet

{fmt_ui(_top_share)}% af samtlige forfattere er samlet på **{_top_fac}** ({int(_top_val)}).
Hvorimod **{_bot_fac}** bidrager mindst med **{int(_bot_val)}** unikke forfattere. 
Bemærk her forskelle i publikationspraksis (se mere i Datagrundlag-fanen).

**Gennemsnitligt antal forfattere** er beregnet som det samlede 
antal forfattere divideret med antallet af enheder under fakultetet - dvs. antallet af unikke 
kombinationer af institut og stillingsgruppe i den valgte visning. Et højt gennemsnit indikerer, 
at de enkelte enheder er relativt store. Gennemsnitligt antal forfattere er højest 
ved **{_avg_top}** ({fmt_ui(_avg_top_val)}) og lavest ved **{_avg_bot}** ({fmt_ui(_avg_bot_val)}). 
""")

    # Forfatterbidrag-plots
    _tab_ft, _tab_fp, _tab_fa = st.tabs(["Antal", "Andel (%)", "Gns. antal forfattere"])
    with _tab_ft:
        _fig_ft = go.Figure(go.Bar(
            y=_facs_ord, x=_tots_ord, orientation="h",
            marker_color=_fac_colors,
            text=[f"{v}" for v in _tots_ord],
            textposition="inside",
        ))
        _fig_ft.update_layout(
            xaxis_title="Forfatterantal",
            height=max(300, 50 * len(_facs_ord)),
            margin=dict(l=80, t=50, r=80),
            yaxis=dict(autorange="reversed"),
            title = f"Fakulteternes forfattere, {year}"
        )
        st.plotly_chart(_fig_ft, width="stretch", key=f"fac_tot_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
        )
        if filter_caption:
            st.caption(filter_caption)

    with _tab_fp:
        _pct_vals = [round(100 * v / _grand_tot, 1) for v in _tots_ord]
        _fig_fp = go.Figure(go.Bar(
            y=_facs_ord, x=_pct_vals, orientation="h",
            marker_color=_fac_colors,
            text=[f"{fmt_ui(v)}%" for v in _pct_vals],
            textposition="inside",
        ))
        _fig_fp.update_layout(
            xaxis=dict(title="Andel (%)", range=[0, 100]),
            height=max(300, 50 * len(_facs_ord)),
            margin=dict(l=80, t=50, r=80),
            yaxis=dict(autorange="reversed"),
            title = f"Fakulteternes andel af samlet forfatterantal, {year}"
        )
        st.plotly_chart(_fig_fp, width="stretch", key=f"fac_pct_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
        )
        if filter_caption:
            st.caption(filter_caption)

    with _tab_fa:
        _fig_fa = go.Figure(go.Bar(
            y=_facs_ord, x=_avgs_ord, orientation="h",
            marker_color=_fac_colors,
            text=[f"{fmt_ui(v)}" for v in _avgs_ord],
            textposition="inside",
        ))
        _fig_fa.update_layout(
            xaxis_title="Gns. antal forfattere",
            height=max(300, 50 * len(_facs_ord)),
            margin=dict(l=80, t=50, r=80),
            yaxis=dict(autorange="reversed"),
            title = f"Fakulteternes gennemsnitlige forfatterantal, {year}"
        )
        st.plotly_chart(_fig_fa, width="stretch", key=f"fac_avg_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
        )
        if filter_caption:
            st.caption(filter_caption)
    
    # ── Tabel ─────────────────────────────────────────────────────────────────
    _fac_summary_rows = [
        {
            "Fakultet":              fac,
            "Antal forfattere": int(fac_tot_size.get(fac, 0)),
            "Andel (%)":             round(100 * fac_tot_size.get(fac, 0) / _grand_tot, 1),
            "Gns. forfatterantal":  round(fac_avg_size.get(fac, 0), 1),
        }
        for fac in _facs_ord
    ]
    _fac_summary_schema = [
        ("Fakultet",               pa.string()),
        ("Antal forfattere",  pa.int64()),
        ("Andel (%)",              pa.float64()),
        ("Gns. forfatterantal",   pa.float64()),
    ]
    
    with st.expander("Se tabel"):
        st.dataframe(build_table(_fac_summary_rows, _fac_summary_schema), hide_index=True, width="stretch")
        st.download_button(
            "Download (.xlsx)",
            data=rows_to_excel_bytes(_fac_summary_rows, [n for n, _ in _fac_summary_schema]),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl__fac_bidrag_{year}_{mode}",
        )

    # Forfatterpar-plot (kun modes hvor det giver mening)
    _fac_ew_all_yr = {}
    if all_years_data and year in all_years_data:
        _fi = all_years_data[year].get("fac_intra_ew_all", {})
        _fx = all_years_data[year].get("fac_inter_ew_all", {})
        _fac_ew_all_yr = {f: _fi.get(f, 0.0) + _fx.get(f, 0.0) for f in set(_fi) | set(_fx)}

    st.markdown("---")

    # Forfatterpar-plot (kun modes hvor det giver mening)
    if fac_ew and _grand_ew:
        st.markdown(
f"""
#### Forfatterpar per fakultet i {year}

**Forfatterpar** opgør det samlede samarbejdsomfang for hvert fakultet.
Et samarbejde mellem to fakulteter tælles én gang hos hvert af dem, hvilket betyder,
at summen af forfatterpar på tværs af fakulteter svarer til **to gange** det samlede
antal unikke forfatterpar i netværket.

Målet afspejler samarbejdsaktivitet relativt til fakulteternes størrelse, men siger
ikke noget om, hvor mange forskellige samarbejdspartnere et fakultet har – kun
hvor intensivt det indgår i sampublicering.

Forfatterpar per forfatter sætter forfatterpar i forhold til forfatterantal og gør det muligt
at sammenligne samarbejdsintensitet på tværs af fakulteter med meget forskellige
størrelser.

En høj rate indikerer, at de enkelte forskere i gennemsnit indgår i mange
samarbejdsrelationer, men raten påvirkes også af publikationspraksis og typiske
forfatterkonstellationer på fakultetet.

**Eksempel**: et fakultet har 100 unikke forfattere og indgår i 300 forfatterpar. Sampubliceringsraten
er da 300 / 100 = 3,0, hvilket betyder, at den gennemsnitlige forfatter indgår i tre forfatterpar. 
Et andet fakultete med 50 forfattere og 200 forfatterpar har en højere rate på 4,0 - selv om 
det samlede samarbejdsomfang er lavere.

"""
+ _ew_top_txt if _ew_top_txt else "")

        _use_fac_all = st.toggle(
            "Inkludér intra-enhed forfatterpar (f.eks. to professorer på samme institut)",
            value=False,
            key=f"fac_ew_bar_all_{year}_{mode}",
            help="Til: alle forfatterpar inkl. par inden for samme enhed. Fra: kun netværkskanter.",
            disabled=not show_intra
        )
        if _use_fac_all:
            st.caption("Alle forfatterpar inkl. par inden for samme enhed (f.eks. to mandlige professorer på Kemisk Institut).")
        else:
            st.caption("Kun forfatterpar *mellem* de valgte organisatoriske enheder - som vist i netværket.")

        _active_fac_ew = _fac_ew_all_yr if (_use_fac_all and _fac_ew_all_yr) else fac_ew
        if _use_fac_all:
            # Intra-toggle på: brug sum/2 af aktiv ew.
            _grand_active_ew = (sum(_active_fac_ew.values()) / 2) or 1
        elif total_pubs:
            # Intra-toggle fra, ingen lokalt filter på fakultets-fanen: brug global total_pubs.
            _grand_active_ew = total_pubs
        else:
            _grand_active_ew = (sum(_active_fac_ew.values()) / 2) or 1
        _ews_ord_active = [_active_fac_ew.get(f, 0) for f in _facs_ord]
        _ew_ratio_ord = [round(_active_fac_ew.get(f, 0) / (fac_tot_size.get(f) or 1), 3) for f in _facs_ord]

        _pct_ews = [round(100 * v / _grand_active_ew, 1) for v in _ews_ord_active]
        _tab_ew_abs, _tab_ew_pct, _tab_ew_ratio = st.tabs(["Antal", "Andel (%)", "Forfatterpar per forfatter"])

        with _tab_ew_abs:
            _fig_ew = go.Figure(go.Bar(
                y=_facs_ord, x=_ews_ord_active, orientation="h",
                marker_color=_fac_colors,
                text=[f"{fmt_ui(v)}" for v in _ews_ord_active],
                textposition="inside",
            ))
            _fig_ew.update_layout(
                xaxis_title="Antal forfatterpar",
                height=max(300, 50 * len(_facs_ord)),
                margin=dict(l=80, t=50, r=80),
                yaxis=dict(autorange="reversed"),
                title=f"Forfatterpar per fakultet{' (inkl. intra-enhed)' if _use_fac_all else ''}, {year}",
            )
            if filter_caption:
                st.caption(filter_caption)
            st.plotly_chart(_fig_ew, width="stretch", key=f"fac_ew_abs_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}})

        with _tab_ew_pct:
            _fig_ew_p = go.Figure(go.Bar(
                y=_facs_ord, x=_pct_ews, orientation="h",
                marker_color=_fac_colors,
                text=[f"{fmt_ui(v)}%" for v in _pct_ews],
                textposition="inside",
            ))
            _fig_ew_p.update_layout(
                xaxis=dict(title="Andel (%)", range=[0, 100]),
                height=max(300, 50 * len(_facs_ord)),
                margin=dict(l=80, t=50, r=80),
                yaxis=dict(autorange="reversed"),
                title=f"Forfatterpar per fakultet{' (inkl. intra-enhed)' if _use_fac_all else ''} - andel (%), {year}",
            )
            st.caption("Vær opmærksom på, at procenter kan overstige 100%. Se **Læsning af procenter** i Oversigt-fanen.")
            if filter_caption:
                st.caption(filter_caption)
            st.plotly_chart(_fig_ew_p, width="stretch", key=f"fac_ew_pct_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}})

        with _tab_ew_ratio:
            _fig_ew_r = go.Figure(go.Bar(
                y=_facs_ord, x=_ew_ratio_ord, orientation="h",
                marker_color=_fac_colors,
                text=[f"{fmt_ui(v,3)}" for v in _ew_ratio_ord],
                textposition="inside",
            ))
            _fig_ew_r.update_layout(
                xaxis_title="Forfatterpar per forfatter",
                height=max(300, 50 * len(_facs_ord)),
                margin=dict(l=80, t=50, r=80),
                yaxis=dict(autorange="reversed"),
                title=f"Forfatterpar per forfatter per fakultet{' (inkl. intra-enhed)' if _use_fac_all else ''}, {year}",
            )
            if filter_caption:
                st.caption(filter_caption)
            st.plotly_chart(_fig_ew_r, width="stretch", key=f"fac_ew_ratio_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}})

    # Tabel
    _fac_summary_rows = [
        {"Fakultet": f,
         "Forfatterpar (netværk)":       round(fac_ew.get(f, 0), 1) if fac_ew else 0,
         "Forfatterpar (inkl. intra)":   round(_fac_ew_all_yr.get(f, 0), 1) if _fac_ew_all_yr else 0}
        for f in _facs_ord
    ]
    _fac_summary_schema = [
        ("Fakultet",                    pa.string()),
        ("Forfatterpar (netværk)",      pa.float64()),
        ("Forfatterpar (inkl. intra)",  pa.float64()),
    ]
    with st.expander("Se tabel"):
        st.dataframe(build_table(_fac_summary_rows, _fac_summary_schema), width="stretch", hide_index=True)
        st.download_button(
            "Download (.xlsx)",
            data=rows_to_excel_bytes(_fac_summary_rows, [n for n, _ in _fac_summary_schema]),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_fac_bidrag_{year}_{mode}",
        )

    st.markdown("---")

    # ── Samarbejdsmønstre på tværs af fakulteter (heatmap) ───────────────────
    st.markdown(
f"""#### Samarbejdsmønstre på tværs af fakulteter i {year}

Heatmappet viser, **hvilke fakulteter der publicerer sammen** - og hvor intensivt.

- **Antal forfatterpar** viser den rå trafik mellem to fakulteter. Store fakulteter dominerer naturligt.
- **Forfatterpar per forfatter** svarer på spørgsmålet: *hvor ofte er en forfatter fra kolonnen 
med, når nogen fra rækken publicerer?* Matricen er ikke symmetrisk - cellen (SUND, SCIENCE) og cellen 
(SCIENCE, SUND) svarer på to forskellige spørgsmål. Hvis SUND har 15% af sine par med SCIENCE, mens 
SCIENCE kun har 8% med SUND, er SUND mere afhængig af SCIENCE end omvendt.
- **Andel (%) per række** viser for hvert fakultet, hvor dets samarbejder fordeler sig. Hver række 
summerer til 100%.

**Bemærk**: ratioen kan være større end 1. Én forfatter kan indgå i flere par - hvis en 
SUND-forsker er medforfatter på en artikle med tre SCIENCE-forskere, genererer det én forfatter for 
SUND men tre par med SCIENCE. Store forfatterskaber (f.eks. artikler med 100+ forfattere) 
kan derfor skævvride billedet.
""")

    _use_fac_hm_all = st.toggle(
        "Inkludér intra-fakultet forfatterpar (f.eks. to forskere på samme SUND-institut)",
        value=False,
        key=f"fac_hm_all_{year}_{mode}",
        help="Til: alle forfatterpar inkl. par inden for samme fakultet. Fra: kun par mellem fakulteter.",
        disabled=not show_intra
    )

    # Byg par-matrix fra kanterne. Halveret vægt (w/2) matcher bar-chart-ratioen ovenfor.
    _fac_pair_matrix: dict[tuple[str, str], float] = {}
    for u, v, w, *_ in (edges_keep or []):
        fu = (node_meta or {}).get(u, {}).get("fac", "")
        fv = (node_meta or {}).get(v, {}).get("fac", "")
        if not fu or not fv:
            continue
        if fu == fv and not _use_fac_hm_all:
            continue  # spring intra-fakultet par over hvis toggle er slået fra
        _fac_pair_matrix[(fu, fv)] = _fac_pair_matrix.get((fu, fv), 0.0) + w
        if fu != fv:
            _fac_pair_matrix[(fv, fu)] = _fac_pair_matrix.get((fv, fu), 0.0) + w

    # Hvis intra-fakultet toggle er på, brug pre-computede intra-par på diagonalen
    if _use_fac_hm_all and all_years_data and year in all_years_data:
        _fac_intra_src = all_years_data[year].get("fac_intra_ew_all", {})
        for f, v in _fac_intra_src.items():
            if f in fac_tot_size:
                # Overskriv diagonalen med den pre-computede værdi (bidrag fra edges_keep er inkonsistent for intra)
                _fac_pair_matrix[(f, f)] = v

    _hm_facs = [f for f in _facs_ord if f in fac_tot_size]
    _intra_suffix_fac_hm = ' (inkl. intra-fakultet)' if _use_fac_hm_all else ''

    if len(_hm_facs) < 2 or not _fac_pair_matrix:
        st.error("Ikke tilstrækkeligt data til at vise samarbejdsmatricen for det valgte udsnit.")
    else:
        _z_abs_fac = [[round(_fac_pair_matrix.get((fi, fj), 0.0), 1) for fj in _hm_facs] for fi in _hm_facs]
        _z_ratio_fac = [
            [round(_fac_pair_matrix.get((fi, fj), 0.0) / (fac_tot_size.get(fi) or 1), 3) for fj in _hm_facs]
            for fi in _hm_facs
        ]
        _z_row_pct_fac = []
        for fi in _hm_facs:
            _row_total = sum(_fac_pair_matrix.get((fi, fj), 0.0) for fj in _hm_facs) or 1
            _z_row_pct_fac.append([
                round(100 * _fac_pair_matrix.get((fi, fj), 0.0) / _row_total, 2)
                for fj in _hm_facs
            ])

        _tab_hm_abs_fac, _tab_hm_ratio_fac, _tab_hm_pct_fac = st.tabs(
            ["Antal forfatterpar", "Forfatterpar per forfatter", "Andel (%) per række"]
        )

        def _fac_heatmap_fig(z, title, colorbar_title, text_fmt, hover_template):
            _text = [[text_fmt(v) for v in row] for row in z]
            fig = go.Figure(go.Heatmap(
                z=z, x=_hm_facs, y=_hm_facs,
                colorscale=[[0, "#f0f4f8"], [1, "#122947"]],
                text=_text, texttemplate="%{text}",
                textfont=dict(size=12),
                hovertemplate=hover_template,
                colorbar=dict(title=colorbar_title),
            ))
            fig.update_layout(
                title=title,
                xaxis=dict(title="", tickangle=0),
                yaxis=dict(title="", autorange="reversed"),
                height=70 * len(_hm_facs),
                margin=dict(l=100, t=60, r=80, b=80),
            )
            return fig

        with _tab_hm_abs_fac:
            st.plotly_chart(
                _fac_heatmap_fig(
                    _z_abs_fac,
                    f"Antal forfatterpar mellem fakulteter{_intra_suffix_fac_hm}, {year}",
                    "Antal par",
                    lambda v: fmt_ui(v) if v else "",
                    "<b>%{y}</b> og <b>%{x}</b><br>"
                    "Antal forfatterpar: <b>%{z}</b>"
                    "<extra></extra>",
                ),
                width="stretch",
                key=f"fac_hm_abs_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}},
            )
            if filter_caption:
                st.caption(filter_caption)
            st.caption(
                "Symmetrisk matrix: par(A,B) = par(B,A). Diagonalen viser intra-fakultet-samarbejde "
                "og vises kun, hvis intra-fakultet-toggle er aktiveret."
            )

        with _tab_hm_ratio_fac:
            st.plotly_chart(
                _fac_heatmap_fig(
                    _z_ratio_fac,
                    f"Forfatterpar per forfatter mellem fakulteter{_intra_suffix_fac_hm}, {year}",
                    "Par per bidrag",
                    lambda v: fmt_ui(v, 2) if v else "",
                    "Per forfatter fra <b>%{y}</b><br>"
                    "dannes i gennemsnit <b>%{z}</b> par med <b>%{x}</b>"
                    "<extra></extra>",
                ),
                width="stretch",
                key=f"fac_hm_ratio_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}},
            )
            if filter_caption:
                st.caption(filter_caption)
            st.caption(
                "**Asymmetrisk matrix.** Rækken normaliseres med fakultetets forfatterantal. "
                "Høje værdier indikerer intensiv samarbejdsrelation relativt til fakultetets størrelse."
            )

        with _tab_hm_pct_fac:
            st.plotly_chart(
                _fac_heatmap_fig(
                    _z_row_pct_fac,
                    f"Fordeling af samarbejdspartnere per fakultet{_intra_suffix_fac_hm}, {year}",
                    "Andel (%)",
                    lambda v: f"{fmt_ui(v, 1)}%" if v else "",
                    "<b>%{z:.2f}%</b> af <b>%{y}</b>'s forfatterpar<br>"
                    "er med <b>%{x}</b>"
                    "<extra></extra>",
                ),
                width="stretch",
                key=f"fac_hm_pct_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}},
            )
            if filter_caption:
                st.caption(filter_caption)
            st.caption(
                "Hver række summerer til 100%. Viser hvor et fakultets samarbejdspar *fordeler sig* "
                "på tværs af andre fakulteter - uafhængigt af fakultetets størrelse."
            )

        # ── Tabel (langt format) ──────────────────────────────────────────────
        _fac_hm_rows = []
        for fi in _hm_facs:
            _row_total = sum(_fac_pair_matrix.get((fi, fj), 0.0) for fj in _hm_facs) or 1
            _bidrag_i = fac_tot_size.get(fi) or 1
            for fj in _hm_facs:
                _pairs = _fac_pair_matrix.get((fi, fj), 0.0)
                if _pairs == 0:
                    continue
                _fac_hm_rows.append({
                    "Fra (række)":                 fi,
                    "Til (kolonne)":               fj,
                    "Forfatterpar":                round(_pairs, 1),
                    "Forfatterpar per forfatter":     round(_pairs / _bidrag_i, 3),
                    "Andel af rækkens par (%)":    round(100 * _pairs / _row_total, 2),
                })
        _fac_hm_schema = [
            ("Fra (række)",              pa.string()),
            ("Til (kolonne)",            pa.string()),
            ("Forfatterpar",             pa.float64()),
            ("Forfatterpar per forfatter",  pa.float64()),
            ("Andel af rækkens par (%)", pa.float64()),
        ]
        with st.expander("Se tabel"):
            st.dataframe(build_table(_fac_hm_rows, _fac_hm_schema), hide_index=True, width="stretch")
            st.download_button(
                "Download (.xlsx)",
                data=rows_to_excel_bytes(_fac_hm_rows, [n for n, _ in _fac_hm_schema]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_fac_hm_{year}_{mode}",
            )



    st.markdown("---")

    # ── Sammenlign år ─────────────────────────────────────────────────────────
    st.markdown(
f"""### Udvikling over tid

Fanerne nedenfor viser udviklingen i fakulteternes **forfatterantal**,
**forfatterpar** og **rækkevidde** på tværs af de tilgængelige år.

Tidsserierne gør det muligt at vurdere, om forskelle mellem fakulteter er stabile
over tid eller udtryk for mere langsigtede strukturelle ændringer i
samarbejdsmønstrene.
""")
    _facs_in_data    = sorted({f for s in (all_years_data or {}).values() for f in s.get("fac_tot", {})})
    _facs_ew_in_data = sorted({f for s in (all_years_data or {}).values() for f in s.get("fac_ew", {})})
    years_sorted_fac = sorted((all_years_data or {}).keys())

    _tab_yr_tot, _tab_yr_ew = st.tabs(["Forfatterantal per fakultet", "Forfatterpar per fakultet"])

    with _tab_yr_tot:
        _render_year_comparison(
            all_years_data,
            series=[(fac, "fac_tot", fac) for fac in _facs_in_data],
            title=f"Udvikling i fakulteternes forfatterantal, {years_sorted_fac[0]}-{years_sorted_fac[-1]}",
            colors=faculty_base_colors,
            key_suffix=f"fac_tot_{year}_{mode}",
            show_table=True,
        )

    with _tab_yr_ew:
        _use_fac_ew_all = st.toggle(
            "Inkludér intra-enhed forfatterpar (f.eks. to professorer på samme institut)",
            value=False,
            key=f"fac_ew_all_{year}_{mode}",
            help="Til: alle forfatterpar inkl. par inden for samme enhed. Fra: kun netværkskanter (par mellem noder).",
            disabled=not show_intra
        )
        if _use_fac_ew_all:
            st.caption("Alle forfatterpar inkl. par inden for samme enhed (f.eks. to mandlige professorer på Kemisk Institut).")
        else:
            st.caption("Kun forfatterpar *mellem* de valgte organisatoriske enheder - som vist i netværket.")

        _fac_ew_key = "fac_ew_all" if _use_fac_ew_all else "fac_ew"
        _facs_ew_plot = sorted({
            f for s in (all_years_data or {}).values()
            for f in s.get(_fac_ew_key, {})
        })

        _render_year_comparison(
            all_years_data,
            series=[(fac, _fac_ew_key, fac) for fac in _facs_ew_plot],
            title=f"Udvikling i antal forfatterpar per fakultet{' (inkl. intra-enhed)' if _use_fac_ew_all else ''}, {years_sorted_fac[0]}-{years_sorted_fac[-1]}",
            yaxis_label="Antal forfatterpar",
            colors=faculty_base_colors,
            key_suffix=f"fac_ew_{year}_{mode}",
            show_table=True,
        )
  
    if fac_ew and _grand_ew:
        _fac_ew_all = None
        if all_years_data and year in all_years_data:
            _fi = all_years_data[year].get("fac_intra_ew_all", {})
            _fx = all_years_data[year].get("fac_inter_ew_all", {})
            _fac_ew_all = {f: _fi.get(f, 0.0) + _fx.get(f, 0.0) for f in set(_fi) | set(_fx)}

def render_tab_institutter(year, mode, inst_tot_size, inst_avg_size, institut_fakultets_map, edges_keep, node_meta, all_years_data=None, inst_ew=None, faculty_base_colors=None, size_map=None, total_pubs=None, filter_caption=None, show_intra=True):
    if not inst_tot_size or not inst_avg_size:
        st.error("Ingen institutsdata for det valgte filter.")
        return

    st.markdown(
f"""
### Institutternes bidrag til sampublicering

Institutfanen analyserer, hvordan sampubliceringsaktiviteten på KU fordeler sig
mellem institutter i det valgte år.

Sammenlignet med fakultetsniveauet giver institutniveauet et **mere detaljeret
billede af forskningssamarbejdet**, men er samtidig mere følsomt over for lokale
forhold som forskningsorganisering, laboratorier, centre og projektstrukturer.
Resultater på dette niveau viser derfor **strukturelle mønstre i samarbejde**
snarere end generelle vurderinger af institutters performance.

Fanen belyser forskelle i samarbejdsomfang, samarbejdsintensitet og rækkevidde
og supplerer de overordnede analyser ved at vise, hvor variationen inden for
fakulteterne faktisk opstår.

Opgørelserne i fanen bygger på de samme grundlæggende mål som i de øvrige
organisatoriske faner:

- **Forfatterantal** viser antallet af unikke KU‑VIP‑forskere, der er knyttet
  til hvert institut.
- **Forfatterpar** viser, hvor mange forfatterpar et institut indgår i. Et
  samarbejde mellem to institutter tælles én gang hos hvert institut.
- **Sampubliceringsrate** sætter forfatterpar i forhold til forfatterantal og
  muliggør sammenligning af samarbejdsintensitet mellem institutter med meget
  forskellig størrelse.

Brug filtret nedenfor til at zoome ind på et enkelt fakultet. 
""")

    if size_map is None:
        size_map = inst_tot_size

    # ── Farver ────────────────────────────────────────────────────────────────
    _inst_color_map = {}
    if faculty_base_colors:
        _insts_by_fac: dict[str, list] = {}
        for m in node_meta.values():
            if m.get("inst") and m.get("fac"):
                _insts_by_fac.setdefault(m["fac"], [])
                if m["inst"] not in _insts_by_fac[m["fac"]]:
                    _insts_by_fac[m["fac"]].append(m["inst"])
        for fac, insts in _insts_by_fac.items():
            insts_sorted = sorted(insts)
            k = max(1, len(insts_sorted))
            base = faculty_base_colors.get(fac, "#122947")
            for rank, inst in enumerate(insts_sorted):
                t = rank / max(1, k - 1)
                lf = 0.5 + 1.8 * t
                sf = 1.3 - 0.7 * t
                _inst_color_map[inst] = adjust_color(base, lf, sf)

    # ── Lokalt fakultetsfilter ────────────────────────────────────────────────
    _all_facs_in_tab = sorted({
        institut_fakultets_map.get(inst, "")
        for inst in inst_tot_size
        if institut_fakultets_map.get(inst, "")
    })
    _fac_filter = st.multiselect(
        "**Filtrer på fakultet**",
        options=_all_facs_in_tab,
        default=[],
        key=f"inst_tab_fac_filter_{year}_{mode}",
        placeholder="Alle fakulteter",
    )
    def _fac_ok(inst):
        return not _fac_filter or institut_fakultets_map.get(inst, "") in _fac_filter

    # ── Beregnede størrelser ──────────────────────────────────────────────────
    _insts_filt     = [i for i in sorted(inst_tot_size, key=lambda i: -inst_tot_size[i]) if _fac_ok(i)]
    _insts_avg_filt = [i for i in sorted(inst_avg_size, key=lambda i: -inst_avg_size[i]) if _fac_ok(i)]
    _tots           = [int(inst_tot_size[i]) for i in _insts_filt]
    _avgs           = [round(inst_avg_size[i], 1) for i in _insts_avg_filt]
    _inst_ew_filt   = {k: v for k, v in (inst_ew or {}).items() if _fac_ok(k)}
    _inst_tot_filt  = {k: v for k, v in inst_tot_size.items() if _fac_ok(k)}
    _grand_tot      = sum(_tots) or 1
    if total_pubs and not _fac_filter:
        _grand_ew = total_pubs
    else:
        _grand_ew = (sum(_inst_ew_filt.values()) / 2) if _inst_ew_filt else 0
    _inst_colors    = [_inst_color_map.get(i, "#122947") for i in _insts_filt]

    _sorted_tot = sorted(_inst_tot_filt.items(), key=lambda x: -x[1])
    _top_inst, _top_val = _sorted_tot[0]
    _bot_inst, _bot_val = _sorted_tot[-1]
    _top_share = 100 * _top_val / _grand_tot
    _sorted_avg = sorted({k: v for k, v in inst_avg_size.items() if _fac_ok(k)}.items(), key=lambda x: -x[1])
    _avg_top, _avg_top_val = _sorted_avg[0]
    _avg_bot, _avg_bot_val = _sorted_avg[-1]

    # ── Tekst-summary ─────────────────────────────────────────────────────────
    _ew_summary = ""
    if _inst_ew_filt and _grand_ew:
        _ew_top, _ew_top_val = max(_inst_ew_filt.items(), key=lambda x: x[1])
        _ew_share = 100 * _ew_top_val / _grand_ew
        _ratio_sorted = sorted(
            [(k, _inst_ew_filt[k] / _inst_tot_filt[k]) for k in _inst_tot_filt if _inst_tot_filt.get(k) and _inst_ew_filt.get(k)],
            key=lambda x: -x[1],
        )
        _rat_top, _rat_top_val = _ratio_sorted[0]
        _rat_bot, _rat_bot_val = _ratio_sorted[-1]

    st.markdown("---")

    st.markdown(
f"""#### Institutternes forfattere i {year}

Figuren viser fordelingen af **unikke KU‑VIP‑forfattere** på institutniveau.
Store institutter vil naturligt bidrage med flere forfattere, men fordelingen
giver samtidig indblik i, hvor forskningsaktiviteten er koncentreret, og hvor
den er mere fragmenteret.
 
**{_top_inst}** bidrager mest med **{int(_top_val)}** ({fmt_ui(_top_share)}% af samtlige bidrag), 
mens **{_bot_inst}** bidrager mindst med **{int(_bot_val)}**. 
""" + (
f"""
**Gennemsnitligt forfatterantal** er højest ved **{_avg_top}** ({fmt_ui(_avg_top_val)}) og 
lavest ved **{_avg_bot}** ({fmt_ui(_avg_bot_val)}). Gennemsnittet er beregnet som det samlede 
antal forfattere divideret med antallet af noder under instituttet - dvs. antallet af unikke 
stillingsgrupper i den valgte visning. Et højt gennemsnit indikerer, at de enkelte 
stillingsgrupper ved instituttet er relativt store.
""" if "G" in mode else ""))

    # ── Forfatterbidrag-plots ─────────────────────────────────────────────────
    _inst_tab_labels = ["Antal", "Andel (%)"]
    if "G" in mode:
        _inst_tab_labels.append("Gns. forfatterantal")
    _inst_tabs = st.tabs(_inst_tab_labels)
    _tab_it = _inst_tabs[0]
    _tab_ip = _inst_tabs[1]
    _tab_ia = _inst_tabs[2] if "G" in mode else None

    with _tab_it:
        _fig_it = go.Figure(go.Bar(
            y=_insts_filt, x=_tots, orientation="h",
            marker_color=_inst_colors,
            text=[f"{v}" for v in _tots],
            textposition="inside",
        ))
        _fig_it.update_layout(
            title = f"Forfatterantal per institut, {year}",
            xaxis_title="Forfatterantal",
            height=max(350, 35 * len(_insts_filt)),
            margin=dict(l=160, t=50, r=80),
            yaxis=dict(autorange="reversed"),
        )
        st.plotly_chart(_fig_it, width="stretch", key=f"inst_tot__{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
        )
        if filter_caption:
            st.caption(filter_caption)

    with _tab_ip:
        _pct_vals = [round(100 * v / _grand_tot, 1) for v in _tots]
        _fig_ip = go.Figure(go.Bar(
            y=_insts_filt, x=_pct_vals, orientation="h",
            marker_color=_inst_colors,
            text=[f"{fmt_ui(v)}%" for v in _pct_vals],
            textposition="inside",
        ))
        _fig_ip.update_layout(
            title=f"Institutternes andel af samlet forfatterantal, {year}",
            xaxis=dict(title="Andel (%)", range=[0, 100]),
            height=max(350, 35 * len(_insts_filt)),
            margin=dict(l=160, t=50, r=80),
            yaxis=dict(autorange="reversed"),
        )
        st.plotly_chart(_fig_ip, width="stretch", key=f"inst_pct_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
        )
        if filter_caption:
            st.caption(filter_caption)
    
    if _tab_ia:
        with _tab_ia:
            _fig_ia = go.Figure(go.Bar(
                y=_insts_avg_filt, x=_avgs, orientation="h",
                marker_color=[_inst_color_map.get(i, "#122947") for i in _insts_avg_filt],
                text=[f"{fmt_ui(v)}" for v in _avgs],
                textposition="inside",
            ))
            _fig_ia.update_layout(
                title=f"Institutternes gennemsnitlige forfatterantal, {year}",
                xaxis_title="Gns. forfatterantal",
                height=max(350, 35 * len(_insts_avg_filt)),
                margin=dict(l=160, t=50, r=80),
                yaxis=dict(autorange="reversed"),
            )
            st.plotly_chart(_fig_ia, width="stretch", key=f"inst_avg_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}})
            if filter_caption:
                st.caption(filter_caption)

    # ── Tabel ─────────────────────────────────────────────────────────────────
    _inst_summary_rows = [
        {
            "Fakultet":                institut_fakultets_map.get(inst, ""),
            "Institut":                inst,
            "Antal forfattere":  int(_inst_tot_filt.get(inst, 0)),
            "Andel (%)":              round(100 * _inst_tot_filt.get(inst, 0) / _grand_tot, 1),
            **( {"Gns. forfatterantal": round(inst_avg_size.get(inst, 0), 1)} if "G" in mode else {} ),
        }
        for inst in _insts_filt
    ]
    _inst_summary_schema = [
        ("Fakultet",               pa.string()),
        ("Institut",               pa.string()),
        ("Antal forfattere", pa.int64()),
        ("Andel (%)",              pa.float64()),
        *( [("Gns. forfatterantal", pa.float64())] if "G" in mode else [] ),
    ]
    with st.expander("Se tabel"):
        st.dataframe(build_table(_inst_summary_rows, _inst_summary_schema), hide_index=True, width="stretch")
        st.download_button(
            "Download (.xlsx)",
            data=rows_to_excel_bytes(_inst_summary_rows, [n for n, _ in _inst_summary_schema]),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl__inst_bidrag_{year}_{mode}",
        )
    
 
    # ── Forfatterpar-plots ────────────────────────────────────────────────────
    _inst_ew_all_yr = {}
    _active_inst_ew = _inst_ew_filt
    
    st.markdown("---")

    if _inst_ew_filt and _grand_ew:
        st.markdown(
f"""#### Forfatterpar per institut i {year}

**Forfatterpar** viser det samlede samarbejdsomfang for hvert institut.
Et samarbejde mellem to institutter tælles én gang hos hvert institut,
hvilket betyder, at summen af forfatterpar på tværs af institutter svarer
til **to gange** det samlede antal unikke forfatterpar i netværket.

Målet afspejler, hvor intensivt et institut indgår i sampublicering, men
siger ikke noget om, hvor mange forskellige samarbejdspartnere instituttet
har – kun hvor omfattende samarbejdet er.

**Eksempel**: et institut har 100 forfattere, som indgår i 300 forfatterpar. Sampubliceringsraten er da 
300 / 100 = 3,0, hvilket betyder, at den gennemsnitlige forfatter indgår i tre forfatterpar. 
Et andet institut med 50 forfattere og 200 forfatterpar har en højere rate på 4,0 - selvom 
det samlede samarbejdsomfang er lavere. 
""")

        _inst_ew_all_yr = {}
        if all_years_data and year in all_years_data:
            _ii = all_years_data[year].get("inst_intra_ew_all", {})
            _ix = all_years_data[year].get("inst_inter_ew_all", {})
            _inst_ew_all_yr = {i: _ii.get(i, 0.0) + _ix.get(i, 0.0) for i in set(_ii) | set(_ix)}

        _use_inst_all = st.toggle(
            "Inkludér intra-enhed forfatterpar (f.eks. to professorer på samme institut)",
            value=False,
            key=f"inst_ew_bar_all_{year}_{mode}",
            help="Til: alle forfatterpar inkl. par inden for samme enhed. Fra: kun netværkskanter.",
            disabled=not show_intra
        )
        if _use_inst_all:
            st.caption("Alle forfatterpar inkl. par inden for samme enhed (f.eks. to mandlige professorer på Kemisk Institut).")
        else:
            st.caption("Kun forfatterpar *mellem* de valgte organisatoriske enheder - som vist i netværket.")

        _active_inst_ew = {k: v for k, v in _inst_ew_all_yr.items() if _fac_ok(k)} \
            if (_use_inst_all and _inst_ew_all_yr) else _inst_ew_filt
        if _use_inst_all:
            # Intra-toggle på: ingen global total_pubs-pendant inkl. intra findes,
            # så brug sum/2 af aktiv ew (den dobbelttalte sum svarer til 2 × antal par inkl. intra).
            _grand_active_inst_ew = (sum(_active_inst_ew.values()) / 2) or 1
        elif total_pubs and not _fac_filter:
            # Intra-toggle fra, intet lokalt filter: brug global total_pubs.
            _grand_active_inst_ew = total_pubs
        else:
            # Intra-toggle fra, men lokalt fakultetsfilter aktivt: lokal total = sum/2.
            _grand_active_inst_ew = (sum(_active_inst_ew.values()) / 2) or 1

        _ews_ord      = [round(_active_inst_ew.get(i, 0), 1) for i in _insts_filt]
        _ew_ratio_ord = [round(_active_inst_ew.get(i, 0) / (_inst_tot_filt.get(i) or 1), 3) for i in _insts_filt]

        _pct_ews_inst = [round(100 * v / _grand_active_inst_ew, 1) for v in _ews_ord]
        _intra_suffix = ' (inkl. intra-enhed)' if _use_inst_all else ''
        _tab_ew_abs, _tab_ew_pct, _tab_ew_ratio = st.tabs(["Antal", "Andel (%)", "Forfatterpar per forfatter"])

        with _tab_ew_abs:
            _fig_ew = go.Figure(go.Bar(
                y=_insts_filt, x=_ews_ord, orientation="h",
                marker_color=_inst_colors,
                text=[f"{fmt_ui(v)}" for v in _ews_ord],
                textposition="inside",
            ))
            _fig_ew.update_layout(
                title=f"Forfatterpar per institut{_intra_suffix}, {year}",
                xaxis_title="Antal forfatterpar",
                height=max(350, 35 * len(_insts_filt)),
                margin=dict(l=160, t=50, r=80),
                yaxis=dict(autorange="reversed"),
            )
            st.plotly_chart(_fig_ew, width="stretch", key=f"inst_ew_abs_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}})
            if filter_caption:
                st.caption(filter_caption)

        with _tab_ew_pct:
            _fig_ew_p = go.Figure(go.Bar(
                y=_insts_filt, x=_pct_ews_inst, orientation="h",
                marker_color=_inst_colors,
                text=[f"{fmt_ui(v)}%" for v in _pct_ews_inst],
                textposition="inside",
            ))
            _fig_ew_p.update_layout(
                title=f"Forfatterpar per institut{_intra_suffix} - andel (%), {year}",
                xaxis=dict(title="Andel (%)", range=[0, 100]),
                height=max(350, 35 * len(_insts_filt)),
                margin=dict(l=160, t=50, r=80),
                yaxis=dict(autorange="reversed"),
            )
            st.caption("Vær opmærksom på, at procenter kan overstige 100%. Se **Læsning af procenter** i Oversigt-fanen.")
            st.plotly_chart(_fig_ew_p, width="stretch", key=f"inst_ew_pct_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}})
            if filter_caption:
                st.caption(filter_caption)

        with _tab_ew_ratio:
            _fig_ew_r = go.Figure(go.Bar(
                y=_insts_filt, x=_ew_ratio_ord, orientation="h",
                marker_color=_inst_colors,
                text=[f"{fmt_ui(v,3)}" for v in _ew_ratio_ord],
                textposition="inside",
            ))
            _fig_ew_r.update_layout(
                title=f"Forfatterpar per forfatterantal per institut{_intra_suffix}, {year}",
                xaxis_title="Forfatterpar per forfatter",
                height=max(350, 35 * len(_insts_filt)),
                margin=dict(l=160, t=50, r=80),
                yaxis=dict(autorange="reversed"),
            )
            st.plotly_chart(_fig_ew_r, width="stretch", key=f"inst_ew_ratio_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}})
            if filter_caption:
                st.caption(filter_caption)

    # ── Tabel ─────────────────────────────────────────────────────────────────
    _inst_summary_rows = [
        {"Fakultet":                      institut_fakultets_map.get(inst, ""),
         "Institut":                      inst,
         "Forfatterpar (netværk)":        round(_inst_ew_filt.get(inst, 0), 1),
         "Forfatterpar (inkl. intra)":    round(_inst_ew_all_yr.get(inst, 0), 1) if _inst_ew_all_yr else 0}
        for inst in _insts_filt
    ]
    _inst_summary_schema = [
        ("Fakultet",                    pa.string()),
        ("Institut",                    pa.string()),
        ("Forfatterpar (netværk)",      pa.float64()),
        ("Forfatterpar (inkl. intra)",  pa.float64()),
    ]
    with st.expander("Se tabel"):
        st.dataframe(build_table(_inst_summary_rows, _inst_summary_schema), width="stretch", hide_index=True)
        st.download_button(
            "Download (.xlsx)",
            data=rows_to_excel_bytes(_inst_summary_rows, [n for n, _ in _inst_summary_schema]),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_inst_bidrag_{year}_{mode}",
        )

    st.markdown("---")

    # ── Institutternes rækkevidde ─────────────────────────────────────────────
    st.markdown(
f"""#### Institutternes rækkevidde i {year}

**Rækkevidde** måler, hvor mange forskellige institutter et institut
samarbejder med – uanset hvor intensivt samarbejdet er.

En høj rækkevidde peger på et relativt bredt samarbejdsmønster på tværs af
KU, mens lav rækkevidde indikerer, at samarbejdet er koncentreret omkring
få, faste partnere.

**Eksempel**: et institut samarbejder med ti andre institutter, men kun i meget begrænset omfang. 
Et andet andet institut samarbejder med tre faste partnere. Det første institut har højere rækkevidde, 
selvom det andet samlet set kan have flere forfatterpar. 
""")

    _inst_partners_reach: dict[str, set] = {}
    _inst_fac_reach: dict[str, set] = {}
    _inst_intra_ew: dict[str, float] = {}
    _inst_inter_ew: dict[str, float] = {}
    for u, v, w, *_ in edges_keep:
        iu = node_meta.get(u, {}).get("inst", "")
        iv = node_meta.get(v, {}).get("inst", "")
        fu = node_meta.get(u, {}).get("fac", "")
        fv = node_meta.get(v, {}).get("fac", "")
        if iu and iv and iu != iv:
            _inst_partners_reach.setdefault(iu, set()).add(iv)
            _inst_partners_reach.setdefault(iv, set()).add(iu)
        if iu and fv and fu != fv:
            _inst_fac_reach.setdefault(iu, set()).add(fv)
        if iv and fu and fu != fv:
            _inst_fac_reach.setdefault(iv, set()).add(fu)
        mu, mv = node_meta.get(u, {}), node_meta.get(v, {})
        cross = mu.get("fac", "") != mv.get("fac", "")
        for inst in [mu.get("inst", ""), mv.get("inst", "")]:
            if not inst: continue
            if cross:
                _inst_inter_ew[inst] = _inst_inter_ew.get(inst, 0.0) + w
            else:
                _inst_intra_ew[inst] = _inst_intra_ew.get(inst, 0.0) + w

    _all_insts_reach = sorted({
        node_meta.get(nid, {}).get("inst", "")
        for nid in node_meta
        if node_meta.get(nid, {}).get("inst") and _fac_ok(node_meta.get(nid, {}).get("inst", ""))
    })
    _all_facs_reach  = sorted({node_meta.get(nid, {}).get("fac", "") for nid in node_meta if node_meta.get(nid, {}).get("fac")})
    _max_inst_reach  = len(_all_insts_reach) - 1
    _max_fac_reach   = len(_all_facs_reach) - 1

    _inst_reach_sorted = sorted(
        [(i, len(_inst_partners_reach.get(i, set()))) for i in _all_insts_reach], key=lambda x: -x[1])
    _inst_fac_reach_sorted = sorted(
        [(i, len(_inst_fac_reach.get(i, set()))) for i in _all_insts_reach], key=lambda x: -x[1])

    _show_inter_tab = bool(edges_keep and mode in ("FI", "FIG", "IG"))
    _reach_tab_labels = ["Samarbejde med andre institutter", "Samarbejde med andre fakulteter"]
    if _show_inter_tab:
        _reach_tab_labels.append("Inter-fakultet andel (%)")
    
    if not any(n > 0 for _, n in _inst_reach_sorted):
        st.error("Ingen forfatterpar i det valgte udsnit - rækkevidde kan ikke beregnes.")
    else:
        _reach_tabs = st.tabs(_reach_tab_labels)

        with _reach_tabs[0]:
            st.markdown("Antal *forskellige* institutter hvert institut sampublicerer med.")
            if any(n > 0 for _, n in _inst_reach_sorted):
                _fig_inst_reach = go.Figure(go.Bar(
                    y=[i for i, _ in _inst_reach_sorted],
                    x=[n for _, n in _inst_reach_sorted],
                    orientation="h",
                    marker_color=[_inst_color_map.get(i, "#122947") for i, _ in _inst_reach_sorted],
                    text=[f"{n} / 48" for _, n in _inst_reach_sorted],
                    textposition="inside",
                ))
                _fig_inst_reach.update_layout(
                    title=f"Samarbejde med andre institutter, {year}",
                    xaxis=dict(title="Antal samarbejdsinstitutter", range=[0, _max_inst_reach + 0.5], dtick=1),
                    height=max(300, 28 * len(_inst_reach_sorted)),
                    margin=dict(l=200, t=50, r=80),
                    yaxis=dict(autorange="reversed"),
                )
                st.plotly_chart(_fig_inst_reach, width="stretch", key=f"inst_reach_{year}_{mode}",
                    config={"toImageButtonOptions": {"format": "png", "scale": 3}})
                if filter_caption:
                    st.caption(filter_caption)

        with _reach_tabs[1]:
            st.markdown("Antal *forskellige* fakulteter hvert institut sampublicerer med.")
            if any(n > 0 for _, n in _inst_fac_reach_sorted):
                _fig_fac_reach = go.Figure(go.Bar(
                    y=[i for i, _ in _inst_fac_reach_sorted],
                    x=[n for _, n in _inst_fac_reach_sorted],
                    orientation="h",
                    marker_color=[_inst_color_map.get(i, "#122947") for i, _ in _inst_fac_reach_sorted],
                    text=[f"{n} / {_max_fac_reach}" for _, n in _inst_fac_reach_sorted],
                    textposition="inside",
                ))
                _fig_fac_reach.update_layout(
                    title=f"Samarbejde med andre fakulteter, {year}",
                    xaxis=dict(title="Antal samarbejdsfakulteter", range=[0, _max_fac_reach + 0.5], dtick=1),
                    height=max(300, 28 * len(_inst_fac_reach_sorted)),
                    margin=dict(l=200, t=50, r=80),
                    yaxis=dict(autorange="reversed"),
                )
                st.plotly_chart(_fig_fac_reach, width="stretch", key=f"inst_fac_reach_{year}_{mode}",
                    config={"toImageButtonOptions": {"format": "png", "scale": 3}})
                if filter_caption:
                    st.caption(filter_caption)

        if _show_inter_tab:
            with _reach_tabs[2]:
                st.markdown(
                    "Andelen af hvert instituts **tværgående forfatterpar**, der involverer et institut fra et **andet fakultet**. "
                    "Institutter med en høj andel samarbejder i særlig grad på tværs af fakultetsgrænser."
                )
                _all_insts_inter = sorted(set(_inst_intra_ew) | set(_inst_inter_ew))
                _inter_rows = []
                for inst in _all_insts_inter:
                    intra = _inst_intra_ew.get(inst, 0.0)
                    inter = _inst_inter_ew.get(inst, 0.0)
                    total = intra + inter
                    _inter_rows.append({
                        "Fakultet":            institut_fakultets_map.get(inst, ""),
                        "Institut":            inst,
                        "Inter-fakultet (%)":  round(100 * inter / total, 1) if total else 0.0,
                        "Inter-fakultet (fp)": round(inter, 1),
                        "Total forfatterpar":  round(total, 1),
                    })
                if _fac_filter:
                    _inter_rows = [r for r in _inter_rows if r["Fakultet"] in _fac_filter]
                _inter_rows.sort(key=lambda r: -r["Inter-fakultet (%)"])
                if _inter_rows:
                    _fig_inter = go.Figure(go.Bar(
                        y=[r["Institut"] for r in _inter_rows],
                        x=[r["Inter-fakultet (%)"] for r in _inter_rows],
                        orientation="h",
                        marker_color=[_inst_color_map.get(r["Institut"], "#122947") for r in _inter_rows],
                        text=[f"{fmt_ui(r['Inter-fakultet (%)'])}%  ({fmt_ui(r['Inter-fakultet (fp)'],0)} fp)" for r in _inter_rows],
                        textposition="inside",
                    ))
                    _fig_inter.update_layout(
                        title=f"Inter-fakultet andel per institut, {year}",
                        xaxis_title="Andel inter-fakultet forfatterpar (%)",
                        xaxis_range=[0, 100],
                        height=max(400, 28 * len(_inter_rows)),
                        margin=dict(l=200, r=80, t=50),
                        yaxis=dict(autorange="reversed"),
                    )
                    st.plotly_chart(_fig_inter, width="stretch", key=f"inst_inter_{year}_{mode}",
                        config={"toImageButtonOptions": {"format": "png", "scale": 3}})
                    if filter_caption:
                        st.caption(filter_caption)
    
    # ── Rækkevidde-tabel (udenfor faner) ──────────────────────────────────────
    _reach_rows = []
    for inst in _all_insts_reach:
        row = {
            "Fakultet":                    institut_fakultets_map.get(inst, ""),
            "Institut":                    inst,
            "Samarbejdsinstitutter":       len(_inst_partners_reach.get(inst, set())),
            "Samarbejdsfakulteter":        len(_inst_fac_reach.get(inst, set())),
        }
        if _show_inter_tab:
            intra = _inst_intra_ew.get(inst, 0.0)
            inter = _inst_inter_ew.get(inst, 0.0)
            total = intra + inter
            row["Inter-fakultet (%)"]  = round(100 * inter / total, 1) if total else 0.0
            row["Inter-fakultet (fp)"] = round(inter, 1)
            row["Total forfatterpar"]  = round(total, 1)
        _reach_rows.append(row)

    _reach_schema = [
        ("Fakultet",             pa.string()),
        ("Institut",             pa.string()),
        ("Samarbejdsinstitutter", pa.int64()),
        ("Samarbejdsfakulteter",  pa.int64()),
        *( [("Inter-fakultet (%)",  pa.float64()),
            ("Inter-fakultet (fp)", pa.float64()),
            ("Total forfatterpar",  pa.float64())] if _show_inter_tab else [] ),
    ]
    with st.expander("Se tabel"):
        st.dataframe(build_table(_reach_rows, _reach_schema), hide_index=True, width="stretch")
        st.download_button(
            "Download (.xlsx)",
            data=rows_to_excel_bytes(_reach_rows, [n for n, _ in _reach_schema]),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_inst_raekkevidde_{year}_{mode}",
        )
    

    st.markdown("---")

    # ── Samarbejdsmønstre på tværs af institutter (heatmap) ──────────────────
    st.markdown(
f"""#### Samarbejdsmønstre på tværs af institutter i {year}

Heatmappet viser, **hvilke institutter der publicerer sammen** - og hvor intensivt.

- **Antal forfatterpar** viser den rå trafik mellem to institutter. Store institutter dominerer naturligt.
- **Forfatterpar per forfatter** svarer på spørgsmålet: *hvor ofte er en forfatter fra kolonnen 
med, når nogen fra rækken publicerer?* Matricen er ikke symmetrisk - cellen (A, B) og cellen (B, A) 
svarer på to forskellige spørgsmål. Et lille institut kan have høj andel af sine par med et stort 
institut, uden at det omvendte gælder.
- **Andel (%) per række** viser for hvert institut, hvor dets samarbejder fordeler sig. Hver række 
summerer til 100%.

**Bemærk**: ratioen kan være større end 1. Ét forfatter kan indgå i flere par - hvis en 
forsker er medforfatter på et paper med tre forskere fra et andet institut, genererer det ét bidrag 
men tre par med det andet institut. Store forfatterskaber (f.eks. artikler med 100+ 
forfattere) kan derfor skævvride billedet.
""")

    # Bestem hvilke institutter der skal indgå i heatmappet.
    # Hvis brugeren har filtreret på fakultet(er) ovenfor, bruges alle institutter fra de valgte fakulteter.
    # Ellers skal brugeren vælge institutter manuelt, da et 50+ × 50+ heatmap er ulæseligt.
    _hm_inst_candidates = [i for i in inst_tot_size if _fac_ok(i)]

    if _fac_filter:
        # Fakultets-filter er aktivt: brug alle institutter under de valgte fakulteter
        _hm_insts_selected = sorted(_hm_inst_candidates, key=lambda i: -inst_tot_size.get(i, 0))
        st.caption(
            f"Heatmappet viser alle institutter under **{', '.join(_fac_filter)}** "
            f"({len(_hm_insts_selected)} institutter)."
        )
    else:
        # Intet fakultets-filter: lad brugeren vælge institutter direkte
        st.markdown(
            "Uden et fakultets-filter er der for mange institutter til at vise meningsfuldt. "
            "Vælg enten et fakultet i filtret ovenfor, eller vælg institutter direkte her:"
        )
        _default_top_n = sorted(_hm_inst_candidates, key=lambda i: -inst_tot_size.get(i, 0))[:10]
        _hm_insts_selected = st.multiselect(
            "**Vælg institutter til heatmap**",
            options=sorted(_hm_inst_candidates),
            default=_default_top_n,
            key=f"inst_hm_select_{year}_{mode}",
            help="Default: de 10 institutter med fleste forfattere. Ryd feltet og tilføj dine egne valg.",
        )

    _use_inst_hm_all = st.toggle(
        "Inkludér intra-institut forfatterpar (f.eks. to professorer på samme institut)",
        value=False,
        key=f"inst_hm_all_{year}_{mode}",
        help="Til: alle forfatterpar inkl. par inden for samme institut. Fra: kun par mellem institutter.",
        disabled=not show_intra
    )

    # Byg par-matrix fra kanterne. Halveret vægt (w/2) matcher bar-chart-ratioen ovenfor.
    _inst_pair_matrix: dict[tuple[str, str], float] = {}
    _selected_set = set(_hm_insts_selected)
    for u, v, w, *_ in (edges_keep or []):
        iu = (node_meta or {}).get(u, {}).get("inst", "")
        iv = (node_meta or {}).get(v, {}).get("inst", "")
        if not iu or not iv:
            continue
        if iu not in _selected_set or iv not in _selected_set:
            continue  # Begge endepunkter skal være blandt de valgte institutter
        if iu == iv and not _use_inst_hm_all:
            continue  # spring intra-institut par over hvis toggle er slået fra
        _inst_pair_matrix[(iu, iv)] = _inst_pair_matrix.get((iu, iv), 0.0) + w
        if iu != iv:
            _inst_pair_matrix[(iv, iu)] = _inst_pair_matrix.get((iv, iu), 0.0) + w

    # Hvis intra-institut toggle er på, brug pre-computede intra-par på diagonalen
    if _use_inst_hm_all and all_years_data and year in all_years_data:
        _inst_intra_src = all_years_data[year].get("inst_intra_ew_all", {})
        for i, v in _inst_intra_src.items():
            if i in _selected_set:
                _inst_pair_matrix[(i, i)] = v

    _intra_suffix_inst_hm = ' (inkl. intra-institut)' if _use_inst_hm_all else ''

    if len(_hm_insts_selected) < 2:
        st.error("Vælg mindst to institutter for at vise samarbejdsmatricen.")
    elif not _inst_pair_matrix:
        st.error("Ingen forfatterpar mellem de valgte institutter i det valgte udsnit.")
    else:
        # Sortér valgte institutter efter størrelse (størst først) for konsistent visning
        _hm_insts_ordered = sorted(_hm_insts_selected, key=lambda i: -inst_tot_size.get(i, 0))

        _z_abs_inst = [[round(_inst_pair_matrix.get((ii, ij), 0.0), 1) for ij in _hm_insts_ordered] for ii in _hm_insts_ordered]
        _z_ratio_inst = [
            [round(_inst_pair_matrix.get((ii, ij), 0.0) / (inst_tot_size.get(ii) or 1), 3) for ij in _hm_insts_ordered]
            for ii in _hm_insts_ordered
        ]
        _z_row_pct_inst = []
        for ii in _hm_insts_ordered:
            _row_total = sum(_inst_pair_matrix.get((ii, ij), 0.0) for ij in _hm_insts_ordered) or 1
            _z_row_pct_inst.append([
                round(100 * _inst_pair_matrix.get((ii, ij), 0.0) / _row_total, 2)
                for ij in _hm_insts_ordered
            ])

        _tab_hm_abs_inst, _tab_hm_ratio_inst, _tab_hm_pct_inst = st.tabs(
            ["Antal forfatterpar", "Forfatterpar per forfatter", "Andel (%) per række"]
        )

        def _inst_heatmap_fig(z, title, colorbar_title, text_fmt, hover_template):
            _text = [[text_fmt(v) for v in row] for row in z]
            # Højde skaleres med antal valgte institutter; bredde låst til container
            _n = len(_hm_insts_ordered)
            fig = go.Figure(go.Heatmap(
                z=z, x=_hm_insts_ordered, y=_hm_insts_ordered,
                colorscale=[[0, "#f0f4f8"], [1, "#122947"]],
                text=_text, texttemplate="%{text}",
                textfont=dict(size=10 if _n > 15 else 11),
                hovertemplate=hover_template,
                colorbar=dict(title=colorbar_title),
            ))
            fig.update_layout(
                title=title,
                xaxis=dict(title="", tickangle=-45),
                yaxis=dict(title="", autorange="reversed"),
                height=max(400, 60 * _n),
                margin=dict(l=250, t=60, r=80, b=200),
            )
            return fig

        with _tab_hm_abs_inst:
            st.plotly_chart(
                _inst_heatmap_fig(
                    _z_abs_inst,
                    f"Antal forfatterpar mellem institutter{_intra_suffix_inst_hm}, {year}",
                    "Antal par",
                    lambda v: fmt_ui(v) if v else "",
                    "<b>%{y}</b> og <b>%{x}</b><br>"
                    "Antal forfatterpar: <b>%{z}</b>"
                    "<extra></extra>",
                ),
                width="stretch",
                key=f"inst_hm_abs_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}},
            )
            if filter_caption:
                st.caption(filter_caption)

        with _tab_hm_ratio_inst:
            st.plotly_chart(
                _inst_heatmap_fig(
                    _z_ratio_inst,
                    f"Forfatterpar per forfatter mellem institutter{_intra_suffix_inst_hm}, {year}",
                    "Par per bidrag",
                    lambda v: fmt_ui(v, 2) if v else "",
                    "Per forfatter fra <b>%{y}</b><br>"
                    "dannes i gennemsnit <b>%{z}</b> par med <b>%{x}</b>"
                    "<extra></extra>",
                ),
                width="stretch",
                key=f"inst_hm_ratio_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}},
            )
            if filter_caption:
                st.caption(filter_caption)

        with _tab_hm_pct_inst:
            st.plotly_chart(
                _inst_heatmap_fig(
                    _z_row_pct_inst,
                    f"Fordeling af samarbejdspartnere per institut{_intra_suffix_inst_hm}, {year}",
                    "Andel (%)",
                    lambda v: f"{fmt_ui(v, 1)}%" if v else "",
                    "<b>%{z:.2f}%</b> af <b>%{y}</b>s forfatterpar<br>"
                    "er med <b>%{x}</b>"
                    "<extra></extra>",
                ),
                width="stretch",
                key=f"inst_hm_pct_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}},
            )
            if filter_caption:
                st.caption(filter_caption)

        # ── Tabel (langt format) ──────────────────────────────────────────────
        _inst_hm_rows = []
        for ii in _hm_insts_ordered:
            _row_total = sum(_inst_pair_matrix.get((ii, ij), 0.0) for ij in _hm_insts_ordered) or 1
            _bidrag_i = inst_tot_size.get(ii) or 1
            for ij in _hm_insts_ordered:
                _pairs = _inst_pair_matrix.get((ii, ij), 0.0)
                if _pairs == 0:
                    continue
                _inst_hm_rows.append({
                    "Fra (række)":                 ii,
                    "Til (kolonne)":               ij,
                    "Forfatterpar":                round(_pairs, 1),
                    "Par per forfatter":     round(_pairs / _bidrag_i, 3),
                    "Andel af rækkens par (%)":    round(100 * _pairs / _row_total, 2),
                })
        _inst_hm_schema = [
            ("Fra (række)",              pa.string()),
            ("Til (kolonne)",            pa.string()),
            ("Forfatterpar",             pa.float64()),
            ("Par per forfatter",  pa.float64()),
            ("Andel af rækkens par (%)", pa.float64()),
        ]
        with st.expander("Se tabel"):
            st.dataframe(build_table(_inst_hm_rows, _inst_hm_schema), hide_index=True, width="stretch")
            st.download_button(
                "Download (.xlsx)",
                data=rows_to_excel_bytes(_inst_hm_rows, [n for n, _ in _inst_hm_schema]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_inst_hm_{year}_{mode}",
            )




    st.markdown("---")
    
    # ── Sammenlign år ─────────────────────────────────────────────────────────
    st.markdown(
f"""### Udvikling over tid

Afsnittet nedenfor viser udviklingen i institutternes **forfatterantal**,
**forfatterpar** og **samarbejde** på tværs af de tilgængelige år.

Tidsserierne gør det muligt at vurdere, om institutters samarbejdsmønstre
er stabile over tid eller ændrer sig som følge af organisatoriske,
strategiske eller faglige forskydninger.
""")
    _insts_in_data    = sorted({i for s in (all_years_data or {}).values() for i in s.get("inst_tot", {}) if _fac_ok(i)})
    _insts_ew_in_data = sorted({i for s in (all_years_data or {}).values() for i in s.get("inst_ew", {}) if _fac_ok(i)})
    years_sorted_inst = sorted((all_years_data or {}).keys())

    _tab_yr_tot, _tab_yr_ew, _tab_yr_reach_inst, _tab_yr_reach_fac = st.tabs([
        "Forfatterantal per institut",
        "Forfatterpar per institut",
        "Samarbejdsinstitutter",
        "Samarbejdsfakulteter",
    ])

    with _tab_yr_tot:
        _render_year_comparison(
            all_years_data,
            series=[(inst, "inst_tot", inst) for inst in _insts_in_data],
            title=f"Forfatterantal per institut, {years_sorted_inst[0]}–{years_sorted_inst[-1]}",
            colors=_inst_color_map,
            key_suffix=f"inst__tot_{year}_{mode}",
            show_table=True,
        )

    with _tab_yr_ew:
        _use_inst_ew_all = st.toggle(
            "Inkludér intra-enhed forfatterpar (f.eks. to professorer på samme institut)",
            value=False,
            key=f"inst_ew_all_{year}_{mode}",
            help="Til: alle forfatterpar inkl. par inden for samme enhed. Fra: kun netværkskanter.",
            disabled=not show_intra
        )
        if _use_inst_ew_all:
            st.caption("Alle forfatterpar inkl. par inden for samme enhed (f.eks. to mandlige professorer på Kemisk Institut).")
        else:
            st.caption("Kun forfatterpar *mellem* de valgte organisatoriske enheder - som vist i netværket.")

        _inst_ew_key = "inst_ew_all" if _use_inst_ew_all else "inst_ew"
        _insts_ew_plot = sorted({
            i for s in (all_years_data or {}).values()
            for i in s.get(_inst_ew_key, {}) if _fac_ok(i)
        })

        _render_year_comparison(
            all_years_data,
            series=[(inst, _inst_ew_key, inst) for inst in _insts_ew_plot],
            title=f"Forfatterpar per institut{' (inkl. intra-enhed)' if _use_inst_ew_all else ''}, {years_sorted_inst[0]}–{years_sorted_inst[-1]}",
            yaxis_label="Antal forfatterpar",
            colors=_inst_color_map,
            key_suffix=f"inst_ew_{year}_{mode}",
            show_table=True,
        )

    with _tab_yr_reach_inst:
        st.markdown("Antal *forskellige* institutter hvert institut sampublicerer med, per år.")
        _insts_reach_in_data = sorted({
            i for s in (all_years_data or {}).values()
            for i in s.get("inst_reach", {}) if _fac_ok(i)
        })
        _render_year_comparison(
            all_years_data,
            series=[(inst, "inst_reach", inst) for inst in _insts_reach_in_data],
            title=f"Samarbejde med andre institutter, {years_sorted_inst[0]}–{years_sorted_inst[-1]}",
            yaxis_label="Antal samarbejdsinstitutter",
            colors=_inst_color_map,
            key_suffix=f"inst_reach_{year}_{mode}",
            show_table=True,
        )

    with _tab_yr_reach_fac:
        st.markdown("Antal *forskellige* fakulteter hvert institut sampublicerer med, per år.")
        _insts_fac_reach_in_data = sorted({
            i for s in (all_years_data or {}).values()
            for i in s.get("inst_fac_reach", {}) if _fac_ok(i)
        })
        _render_year_comparison(
            all_years_data,
            series=[(inst, "inst_fac_reach", inst) for inst in _insts_fac_reach_in_data],
            title=f"Institutternes samarbejde med andre fakulteter, {years_sorted_inst[0]}–{years_sorted_inst[-1]}",
            yaxis_label="Antal samarbejdsfakulteter",
            colors=_inst_color_map,
            key_suffix=f"inst_fac_reach_{year}_{mode}",
            show_table=True,
            height=380,
        )

    _insts_fac_reach_in_data = sorted({
        i for s in (all_years_data or {}).values()
        for i in s.get("inst_fac_reach", {}) if _fac_ok(i)
    })
    _yr_inst_rows = []
    for yr in years_sorted_inst:
        snap = (all_years_data or {}).get(yr, {})
        row = {"År": yr}
        for i in _insts_in_data:
            row[f"{i} (bidrag)"] = snap.get("inst_tot", {}).get(i, 0)
        for i in _insts_ew_in_data:
            row[f"{i} (forfatterpar)"] = round(snap.get("inst_ew", {}).get(i, 0.0), 1)
        for i in _insts_reach_in_data:
            row[f"{i} (rækkevidde inst)"] = snap.get("inst_reach", {}).get(i, 0)
        for i in _insts_fac_reach_in_data:
            row[f"{i} (rækkevidde fak)"] = snap.get("inst_fac_reach", {}).get(i, 0)
        _yr_inst_rows.append(row)
    _yr_inst_col_names = (
        ["År"]
        + [f"{i} (bidrag)"          for i in _insts_in_data]
        + [f"{i} (forfatterpar)"    for i in _insts_ew_in_data]
        + [f"{i} (rækkevidde inst)" for i in _insts_reach_in_data]
        + [f"{i} (rækkevidde fak)"  for i in _insts_fac_reach_in_data]
    )




def render_pos_chart(pos_counts: dict, tab_key: str, year: int, mode: str, grp_tot_size: dict = None):
    stillinger_sorted = sorted(pos_counts.keys(), key=lambda g: HIERARKI.get(g, 999))
    if not stillinger_sorted:
        st.error("Ingen data i det valgte udsnit.")
        return

    _pos_colors = ["#bac7d9", "#425570", "#901a1E"]
    pos_labels = [("first", "Førsteforfatter"), ("middle", "Mellemforfatter"), ("last", "Sidsteforfatter")]

    _tab_abs, _tab_pct, _tab_rate = st.tabs(["Antal", "Andel (%)", "Rate (per forfatter)"])

    with _tab_abs:
        fig = go.Figure()
        for (pos_name, label), color in zip(pos_labels, _pos_colors):
            fig.add_trace(go.Bar(
                name=label,
                y=stillinger_sorted,
                x=[pos_counts.get(g, {}).get(pos_name, 0) for g in stillinger_sorted],
                orientation="h",
                marker_color=color,
                text=[pos_counts.get(g, {}).get(pos_name, 0) for g in stillinger_sorted],
                textposition="inside",
            ))
        fig.update_layout(
            title=f"Forfatterpositioner per stillingsgruppe, {year}",
            barmode="stack", xaxis_title="Antal publikationer",
            height=max(350, 40 * len(stillinger_sorted)),
            legend_title="Forfatterposition",
            margin=dict(l=160, t=50, r=20),
            yaxis=dict(autorange="reversed"),
        )
        st.plotly_chart(fig, width="stretch", key=f"pos_abs_{tab_key}_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
    )
        

    with _tab_pct:
        fig_pct = go.Figure()
        for (pos_name, label), color in zip(pos_labels, _pos_colors):
            _vals = []
            for g in stillinger_sorted:
                tot = pos_counts.get(g, {}).get("total", 0) or 1
                _vals.append(round(100 * pos_counts.get(g, {}).get(pos_name, 0) / tot, 1))
            fig_pct.add_trace(go.Bar(
                name=label,
                y=stillinger_sorted,
                x=_vals,
                orientation="h",
                marker_color=color,
                text=[f"{fmt_ui(v)}%" for v in _vals],
                textposition="inside",
            ))
        fig_pct.update_layout(
            title=f"Forfatterpositioner per stillingsgruppe - andel (%), {year}",
            barmode="stack",
            xaxis=dict(title="Andel (%)", range=[0, 100]),
            height=max(350, 40 * len(stillinger_sorted)),
            legend_title="Forfatterposition",
            margin=dict(l=160, t=50, r=20),
            yaxis=dict(autorange="reversed"),
        )
        st.plotly_chart(fig_pct, width="stretch", key=f"pos_pct_{tab_key}_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
    )

    with _tab_rate:
        st.markdown("Antal gange en typisk forfatter i gruppen optræder i hver forfatterposition - beregnet som antal positioner divideret med antal unikke forfattere i gruppen.")
        if not grp_tot_size:
            st.error("Forfatterantal per stillingsgruppe ikke tilgængeligt - rate kan ikke beregnes.")
        else:
            fig_rate = go.Figure()
            for (pos_name, label), color in zip(pos_labels, _pos_colors):
                _vals = []
                for g in stillinger_sorted:
                    tot_authors = grp_tot_size.get(g, 0) or 1
                    _vals.append(round(pos_counts.get(g, {}).get(pos_name, 0) / tot_authors, 3))
                fig_rate.add_trace(go.Bar(
                    name=label,
                    y=stillinger_sorted,
                    x=_vals,
                    orientation="h",
                    marker_color=color,
                    text=[f"{fmt_ui(v,3)}" for v in _vals],
                    textposition="inside",
                    hovertemplate="<b>%{y}</b><br>" + label + ": %{fmt_ui(x,3)} per forsker<extra></extra>",
                ))
            fig_rate.update_layout(
                title=f"Forfatterpositioner per stillingsgruppe - rate per forsker, {year}",
                barmode="stack",
                xaxis_title="Positioner per forsker",
                height=max(350, 40 * len(stillinger_sorted)),
                legend_title="Forfatterposition",
                margin=dict(l=160, t=50, r=20),
                yaxis=dict(autorange="reversed"),
            )
            st.plotly_chart(fig_rate, width="stretch", key=f"pos_rate_{tab_key}_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
    )

    rows = [
        {
            "Stillingsgruppe":     grp,
            "Førsteforfatter":     pos_counts.get(grp, {}).get("first",  0),
            "Førsteforfatter (%)": round(100 * pos_counts.get(grp, {}).get("first",  0) / (pos_counts.get(grp, {}).get("total") or 1), 1),
            "Mellemforfatter":     pos_counts.get(grp, {}).get("middle", 0),
            "Mellemforfatter (%)": round(100 * pos_counts.get(grp, {}).get("middle", 0) / (pos_counts.get(grp, {}).get("total") or 1), 1),
            "Sidsteforfatter":     pos_counts.get(grp, {}).get("last",   0),
            "Sidsteforfatter (%)": round(100 * pos_counts.get(grp, {}).get("last",   0) / (pos_counts.get(grp, {}).get("total") or 1), 1),
            "Total deltagelse":    pos_counts.get(grp, {}).get("total",  0),
        }
        for grp in stillinger_sorted
    ]
    schema = [
        ("Stillingsgruppe",    pa.string()),
        ("Førsteforfatter",    pa.int64()),  ("Førsteforfatter (%)",  pa.float64()),
        ("Mellemforfatter",    pa.int64()),  ("Mellemforfatter (%)",  pa.float64()),
        ("Sidsteforfatter",    pa.int64()),  ("Sidsteforfatter (%)",  pa.float64()),
        ("Total deltagelse",   pa.int64()),
    ]
    with st.expander("Se tabel"):
        st.dataframe(build_table(rows, schema), width="stretch", hide_index=True)



def render_tab_stillingsgrupper(year, mode, all_groups, grp_tot_size, grp_avg_size,
                                 selected_facs, selected_insts, selected_grps, 
                                 selected_genders, selected_citizenships,
                                 all_years_data=None, grp_ew=None, edges_keep=None, 
                                 node_meta=None, forfatterpositioner=None,
                                 inst_to_fac = None, total_pubs=None, filter_caption=None, show_intra=True, show_inter=True):
    st.subheader("Stillingsgruppernes bidrag til sampublicering")
    st.markdown(
f""" 
Fanen kortlægger, hvordan sampubliceringsaktiviteten på KU fordeler sig på tværs af karrieretrin i {year}. 

Analysen belyser ikke blot forskelle i samarbejdsomfang, men også forskelle i
rolle og position i forskningssamarbejde – herunder hvordan ph.d.-studerende,
postdocs og faste videnskabelige medarbejdere typisk indgår i sampublicering.

Fanen fokuserer på **strukturelle mønstre i samarbejde på tværs af karrieretrin**
og giver et supplement til de organisatoriske analyser ved at vise, hvordan
sampublicering fordeler sig horisontalt i organisationen snarere end vertikalt
mellem enheder.

Opgørelserne i fanen bygger på tre gennemgående mål:

- **Forfatterantal** viser antallet af unikke KU‑VIP‑forskere i hver
  stillingsgruppe.
- **Forfatterpar** viser, hvor mange forfatterpar en stillingsgruppe indgår i.
  Et samarbejde mellem to stillingsgrupper tælles én gang hos hver gruppe.
- **Sampubliceringsrate** sætter forfatterpar i forhold til forfatterantal og
  gør det muligt at sammenligne samarbejdsintensitet på tværs af karrieretrin
  med forskellig størrelse.

Brug filtret nedenfor til at zoome ind på et enkelt fakultet eller institut.
""")

    if not show_intra and not show_inter:
        st.error("Ingen forfatterpar er valgt. Vælg mindst én par-type (intra eller inter) under **Organisation** i sidepanelet.")
        return

     # ── Lokale filtre ─────────────────────────────────────────────────────────
    _fac_to_insts: dict[str, list] = {}
    _facs_only: list = []
    if node_meta:
        for m in node_meta.values():
            f, i = m.get("fac", ""), m.get("inst", "")
            if f and i:
                if i not in _fac_to_insts.setdefault(f, []):
                    _fac_to_insts[f].append(i)
            elif f:
                if f not in _facs_only:
                    _facs_only.append(f)
    _all_facs_grp = sorted(set(list(_fac_to_insts.keys()) + _facs_only))
    _show_local_filter = bool(_all_facs_grp) and (fac_in_mode(mode) or base_mode(mode) == "G")
    _show_inst_filter  = bool(_fac_to_insts) and (inst_in_mode(mode) or grp_in_mode(mode) or base_mode(mode) == "G")

    _grp_fac_filter = st.multiselect(
        "**Filtrer på fakultet**",
        options=_all_facs_grp if _show_local_filter else [],
        default=[],
        key=f"grp_tab_fac_filter_{year}_{mode}_v2",
        placeholder="Alle fakulteter" if _show_local_filter else "Ikke tilgængeligt i valgte filtre",
        disabled=not _show_local_filter,
    )
    if _show_inst_filter:
        _inst_opts = sorted({
            i
            for f, insts in _fac_to_insts.items()
            for i in insts
            if not _grp_fac_filter or f in _grp_fac_filter
        })
        _grp_inst_filter = st.multiselect(
            "**Filtrer på institut**",
            options=_inst_opts,
            default=[],
            key=f"grp_tab_inst_filter_{year}_{mode}_v2",
            placeholder="Alle institutter",
        )
    else:
        _grp_inst_filter = []
    
    def _grp_node_ok(nid):
        m = (node_meta or {}).get(nid, {})
        if _grp_fac_filter and m.get("fac", "") not in _grp_fac_filter:
            return False
        if _grp_inst_filter and m.get("inst", "") not in _grp_inst_filter:
            return False
        return True
    _edges_grp = (
        [(u, v, w, *r) for u, v, w, *r in (edges_keep or [])
         if _grp_node_ok(u) or _grp_node_ok(v)]
        if (_grp_fac_filter or _grp_inst_filter) else (edges_keep or [])
    )
    _nodes_grp = {
        nid for edge in _edges_grp for nid in (edge[0], edge[1])
    } | {
        nid for nid in (node_meta or {})
        if (node_meta or {}).get(nid, {}).get("type") == "grp"
        and _grp_node_ok(nid)
    }

    if _grp_fac_filter or _grp_inst_filter:
        _grp_counts: dict[str, list] = {}
        for nid in _nodes_grp:
            m = (node_meta or {}).get(nid, {})
            g = m.get("grp", "")
            if not g or not _grp_node_ok(nid): continue
            _grp_counts.setdefault(g, []).append(m.get("size", 0))
        _grp_tot_filtered: dict[str, float] = {g: sum(s) for g, s in _grp_counts.items()}
        _grp_avg_filtered: dict[str, float] = {g: sum(s)/len(s) for g, s in _grp_counts.items() if s}
        _grp_ew_filtered:  dict[str, float] = {}
        for u, v, w, *_ in _edges_grp:
            for n in (u, v):
                if not _grp_node_ok(n): continue
                g = (node_meta or {}).get(n, {}).get("grp", "")
                if g:
                    _grp_ew_filtered[g] = _grp_ew_filtered.get(g, 0.0) + w
        grp_tot_size = _grp_tot_filtered
        grp_avg_size = _grp_avg_filtered
        grp_ew       = _grp_ew_filtered
    else:
        grp_ew = grp_ew or {}

    if not grp_tot_size:
        st.error("Ingen forfattere inden for det valgte filter.")
        return

    if _grp_fac_filter or _grp_inst_filter:
        _years_for_comparison: dict = {}
        for yr, snap in (all_years_data or {}).items():
            _merged_tot: dict[str, int] = {}
            _merged_ew:  dict[str, float] = {}
            if _grp_inst_filter:
                for inst in _grp_inst_filter:
                    for g, v in snap.get("inst_grp_tot", {}).get(inst, {}).items():
                        _merged_tot[g] = _merged_tot.get(g, 0) + v
                    for g, v in snap.get("inst_grp_ew", {}).get(inst, {}).items():
                        _merged_ew[g] = _merged_ew.get(g, 0.0) + v
            else:
                for fac in _grp_fac_filter:
                    for g, v in snap.get("fac_grp_tot", {}).get(fac, {}).items():
                        _merged_tot[g] = _merged_tot.get(g, 0) + v
                    for g, v in snap.get("fac_grp_ew", {}).get(fac, {}).items():
                        _merged_ew[g] = _merged_ew.get(g, 0.0) + v
            _years_for_comparison[yr] = {**snap, "grp_tot": _merged_tot, "grp_ew": _merged_ew}
    else:
        _years_for_comparison = all_years_data

    _grand_tot   = sum(grp_tot_size.values()) or 1
    _top_grp, _top_val = max(grp_tot_size.items(), key=lambda x: x[1])
    _bot_grp, _bot_val = min(grp_tot_size.items(), key=lambda x: x[1])
    _top_share   = 100 * _top_val / _grand_tot
    _avg_top, _avg_top_val = max(grp_avg_size.items(), key=lambda x: x[1])
    _avg_bot, _avg_bot_val = min(grp_avg_size.items(), key=lambda x: x[1])
    _grps_hier   = sorted(grp_tot_size.keys(), key=lambda g: HIERARKI.get(g, 999))
    _tots_hier   = [int(grp_tot_size[g]) for g in _grps_hier]
    _avgs_hier   = [round(grp_avg_size.get(g, 0), 1) for g in _grps_hier]
    _tot_grand   = sum(_tots_hier) or 1
    grp_colors   = stillingsgruppe_colors({})
    
    st.markdown("---")

    st.markdown(
f"""
#### Stillingsgruppernes forfatterantal i {year}

Figuren viser fordelingen af **KU‑VIP‑forfattere** på tværs af
stillingsgrupperne. Store stillingsgrupper vil naturligt dominere i absolutte
tal, men fordelingen giver samtidig et indblik i, hvor forskningsaktiviteten er
koncentreret på tværs af karrieretrin. Generelt udgør **{str(_top_grp).lower()}er** flest forfattere med **{int(_top_val)}** 
({fmt_ui(_top_share)}% af samtlige bidrag i udsnittet), mens **{_bot_grp}** har færrest med 
**{int(_bot_val)}** unikke forfattere. 

Målt som gennemsnitligt forfatterantal per node giver opgørelsen desuden et
billede af, hvor store de underliggende organisatoriske enheder typisk er
for hver stillingsgruppe. **{str(_avg_top)}er** topper (**{fmt_ui(_avg_top_val)}**), 
mens **{str(_avg_bot).lower()}er** ligger lavest (**{fmt_ui(_avg_bot_val)}**).
""")

   # ── Forfatterbidrag-plots ─────────────────────────────────────────────────
    _tab_gt, _tab_gp, _tab_ga = st.tabs(["Antal", "Andel (%)", "Gns. forfatterantal"])

    with _tab_gt:
        _fig_tot = go.Figure(go.Bar(
            y=_grps_hier, x=_tots_hier, orientation="h",
            marker_color=[grp_colors.get(g, "#122947") for g in _grps_hier],
            text=[f"{v}" for v in _tots_hier],
            textposition="inside",
        ))
        _fig_tot.update_layout(
            xaxis_title="Forfatterantal",
            height=max(350, 40 * len(_grps_hier)),
            margin=dict(l=160, t=50, r=80),
            yaxis=dict(autorange="reversed"),
            title=f"Stillingsgruppernes forfatterantal, {year}",
        )
        st.plotly_chart(_fig_tot, width="stretch", key=f"grp_tot_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
        )
        if filter_caption:
            st.caption(filter_caption)

    with _tab_gp:
        _pct_vals = [round(100 * v / _tot_grand, 1) for v in _tots_hier]
        _fig_pct = go.Figure(go.Bar(
            y=_grps_hier, x=_pct_vals, orientation="h",
            marker_color=[grp_colors.get(g, "#122947") for g in _grps_hier],
            text=[f"{fmt_ui(v)}%" for v in _pct_vals],
            textposition="inside",
        ))
        _fig_pct.update_layout(
            xaxis=dict(title="Andel (%)", range=[0, 100]),
            height=max(350, 40 * len(_grps_hier)),
            margin=dict(l=160, t=50, r=80),
            yaxis=dict(autorange="reversed"),
            title=f"Stillingsgruppernes forfatterantal, {year}"
        )
        st.plotly_chart(_fig_pct, width="stretch", key=f"grp_pct_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
        )
        if filter_caption:
            st.caption(filter_caption)

    with _tab_ga:
        _fig_avg = go.Figure(go.Bar(
            y=_grps_hier, x=_avgs_hier, orientation="h",
            marker_color=[grp_colors.get(g, "#122947") for g in _grps_hier],
            text=[f"{fmt_ui(v)}" for v in _avgs_hier],
            textposition="inside",
        ))
        _fig_avg.update_layout(
            xaxis_title="Gns. forfatterantal",
            height=max(350, 40 * len(_grps_hier)),
            margin=dict(l=160, t=50, r=80),
            yaxis=dict(autorange="reversed"),
            title=f"Stillingsgruppernes gennemsnitlige forfatterantal, {year}"
        )
        st.plotly_chart(_fig_avg, width="stretch", key=f"grp_avg_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
        )
        if filter_caption:
            st.caption(filter_caption)
    
    # ── Tabel ─────────────────────────────────────────────────────────────────
    # Byg én række per grp+fac (eller grp+inst) kombination
    _grp_fac_map: dict[str, set] = {}
    _grp_inst_map: dict[str, set] = {}
    for nid, m in node_meta.items():
        if not _grp_node_ok(nid): continue
        g = m.get("grp", "")
        if not g: continue
        if m.get("fac"): _grp_fac_map.setdefault(g, set()).add(m["fac"])
        if m.get("inst"): _grp_inst_map.setdefault(g, set()).add(m["inst"])

    _grp_bid_rows = []
    for g in _grps_hier:
        row = {"Stillingsgruppe": g}
        if _grp_fac_filter:
            row["Fakultet"] = ", ".join(sorted(_grp_fac_map.get(g, set())))
        if _grp_inst_filter:
            row["Institut"] = ", ".join(sorted(_grp_inst_map.get(g, set())))
        row["Antal unikke forfattere"] = int(grp_tot_size.get(g, 0))
        row["Andel (%)"]              = round(100 * grp_tot_size.get(g, 0) / _tot_grand, 1)
        row["Gns. forfatterantal"]   = round(grp_avg_size.get(g, 0), 1)
        _grp_bid_rows.append(row)

    _grp_bid_schema = (
        ([("Fakultet", pa.string())] if _grp_fac_filter else []) +
        ([("Institut", pa.string())] if _grp_inst_filter else []) +
        [
            ("Stillingsgruppe",        pa.string()),
            ("Antal unikke forfattere", pa.int64()),
            ("Andel (%)",              pa.float64()),
            ("Gns. forfatterantal",   pa.float64()),
        ]
    )
    with st.expander("Se tabel"):
        st.dataframe(build_table(_grp_bid_rows, _grp_bid_schema), hide_index=True, width="stretch")
        st.download_button(
            "Download (.xlsx)",
            data=rows_to_excel_bytes(_grp_bid_rows, [n for n, _ in _grp_bid_schema]),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_grp_bidrag_tot_{year}_{mode}",
        )
    
    st.markdown("---")

    # ── Forfatterpar-plots ────────────────────────────────────────────────────
    st.markdown(
f"""#### Forfatterpar per stillingsgruppe i {year}

**Forfatterpar** viser det samlede samarbejdsomfang for hver stillingsgruppe.
Et samarbejde mellem to forskellige stillingsgrupper tælles én gang hos hver
af dem, hvilket betyder, at summen af forfatterpar på tværs af grupper svarer
til **to gange** det samlede antal unikke forfatterpar i netværket.

Målet afspejler, hvor intensivt de forskellige karrieretrin indgår i
sampublicering, men siger ikke noget om, hvor mange forskellige
samarbejdspartnere den enkelte gruppe har.

Sampubliceringsraten normaliserer for stillingsgruppernes størrelse og viser,
hvor mange forfatterpar den gennemsnitlige forsker i gruppen indgår i.

En høj rate indikerer intensivt samarbejde, men påvirkes også af
publikationspraksis og typiske samarbejdskonstellationer for det pågældende
karrieretrin.
""")
    _grp_ew_all_yr = {}
    if grp_ew:
        if total_pubs and not (_grp_fac_filter or _grp_inst_filter):
            _grand_ew = total_pubs
        else:
            _grand_ew = (sum(grp_ew.values()) / 2) if grp_ew else 1
        _ews_hier     = [round(grp_ew.get(g, 0), 1) for g in _grps_hier]
        _ew_ratio_hier = [round(grp_ew.get(g, 0) / (grp_tot_size.get(g) or 1), 3) for g in _grps_hier]

        _grp_ew_all_yr = {}
        if all_years_data and year in all_years_data:
            _gi = all_years_data[year].get("grp_intra_ew_all", {})
            _gx = all_years_data[year].get("grp_inter_ew_all", {})
            _grp_ew_all_yr = {g: _gi.get(g, 0.0) + _gx.get(g, 0.0) for g in set(_gi) | set(_gx)}

        _use_grp_all = st.toggle(
            "Inkludér intra-enhed forfatterpar (f.eks. to professorer på samme institut)",
            value=False,
            key=f"grp_ew_bar_all_{year}_{mode}",
            help="Til: alle forfatterpar inkl. par inden for samme enhed. Fra: kun netværkskanter.",
            disabled=not show_intra
        )
        if _use_grp_all:
            st.caption("Alle forfatterpar inkl. par inden for samme enhed (f.eks. to mandlige professorer på Kemisk Institut).")
        else:
            st.caption("Kun forfatterpar *mellem* de valgte organisatoriske enheder - som vist i netværket.")

        _active_grp_ew = _grp_ew_all_yr if (_use_grp_all and _grp_ew_all_yr) else (grp_ew or {})
        if _use_grp_all:
            # Intra-toggle på: brug sum/2 af aktiv ew.
            _grand_active_grp_ew = (sum(_active_grp_ew.values()) / 2) or 1
        elif total_pubs and not (_grp_fac_filter or _grp_inst_filter):
            # Intra-toggle fra, intet lokalt filter: brug global total_pubs.
            _grand_active_grp_ew = total_pubs
        else:
            # Intra-toggle fra, men lokalt filter aktivt: lokal total = sum/2.
            _grand_active_grp_ew = (sum(_active_grp_ew.values()) / 2) or 1
        _ews_hier      = [round(_active_grp_ew.get(g, 0), 1) for g in _grps_hier]
        _ew_ratio_hier = [round(_active_grp_ew.get(g, 0) / (grp_tot_size.get(g) or 1), 3) for g in _grps_hier]

        _pct_ews_grp = [round(100 * v / _grand_active_grp_ew, 1) for v in _ews_hier]
        _intra_suffix_grp = ' (inkl. intra-enhed)' if _use_grp_all else ''
        _tab_ew_abs, _tab_ew_pct, _tab_ew_ratio = st.tabs(["Antal", "Andel (%)", "Forfatterpar per forfatter"])

        with _tab_ew_abs:
            _fig_ew = go.Figure(go.Bar(
                y=_grps_hier, x=_ews_hier, orientation="h",
                marker_color=[grp_colors.get(g, "#122947") for g in _grps_hier],
                text=[f"{fmt_ui(v)}" for v in _ews_hier],
                textposition="inside",
            ))
            _fig_ew.update_layout(
                title=f"Forfatterpar per stillingsgruppe{_intra_suffix_grp}, {year}",
                xaxis_title="Antal forfatterpar",
                height=max(350, 40 * len(_grps_hier)),
                margin=dict(l=160, t=50, r=80),
                yaxis=dict(autorange="reversed"),
            )
            st.caption("Vær opmærksom på, at procenter kan overstige 100%. Se **Læsning af procenter** i Oversigt-fanen.")
            st.plotly_chart(_fig_ew, width="stretch", key=f"grp_ew_abs_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}})
            if filter_caption:
                st.caption(filter_caption)

        with _tab_ew_pct:
            _fig_ew_p = go.Figure(go.Bar(
                y=_grps_hier, x=_pct_ews_grp, orientation="h",
                marker_color=[grp_colors.get(g, "#122947") for g in _grps_hier],
                text=[f"{fmt_ui(v)}%" for v in _pct_ews_grp],
                textposition="inside",
            ))
            _fig_ew_p.update_layout(
                title=f"Forfatterpar per stillingsgruppe{_intra_suffix_grp} - andel (%), {year}",
                xaxis=dict(title="Andel (%)", range=[0, 100]),
                height=max(350, 40 * len(_grps_hier)),
                margin=dict(l=160, t=50, r=80),
                yaxis=dict(autorange="reversed"),
            )
            st.plotly_chart(_fig_ew_p, width="stretch", key=f"grp_ew_pct_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}})
            if filter_caption:
                st.caption(filter_caption)

        with _tab_ew_ratio:
            _fig_ew_r = go.Figure(go.Bar(
                y=_grps_hier, x=_ew_ratio_hier, orientation="h",
                marker_color=[grp_colors.get(g, "#122947") for g in _grps_hier],
                text=[f"{fmt_ui(v,3)}" for v in _ew_ratio_hier],
                textposition="inside",
            ))
            _fig_ew_r.update_layout(
                title=f"Forfatterpar per forfatter per stillingsgruppe{_intra_suffix_grp}, {year}",
                xaxis_title="Forfatterpar per forfatter",
                height=max(350, 40 * len(_grps_hier)),
                margin=dict(l=160, t=50, r=80),
                yaxis=dict(autorange="reversed"),
            )
            st.plotly_chart(_fig_ew_r, width="stretch", key=f"grp_ew_ratio_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}})
            if filter_caption:
                st.caption(filter_caption)

        # ── Tabel ─────────────────────────────────────────────────────────────────
        _grp_fac_map: dict[str, set] = {}
        _grp_inst_map: dict[str, set] = {}
        for nid, m in (node_meta or {}).items():
            if not _grp_node_ok(nid): continue
            g = m.get("grp", "")
            if not g: continue
            if m.get("fac"):  _grp_fac_map.setdefault(g, set()).add(m["fac"])
            if m.get("inst"): _grp_inst_map.setdefault(g, set()).add(m["inst"])

        _grp_summary_rows = []
        for g in _grps_hier:
            row = {}
            if _grp_fac_filter:
                row["Fakultet"] = ", ".join(sorted(_grp_fac_map.get(g, set())))
            if _grp_inst_filter:
                row["Institut"] = ", ".join(sorted(_grp_inst_map.get(g, set())))
            row.update({
                "Stillingsgruppe":            g,
                "Samlet antal forfattere":     int(grp_tot_size[g]),
                "Andel (%)":                  round(100 * grp_tot_size[g] / _tot_grand, 1),
                "Gns. forfatterantal":       round(grp_avg_size.get(g, 0), 1),
                "Forfatterpar (netværk)":     round((grp_ew or {}).get(g, 0), 1),
                "Forfatterpar (inkl. intra)": round(_grp_ew_all_yr.get(g, 0), 1) if _grp_ew_all_yr else 0,
            })
            _grp_summary_rows.append(row)

        _grp_schema = (
            ([("Fakultet", pa.string())] if _grp_fac_filter else []) +
            ([("Institut", pa.string())] if _grp_inst_filter else []) +
            [
                ("Stillingsgruppe",               pa.string()),
                ("Samlet antal forfattere",        pa.int64()),
                ("Andel (%)",                     pa.float64()),
                ("Gns. forfatterantal",          pa.float64()),
                ("Forfatterpar (netværk)",        pa.float64()),
                ("Forfatterpar (inkl. intra)",    pa.float64()),
            ]
        )
        with st.expander("Se tabel"):
            st.dataframe(build_table(_grp_summary_rows, _grp_schema), hide_index=True, width="stretch")
            st.download_button(
                "Download (.xlsx)",
                data=rows_to_excel_bytes(_grp_summary_rows, [n for n, _ in _grp_schema]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_grp_bidrag_{year}_{mode}",
            )
    else:
        st.error("Ingen forfatterpar i valgte filtre.")
    
    st.markdown("---")

    # ── Stillingsgruppernes rækkevidde ────────────────────────────────────────
    st.markdown(
f"""#### Stillingsgruppernes rækkevidde i {year}

**Rækkevidde** måler, hvor mange forskellige stillingsgrupper en gruppe
samarbejder med – uanset hvor intensivt samarbejdet er.

En bred rækkevidde indikerer, at gruppen indgår i samarbejde på tværs af mange
karrieretrin, mens lav rækkevidde peger på mere koncentrerede
samarbejdsmønstre. Samarbejde udelukkende inden for samme stillingsgruppe
giver rækkevidde 0, selv hvis samarbejdet er intensivt.
""")

    _grp_partners_reach: dict[str, set] = {}
    for u, v, w, *_ in (edges_keep or []):
        gu = node_meta.get(u, {}).get("grp", "") if node_meta else ""
        gv = node_meta.get(v, {}).get("grp", "") if node_meta else ""
        if gu and gv and gu != gv:
            _grp_partners_reach.setdefault(gu, set()).add(gv)
            _grp_partners_reach.setdefault(gv, set()).add(gu)

    _all_grps_reach = sorted(
        {g for g in (grp_tot_size or {}) if g in HIERARKI},
        key=lambda g: HIERARKI.get(g, 999),
    )
    _all_grps_in_edges = {
        node_meta.get(nid, {}).get("grp", "")
        for edge in (edges_keep or [])
        for nid in [edge[0], edge[1]]
        if node_meta and node_meta.get(nid, {}).get("grp", "")
    }
    _max_grp_reach = max(len(_all_grps_in_edges) - 1, 1)

    _grp_reach_sorted = sorted(
        [(g, len(_grp_partners_reach.get(g, set()))) for g in _all_grps_reach],
        key=lambda x: -x[1],
    )

    if any(n > 0 for _, n in _grp_reach_sorted):
        _fig_grp_reach = go.Figure(go.Bar(
            y=[g for g, _ in _grp_reach_sorted],
            x=[n for _, n in _grp_reach_sorted],
            orientation="h",
            marker_color=[grp_colors.get(g, "#122947") for g, _ in _grp_reach_sorted],
            text=[f"{n} / {_max_grp_reach}" for _, n in _grp_reach_sorted],
            textposition="inside",
        ))
        _fig_grp_reach.update_layout(
            xaxis=dict(title="Antal samarbejdsstillingsgrupper", range=[0, _max_grp_reach + 0.5], dtick=1),
            height=max(300, 50 * len(_grp_reach_sorted)),
            margin=dict(l=180, t=50, r=80),
            yaxis=dict(autorange="reversed"),
            title=f"Stillingsgruppernes rækkevidde, {year}"
        )
        st.plotly_chart(_fig_grp_reach, width="stretch", key=f"grp_reach_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
        )
        if filter_caption:
            st.caption(filter_caption)
    
    st.markdown("---")

    # ── Samarbejdsintensitet på tværs af stillingsgrupper (heatmap) ──────────
    st.markdown(
f"""#### Samarbejdsmønstre på tværs af stillingsgrupper i {year}

Heatmappet nedenfor viser, **hvilke stillingsgrupper der publicerer sammen** - og hvor intensivt.

- **Antal forfatterpar** viser den rå trafik mellem to grupper. Det er et volumenmål, og store grupper
  dominerer naturligt.
- **Forfatterpar per forfatter** svarer på spørgsmålet: hvor ofte er en forfatter fra kolonnen 
med, når nogen fra rækken publicerer? Matricen er ikke symmetrisk, fordi cellen (ph.d.,professor) og
cellen (professor, ph.d.) svarer på to forskellige spørgsmål - set fra to forskellige perspektiver.
- **Andel (%) per række** viser for hver stillingsgruppegruppe, hvor dens samarbejder fordeler sig. En række summerer til 100%.

**Bemærk:** Store forfatterskaber (f.eks. artikler med 100+ forfattere) kan skævvride billedet,
da de bidrager med rigtig mange forfatterpar fra én publikation. 
""")

    # Byg par-matrix fra kanterne
    _pair_matrix: dict[tuple[str, str], float] = {}
    for u, v, w, *_ in (_edges_grp or []):
        gu = (node_meta or {}).get(u, {}).get("grp", "")
        gv = (node_meta or {}).get(v, {}).get("grp", "")
        if not gu or not gv:
            continue
        # Par-vægt fordeles ligeligt på de to endepunkter (matcher øvrig logik)
        _pair_matrix[(gu, gv)] = _pair_matrix.get((gu, gv), 0.0) + w
        if gu != gv:
            _pair_matrix[(gv, gu)] = _pair_matrix.get((gv, gu), 0.0) + w

    # Hvis intra-enhed er slået til, flet intra-par ind på diagonalen
    if _use_grp_all and all_years_data and year in all_years_data:
        _intra_src = all_years_data[year].get("grp_intra_ew_all", {})
        for g, v in _intra_src.items():
            if g in grp_tot_size:
                # Intra-par skal på diagonalen; undgå dobbelt-tælling hvis edges_keep allerede har dem
                _pair_matrix[(g, g)] = _pair_matrix.get((g, g), 0.0) + v

    _hm_grps = _grps_hier  # allerede sorteret efter HIERARKI
    if len(_hm_grps) < 2 or not _pair_matrix:
        st.error("Ikke tilstrækkeligt data til at vise samarbejdsmatrix for det valgte udsnit.")
    else:
        _z_abs = [[round(_pair_matrix.get((gi, gj), 0.0), 1) for gj in _hm_grps] for gi in _hm_grps]
        _z_ratio = [
            [round(_pair_matrix.get((gi, gj), 0.0) / (grp_tot_size.get(gi) or 1), 3) for gj in _hm_grps]
            for gi in _hm_grps
        ]
        _z_row_pct = []
        for gi in _hm_grps:
            _row_total = sum(_pair_matrix.get((gi, gj), 0.0) for gj in _hm_grps) or 1
            _z_row_pct.append([
            round(100 * _pair_matrix.get((gi, gj), 0.0) / _row_total, 2)
            for gj in _hm_grps
        ])

        _intra_suffix_hm = ' (inkl. intra-enhed)' if _use_grp_all else ''

        _tab_hm_abs, _tab_hm_ratio, _tab_hm_pct = st.tabs(
            ["Antal forfatterpar", "Forfatterpar per forfatter", "Andel (%) per række"]
        )

        def _heatmap_fig(z, title, colorbar_title, text_fmt, hover_template):
            _text = [[text_fmt(v) for v in row] for row in z]
            fig = go.Figure(go.Heatmap(
                z=z, x=_hm_grps, y=_hm_grps,
                colorscale=[[0, "#f0f4f8"], [1, "#122947"]],
                text=_text, texttemplate="%{text}",
                textfont=dict(size=11),
                hovertemplate=hover_template,
                colorbar=dict(title=colorbar_title),
            ))
            fig.update_layout(
                title=title,
                xaxis=dict(title="", tickangle=-30),
                yaxis=dict(title="", autorange="reversed"),
                height=max(400, 55 * len(_hm_grps)),
                margin=dict(l=180, t=60, r=80, b=120),
            )
            return fig

        with _tab_hm_abs:
            st.plotly_chart(
                _heatmap_fig(
                    _z_abs,
                    f"Antal forfatterpar mellem stillingsgrupper{_intra_suffix_hm}, {year}",
                    "Antal par",
                    lambda v: fmt_ui(v) if v else "",
                    "<b>%{y}</b> og <b>%{x}</b><br>"
                    "Antal forfatterpar: <b>%{z}</b>"
                    "<extra></extra>",
                ),
                width="stretch",
                key=f"grp_hm_abs_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}},
            )
            if filter_caption:
                st.caption(filter_caption)

        with _tab_hm_ratio:
            st.plotly_chart(
                _heatmap_fig(
                    _z_ratio,
                    f"Forfatterpar per forfatter, stillingsgrupper{_intra_suffix_hm}, {year}",
                    "Par per bidrag",
                    lambda v: fmt_ui(v, 2) if v else "",
                    "Per forfatter fra <b>%{y}</b><br>"
                    "dannes i gennemsnit <b>%{z}</b><br> forfatterpar " 
                    "med <b>%{x}</b><br>"
                    "<extra></extra>",
                ),
                width="stretch",
                key=f"grp_hm_ratio_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}},
            )
            if filter_caption:
                st.caption(filter_caption)

        with _tab_hm_pct:
            st.plotly_chart(
                _heatmap_fig(
                    _z_row_pct,
                    f"Fordeling af samarbejdspartnere per stillingsgruppe{_intra_suffix_hm}, {year}",
                    "Andel (%)",
                    lambda v: f"{fmt_ui(v, 2)}%" if v else "",
                    "<b>%{z}%</b> af <b>%{y}</b>ers forfatterpar<br>"
                    "er med <b>%{x}</b>"
                    "<extra></extra>",
                ),
                width="stretch",
                key=f"grp_hm_pct_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}},
            )
            if filter_caption:
                st.caption(filter_caption)

        # ── Tabel (lang format, så man kan filtrere og downloade) ─────────────
        _hm_rows = []
        for gi in _hm_grps:
            _row_total = sum(_pair_matrix.get((gi, gj), 0.0) for gj in _hm_grps) or 1
            _bidrag_i = grp_tot_size.get(gi) or 1
            for gj in _hm_grps:
                _pairs = _pair_matrix.get((gi, gj), 0.0)
                if _pairs == 0:
                    continue
                _hm_rows.append({
                    "Fra (række)":                  gi,
                    "Til (kolonne)":                gj,
                    "Forfatterpar":                 round(_pairs, 1),
                    "Par per forfatter":      round(_pairs / _bidrag_i, 3),
                    "Andel af rækkens par (%)":    round(100 * _pairs / _row_total, 1),
                })
        _hm_schema = [
            ("Fra (række)",              pa.string()),
            ("Til (kolonne)",            pa.string()),
            ("Forfatterpar",             pa.float64()),
            ("Par per forfatter",  pa.float64()),
            ("Andel af rækkens par (%)", pa.float64()),
        ]
        with st.expander("Se tabel"):
            st.dataframe(build_table(_hm_rows, _hm_schema), hide_index=True, width="stretch")
            st.download_button(
                "Download (.xlsx)",
                data=rows_to_excel_bytes(_hm_rows, [n for n, _ in _hm_schema]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_grp_hm_{year}_{mode}",
            )

    st.markdown("---")

    # ── Sammenlign år ─────────────────────────────────────────────────────────
    st.markdown(
f"""### Udvikling over tid

Afsnittet nedenfor viser udviklingen i stillingsgruppernes **forfatterantal**,
**forfatterpar** og **rækkevidde** på tværs af de tilgængelige år.

Tidsserierne gør det muligt at vurdere, om samarbejdsmønstre på tværs af
karrieretrin er stabile, eller om der sker forskydninger i, hvordan
forskere på forskellige niveauer indgår i sampublicering over tid.
""")
    _grps_in_data = sorted(
        {g for s in (all_years_data or {}).values() for g in s.get("grp_tot", {})},
        key=lambda g: HIERARKI.get(g, 999),
    )
    _grps_ew_in_data = sorted(
        {g for s in (all_years_data or {}).values() for g in s.get("grp_ew", {})},
        key=lambda g: HIERARKI.get(g, 999),
    )
    years_sorted_grp = sorted((all_years_data or {}).keys())

    _tab_yr_tot, _tab_yr_ew, _tab_yr_reach = st.tabs(["Forfatterantal per stillingsgruppe", "Forfatterpar per stillingsgruppe", "Stillingsgruppernes rækkevidde"])

    with _tab_yr_tot:
        _render_year_comparison(
            _years_for_comparison,
            series=[(grp, "grp_tot", grp) for grp in _grps_in_data],
            title=f"Udvikling i stillingsgruppernes forfatterantal, {years_sorted_grp[0]}-{years_sorted_grp[-1]}",
            colors=grp_colors,
            key_suffix=f"grp_tot_{year}_{mode}",
            show_table=True,
        )

    with _tab_yr_ew:
        _use_grp_ew_all = st.toggle(
            "Inkludér intra-enhed forfatterpar (f.eks. to professorer på samme institut)",
            value=False,
            key=f"grp_ew_all_{year}_{mode}",
            help="Til: alle forfatterpar inkl. par inden for samme enhed. Fra: kun netværkskanter.",
            disabled=not show_intra
        )
        if _use_grp_ew_all:
            st.caption("Alle forfatterpar inkl. par inden for samme enhed (f.eks. to mandlige professorer på Kemisk Institut).")
        else:
            st.caption("Kun forfatterpar *mellem* de valgte organisatoriske enheder - som vist i netværket.")

        _grp_ew_key = "grp_ew_all" if _use_grp_ew_all else "grp_ew"
        _grps_ew_plot = sorted(
            {g for s in (_years_for_comparison or {}).values() for g in s.get(_grp_ew_key, {})},
            key=lambda g: HIERARKI.get(g, 999),
        )
        _render_year_comparison(
            _years_for_comparison,
            series=[(grp, _grp_ew_key, grp) for grp in _grps_ew_plot],
            title=f"Udvikling i antal forfatterpar per stillingsgruppe{' (inkl. intra-enhed)' if _use_grp_ew_all else ''}, {years_sorted_grp[0]}-{years_sorted_grp[-1]}",
            yaxis_label="Antal forfatterpar",
            colors=grp_colors,
            key_suffix=f"grp_ew_{year}_{mode}",
            show_table=True,
        )
    
    with _tab_yr_reach:
        _grps_reach_in_data = sorted(
            {g for s in (_years_for_comparison or {}).values() for g in s.get("grp_reach", {})},
            key=lambda g: HIERARKI.get(g, 999),
        )
        _render_year_comparison(
            _years_for_comparison,
            series=[(grp, "grp_reach", grp) for grp in _grps_reach_in_data],
            title=f"Udvikling i stillingsgruppernes rækkevidde, {years_sorted_grp[0]}-{years_sorted_grp[-1]}",
            yaxis_label="Antal samarbejdspartnere",
            colors=grp_colors,
            key_suffix=f"grp_reach_{year}_{mode}",
            show_table=True,
        )


    st.markdown("---")

    # ── Forfatterpositioner ───────────────────────────────────────────────────
    st.subheader(f"Stillingsgruppers placering i forfatterrækkefølger, {year}")
    st.markdown(
f"""
Figurerne nedenfor viser fordelingen af **førsteforfatter**, **mellemforfatter**
og **sidsteforfatter** fordelt på stillingsgruppe.

Forfatterposition er bestemt af rækkefølgen i CURIS‑registreringen:
førsteforfatter er den først listede forfatter, og sidsteforfatter den sidst
listede. Opgørelsen omfatter alle KU‑publikationer med mindst to KU‑VIP‑forfattere;
soloartikler indgår ikke, da forfatterposition ikke er meningsfuld uden
medforfattere.

Normer for forfatterrækkefølge varierer betydeligt på tværs af fagområder.
På natur‑ og sundhedsvidenskabelige områder angiver sidstepladsen ofte den
ansvarlige seniorforsker, mens humanistiske og samfundsvidenskabelige traditioner
oftere følger alfabetisk eller bidragsbaseret rækkefølge.
"""
    )

    yr_data = (forfatterpositioner or {}).get(str(year), {})

    def _merge_pos(level_dict: dict, keys: list) -> dict:
        counts = defaultdict(lambda: defaultdict(int))
        for key in keys:
            for grp, c in level_dict.get(key, {}).items():
                for k, v in c.items():
                    counts[grp][k] += v
        return {grp: dict(c) for grp, c in counts.items()}

    active_facs  = (
        _grp_fac_filter if _grp_fac_filter
        else list(selected_facs) if selected_facs
        else [f for f in FAC_ORDER if f in yr_data.get("fac", {})]
    )
    active_insts = (
        _grp_inst_filter if _grp_inst_filter
        else list(selected_insts) if selected_insts
        else sorted(yr_data.get("inst", {}).keys())
    )
    if _grp_inst_filter:
        pos_data = _merge_pos(yr_data.get("inst", {}), _grp_inst_filter)
    elif mode in ("FI", "FIG") and inst_to_fac:
        fac_set = set(active_facs)
        active_insts = [
            i for i in active_insts
            if FAC_ABBRS.get(inst_to_fac.get(i, ""), inst_to_fac.get(i, "")) in fac_set
        ]
        pos_data = _merge_pos(yr_data.get("inst", {}), active_insts)
    elif fac_in_mode(mode) and not inst_in_mode(mode):
        pos_data = _merge_pos(yr_data.get("fac", {}), active_facs)
    elif base_mode(mode) == "G":
        pos_data = yr_data.get("ku", {})
    else:
        pos_data = _merge_pos(yr_data.get("inst", {}), active_insts)

    render_pos_chart(pos_data, "ku", year, mode, grp_tot_size=grp_tot_size)

    # ── Forfatterpositioner over tid ──────────────────────────────────────────
    if forfatterpositioner and len(forfatterpositioner) >= 2:
        st.markdown("#### Forfatterrækkefølge over tid")
        st.markdown(
            "Figuren viser, hvordan stillingsgruppernes placering i forfatterrækkefølger "
            "har udviklet sig over tid. En stigende førsteforfatter-andel for en gruppe "
            "kan indikere øget faglig lederskab i publikationsprocessen."
        )

        _pos_years  = sorted(forfatterpositioner.keys())
        _pos_grps   = sorted(HIERARKI.keys())
        _pos_labels = [("first", "Førsteforfatter"), ("middle", "Mellemforfatter"), ("last", "Sidsteforfatter")]
        _pos_colors = ku_color_sequence(3)

        # Byg data: for hvert år og gruppe, hent pos-counts fra samme niveau som render_pos_chart
        def _get_pos_yr(yr):
            _d = (forfatterpositioner or {}).get(yr, {})
            if _grp_inst_filter:
                return _merge_pos(_d.get("inst", {}), _grp_inst_filter)
            elif mode in ("FI", "FIG") and inst_to_fac and _grp_fac_filter:
                _fac_set = set(_grp_fac_filter)
                _ai = [i for i in sorted(_d.get("inst", {}).keys())
                       if FAC_ABBRS.get(inst_to_fac.get(i, ""), inst_to_fac.get(i, "")) in _fac_set]
                return _merge_pos(_d.get("inst", {}), _ai)
            elif mode in ("FI", "FIG") and inst_to_fac:
                _fac_set = set(active_facs)
                _ai = [i for i in sorted(_d.get("inst", {}).keys())
                       if FAC_ABBRS.get(inst_to_fac.get(i, ""), inst_to_fac.get(i, "")) in _fac_set]
                return _merge_pos(_d.get("inst", {}), _ai)
            elif fac_in_mode(mode) and not inst_in_mode(mode):
                return _merge_pos(_d.get("fac", {}), _grp_fac_filter or active_facs)
            elif base_mode(mode) == "G":
                return _d.get("ku", {})
            else:
                return _merge_pos(_d.get("inst", {}), sorted(_d.get("inst", {}).keys()))

        _pos_sel = st.radio(
            "**Vis forfatterposition:**",
            options=["Førsteforfatter", "Mellemforfatter", "Sidsteforfatter"],
            horizontal=True,
            key=f"pos_tid_radio_{year}_{mode}",
        )
        _pos_name_sel = {"Førsteforfatter": "first", "Mellemforfatter": "middle", "Sidsteforfatter": "last"}[_pos_sel]

        _tab_pos_abs, _tab_pos_pct = st.tabs(["Antal", "Andel (%)"])

        for _tab, _normalise in [(_tab_pos_pct, True), (_tab_pos_abs, False)]:
            with _tab:
                _fig_pos_tid = go.Figure()
                for grp in _pos_grps:
                    color = grp_colors.get(grp, "#122947")
                    _y = []
                    for yr in _pos_years:
                        _pd = _get_pos_yr(yr)
                        _tot = _pd.get(grp, {}).get("total", 0) or 1
                        _val = _pd.get(grp, {}).get(_pos_name_sel, 0)
                        _y.append(round(100 * _val / _tot, 1) if _normalise else _val)
                    _fig_pos_tid.add_trace(go.Scatter(
                        x=_pos_years, y=_y,
                        name=grp,
                        mode="lines+markers",
                        line=dict(width=2, color=color),
                        marker=dict(size=6),
                    ))
                if _normalise:
                    st.markdown(f"Andelen af en stillingsgruppes forfatterskaber, der er som {_pos_sel}.")
                    _fig_pos_tid.update_layout(
                        yaxis_title="Andel (%)" if _normalise else "Antal",
                        xaxis=dict(tickmode="array", tickvals=_pos_years, dtick=1),
                        height=480,
                        legend=dict(orientation="v", yanchor="top", y=1, xanchor="left", x=1.02),
                        margin=dict(t=50, b=20, r=200),
                        title = f"Udvikling i andel af {_pos_sel}skaber, {_pos_years[0]}-{_pos_years[-1]}"
                    )
                else:
                    _fig_pos_tid.update_layout(
                        yaxis_title="Andel (%)" if _normalise else "Antal",
                        xaxis=dict(tickmode="array", tickvals=_pos_years, dtick=1),
                        height=480,
                        legend=dict(orientation="v", yanchor="top", y=1, xanchor="left", x=1.02),
                        margin=dict(t=50, b=20, r=200),
                        title = f"Udvikling i antal af {_pos_sel}skaber, {_pos_years[0]}-{_pos_years[-1]}"
                    )
                st.plotly_chart(_fig_pos_tid, width="stretch",
                                key=f"pos_tid_{'pct' if _normalise else 'abs'}_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
                )
                if filter_caption:
                    st.caption(filter_caption)

        # Tabel
        _pos_tid_rows = []
        for pos_name, pos_label in _pos_labels:
            for grp in _pos_grps:
                row = {"Position": pos_label, "Stillingsgruppe": grp}
                for yr in _pos_years:
                    _pd = _get_pos_yr(yr)
                    _tot = _pd.get(grp, {}).get("total", 0) or 1
                    _val = _pd.get(grp, {}).get(pos_name, 0)
                    row[str(yr)] = round(100 * _val / _tot, 1)
                _pos_tid_rows.append(row)

        _pos_tid_cols = ["Position", "Stillingsgruppe"] + [str(yr) for yr in _pos_years]
        _pos_tid_schema = (
            [("Position", pa.string()), ("Stillingsgruppe", pa.string())] +
            [(str(yr), pa.float64()) for yr in _pos_years]
        )
        with st.expander("Se tabel"):
            st.dataframe(build_table(_pos_tid_rows, _pos_tid_schema), hide_index=True, width="stretch")
            st.download_button(
                "Download (.xlsx)",
                data=rows_to_excel_bytes(_pos_tid_rows, _pos_tid_cols),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_pos_tid_{year}_{mode}",
            )
    



def render_tab_centralitet(year, mode, faculty_wd_sorted, faculty_bs_sorted,
                            inst_wd_sorted, inst_bs_sorted,
                            grp_wd_sorted, grp_bs_sorted, node_meta,
                            grp_node_wd_sorted=None, grp_node_bs_sorted=None,
                            faculty_base_colors=None, grp_colors=None,
                            all_years_data=None, svg_centralitet=None, filter_caption=None):
    st.subheader("Nøgleaktører i sampubliceringsnetværket")
    if "S" in mode or "N" in mode: 
        st.error("Analyserne nedenfor baseres på netværksplottet - uden nogen diversitetsdimensioner.")

    
    st.markdown(
f""" 
Fanen viser, hvilke enheder der spiller den første rolle i sampubliceringsnetværket - både hvem der samarbejder mest, og hvem 
der binder netværket sammen. Analysen beskriver **struktur**, ikke kvalitet eller strategiske valg. 

**Samlet samarbejdsomfang** (centralitet) tæller, hvor mange forfatterpar en enhed indgår i. Det er et mål for aktivitet - et stort institut med 
mange forfattere vil naturligt score højt, uanset om samarbejdet er bredt eller koncentreret. 

**Brobyggerrolle** (*betweenness* centralitet) måler noget andet: hvor ofte en enhed befinder sig på den korteste vel mellem to
andre enheder, som ellers ikke samarbejder, i netværket. En enhed med høj brobyggerrolle er strukturel vigtig - ikke fordi den samarbejde mest, men fordi mange forbindelser 
*løber igennem* den. 

**Eksempel:** antag, at SUND og HUM sjældent samarbejder direkte, men begge samarbejder med et bestemt institut på SCIENCE. 
Det SCIENCE-institut har da en høj brobyggerrolle, fordi det er forbindelsesleddet. Fjernes det, mister SUND og HUM forbindelseslinjen
til hinanden. Et institut kan altså have relativt få forfatterpar og stadig være strukturelt uundværligt. 

Som tommelfingerregel for brobyggerrollen (normaliseret):

- **> 0.3** - stærk strukturel position; enheden er svær at erstatte
- **0.1-0.3** - moderat brobyggerrolle
- **< 0.1** - begrænset strukturel betydninge i det viste netværk

Bemærk, at begge mål afspejler det valgte udsnit og år. Lav centralitet betyder ikke, at enheden i isoleret i praksis - den kan samarbejde
intensivt med få, faste partnere eller primært eksternt.
""")

    if svg_centralitet:
        st.markdown(f'<div style="max-width:800px;">{svg_centralitet}</div>', unsafe_allow_html=True)
        st.caption("Illustration af de to centralitetsmål. En mørkerød node markerer enheden med den højeste værdi i hvert mål.")
    
    with st.expander("Sådan beregnes brobyggerrollen"):
        st.markdown(
f"""
Målet opgør den **andel af de korteste forbindelsesveje mellem alle forfatterpar**, 
som passerer gennem den pågældende enhed ([Brandes 2008](https://doi.org/10.1016/j.socnet.2007.11.001)).

Formelt beregnes brobyggerrollen ($c_B$) for en enhed $v$ som:

$$ c_B(v) = \\sum_{{s \\neq v \\neq t \\in V}} \\frac{{\\sigma(s,t|v)}}{{\\sigma(s,t)}} $$

- $s$ og $t$ er to forskellige enheder i netværket $V$ 
- $\\sigma(s,t)$ er antallet af korteste forbindelsesveje mellem $s$ og $t$ 
- $\\sigma(s,t|v)$ er antallet af disse korteste forbindelsesveje, der passerer 
gennem enhed $v$

Målet **normaliseres**, så værdierne kan sammenlignes på tværs af netværk af forskellig
størrelse. Den viste værdi angiver derfor enhedens **relative strukturelle betydning**
i netværket - ikke et absolut antal forbindelser. 
""")

    st.markdown(
"""
Brug filtret nedenfor til at afgrænse analysen til de organisatoriske niveauer, der er relevante for 
analysen. Ved mange enheder kan det være en fordel at fokusere på ét niveau ad gangen.
""")


    def _cent_charts(label_key, wd_sorted, bs_sorted, extra_col=None, extra_map=None, color_map=None, ew_key=None, bs_key=None):
        """Render bar charts + table for one centralitet level."""
        if not wd_sorted:
            return
        wd_map = dict(wd_sorted)
        bs_map = dict(bs_sorted) if bs_sorted else {}
        keys   = [k for k, _ in wd_sorted]

        # Samlet samarbejdsomfang bar
        st.markdown(f"#### {label_key}, {year}")

        #tab_weg, tab_bet = st.tabs(["Samlet samarbejdsomfang", "Betweenness centralitet"])
        tab_weg, tab_bet = st.tabs(["Samlet samarbejdsomfang", "Brobyggerrolle"])
        with tab_weg:
            _fig_wd = go.Figure(go.Bar(
                y=keys, x=[wd_map[k] for k in keys], orientation="h",
                marker_color=[color_map.get(k, "#122947") for k in keys] if color_map else "#122947",
                text=[f"{fmt_ui(wd_map[k])}" for k in keys], textposition="inside",
            ))
            _fig_wd.update_layout(
                xaxis_title="Samlet samarbejdsomfang", yaxis=dict(autorange="reversed"),
                height=max(300, 32 * len(keys)), margin=dict(l=160, r=80, t=50),
                title = f"Samlet samarbejdsomfang: {label_key}, {year}"
            )
            st.plotly_chart(_fig_wd, width='stretch',
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
            )
            if filter_caption:
                st.caption(filter_caption)

        with tab_bet:
            # Betweenness bar
            if bs_map:
                #st.markdown(f"**Betweenness centralitet - {label_key}**")
                _bs_sorted_keys = sorted(keys, key=lambda k: -bs_map.get(k, 0))
                _fig_bs = go.Figure(go.Bar(
                    y=_bs_sorted_keys,
                    x=[bs_map.get(k, 0) for k in _bs_sorted_keys],
                    orientation="h",
                    marker_color=[color_map.get(k, "#122947") for k in _bs_sorted_keys] if color_map else "#122947",
                    text=[f"{fmt_ui(bs_map.get(k,0),4)}" for k in _bs_sorted_keys],
                    textposition="inside",
                ))
                _fig_bs.update_layout(
                    xaxis_title="Brobyggerrolle (normaliseret)",
                    yaxis=dict(autorange="reversed"),
                    height=max(300, 32 * len(keys)), margin=dict(l=160, r=80, t=50),
                    title=f"Brobyggerrolle: {label_key}, {year}"
                )
                st.plotly_chart(_fig_bs, width='stretch',
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
                )
                if filter_caption:
                    st.caption(filter_caption)

        if bs_map and len(keys) >= 4:
            _n = max(len(keys) - 1, 1)
            _wd_max = max((v for _, v in wd_sorted), default=1) or 1
            _bs_max = max((v for _, v in bs_map.items()), default=1) or 1
            _wd_norm = {k: v / _wd_max for k, v in wd_sorted}
            _bs_norm = {k: v / _bs_max for k, v in bs_map.items()}
            _rank_diff = {k: _bs_norm[k] - _wd_norm[k] for k in keys}

            _biggest_broker = max(keys, key=lambda k: _rank_diff[k])
            _biggest_active = min(keys, key=lambda k: _rank_diff[k])

            if abs(_rank_diff[_biggest_broker]) >= 0.05 or abs(_rank_diff[_biggest_active]) >= 0.05:
                if _rank_diff[_biggest_broker] >= 0.05:
                    st.markdown(
f"""
**{_biggest_broker}** har en relativt høj brobyggerrolle sammenlignet med sit samlede samarbejdsomfang - 
enheden fungerer som forbindelsesled på tværs af organisatoriske enheder, uden nødvendigvis at være den mest aktive samarbejdspartner."""
                        )
                if _rank_diff[_biggest_active] <= -0.05:
                    st.markdown(
f"""**{_biggest_active}** har et højt samlet samarbejdsomfang, med en relativt lav brobyggerrolle - 
enheden samarbejder meget, men primært inden for sin egen enhed."""
                        )
            else:
                st.markdown(
"""De to mål følger hinanden tæt - der er ingen enheder, der skiller sig markant ud ved at 
have en væsentligt højere brobyggerrolle end samarbejdsomfang eller omvendt.""")

            st.markdown(
f""" Samarbejdsomfang er normaliseret til intervallet 0-1, så den mest aktive enhed får værdien 1. 
Forskellen viser altså den *relative* forskel mellem de to rolle - ikke absolutte tal.
""")

        # Full table
        _rows = []
        _has_norm = bs_map and len(keys) >= 4
        for k in keys:
            row = {label_key: k, "Samlet samarbejdsomfang": float(wd_map.get(k, 0)),
                   "Brobyggerrolle": float(bs_map.get(k, 0))}
            if extra_col and extra_map:
                row[extra_col] = extra_map.get(k, "")
            if _has_norm:
                row["Samarbejdsomfang (normaliseret)"] = round(_wd_norm.get(k, 0), 3)
                row["Brobyggerrolle (normaliseret)"]   = round(_bs_norm.get(k, 0), 3)
                row["Brobygger vs. aktivitet (+ = mere brobygger)"] = round(_rank_diff.get(k, 0), 3)
            _rows.append(row)
        _schema = ([(label_key, pa.string())] +
                   ([(extra_col, pa.string())] if extra_col else []) +
                   [("Samlet samarbejdsomfang", pa.float64()), ("Brobyggerrolle", pa.float64())] +
                   ([("Samarbejdsomfang (normaliseret)", pa.float64()),
                     ("Brobyggerrolle (normaliseret)",   pa.float64()),
                     ("Brobygger vs. aktivitet (+ = mere brobygger)", pa.float64())] if _has_norm else []))
        
        with st.expander(f"Se tabel"):
            st.dataframe(build_table(_rows, _schema), width="stretch", hide_index=True)
            st.download_button("Download (.xlsx)",
                               data=rows_to_excel_bytes(_rows, [n for n, _ in _schema]),
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key=f"brobyg_norm_{label_key}_{year}_{mode}")

        # ____Udvikling over tid_____
        if all_years_data and len(all_years_data) >= 2 and (ew_key or bs_key):
            st.markdown("##### Udvikling over tid")
            years_sorted = sorted(all_years_data.keys())

            _tid_tabs = []
            if ew_key: _tid_tabs.append("Samarbejdsomfang")
            if bs_key: _tid_tabs.append("Brobygning")
            _tid_tab_objs = st.tabs(_tid_tabs)

            for tab_name, _tab in zip(_tid_tabs, _tid_tab_objs):
                with _tab:
                    if tab_name == "Samarbejdsomfang":
                        st.markdown(
f"""Figuren viser, hvordan **samarbejdsomfanget** for {label_key.lower()}ne 
har udviklet sig over tid. En stigende kurve betyder, at enheden indgår i flere 
sampubliceringer - ikke nødvendigvis med flere forskellige partnere.""")
                    else:
                        st.markdown(
f"""Figuren viser udviklingen i **brobyggerrollen** for {label_key.lower()}ne over tid. 
En stigende kurve betyder, at enheden i stigende grad fungerer som forbindelsesled 
mellem grupper, der ellers ikke samarbejder direkte - og dermed bliver mere strukturelt 
vigtig for netværkets sammenhæng. En faldende kurve kan betyde, at andre enheder har 
overtaget brobyggerrollen, eller at netværket generelt er blevet mere sammenhængende.""")
                    _key = ew_key if tab_name == "Samarbejdsomfang" else bs_key
                    _units_tid = sorted({
                        u for snap in all_years_data.values()
                        for u in snap.get(_key, {})
                    })
                    _fig_tid = go.Figure()
                    for u in _units_tid:
                        _y = [all_years_data[yr].get(_key, {}).get(u, 0) for yr in years_sorted]
                        _color = (color_map or {}).get(u, "#122947")
                        _fig_tid.add_trace(go.Scatter(
                            x=years_sorted, y=_y,
                            name=u,
                            mode="lines+markers",
                            line=dict(width=2, color=_color),
                            marker=dict(size=7, color=_color),
                        ))
                    _yaxis_title = (
                        "Samlet samarbejdsomfang (forfatterpar)"
                        if tab_name == "Samarbejdsomfang"
                        else "Brobyggerrolle (aggregeret betweenness)"
                    )
                    _fig_tid.update_layout(
                        yaxis_title=_yaxis_title,
                        xaxis=dict(tickmode="array", tickvals=years_sorted, dtick=1),
                        height=420,
                        title=f"Udvikling i {_yaxis_title.lower()}, {years_sorted[0]}–{years_sorted[-1]}",
                        margin=dict(t=50),
                        legend=dict(orientation="v", yanchor="top", y=1, xanchor="left", x=1.02),
                    )
                    st.plotly_chart(_fig_tid, width="stretch", key=f"cent_tid_{tab_name}_{label_key}_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
                    )
                    if filter_caption:
                        st.caption(filter_caption)

                    # Tabel
                    _tbl_rows = [
                        {label_key: u, **{str(yr): round(all_years_data[yr].get(_key, {}).get(u, 0), 1) for yr in years_sorted}}
                        for u in _units_tid
                    ]
                    _tbl_schema = (
                        [(label_key, pa.string())] +
                        [(str(yr), pa.float64()) for yr in years_sorted]
                    )
                    with st.expander("Se tabel"):
                        st.dataframe(build_table(_tbl_rows, _tbl_schema), hide_index=True, width="stretch")
                        st.download_button(
                            "Download (.xlsx)",
                            data=rows_to_excel_bytes(_tbl_rows, [n for n, _ in _tbl_schema]),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_cent_tid_{tab_name}_{label_key}_{year}_{mode}",
                        )


    _available = []
    if faculty_wd_sorted:
        _available.append("Fakulteter")
    if inst_wd_sorted:
        _available.append("Institutter")
    if grp_wd_sorted:
        _available.append("Stillingsgrupper (aggregeret)")
    if grp_node_wd_sorted:
        _available.append("Stillingsgrupper (node-niveau)")

    if not _available:
        st.error("Ingen centralitetsdata tilgængelig for det valgte udsnit.")
        return

    _state_key = f"cent_level_filter_{mode}"
    if _state_key not in st.session_state:
        st.session_state[_state_key] = _available
    # Behold kun valg der stadig er tilgængelige i denne mode
    _default = [v for v in st.session_state[_state_key] if v in _available] or _available
    _selected = st.multiselect(
        "**Vis nøgleaktører for:**",
        options=_available,
        default=_default,
        key=f"cent_level_filter_{year}_{mode}",
        on_change=lambda: st.session_state.update(
            {_state_key: st.session_state[f"cent_level_filter_{year}_{mode}"]}
        ),
    )

    if not _selected:
        st.error("Vælg mindst ét niveau ovenfor.")
        return

    _inst_to_fac = {m.get("inst"): m.get("fac") for m in node_meta.values() if m.get("inst")}

    # Byg institut-farvekort samme metode som render_tab_institutter
    _inst_color_map = {}
    if faculty_base_colors:
        _insts_by_fac: dict[str, list] = {}
        for m in node_meta.values():
            if m.get("inst") and m.get("fac"):
                _insts_by_fac.setdefault(m["fac"], [])
                if m["inst"] not in _insts_by_fac[m["fac"]]:
                    _insts_by_fac[m["fac"]].append(m["inst"])
        for fac, insts in _insts_by_fac.items():
            insts_sorted = sorted(insts)
            k = max(1, len(insts_sorted))
            base = faculty_base_colors.get(fac, "#122947")
            for rank, inst in enumerate(insts_sorted):
                t = rank / max(1, k - 1)
                lf = 0.5 + 1.8 * t
                sf = 1.3 - 0.7 * t
                _inst_color_map[inst] = adjust_color(base, lf, sf)

    if "Fakulteter" in _selected:
        st.markdown("---")
        _cent_charts("Fakulteter", faculty_wd_sorted, faculty_bs_sorted,
                     color_map=faculty_base_colors, ew_key="fac_ew", bs_key="fac_bs")

    if "Institutter" in _selected:
        st.markdown("---")
        _cent_charts("Institutter", inst_wd_sorted, inst_bs_sorted,
                     extra_col="Fakultet", extra_map=_inst_to_fac,
                     color_map=_inst_color_map, ew_key="inst_ew", bs_key="inst_bs")

    if "Stillingsgrupper (aggregeret)" in _selected:
        st.markdown("---")
        _cent_charts("Stillingsgrupper", grp_wd_sorted, grp_bs_sorted,
                     color_map=grp_colors, ew_key="grp_ew", bs_key="grp_bs")

    if "Stillingsgrupper (node-niveau)" in _selected:
        st.markdown("---")
        _node_color_map = {}
        for k, _ in (grp_node_wd_sorted or []):
            parts = [p.strip() for p in k.split("|")]
            # Label-format: "Stillingsgruppe | Institut | Fakultet"
            # Prøv fakultet (sidste del) → institut (næstsidste) → fallback
            fac = parts[-1] if len(parts) >= 3 else (parts[-1] if parts else "")
            inst = parts[-2] if len(parts) >= 2 else ""
            color = (
                (faculty_base_colors or {}).get(fac)
                or _inst_color_map.get(inst)
                or "#122947"
            )
            _node_color_map[k] = color
        _cent_charts("Stillingsgrupper (node-niveau)",
                     grp_node_wd_sorted, grp_node_bs_sorted or [],
                     color_map=_node_color_map)
        st.markdown(
f"""
Udvikling over tid vises ikke for stillingsgrupper på node-niveau, da de sammensatte 
kombinationer af stillingsgruppe, institut og fakultet varierer fra år til år og ikke giver 
et meningsfuldt sammenligningsgrundlag over tid. Brug fanen *Stillingsgrupper* (aggregeret) 
for tidsserier på tværs af år.
""")
    
def compute_modularity_pre_for_key(snap: dict, comm_key: str) -> float:
    """Genberegn foruddefineret modularitet for et snapshot med en given comm_key."""
    import networkx as nx
    nodes_snap = snap.get("nodes_keep", [])
    edges_snap = snap.get("edges_keep", [])
    node_meta_snap = snap.get("node_meta", {})
    if not nodes_snap or not edges_snap:
        return float("nan")
    G = nx.Graph()
    for u in nodes_snap: G.add_node(u)
    for u, v, w, *_ in edges_snap: G.add_edge(u, v, weight=w)
    _connected = {n for n in G.nodes() if G.degree(n) > 0}
    G_conn = G.subgraph(_connected).copy()
    comms = {}
    for nid in nodes_snap:
        gl = node_meta_snap.get(nid, {}).get(comm_key, "")
        if gl:
            comms.setdefault(gl, []).append(nid)
    filtered = [[n for n in c if n in _connected] for c in comms.values()]
    filtered = [c for c in filtered if c]
    if not filtered or G_conn.number_of_edges() == 0:
        return float("nan")
    filtered = [c for c in filtered if len(c) >= 2]
    if len(filtered) < 2:
        return float("nan")
    try:
        val = modq(G_conn, filtered, weight="weight")
        return val if val >= 0 else float("nan")
    except Exception:
        return float("nan")

def render_tab_netvaerksstruktur(year, mode, density, modularity_pre, modularity_greedy,
                                  n_comms, communities_dict, greedy_comms, comm_key,
                                  edges_keep=None, node_meta=None, all_years_data=None, filter_caption=None):
    abbrs = {"fac": "fakulteter", "inst": "institutter", "grp": "stillingsgrupper"}
    abbrs_singular = {"fac": "fakultet", "inst": "institut", "grp": "stillingsgruppe"}

     # Klyngeniveau-vælger
    _available_keys = [
        k for k, active in [
            ("fac",  "F" in mode),
            ("inst", "I" in mode),
            ("grp",  "G" in mode),
        ] if active
    ]

    st.subheader("Samarbejdsmønstre")
    if "S" in mode or "N" in mode: 
        st.error("Analyserne nedenfor baseres på netværksplottet - uden nogen diversitetsdimensioner.")

    st.markdown(
f"""
Fanen analyserer **strukturen i KU’s sampubliceringsnetværk** i {year}.
Fokus er ikke på enkeltrelationer mellem forskere eller enheder, men på,
hvordan samarbejdet samlet set **organiserer sig** på tværs af den formelle
organisationsstruktur.

Analysen gør det muligt at skelne mellem samarbejde, der foregår
i afgrænsede, specialiserede miljøer, og samarbejde, der binder KU sammen
på tværs af institutter og fakulteter i det viste sampubliceringsnetværk. Det har betydning for, hvordan viden
cirkulerer, hvor robust samarbejdsstrukturen er, og hvor let nye forbindelser
kan opstå.

Analyserne kan baseres på forskellige organisatoriske niveauer og omfatter:
- Netværkstæthed og modularitet som samlede strukturmål  
- Sammenligning af formelle og datadrevne samarbejdsklynger  
- Sampubliceringsstyrke mellem klynger målt i forfatterpar
"""
    )
    if len(_available_keys) > 1:
        _key_labels = {"fac": "Fakultet", "inst": "Institut", "grp": "Stillingsgruppe"}
        comm_key = st.radio(
                "**Klyngestrukturen baseres på:**",
                options=_available_keys,
                format_func=lambda k: _key_labels[k],
                horizontal=True,
                index=_available_keys.index(comm_key) if comm_key in _available_keys else 0,
                key=f"comm_key_radio_{year}_{mode}",
            )

    st.markdown(
"""
Analyserne i denne fane beskriver sampubliceringsmønstre - ikke årsager eller strategiske hensigter. Resultaterne
bør derfor læses som et strukturelt supplement til kvalitative vurderinger af samarbejde og organisering. 
""")

    st.markdown("---")

    st.markdown(
f"""
#### Netværksstruktur i {year}


**Netværkstæthed** er et overordnet mål for, hvor mange af de mulige forbindelser
i netværket der faktisk er realiseret.

En høj tæthed indikerer, at mange enheder samarbejder med hinanden, mens
lav tæthed afspejler et mere opdelt netværk. Tæthed siger dog **ikke noget om
kvaliteten eller tværgående karakteren af samarbejdet** – et netværk kan være
tæt uden at binde forskellige miljøer sammen, hvis samarbejdet foregår i mange
små, isolerede grupper.

Tæthed bør derfor altid læses sammen med modularitet og klyngestruktur.

""")

    st.metric("Netværkstæthed", f"{fmt_ui(density,3)}")

    st.markdown(
f"""
**Modularitet** beskriver, i hvilken grad et netværk er opdelt i adskilte klynger 
sammenlignet med, hvad man vil forvente tilfældigt. Begge modularitetsmål beregnes 
**kun på den forbundne del af netværket** - dvs. enheder uden sampubliceringer (isolerede 
noder) indgår ikke.

Typiske fortolkningsgrænser: 

- **> 0.5** antyder et netværk, der karakteriseret ved tydeligte adskilte grupper (klynger)
- **0.3 - 0.5** er moderat til stærk klyngestruktur
- **< 0.3** indikerer et mere sammenhængende netværk uden tydelig klyngestruktur
""")

    if len(_available_keys) > 1:
        # Genberegn communities_dict med det valgte niveau
        communities_dict = {}
        for nid, m in (node_meta or {}).items():
            gl = m.get(comm_key, "")
            if gl:
                communities_dict.setdefault(gl, []).append(nid)
        # Genberegn modularity_pre med valgt comm_key
        _G2_recompute = nx.Graph()
        for nid in (node_meta or {}):
            _G2_recompute.add_node(nid)
        for u, v, w, *_ in (edges_keep or []):
            _G2_recompute.add_edge(u, v, weight=w)
        modularity_pre = _compute_mod_pre(
            _G2_recompute,
            list(_G2_recompute.nodes()),
            node_meta or {},
            comm_key,
        )

    colA, colB, colC = st.columns(3)
    with colA:
        if mode in ("I", "F", "G"):
            st.metric("Modularitet (foruddefinerede klynger)", "n/a",
            help = "Modularitet er kun meningsfuldt, når der er flere noder per klynge")
        else:
            _mod_pre_str = fmt_ui(modularity_pre, 3) if not np.isnan(modularity_pre) else "n/a"
            _mod_pre_help = (
                f"Klyngeniveau: {len(communities_dict)} {abbrs.get(comm_key)}. "
                + ("En negativ værdi betyder, at de foruddefinerede klynger afspejler netværkets struktur dårligere end en tilfældig opdeling - "
                "hvilket typisk sker, når klyngerne (f.eks. stillingsgrupper) går på tværs af de samarbejdsmønstre, der faktisk præger netværket."
                if not np.isnan(modularity_pre) and modularity_pre < 0 else "")
            )
            st.metric("Modularitet (foruddefinerede klynger)", _mod_pre_str, help=_mod_pre_help)
    with colB:
        if mode in ("I", "F", "G"):
            st.metric("Modularitet (Greedy)", "n/a",
            help = "Modularitet er kun meningsfuldt, når der er flere noder per klynge")
        else:
            st.metric("Modularitet (greedy)",
                    f"{fmt_ui(modularity_greedy,3)}" if not np.isnan(modularity_greedy) else "n/a")
    with colC:
        if mode in ("I", "F", "G"):
            st.metric("Antal Greedy-klynger", "n/a",
            help = "Modularitet er kun meningsfuldt, når der er flere noder per klynge")
        else:
            st.metric("Antal Greedy-klynger", n_comms)


    greedy_comms_labeled = [
            sorted([
                " | ".join(p for p in (node_meta.get(nid, {}).get("fac", ""),
                                        node_meta.get(nid, {}).get("inst", ""),
                                        node_meta.get(nid, {}).get("grp", "")) if p)
                for nid in comm        
            ])
            for comm in greedy_comms
        ]

    greedy_rows = [{
        "Klynge (Greedy)": f"Greedy {i}",
        "Antal noder": len(c),
        "Noder": ", ".join(c)}
        for i, c in enumerate(greedy_comms_labeled, 1)]

    schema = [
        ("Klynge (Greedy)", pa.string()),
        ("Antal noder",     pa.int64()),
        ("Noder",           pa.string()),
    ]

    with st.expander("Sådan beregnes modularitet"):
        st.markdown(
"""
Modularitet $Q$ måler, om netværkets klyngeopdeling har **flere interne forfatterpar end forventet**
i et tilfædigt netværk med samme antal forbindelser per node: 

$$Q = \\frac{1}{2m} \\sum_{i,j} \\left[ A_{ij} - \\frac{k_i k_j}{2m} \\right] \\delta(c_i, c_j)$$

- $m$ er det samlede antal **forfatterpar** i netværket
- $A_{ij}$ er antallet af **forfatterpar** mellem node $i$ og $j$
- $\\frac{k_i k_j}{2m}$ er den forventede antal **forfatterpar** mellem $i$ og $j$ i et tilfældigt netværk, 
hvor nodernes samlede antal forbindelser er bevaret
- $\\delta(c_i, c_j)$ er lig 1, hvis $i$ og $j$ er i samme klynge, ellers lig 0

$Q$ er altid i intervallet $[-1, 1]$.
""")

    _kommentar = ""
    if not np.isnan(modularity_greedy) and not np.isnan(modularity_pre):
        _diff = round(modularity_greedy - modularity_pre, 3)
        _diff_str = f"+{_diff}" if _diff > 0 else str(_diff)
        _kommentar = ""
        
        _level_label = {
            "fac": "fakultetsbaserede klynger",
            "inst": "institutbaserede klynger",
            "grp": "stillingsgruppebaserede klynger"}[comm_key]

        if _diff > 0.15:
            _kommentar = (
                f"For **netværket i {year}**, analyseret med {_level_label}, er differencen på **{_diff_str}** substantiel og indikerer, at valgte klyngeopdeling "
                f"organisationsstruktur er en **suboptimal** beskrivelse af sampubliceringsmønstrene. "
                f"Der eksisterer altså stærkere naturlige samarbejdsklynger end dem, der følger af "
                f"den formelle struktur - f.eks. på tværs af institutter eller fakulteter."
            )
        elif _diff > 0.05:
            _kommentar = (
                f"For **netværket i {year}**, analyseret med {_level_label}, er differencen på **{_diff_str}** moderat og indikerer, at valgte klyngeopdeling "
                f"organisationsstruktur delvist afspejler sampubliceringsmønstrene, men at der "
                f"eksisterer samarbejdsrelationer der går på tværs af strukturen."
            )
        else:
            _kommentar = (
                f"For **netværket i {year}**, analyseret med {_level_label}, er differencen på **{_diff_str}** lille, hvilket indikerer at valgte klyngeopdeling "
                f"organisationsstruktur er en god beskrivelse af sampubliceringsmønstrene."
            )

    # Node-count bar per cluster
    st.markdown(
f"""
##### Antal noder per klynge

For at beregne modulariteten opdeles noderne på to forskellige måder: 

- **Foruddefinerede klynger** måler, om KU's organisatoriske struktur (fakulteter, institutter eller 
stillingsgrupper afhængigigt af den valgte visning) afspejler de faktiske sampubliceringsmønstre. 
En høj værdi betyder, at forskere primært publicerer med kolleger fra samme organisatoriske enhed. 
- **Greedy-klynger** er baseret på Greedy-algoritmen, som finder opdelingen, der *maksimerer* modulariteten - 
og fungerer som en øvre reference for, hvor stærk en klyngestruktur netværket kan have
([Newman 2004](https://journals.aps.org/pre/abstract/10.1103/PhysRevE.70.066111)).

Sammenligningen mellem foruddefinerede klynger og Greedy-klynger giver et billede af, 
**hvor godt KU's formelle organisationsstruktur stemmer overens med de faktiske sampubliceringsmønstre**.

Hvis Greedy-klyngerne ligner de organisatoriske enheder, tyder det på, at samarbejdet
følger den formelle struktur. Store forskelle indikerer derimod, at forskningssamrbejdet organiserer
sig på tværs af institutter eller fakulteter - f.eks. omkring fælles metoder, 
temaer eller infrastrukturer. 

**Eksempel**: antag, at fakulteter bruges som foruddefinerede klynger og giver en modularitet
på 0,25. Greedy-algoritmen finder en alternativ opdeling med modularitet 0,45. Forskellen på
0,20 indikerer, at samarbejdet organiserer sig i klynger, der afviger markant fra fakultetsstrukturen - 
f.eks. på tværs af organisatoriske skel omkring fælles metoder eller temaer. 

{_kommentar}
""")

    _tab_pre, _tab_greedy = st.tabs(["Foruddefinerede klynger", "Greedy-klynger"])

    with _tab_pre:
        _comm_sizes = {g: len(m) for g, m in sorted(communities_dict.items(), key=lambda x: x[0])}
        _fig_ns = go.Figure(go.Bar(
            x=list(_comm_sizes.keys()), y=list(_comm_sizes.values()),
            marker_color="#122947",
            text=list(_comm_sizes.values()), textposition="inside",
        ))
        _fig_ns.update_layout(
            title=f"Foruddefinerede klynger: antal noder per klynge, {year}",
            xaxis_title="Klynger", 
            yaxis_title="Antal noder",
            height=320, margin=dict(t=50),
        )
        st.plotly_chart(_fig_ns, width='stretch',
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
        )
        if filter_caption:
            st.caption(filter_caption)

        with st.expander("Se klyngetabel med foruddefinerede klynger"):
            _klynge_col = {"fac": "Fakultet", "inst": "Institut", "grp": "Stillingsgruppe"}.get(comm_key, "Enhed")
            _schema_pre = [(_klynge_col, pa.string()), ("Antal noder", pa.int64())]
            _table_rows_pre = [
                {_klynge_col: g, "Antal noder": len(m)}
                for g, m in sorted(communities_dict.items(), key=lambda x: -len(x[1]))
            ]
            st.dataframe(build_table(_table_rows_pre, _schema_pre), hide_index=True, width="stretch")
            st.download_button(
                "Download (.xlsx)",
                data=rows_to_excel_bytes(_table_rows_pre, [n for n, _ in _schema_pre]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"klynge_{year}_{mode}")
    
    with _tab_greedy:
        if greedy_comms_labeled:
            _greedy_sizes = {f"Greedy {i}": len(c) for i, c in enumerate(greedy_comms_labeled, 1)}
            _fig_gs = go.Figure(go.Bar(
                x=list(_greedy_sizes.keys()), y=list(_greedy_sizes.values()),
                marker_color="#122947",
                text=list(_greedy_sizes.values()), textposition="inside",
            ))
            _fig_gs.update_layout(
                title=f"Greedy-klynger: antal noder per klynge, {year}",
                xaxis_title="Klynger", 
                yaxis_title="Antal noder",
                height=320, margin=dict(t=50),
            )
            st.plotly_chart(_fig_gs, width='stretch',
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
            )
            if filter_caption:
                st.caption(filter_caption)
        else:
            st.error("Ingen Greedy-klynger tilgængelige.")
        
        with st.expander("Se klyngetabel med Greedy-klynger"):
            st.dataframe(build_table(greedy_rows, schema), hide_index=True, width="stretch")
            st.download_button(
                "Download (.xlsx)",
                data=rows_to_excel_bytes(greedy_rows, [n for n, _ in schema]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"greedy_klynge_{year}_{mode}")

    # ── Modularitet og tæthed over tid ───────────────────────────────────────
    if all_years_data and len(all_years_data) >= 2:
        st.markdown(
"""##### Udvikling over tid - modularitet og netværkstæthed
I figuren nedenfor vises netværkstætheden samt modulariteterne. År, hvor 
modulariteten er angivet som n/a, er udeladt fra visualiseringen.

Vær opmærksom på, hvilke filtre der er valgt i sidepanelet, da de både påvirker 
beregninger og visning.
""")

        years_sorted = sorted(all_years_data.keys())

        _density_vals  = [all_years_data[y].get("density",           float("nan")) for y in years_sorted]
        _mod_pre_key = f"modularity_pre_{comm_key}"
        _mod_pre_vals = [all_years_data[y].get(_mod_pre_key, float("nan")) for y in years_sorted]
        _mod_gr_vals   = [all_years_data[y].get("modularity_greedy", float("nan")) for y in years_sorted]

        def _clean(lst):
            return [None if (v is None or (isinstance(v, float) and math.isnan(v))) else v for v in lst]

        # Filtrér år væk hvor alle tre værdier er None/nan
        _yr_plot = years_sorted

        def _filter(lst):
            return _clean(lst)

        _fig_mt = go.Figure()
        _fig_mt.add_trace(go.Scatter(
            x=_yr_plot, y=_filter(_density_vals),
            name="Netværkstæthed",
            mode="lines+markers+text",
            text=[f"{fmt_ui(v,3)}" if v is not None else "" for v in _filter(_density_vals)],
            textposition="top center",
            line=dict(color="#122947", width=2),
            marker=dict(size=8),
        ))
        _fig_mt.add_trace(go.Scatter(
            x=_yr_plot, y=_filter(_mod_pre_vals),
            name="Modularitet (foruddefinerede)",
            mode="lines+markers+text",
            text=[f"{fmt_ui(v,3)}" if v is not None else "" for v in _filter(_mod_pre_vals)],
            textposition="top center",
            line=dict(color="#4a7ca8", width=2, dash="dash"),
            marker=dict(size=8),
        ))
        _fig_mt.add_trace(go.Scatter(
            x=_yr_plot, y=_filter(_mod_gr_vals),
            name="Modularitet (greedy)",
            mode="lines+markers+text",
            text=[f"{fmt_ui(v,3)}" if v is not None else "" for v in _filter(_mod_gr_vals)],
            textposition="bottom center",
            line=dict(color="#39641c", width=2, dash="dot"),
            marker=dict(size=8),
        ))
        _fig_mt.update_layout(
            xaxis=dict(tickmode="array", tickvals=_yr_plot, dtick=1),
            yaxis=dict(title="Værdi (0-1)", rangemode="tozero"),
            legend=dict(orientation="v", yanchor="top", y=1, xanchor="left", x=1.02),
            height=420,
            margin=dict(t=50, b=80),
            title=f"Netværkstæthed og modularitet over tid, {years_sorted[0]}–{years_sorted[-1]}",
        )
        st.plotly_chart(_fig_mt, width='stretch',
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
        )
        if filter_caption:
            st.caption(filter_caption)

        with st.expander("Se tabel", expanded=False):
            _metrics = [
                ("Netværkstæthed",               "density",           4),
                ("Modularitet (foruddefinerede)", _mod_pre_key,        4),
                ("Modularitet (greedy)",          "modularity_greedy", 4),
            ]
            _pivot_rows = [
                {"Metrik": label, **{str(y): round(all_years_data[y].get(key, float("nan")), dec) for y in years_sorted}}
                for label, key, dec in _metrics
            ]
            _pivot_schema = (
                [("Metrik", pa.string())] +
                [(str(y), pa.float64()) for y in years_sorted]
            )
            st.dataframe(build_table(_pivot_rows, _pivot_schema), hide_index=True, width="stretch")
            st.download_button(
                "Download (.xlsx)",
                data=rows_to_excel_bytes(_pivot_rows, [n for n, _ in _pivot_schema]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"forud_greedy_{year}_{mode}"
            )

    HEATMAP_MAX = 10

    st.markdown("---")

    # ── Heatmap: forfatterpar mellem klynger ────────────────────────────────────
    st.markdown(f"#### Sampubliceringsstyrke mellem klyngerne i {year}")
    st.markdown(
f"""
Heatmappet viser det samlede antal forfatterpar mellem {abbrs.get(comm_key, comm_key)}.
Diagonalen angiver interne forfatterpar inden for samme {abbrs_singular.get(comm_key, comm_key)}.

Vær opmærksom på, at det kun er forfatterpar på tværs af organisatoriske enheder.
""")
    _units = sorted(communities_dict.keys())
    if len(_units) >= 2 and edges_keep and node_meta:
        # Map each node → its cluster label via comm_key
        _node_to_unit = {
            nid: m.get(comm_key, "")
            for nid, m in node_meta.items()
            if m.get(comm_key, "") in communities_dict.keys()
        }
        # Accumulate edge weight into unit×unit matrix (symmetric)
        _mat = {(a, b): 0.0 for a in _units for b in _units}
        for u, v, w, *_ in edges_keep:
            gu = _node_to_unit.get(u, "")
            gv = _node_to_unit.get(v, "")
            if gu and gv:
                _mat[(gu, gv)] = _mat.get((gu, gv), 0.0) + w
                if gu != gv:
                    _mat[(gv, gu)] = _mat.get((gv, gu), 0.0) + w

        _z = [[_mat.get((a, b), 0.0) for b in _units] for a in _units]

        if len(_units) > HEATMAP_MAX:
            _unit_totals = sorted(
                _units,
                key=lambda a: sum(_mat.get((a, b), 0) for b in _units),
                reverse=True
            )
            _default_units = _unit_totals[:HEATMAP_MAX]
            _selected_units = st.multiselect(
                f"For mange enheder til overskueligt heatmap ({len(_units)}). Vælg, hvilke der vises:",
                options=_units,
                default=_default_units,
                key=f"heatmap_filter_{year}_{mode}",
            )
            _units_plot = _selected_units if _selected_units else _default_units
        else:
            _units_plot = _units

        if _units_plot:
            _z_plot = [[_mat.get((a, b), 0.0) for b in _units_plot] for a in _units_plot]
            _fig_heat = go.Figure(go.Heatmap(
                z=_z_plot, x=_units_plot, y=_units_plot,
                colorscale=[[0, "#f0f4f8"], [1, "#122947"]],
                text=[[f"{fmt_ui(_mat.get((a,b),0),0)}" for b in _units_plot] for a in _units_plot],
                texttemplate="%{text}",
                hovertemplate="%{y} → %{x}: %{z:.0f}<extra></extra>",
                showscale=True,
                colorbar=dict(title="Antal forfatterpar"),
            ))
            _fig_heat.update_layout(
                xaxis_title=abbrs.get(comm_key, comm_key).capitalize(),
                yaxis_title=abbrs.get(comm_key, comm_key).capitalize(),
                yaxis=dict(showgrid=False),
                height=max(380, 60 * len(_units_plot)),
                margin=dict(l=140, b=140, t=50),
                title=f"Sampubliceringsstyrke mellem {abbrs.get(comm_key, comm_key)}, {year}"
            )
            
            st.plotly_chart(_fig_heat, width='stretch',
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
            )
            if filter_caption:
                st.caption(filter_caption)

        _heat_rows = [
                {"Fra": a, "Til": b, "Antal forfatterpar": int(_mat.get((a, b), 0))}
                for a in _units for b in _units
                if a != b and _mat.get((a, b), 0) > 0
            ]
        _heat_schema = [("Fra", pa.string()), ("Til", pa.string()), ("Antal forfatterpar", pa.int64())]
        
        with st.expander("Se tabel"):
            st.dataframe(build_table(_heat_rows, _heat_schema), hide_index=True, width="stretch")
            st.download_button(
                    "Download (.xlsx)",
                    data=rows_to_excel_bytes(_heat_rows, ["Fra", "Til", "Antal forfatterpar"]),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"heat_{year}_{mode}"
                )

    _units = sorted(communities_dict.keys())
    


def render_tab_netværksudvikling(year, mode, all_years_data=None, faculty_base_colors=None, filter_caption=None, show_intra=True):
    st.subheader("Netværksudvikling")
    if "S" in mode or "N" in mode:
        st.error("Analyserne nedenfor baseres på netværksplottet – uden nogen diversitetsdimensioner.")

    st.markdown(
"""
Fanen viser, hvordan KU’s sampubliceringsnetværk **forandrer sig over tid**.

Hvor *Samarbejdsmønstre*-fanen beskriver netværkets struktur i det valgte år, fokuserer
denne fane på bevægelser i samarbejdet: om netværket bliver mere eller mindre
siloopdelt, om samarbejdsrelationer fornyes eller gentages, og om bestemte
enheder systematisk bevæger sig ind og ud af netværket.

Analysen giver dermed et dynamisk perspektiv på sampublicering og viser, om
observerede mønstre er stabile strukturer eller udtryk for kortsigtede udsving.

Fanen belyser netværksudviklingen gennem fire centrale analyser:

- **Siloering over tid**, målt ved udviklingen i intra‑ og inter‑samarbejde  
- **Fornyelse i samarbejdsrelationerne**, dvs. om par fortsætter, forsvinder eller opstår  
- **Udvikling i specifikke samarbejder**, hvor enkelte relationer kan følges over tid  
- **Isolerede enheder over tid**, dvs. hvilke enheder der gentagne gange står uden for netværket  

Alle analyser afspejler de filtre og organisatoriske niveauer, der er valgt i
sidepanelet.
"""
    )

    if not all_years_data or len(all_years_data) < 2:
        st.error("Netværksudvikling kræver mindst to års data. Udvid udvalget i sidepanelet.")
        return

    years_sorted = sorted(all_years_data.keys())

    # ══════════════════════════════════════════════════════════════════════════
    # SCOPE: niveau + enhed (gælder alle sektioner nedenfor)
    # ══════════════════════════════════════════════════════════════════════════
    _available_silo_levels = [
        k for k, active in [
            ("fac",  "F" in mode),
            ("inst", "I" in mode),
            ("grp",  "G" in mode),
        ] if active
    ]
    _level_labels = {"fac": "Fakultet", "inst": "Institut", "grp": "Stillingsgruppe"}
    _level_labels_single = {"fac": "fakultet", "inst": "institut", "grp": "stillingsgruppe"}


    if len(_available_silo_levels) > 1:
        _silo_level = st.radio(
            "**Organisatorisk niveau**",
            options=_available_silo_levels,
            format_func=lambda k: _level_labels[k],
            horizontal=True,
            key=f"netudv_level_{year}_{mode}",
        )
    else:
        _silo_level = _available_silo_levels[0]
        st.caption(f"Organisatorisk niveau: **{_level_labels[_silo_level]}** (låst af valgt mode)")

    # Find enheder på det valgte niveau – union over alle år, så en enhed
    # der falder ud i ét år stadig kan vælges til sin historik
    _unit_set = set()
    for yr in years_sorted:
        _unit_set |= set(all_years_data[yr].get(f"{_silo_level}_intra_ew", {}).keys())
        _unit_set |= set(all_years_data[yr].get(f"{_silo_level}_inter_ew", {}).keys())
    _units_sorted = sorted(_unit_set)

    _scope_unit = st.selectbox(
        "**Enhed**",
        options=["Samlet KU"] + _units_sorted,
        index=0,
        key=f"netudv_scope_unit_{year}_{mode}",
        help=(
                "Vælg en specifik enhed for at zoome ind på dens samarbejdsudvikling. "
                "Valget påvirker alle sektioner nedenfor."
        ),
    )
    _scope_unit_val = None if _scope_unit == "Samlet KU" else _scope_unit

    # Fallback-palette til stillingsgrupper (som ikke har en fakultetsfarve)
    _GRP_PALETTE = [
        "#425570", "#901a1E", "#39641c", "#b8860b", "#5d3a8e",
        "#2c6e7f", "#a04040", "#4a6b2c", "#8b5a2b", "#704070",
    ]

    # Byg institut → fakultet mapping ud fra pair_keys i alle snapshots,
    # så vi kan farve institutter efter deres fakultet
    _inst_to_fac_local: dict[str, str] = {}
    for yr in years_sorted:
        for p in all_years_data[yr].get("top_pairs", {}).keys():
            for _side in p.split(" ↔ "):
                _parts = [_x.strip() for _x in _side.split(" | ")]
                if len(_parts) >= 2 and _parts[0] and _parts[1]:
                    _inst_to_fac_local.setdefault(_parts[1], _parts[0])

    def _color_for_unit(unit: str, level: str, palette_idx: int = 0) -> str:
        """Returnér en farve for en enhed. Fakultet/institut farves efter fakultet;
        stillingsgruppe farves efter en fast rotation."""
        if not faculty_base_colors:
            return _GRP_PALETTE[palette_idx % len(_GRP_PALETTE)]
        if level == "fac":
            return faculty_base_colors.get(unit, "#425570")
        if level == "inst":
            _fac = _inst_to_fac_local.get(unit, "")
            return faculty_base_colors.get(_fac, "#425570")
        # grp-niveau: fast palette
        return _GRP_PALETTE[palette_idx % len(_GRP_PALETTE)]

    # Helper: tjek om en pair_key eller isolated_unit involverer en given enhed
    # pair_key har formen: "FAC_A | INST_A | GRP_A ↔ FAC_B | INST_B | GRP_B"
    # isolated_unit har formen: "FAC | INST | GRP"
    def _unit_at_level(label: str, level: str) -> str:
        """Returnér værdien af level (fac/inst/grp) i et 'fac | inst | grp'-label."""
        _part_idx = {"fac": 0, "inst": 1, "grp": 2}[level]
        _parts = [p.strip() for p in label.split(" | ")]
        return _parts[_part_idx] if _part_idx < len(_parts) else ""

    def _pair_involves_unit(pair_key: str, unit: str, level: str) -> bool:
        for side in pair_key.split(" ↔ "):
            if _unit_at_level(side, level) == unit:
                return True
        return False

    def _pair_other_side(pair_key: str, unit: str, level: str) -> str:
        """Returnér modpartens enhed på samme niveau. Tom streng hvis begge sider er unit."""
        _sides = pair_key.split(" ↔ ")
        if len(_sides) != 2:
            return ""
        _a = _unit_at_level(_sides[0], level)
        _b = _unit_at_level(_sides[1], level)
        if _a == unit and _b != unit:
            return _b
        if _b == unit and _a != unit:
            return _a
        return ""  # intra-unit par – ikke interessant for cross-unit visning

    st.markdown("---")

    # ══════════════════════════════════════════════════════════════════════════
    # 1. SILOERING OVER TID
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("#### Siloering over tid")

    def _pct(snap, num_key, denom_key="total_pubs"):
        tot = snap.get(denom_key, 0)
        return round(100 * snap.get(num_key, 0) / tot, 1) if tot else 0.0

    _intra_fak_pcts  = [_pct(all_years_data[yr], "intra_pubs")      for yr in years_sorted]
    _inter_fak_pcts  = [_pct(all_years_data[yr], "inter_pubs")      for yr in years_sorted]
    _intra_inst_pcts = [_pct(all_years_data[yr], "intra_inst_pubs") for yr in years_sorted]
    _inter_inst_pcts = [_pct(all_years_data[yr], "inter_inst_pubs") for yr in years_sorted]
    _intra_grp_pcts  = [_pct(all_years_data[yr], "intra_grp_pubs")  for yr in years_sorted]
    _inter_grp_pcts  = [_pct(all_years_data[yr], "inter_grp_pubs")  for yr in years_sorted]

    # Dynamisk analytisk tekst (baseret på fakultetsniveau som bredeste mål)
    _first_yr, _last_yr   = years_sorted[0], years_sorted[-1]
    _first_pct, _last_pct = _intra_fak_pcts[0], _intra_fak_pcts[-1]
    _delta     = round(_last_pct - _first_pct, 1)
    _delta_str = f"+{_delta}" if _delta > 0 else str(_delta)
    _trend = (
        "en svagt stigende tendens, hvilket indikerer øget siloering i det viste sampubliceringsnetværk"
        if _delta > 2 else
        "en svagt faldende tendens, hvilket indikerer mere tværgående samarbejde"
        if _delta < -2 else
        "et bemærkelsesværdigt stabilt forhold uden entydige tegn på hverken øget integration eller siloering"
    )

    _inst_trend_str = ""
    if "I" in mode and mode not in ("FI",):
        _first_inst, _last_inst = _intra_inst_pcts[0], _intra_inst_pcts[-1]
        _delta_inst     = round(_last_inst - _first_inst, 1)
        _delta_inst_str = f"+{_delta_inst}" if _delta_inst > 0 else str(_delta_inst)
        _inst_trend = (
            "er intra-institutandelen ligeledes steget"
            if _delta_inst > 2 else
            "er intra-institutandelen faldet"
            if _delta_inst < -2 else
            "er intra-institutandelen ligeledes forblevet stabil"
        )
        _inst_trend_str = (
            f"På **institutniveau** {_inst_trend} fra **{_first_inst}%** til "
            f"**{_last_inst}%** ({_delta_inst_str} pp) i samme periode."
        )

    st.markdown(
f"""

Afsnittet nedenfor viser udviklingen i andelen af **intra‑** og **inter‑samarbejde**
over tid på det valgte organisatoriske niveau.

En stigende intra‑andel indikerer, at samarbejdet i stigende grad foregår
inden for organisatoriske enheder, mens en faldende intra‑andel peger på
mere tværgående samarbejde på tværs af KU.

Udviklingen skal læses som en strukturel indikator: ændringer kan skyldes
organisatoriske forhold, forskydninger i forskningspraksis eller ændrede
rammevilkår for samarbejde. Fra **{_first_yr}** til **{_last_yr}** har intra-fakultetandelen bevæget sig 
fra **{_first_pct}%** til **{_last_pct}%** ({_delta_str} procentpoint), hvilket peger på {_trend}.

{_inst_trend_str}
"""
    )

    _use_all_pairs = st.toggle(
        "Inkludér intra-enhed forfatterpar (f.eks. to professorer på samme institut)",
        value=False,
        key=f"silo_all_pairs_netudv_{year}_{mode}",
        help="Til: alle forfatterpar inkl. par inden for samme enhed. Fra: kun netværkskanter (par mellem noder).",
        disabled=not show_intra
    )
    if _use_all_pairs:
        st.caption("Alle forfatterpar inkl. par inden for samme enhed (f.eks. to mandlige professorer på Kemisk Institut).")
    else:
        st.caption("Kun forfatterpar *mellem* de valgte organisatoriske enheder – som vist i netværket.")

    _s = "_all" if _use_all_pairs else ""

    # Byg _plot_levels mode-drevet
    _plot_levels = []
    if "F" in mode:
        _plot_levels.append(("fac",
                             f"fac_intra_ew{_s}",  f"fac_inter_ew{_s}",
                             f"intra_pubs{_s}",     f"inter_pubs{_s}",
                             "#122947"))
    if "I" in mode:
        _plot_levels.append(("inst",
                             f"inst_intra_ew{_s}", f"inst_inter_ew{_s}",
                             f"intra_inst_pubs{_s}", f"inter_inst_pubs{_s}",
                             "#4a7ca8"))
    if "G" in mode:
        _plot_levels.append(("grp",
                             f"grp_intra_ew{_s}",  f"grp_inter_ew{_s}",
                             f"intra_grp_pubs{_s}", f"inter_grp_pubs{_s}",
                             "#39641c"))

    if not _plot_levels:
        st.caption("Vælg mindst ét organisatorisk niveau i sidepanelet for at se siloeringsudvikling.")
    else:
        def _render_silo_plot(filter_unit=None):
            _fig_silo = go.Figure()
            _tbl_rows = {yr: {"År": yr} for yr in years_sorted}
            _tbl_cols = ["År"]

            for unit_key, _intra_k, _inter_k, intra_abs_key, inter_abs_key, color in _plot_levels:
                if filter_unit is None:
                    _intra_abs = [int(all_years_data[yr].get(intra_abs_key, 0)) for yr in years_sorted]
                    _inter_abs = [int(all_years_data[yr].get(inter_abs_key, 0)) for yr in years_sorted]
                elif unit_key == _silo_level:
                    _intra_abs = [int(all_years_data[yr].get(_intra_k, {}).get(filter_unit, 0))
                                  for yr in years_sorted]
                    _inter_abs = [int(all_years_data[yr].get(_inter_k, {}).get(filter_unit, 0))
                                  for yr in years_sorted]
                else:
                    continue

                _totals  = [i + x or 1 for i, x in zip(_intra_abs, _inter_abs)]
                _y_intra = [round(100 * i / t, 1) for i, t in zip(_intra_abs, _totals)]
                _y_inter = [round(100 - v, 1)     for v in _y_intra]

                _level_label = {"fac": "fakultet", "inst": "institut", "grp": "stillingsgruppe"}[unit_key]

                _fig_silo.add_trace(go.Scatter(
                    x=years_sorted, y=_y_intra,
                    mode="lines+markers+text",
                    text=[f"{v}%" for v in _y_intra],
                    textposition="top center",
                    line=dict(color=color, width=2),
                    marker=dict(size=9),
                    name=f"Intra-{_level_label} (%)",
                ))
                _fig_silo.add_trace(go.Scatter(
                    x=years_sorted, y=_y_inter,
                    mode="lines+markers+text",
                    text=[f"{v}%" for v in _y_inter],
                    textposition="bottom center",
                    line=dict(color=color, width=2, dash="dot"),
                    marker=dict(size=9),
                    name=f"Inter-{_level_label} (%)",
                ))

                _intra_col = f"Intra-{_level_label} (forfatterpar)"
                _inter_col = f"Inter-{_level_label} (forfatterpar)"
                _pct_col   = f"Intra-{_level_label} (%)"
                _tbl_cols += [_intra_col, _inter_col, _pct_col]
                for i, yr in enumerate(years_sorted):
                    _tbl_rows[yr][_intra_col] = _intra_abs[i]
                    _tbl_rows[yr][_inter_col] = _inter_abs[i]
                    _tbl_rows[yr][_pct_col]   = _y_intra[i]

            _title_suffix = filter_unit or "KU samlet"
            _fig_silo.update_layout(
                yaxis_title="Andel af forfatterpar (%)",
                yaxis=dict(range=[0, 100]),
                xaxis=dict(tickmode="array", tickvals=years_sorted, dtick=1),
                legend=dict(orientation="v", yanchor="top", y=1, xanchor="left", x=1.02),
                height=400,
                margin=dict(t=50, b=100),
                title=f"Siloeringsudvikling{' (inkl. intra-enhed)' if _use_all_pairs else ''} – {_title_suffix}, {years_sorted[0]}–{years_sorted[-1]}",
            )
            _fu = filter_unit or "samlet"
            st.plotly_chart(
                _fig_silo, width="stretch",
                key=f"silo_netudv_{_fu}_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}},
            )
            if filter_caption:
                st.caption(filter_caption)

            _pivot_rows = []
            for col in _tbl_cols[1:]:
                _row = {"Metrik": col}
                for yr in years_sorted:
                    _row[str(yr)] = _tbl_rows[yr].get(col, 0)
                _pivot_rows.append(_row)
            _pivot_schema = (
                [("Metrik", pa.string())] +
                [(str(yr), pa.float64()) for yr in years_sorted]
            )
            with st.expander("Se tabel"):
                st.dataframe(build_table(_pivot_rows, _pivot_schema),
                             hide_index=True, width="stretch")
                st.download_button(
                    "Download (.xlsx)",
                    data=rows_to_excel_bytes(_pivot_rows, [n for n, _ in _pivot_schema]),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_silo_netudv_{_fu}_{year}_{mode}",
                )

        _render_silo_plot(filter_unit=_scope_unit_val)

    st.markdown("---")

    # ══════════════════════════════════════════════════════════════════════════
    # 2. FORNYELSE I SAMARBEJDSRELATIONERNE
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("#### Fornyelse i samarbejdsrelationerne")

    st.markdown(
f"""
Selv et netværk med mange samarbejder kan være enten dynamisk eller stivnet.
Et institut eller fakultet kan have mange forfatterpar år efter år uden, at
samarbejdet nødvendigvis udvides til nye relationer.

Figuren nedenfor opdeler hvert års forfatterpar i fire kategorier:

- **Fortsættende par** – par der også eksisterede året før. Den stabile kerne.
- **Genoptagne par** – par der har eksisteret tidligere, forsvandt i mindst ét år, og
  er tilbage nu. Viser at relationer ofte *pulser* snarere end forsvinder permanent.
- **Nye par** – par der aldrig før har publiceret sammen i datasættet.
- **Bortfaldne par** – par der eksisterede året før, men ikke i det aktuelle år.

Fortsættende, genoptagne og nye par udgør tilsammen årets samlede samarbejde,
mens bortfaldne par vises særskilt, da de hører til det foregående års netværk.

En høj andel **nye par** indikerer et netværk i udvidelse, hvor nye relationer
opstår. En høj andel **fortsættende par** peger på stabile, veletablerede
samarbejder, men kan også afspejle, at samarbejdet gentager sig inden for
faste kredse.

**Genoptagne par** viser, at samarbejdsrelationer ofte er pulserende snarere
end permanente, og at midlertidigt bortfald ikke nødvendigvis er et endeligt
ophør.

Det første år i tidsserien har ingen historik og fungerer derfor kun som
referencepunkt. Det andet år kan ikke indeholde genoptagne par, da der kun
findes ét tidligere observationsår.

Opgørelsen bygger på **forfatterpar mellem de organisatoriske enheder**, der
vises i netværket. Forfatterpar inden for samme organisatoriske enhed indgår
ikke i denne analyse.
""")

    # Byg kumulativ historik – filtreret efter scope-enhed hvis valgt
    _seen_before: set[str] = set()       # alle par set i år < idx
    _prev_year_pairs: set[str] = set()   # par set i år idx-1
    _novelty_rows = []

    for idx, yr in enumerate(years_sorted):
        _raw_pairs = {
            p for p, w in all_years_data[yr].get("top_pairs", {}).items() if w > 0
        }
        if _scope_unit_val is None:
            _this_pairs = _raw_pairs
        else:
            _this_pairs = {
                p for p in _raw_pairs
                if _pair_involves_unit(p, _scope_unit_val, _silo_level)
            }

        if idx == 0:
            # Reference-år: ingen historik
            _novelty_rows.append({
                "År": int(yr),
                "Fortsættende par": None,
                "Genoptagne par": None,
                "Nye par": None,
                "Bortfaldne par": None,
                "Total par i år": len(_this_pairs),
                "Andel nye (%)": None,
                "Andel fortsættende (%)": None,
            })
        else:
            _continuing = _this_pairs & _prev_year_pairs
            # Genoptagne = set før, men ikke i umiddelbart foregående år, tilbage nu
            _resumed    = (_this_pairs & _seen_before) - _prev_year_pairs
            _new        = _this_pairs - _seen_before
            _dropped    = _prev_year_pairs - _this_pairs
            _tot        = len(_this_pairs) or 1
            _novelty_rows.append({
                "År": int(yr),
                "Fortsættende par": len(_continuing),
                "Genoptagne par":   len(_resumed),
                "Nye par":          len(_new),
                "Bortfaldne par":   len(_dropped),
                "Total par i år":   len(_this_pairs),
                "Andel nye (%)":          round(100 * len(_new) / _tot, 1),
                "Andel fortsættende (%)": round(100 * len(_continuing) / _tot, 1),
            })
        _seen_before |= _this_pairs
        _prev_year_pairs = _this_pairs

    if _novelty_rows[0]["Total par i år"] == 0 and all(r["Total par i år"] == 0 for r in _novelty_rows):
        st.error(
            f"Ingen forfatterpar fundet for **{_scope_unit_val}** på "
            f"{_level_labels_single[_silo_level]}-niveau i det valgte udsnit."
        )
    else:
        _plot_years = [r["År"] for r in _novelty_rows]
        _cont_vals = [r["Fortsættende par"] if r["Fortsættende par"] is not None else 0 for r in _novelty_rows]
        _res_vals  = [r["Genoptagne par"]   if r["Genoptagne par"]   is not None else 0 for r in _novelty_rows]
        _new_vals  = [r["Nye par"]          if r["Nye par"]          is not None else 0 for r in _novelty_rows]
        _drop_vals = [r["Bortfaldne par"]   if r["Bortfaldne par"]   is not None else 0 for r in _novelty_rows]
        _first_total = _novelty_rows[0]["Total par i år"]

        _fig_nov = go.Figure()
        # Reference-år (første år) som grå søjle uden opdeling
        _fig_nov.add_trace(go.Bar(
            x=[_plot_years[0]],
            y=[_first_total],
            name=f"Alle par (reference)",
            marker=dict(color="#c9c9c9"),
            hovertemplate="År %{x}<br>Total par: %{y}<br><i>ingen historik</i><extra></extra>",
        ))
        # Stablet fra bund: fortsættende → genoptagne → nye
        _fig_nov.add_trace(go.Bar(
            x=_plot_years[1:],
            y=_cont_vals[1:],
            name="Fortsættende par",
            marker=dict(color="#122947"),
            hovertemplate="År %{x}<br>Fortsættende: %{y}<extra></extra>",
        ))
        _fig_nov.add_trace(go.Bar(
            x=_plot_years[1:],
            y=_res_vals[1:],
            name="Genoptagne par",
            marker=dict(color="#6b8caf"),
            hovertemplate="År %{x}<br>Genoptagne: %{y}<extra></extra>",
        ))
        _fig_nov.add_trace(go.Bar(
            x=_plot_years[1:],
            y=_new_vals[1:],
            name="Nye par",
            marker=dict(color="#901a1E"),
            hovertemplate="År %{x}<br>Nye: %{y}<extra></extra>",
        ))
        # Bortfaldne som stiplet linje på sekundær akse
        _fig_nov.add_trace(go.Scatter(
            x=_plot_years[1:],
            y=_drop_vals[1:],
            name="Bortfaldne par (ift. året før)",
            mode="lines+markers",
            line=dict(color="#3d3d3d", width=2, dash="dot"),
            marker=dict(size=8, symbol="x"),
            yaxis="y2",
            hovertemplate="År %{x}<br>Bortfaldne: %{y}<extra></extra>",
        ))
        _nov_title_suffix = _scope_unit_val or "KU samlet"
        _fig_nov.update_layout(
            barmode="stack",
            xaxis=dict(tickmode="array", tickvals=_plot_years, dtick=1, title="År"),
            yaxis=dict(title="Antal forfatterpar"),
            yaxis2=dict(title="Bortfaldne par", overlaying="y", side="right", showgrid=False),
            height=440,
            margin=dict(t=60, r=80),
            legend=dict(orientation="h", yanchor="top", y=-0.2, xanchor="left", x=0),
            title=f"Fortsættende, genoptagne, nye og bortfaldne forfatterpar – {_nov_title_suffix}, {_plot_years[0]}–{_plot_years[-1]}",
        )
        st.plotly_chart(
            _fig_nov, width="stretch",
            key=f"novelty_{_scope_unit or 'samlet'}_{year}_{mode}",
            config={"toImageButtonOptions": {"format": "png", "scale": 3}},
        )
        if filter_caption:
            st.caption(filter_caption)

        # Dynamisk fortolkningstekst
        _valid_rows = [r for r in _novelty_rows if r["Andel nye (%)"] is not None]
        if len(_valid_rows) >= 2:
            _last      = _valid_rows[-1]
            _first_cmp = _valid_rows[0]
            _last_pct_nov  = _last["Andel nye (%)"]
            _first_pct_nov = _first_cmp["Andel nye (%)"]
            _delta_nov     = round(_last_pct_nov - _first_pct_nov, 1)

            if _delta_nov > 5:
                _nov_trend = (
                    f"en **stigende fornyelse**; andelen af nye par er vokset fra "
                    f"**{_first_pct_nov}%** i {_first_cmp['År']} til **{_last_pct_nov}%** i "
                    f"{_last['År']} (+{_delta_nov} pp). Det tyder på et netværk, der "
                    f"fortsat udvider sig til nye konstellationer."
                )
            elif _delta_nov < -5:
                _nov_trend = (
                    f"en **aftagende fornyelse**; andelen af nye par er faldet fra "
                    f"**{_first_pct_nov}%** i {_first_cmp['År']} til **{_last_pct_nov}%** i "
                    f"{_last['År']} ({_delta_nov} pp). Det kan indikere, at samarbejdet "
                    f"i stigende grad foregår i faste, etablerede kredse."
                )
            else:
                _nov_trend = (
                    f"en **stabil fornyelsesrate**; andelen af nye par ligger på "
                    f"**{_last_pct_nov}%** i {_last['År']} mod **{_first_pct_nov}%** i "
                    f"{_first_cmp['År']} ({'+' if _delta_nov >= 0 else ''}{_delta_nov} pp)."
                )

            _last_new  = _last["Nye par"]
            _last_drop = _last["Bortfaldne par"]
            if _last_new > _last_drop:
                _net_txt = (
                    f"I {_last['År']} kom der **{_last_new - _last_drop}** flere nye par "
                    f"til, end der bortfaldt – netværket er i nettovækst."
                )
            elif _last_drop > _last_new:
                _net_txt = (
                    f"I {_last['År']} bortfaldt **{_last_drop - _last_new}** flere par, end "
                    f"der kom nye til – netværket trækker sig sammen."
                )
            else:
                _net_txt = f"I {_last['År']} balancerer nye og bortfaldne par nogenlunde."

            # Tilføj comeback-indsigt hvis genoptagne er betydelige
            _last_resumed = _last["Genoptagne par"] or 0
            _last_total   = _last["Total par i år"] or 1
            _resumed_pct  = round(100 * _last_resumed / _last_total, 1)
            if _last_resumed > 0 and _resumed_pct >= 5:
                _comeback_txt = (
                    f" Bemærk også, at **{_last_resumed}** par ({_resumed_pct}% af årets total) "
                    f"er *genoptagne* – relationer der tidligere har været sat på pause og nu "
                    f"er tilbage."
                )
            else:
                _comeback_txt = ""

            st.markdown(f"Analysen viser {_nov_trend}\n\n{_net_txt}{_comeback_txt}")
        elif len(_valid_rows) == 1:
            st.caption("Kun to år i udvalget – tilføj flere år for at se udvikling over tid.")

        # Tabel
        _nov_years = [r["År"] for r in _novelty_rows]
        _nov_metrics = [
            ("Fortsættende par",       pa.int64()),
            ("Genoptagne par",         pa.int64()),
            ("Nye par",                pa.int64()),
            ("Bortfaldne par",         pa.int64()),
            ("Total par i år",         pa.int64()),
            ("Andel nye (%)",          pa.float64()),
            ("Andel fortsættende (%)", pa.float64()),
        ]
        _nov_pivot = [
            {"Metrik": metrik,
             **{str(r["År"]): (r[metrik] if r[metrik] is not None else 0)
                for r in _novelty_rows}}
            for metrik, _ in _nov_metrics
        ]
        _nov_schema = (
            [("Metrik", pa.string())] +
            [(str(yr), dtype) for yr in _nov_years
             for metrik, dtype in _nov_metrics[:1]]  # brug int64 til år-kolonner
        )
        _nov_schema = [("Metrik", pa.string())] + [(str(yr), pa.float64()) for yr in _nov_years]
        with st.expander("Se tabel"):
            st.dataframe(
                build_table(_nov_pivot, _nov_schema),
                hide_index=True, width="stretch",
            )
            st.caption(
                f"Bemærk: {_plot_years[0]} er referenceår og har ingen historik – "
                "alle fire kategorier er derfor angivet som 0. "
                f"{_plot_years[1] if len(_plot_years) > 1 else ''} kan ikke have "
                "genoptagne par, fordi der kun findes ét tidligere år at sammenligne med."
            )
            st.download_button(
                "Download (.xlsx)",
                data=rows_to_excel_bytes(_nov_pivot, [n for n, _ in _nov_schema]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_novelty_{_scope_unit or 'samlet'}_{year}_{mode}",
            )

    st.markdown("---")

    # ══════════════════════════════════════════════════════════════════════════
    # 3. UDVIKLING I SPECIFIKKE SAMARBEJDER (NY SEKTION)
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("#### Udvikling i konkrete samarbejdsrelationer")

    if _scope_unit_val is None:
        st.markdown(
f"""
I dette afsnit kan udviklingen i samarbejdsrelationer følges mere detaljeret.
Vælg en specifik enhed øverst i fanen for at se, hvordan dens samarbejde med
andre {_level_labels_single[_silo_level]}er på det valgte niveau udvikler sig over tid.

Ved valg af **Samlet KU** vises i stedet de største samarbejdsrelationer på
tværs af alle enheder, hvilket giver et overblik over de mest betydende
relationer i netværket som helhed.

Figuren viser udviklingen i antallet af forfatterpar mellem den valgte enhed
og hver af dens samarbejdspartnere. Hver linje repræsenterer én modpart.

En stigende linje indikerer vækst i samarbejdet, mens en faldende linje peger
på aftagende sampublicering mellem de pågældende enheder.
"""
        )

        # Samlet KU: vis top-N par på tværs af alle niveauet
        _top_n_cross = st.number_input(
            "Antal top-par at vise",
            min_value=3, max_value=20, value=8,
            key=f"cross_topn_samlet_{year}_{mode}",
        )

        # Aggreger vægt per fac-fac (eller inst-inst / grp-grp) par over alle år
        _cross_totals: dict[tuple, float] = {}
        _cross_by_year: dict[tuple, dict] = {}
        for yr in years_sorted:
            for p, w in all_years_data[yr].get("top_pairs", {}).items():
                if w <= 0:
                    continue
                _sides = p.split(" ↔ ")
                if len(_sides) != 2:
                    continue
                _a = _unit_at_level(_sides[0], _silo_level)
                _b = _unit_at_level(_sides[1], _silo_level)
                if not _a or not _b or _a == _b:
                    continue
                _key = tuple(sorted([_a, _b]))
                _cross_totals[_key] = _cross_totals.get(_key, 0.0) + w
                _cross_by_year.setdefault(_key, {})[yr] = _cross_by_year.get(_key, {}).get(yr, 0.0) + w

        if not _cross_totals:
            st.caption(f"Ingen samarbejder mellem {_level_labels_single[_silo_level]}er i det valgte udsnit.")
        else:
            _top_keys = sorted(_cross_totals.items(), key=lambda x: -x[1])[:int(_top_n_cross)]

            _fig_cross = go.Figure()
            for _idx, ((_a, _b), _tot_w) in enumerate(_top_keys):
                _y = [_cross_by_year.get((_a, _b), {}).get(yr, 0) for yr in years_sorted]
                _color = _color_for_unit(_a, _silo_level, palette_idx=_idx)

                _fig_cross.add_trace(go.Scatter(
                    x=years_sorted, y=_y,
                    mode="lines+markers",
                    name=f"{_a} ↔ {_b}",
                    line=dict(color=_color, width=2),
                    marker=dict(size=7, color=_color),
                    hovertemplate=f"{_a} ↔ {_b}<br>År %{{x}}<br>Forfatterpar: %{{y:.0f}}<extra></extra>",
                ))
            _fig_cross.update_layout(
                xaxis=dict(tickmode="array", tickvals=years_sorted, dtick=1, title="År"),
                yaxis=dict(title="Antal forfatterpar"),
                height=440,
                margin=dict(t=60, b=60),
                legend=dict(orientation="v", yanchor="top", y=1, xanchor="left", x=1.02),
                title=f"Top {len(_top_keys)} samarbejder mellem {_level_labels_single[_silo_level]}er, {years_sorted[0]}–{years_sorted[-1]}",
            )
            st.plotly_chart(
                _fig_cross, width="stretch",
                key=f"cross_samlet_{_silo_level}_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}},
            )
            if filter_caption:
                st.caption(filter_caption)

            _cross_rows = []
            for (_a, _b), _tot_w in _top_keys:
                _row = {"Samarbejde": f"{_a} ↔ {_b}", "Total": round(_tot_w, 1)}
                for yr in years_sorted:
                    _row[str(yr)] = round(_cross_by_year.get((_a, _b), {}).get(yr, 0), 1)
                _cross_rows.append(_row)
            _cross_schema = (
                [("Samarbejde", pa.string()), ("Total", pa.float64())] +
                [(str(yr), pa.float64()) for yr in years_sorted]
            )
            with st.expander("Se tabel"):
                st.dataframe(build_table(_cross_rows, _cross_schema),
                             hide_index=True, width="stretch")
                st.download_button(
                    "Download (.xlsx)",
                    data=rows_to_excel_bytes(_cross_rows, [n for n, _ in _cross_schema]),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_cross_samlet_{_silo_level}_{year}_{mode}",
                )

    else:
        # Enhedsfokus: vis _scope_unit_val's samarbejder med hver modpart over tid
        st.markdown(
f"""
Figuren viser, hvordan **{_scope_unit_val}**'s forfatterpar med hver af de andre
{_level_labels_single[_silo_level]}er har udviklet sig over tid. Hver linje er én
modpart. En faldende linje betyder, at samarbejdet med den pågældende modpart aftager;
en stigende linje betyder vækst.
"""
        )

        # Aggreger vægt per modpart per år
        _partner_by_year: dict[str, dict] = {}
        _partner_totals: dict[str, float] = {}
        for yr in years_sorted:
            for p, w in all_years_data[yr].get("top_pairs", {}).items():
                if w <= 0:
                    continue
                _other = _pair_other_side(p, _scope_unit_val, _silo_level)
                if not _other:
                    continue
                _partner_by_year.setdefault(_other, {})[yr] = _partner_by_year.get(_other, {}).get(yr, 0.0) + w
                _partner_totals[_other] = _partner_totals.get(_other, 0.0) + w

        if not _partner_totals:
            st.error(
                f"**{_scope_unit_val}** har ingen forfatterpar med andre "
                f"{_level_labels_single[_silo_level]}er i det valgte udsnit."
            )
        else:
            # Sorter modparter efter samlet vægt – mest samarbejdende øverst
            _partners_sorted = sorted(_partner_totals.items(), key=lambda x: -x[1])

            _fig_part = go.Figure()
            for _idx, (_other, _tot_w) in enumerate(_partners_sorted):
                _y = [_partner_by_year.get(_other, {}).get(yr, 0) for yr in years_sorted]
                _color = _color_for_unit(_other, _silo_level, palette_idx=_idx)

                _fig_part.add_trace(go.Scatter(
                    x=years_sorted, y=_y,
                    mode="lines+markers",
                    name=_other,
                    line=dict(color=_color, width=2),
                    marker=dict(size=7, color=_color),
                    hovertemplate=f"{_scope_unit_val} ↔ {_other}<br>År %{{x}}<br>Forfatterpar: %{{y:.0f}}<extra></extra>",
                ))
            _fig_part.update_layout(
                xaxis=dict(tickmode="array", tickvals=years_sorted, dtick=1, title="År"),
                yaxis=dict(title="Antal forfatterpar"),
                height=440,
                margin=dict(t=60, b=60),
                legend=dict(orientation="v", yanchor="top", y=1, xanchor="left", x=1.02),
                title=f"{_scope_unit_val}s samarbejder med andre {_level_labels_single[_silo_level]}er, {years_sorted[0]}–{years_sorted[-1]}",
            )
            st.plotly_chart(
                _fig_part, width="stretch",
                key=f"cross_unit_{_scope_unit_val}_{_silo_level}_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}},
            )
            if filter_caption:
                st.caption(filter_caption)

            # Dynamisk fortolkning: find største vokser og største faldende
            if len(years_sorted) >= 2:
                _first_yr_cr = years_sorted[0]
                _last_yr_cr = years_sorted[-1]
                _changes = []
                for _other in _partner_totals:
                    _v_first = _partner_by_year.get(_other, {}).get(_first_yr_cr, 0)
                    _v_last  = _partner_by_year.get(_other, {}).get(_last_yr_cr,  0)
                    _changes.append((_other, _v_first, _v_last, _v_last - _v_first))
                _growers = sorted([c for c in _changes if c[3] > 0], key=lambda x: -x[3])
                _decliners = sorted([c for c in _changes if c[3] < 0], key=lambda x: x[3])

                _msg_parts = []
                if _growers:
                    _g = _growers[0]
                    _msg_parts.append(
                        f"Det stærkest **voksende** samarbejde er med **{_g[0]}** "
                        f"(fra {_g[1]:.0f} til {_g[2]:.0f} forfatterpar, +{_g[3]:.0f})."
                    )
                if _decliners:
                    _d = _decliners[0]
                    _msg_parts.append(
                        f"Det stærkest **aftagende** samarbejde er med **{_d[0]}** "
                        f"(fra {_d[1]:.0f} til {_d[2]:.0f} forfatterpar, {_d[3]:.0f})."
                    )
                if _msg_parts:
                    st.markdown(" ".join(_msg_parts))

            # Tabel
            _part_rows = []
            for _other, _tot_w in _partners_sorted:
                _row = {"Modpart": _other, "Total": round(_tot_w, 1)}
                for yr in years_sorted:
                    _row[str(yr)] = round(_partner_by_year.get(_other, {}).get(yr, 0), 1)
                _part_rows.append(_row)
            _part_schema = (
                [("Modpart", pa.string()), ("Total", pa.float64())] +
                [(str(yr), pa.float64()) for yr in years_sorted]
            )
            with st.expander("Se tabel"):
                st.dataframe(build_table(_part_rows, _part_schema),
                             hide_index=True, width="stretch")
                st.download_button(
                    "Download (.xlsx)",
                    data=rows_to_excel_bytes(_part_rows, [n for n, _ in _part_schema]),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_cross_unit_{_scope_unit_val}_{_silo_level}_{year}_{mode}",
                )

    st.markdown("---")

    # ══════════════════════════════════════════════════════════════════════════
    # 4. ER DE SAMME ENHEDER ALTID ISOLEREDE?
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("#### Er de samme enheder altid isolerede?")

    _iso_scope_phrase = (
        "Alle isolerede enheder i det valgte udsnit vises nedenfor."
        if _scope_unit_val is None
        else f"Kun enheder under **{_scope_unit_val}** vises nedenfor."
    )
    st.markdown(
f"""
Afsnittet nedenfor viser, hvilke enheder der gentagne gange er **isolerede**,
dvs. uden sampublicering på tværs af KU i det valgte udsnit.

Enheder vises kun, hvis de har været isolerede i mindst to år, for at adskille
systematiske mønstre fra midlertidige udsving eller dataudfald.

Isolation betyder her fravær af sampublicering på tværs af organisatoriske
enheder – ikke fravær af publikationer generelt.

{_iso_scope_phrase}
"""
    )

    _all_isolated: dict[str, list] = {}
    for yr in years_sorted:
        for unit in all_years_data[yr].get("isolated_units", []):
            if _scope_unit_val is not None:
                if _unit_at_level(unit, _silo_level) != _scope_unit_val:
                    continue
            _all_isolated.setdefault(unit, []).append(yr)

    _persistent = {unit: yrs for unit, yrs in _all_isolated.items() if len(yrs) >= 2}

    if not _persistent:
        if _scope_unit_val is None:
            st.error("Ingen enheder har været isolerede i to eller flere år.")
        else:
            st.error(
                f"Ingen enhed under **{_scope_unit_val}** har været isoleret i to "
                f"eller flere år."
            )
        return

    _iso_sorted = sorted(_persistent.items(), key=lambda x: -len(x[1]))
    _iso_tbl_rows = [
        {
            "Enhed":               unit,
            "Antal isolerede år":  int(len(yrs)),
            "Isoleret i år":       ", ".join(str(y) for y in sorted(yrs)),
            "Konsekvent isoleret": "Ja" if len(yrs) == len(years_sorted) else "Nej",
        }
        for unit, yrs in _iso_sorted
    ]
    _iso_schema = [
        ("Enhed",                pa.string()),
        ("Antal isolerede år",   pa.int64()),
        ("Isoleret i år",        pa.string()),
        ("Konsekvent isoleret",  pa.string()),
    ]

    _iso_units = [r["Enhed"] for r in _iso_tbl_rows]
    _iso_z = [
        [1 if yr in _persistent[unit] else 0 for yr in years_sorted]
        for unit in _iso_units
    ]
    _fig_iso = go.Figure(go.Heatmap(
        z=_iso_z,
        x=[str(yr) for yr in years_sorted],
        y=_iso_units,
        colorscale=[[0, "#f0f4f8"], [1, "#7A131A"]],
        zmin=0, zmax=1,
        text=[["Isoleret" if v else "" for v in row] for row in _iso_z],
        texttemplate="%{text}",
        showscale=False,
        hovertemplate="%{y} – %{x}: %{text}<extra></extra>",
    ))
    _iso_title_suffix = f" – {_scope_unit_val}" if _scope_unit_val else ""
    _fig_iso.update_layout(
        xaxis_title="År",
        height=max(300, 28 * len(_iso_units)),
        margin=dict(l=220, t=50, r=20),
        title=f"Isolerede enheder over tid{_iso_title_suffix}",
    )
    st.plotly_chart(
        _fig_iso, width="stretch",
        key=f"iso_netudv_{_scope_unit or 'samlet'}_{year}_{mode}",
        config={"toImageButtonOptions": {"format": "png", "scale": 3}},
    )
    if filter_caption:
            st.caption(filter_caption)

    with st.expander("Se tabel"):
        st.dataframe(build_table(_iso_tbl_rows, _iso_schema), hide_index=True, width="stretch")
        st.download_button(
            "Download (.xlsx)",
            data=rows_to_excel_bytes(_iso_tbl_rows, [n for n, _ in _iso_schema]),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_iso_netudv_{_scope_unit or 'samlet'}_{year}_{mode}",
        )


def render_tab_datagrundlag(year, mode, all_groups, selected_facs, selected_insts, selected_grps,
                             forfatterantal=None, publikationstyper=None,
                             faculty_base_colors=None, years_sorted=None, pubtype_map=None,
                             all_years_data=None, forfatterantal_dist = None):
    st.subheader("Datagrundlag")

    SUND_LEKTOR_HR = {2021: 331, 2022: 328, 2023: 330, 2024: 334, 2025: 353}

    _sund_lektor_hr  = SUND_LEKTOR_HR.get(year)
    _sund_lektor_app = None
    if all_years_data and year in all_years_data:
        _sund_lektor_app = all_years_data[year].get("sund_lektor_size")
    
    _sund_txt = ""
    if _sund_lektor_hr and _sund_lektor_app:
        _diff_pct = round(100 * (_sund_lektor_app - _sund_lektor_hr) / _sund_lektor_hr, 1)
        _diff_str = f"+{_diff_pct}" if _diff_pct > 0 else str(_diff_pct)
        _sund_txt = (
            f"Som et grelt eksempel er der for lektorer på SUND i {year} en afvigelse på "
            f"**{_diff_str}%** mellem unikke forfattere og antal årsværk fra HR-data."
        )
    
    #st.write(f"DEBUG: year={year}, hr={_sund_lektor_hr}, app={_sund_lektor_app}")
    #st.write(f"DEBUG: all_years_data keys={list((all_years_data or {}).keys())}")
    #if all_years_data and year in all_years_data:
        #st.write(f"DEBUG: snapshot keys={list(all_years_data[year].keys())[:10]}")

    st.markdown(
f"""Netværk og de tilhørende opgørelser bygger på følgende datagrundlag:
- CURIS-publikationer for 2021-2025
    - Særligt 2025-data kan være ufuldstændige, da registrering 
    kan ske med forsinkelse
- VIP-forfattere matchet via HR-data
    - Forfattere identificeres via deres external_id i CURIS og matches til HR-data for det
    pågældende udgivelsesår. Kun forfattere, der er **aktivt ansatte** i HR-data i udgivelsesåret
    medtages. 
    - Hvis en forsker har **flere stillinger** i samme år (f.eks. både lektor og professor ved
    forskellige institutter), tildeles vedkommende den **højst rangerende stilling** med tilhørende
    institut og fakultet. Se rangering i tabellen nedenfor.
    - Antal unikke **forfattere** er altså ikke det samme som antal ansatte, da forfattere opgør forskere med
    mindst én CURIS-publikation i pågældende år.

{_sund_txt}

---

#### Netværk og forfatterpar

Netværket viser (indtil videre) kun **interne KU-sampubliceringer** - samarbejde med forfattere 
fra eksterne institutioner indgår ikke. Et forfatterpar opstår, når to organisatoriske enheder
har mindst én fælles publikation med to eller flere KU-VIP-forfattere. 

Kun forfattere, der kan knyttes til en gyldet KU-enhed vis HR-data, indgår i netværket. Forfattere
uden gyldigt external_id i CURIS tælles ikke i forfattere eller forfatterpar. 

---

#### Stillingsgrupper 

Stillingsgrupper er bestemt af HR-data på udgivelsestidspunktet - ikke af affilieringen i CURIS. 

| Rang | Stillingsgruppe | Eksempler |
|---|---|---|
| 1 | Professor | *Professor, Professor MSO, Klinisk Professor* |
| 2 | Lektor | *Lektor, Seniorforsker* |
| 3 | Adjunkt | *Adjunkt, Tenure Track Adjunkt, Forsker* |
| 4 | Postdoc | *Postdoc* |
| 5 | Ph.d. | *Ph.d.-stipendiat, Ph.d. Studerende* |
| 6 | Stillinger u. adjunktniveau | *Videnskabelig assistent, Ekstern Lektor* |
| 7 | Øvrige VIP (DVIP) | *Gæstelærer, Underviser, International Researcher* |
| 8 | Særlig stilling | *Studieadjunkt* |

Stillingstyper som *TAP, DTAP, pensionister, læger* og *tandlæger* uden VIP-status indgår ikke i 
analysen. Bemærk, at *kliniske lektorer* heller ikke er inkluderede.

---

#### Frasortering og afgrænsning

Analysen omfatter kun publikationer med mindst én KU-affilieret forfatter. Generelt indgår følgende
publikationer ikke i opgørelserne:

- publikationer uden entydig organisationstilknytning,
- publikationer, der er registrerede uden forfattere,
- publikationer, hvor der kun er registreret én KU-forfatter.
""")

    yr_str   = str(year)
    fac_order = [f for f in FAC_ORDER if f in (forfatterantal or {}).get(yr_str, {})]
    _fa       = (forfatterantal or {}).get(yr_str, {})
    _fac_colors = [faculty_base_colors.get(f, "#122947") if faculty_base_colors else "#122947" for f in fac_order]

    # ── Forfatterantal ────────────────────────────────────────────────
    one_counts  = [_fa.get(f, {}).get("1",   0) for f in fac_order]
    many_counts = [_fa.get(f, {}).get("gt1", 0) for f in fac_order]

    ratio_per_fac = {f: one / (one + many) * 100
                     for f, one, many in zip(fac_order, one_counts, many_counts)
                     if (one + many) > 0}

    if ratio_per_fac:
        min_fac = min(ratio_per_fac, key=ratio_per_fac.get)
        max_fac = max(ratio_per_fac, key=ratio_per_fac.get)
        min_val, max_val = ratio_per_fac[min_fac], ratio_per_fac[max_fac]
        abs_one = _fa.get(max_fac, {}).get("1", 0)
    else:
        min_fac = max_fac = None
        min_val = max_val = abs_one = 0

    st.markdown("---")

    st.markdown(
f"""
#### Forfatterantal per publikation

Publikationspraksis varierer betydeligt på tværs af fakulteterne, bl.a.
i forhold til omfanget af co-forfatterskaber og udbredelsen af soloartikler. Det påvirker
både antallet af KU-forfattere per publikation - og dermed antallet af mulige forfatterpar. 

Andelen af publikationer med én KU-forfatter varierer fra **{fmt_ui(min_val,1)}%** ({min_fac}) til
**{fmt_ui(max_val,1)}%** ({max_fac}), svarende til **{abs_one}** frasorterede publikationer
(soloartikler indgår ikke i netværksanalyserne, da de per definition ikke har forfatterpar).
""")
    
    _tab_abs_fa, _tab_pct_fa, _tab_tid_fa = st.tabs(["Antal", "Andel (%)", "Udvikling over tid"])

    with _tab_abs_fa:
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=fac_order, y=one_counts, name="1 KU-forfatter",
            marker_color=[adjust_color(c, 2.2, 0.3) for c in _fac_colors],
            text=[str(v) for v in one_counts], textposition="inside", textfont_color="white",
        ))
        fig.add_trace(go.Bar(
            x=fac_order, y=many_counts, name=">1 KU-forfatter",
            marker_color=_fac_colors,
            text=[str(v) for v in many_counts], textposition="inside", textfont_color="white",
        ))
        fig.update_layout(barmode="stack", yaxis_title="Antal publikationer",
                          legend_title="KU-forfattere", height=380, margin=dict(t=50),
                          title=f"Publikationer per fakultet fordelt på antal KU-forfattere, {year}")
        st.plotly_chart(fig, width="stretch", key=f"dg_forfatter_abs_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
        )

    with _tab_pct_fa:
        _pct_one  = [ratio_per_fac.get(f, 0) for f in fac_order]
        _pct_many = [round(100 - v, 1) for v in _pct_one]
        fig_pct = go.Figure()
        fig_pct.add_trace(go.Bar(
            x=fac_order, y=_pct_one, name="1 KU-forfatter",
            marker_color=[adjust_color(c, 2.2, 0.3) for c in _fac_colors],
            text=[f"{fmt_ui(v,1)}%" for v in _pct_one], textposition="inside", textfont_color="white",
        ))
        fig_pct.add_trace(go.Bar(
            x=fac_order, y=_pct_many, name=">1 KU-forfatter",
            marker_color=_fac_colors,
            text=[f"{fmt_ui(v,1)}%" for v in _pct_many], textposition="inside", textfont_color="white",
        ))
        fig_pct.update_layout(barmode="stack", yaxis=dict(title="Andel (%)", range=[0, 100]),
                               legend_title="KU-forfattere", height=380, margin=dict(t=50),
                               title=f"Andelen af publikationer fordelt på antal KU-forfattere, {year}")
        st.plotly_chart(fig_pct, width="stretch", key=f"dg_forfatter_pct_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
    )

    with _tab_tid_fa:
        if forfatterantal and years_sorted:
            fig_tid = go.Figure()
            for f, col in zip(fac_order, _fac_colors):
                _y_vals = []
                for yr in years_sorted:
                    _d = forfatterantal.get(str(yr), {}).get(f, {})
                    _one, _many = _d.get("1", 0), _d.get("gt1", 0)
                    _tot = _one + _many
                    _y_vals.append(round(_one / _tot * 100, 1) if _tot else 0)
                fig_tid.add_trace(go.Scatter(
                    x=years_sorted, y=_y_vals, name=f,
                    mode="lines+markers+text",
                    text=[f"{v:.1f}%" for v in _y_vals],
                    textposition="top center",
                    line=dict(color=col, width=2), marker=dict(size=7),
                ))
            fig_tid.update_layout(
                yaxis=dict(title="Andel med én KU-forfatter (%)", range=[0, 100]),
                xaxis=dict(tickmode="array", tickvals=years_sorted, dtick=1),
                height=420, margin=dict(t=50, b=120),
                legend_title="Fakulteter",
                title=f"Udvikling i andelen af publikationer med én KU-forfatter, {years_sorted[0]}-{years_sorted[-1]}"
            )
            st.plotly_chart(fig_tid, width="stretch", key=f"dg_forfatter_tid_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
    )
        else:
            st.error("Tidsdata ikke tilgængelig.")
    
    _dist = (forfatterantal_dist or {}).get(yr_str, {})

    if _dist:
        st.markdown(
f"""
Hvor mange KU-VIP-forfattere står der typisk på en publikation? Det er denne metrik, der direkte påvirker netværksanalysen:
en publikation med fire KU-VIP-forfattere genererer seks forfatterpar for én artikel, mens en soloartikel ikke genererer nogen. 

- **Median** viser den typiske publikation. En median på 1 betyder, at over halvdelen af publikationerne har én KU-VIP-forfatter
(resten er KU-interne samarbejder).
- **Gennemsnit** trækkes op af enkelte artikler med mange KU-VIP-forfattere. En stor forskel mellem median og gennemsnit betyder, 
at der findes en mindre gruppe artikler med markant flere forfattere end det typiske niveau - tilstrækkelig til at trække
gennemsnittet op uden at flytte medianen.
- **P95** viser den 95. percentil: 95% af publikationerne har færre KU-VIP-forfattere end dette tal. P95 er mere robust 
end max, fordi den ikke påvirkes af enkelte ekstreme outliers.
- **Max** er det højeste antal KU-VIP-forfattere på en enkelt publikation i {year}.
"""
        )

        _dist_facs = [f for f in fac_order if f in _dist]
        _medians = [_dist[f].get("median", 0) for f in _dist_facs]
        _means   = [_dist[f].get("mean",   0) for f in _dist_facs]
        _p95s    = [_dist[f].get("p95",    0) for f in _dist_facs]
        _maxes   = [_dist[f].get("max",    0) for f in _dist_facs]

        _fac_colors_dist = [faculty_base_colors.get(f, "#122947") for f in _dist_facs]

        fig_dist = go.Figure()
        # Median: lyseste (højest lightness)
        fig_dist.add_trace(go.Bar(
            x=_dist_facs, y=_medians, name="Median",
            marker_color=[adjust_color(c, 2.4, 0.4) for c in _fac_colors_dist],
            text=[fmt_ui(v, 1) for v in _medians], textposition="inside",
        ))
        # Gennemsnit: lys
        fig_dist.add_trace(go.Bar(
            x=_dist_facs, y=_means, name="Gennemsnit",
            marker_color=[adjust_color(c, 1.7, 0.7) for c in _fac_colors_dist],
            text=[fmt_ui(v, 1) for v in _means], textposition="inside",
        ))
        # P95: basisfarve
        fig_dist.add_trace(go.Bar(
            x=_dist_facs, y=_p95s, name="P95",
            marker_color=_fac_colors_dist,
            text=[fmt_ui(v, 0) for v in _p95s], textposition="inside",
        ))
        # Max: mørkeste
        fig_dist.add_trace(go.Bar(
            x=_dist_facs, y=_maxes, name="Max",
            marker_color=[adjust_color(c, 0.5, 1.0) for c in _fac_colors_dist],
            text=[str(v) for v in _maxes], textposition="inside",
        ))
        fig_dist.update_layout(
            barmode="group",
            yaxis=dict(
                title="KU-VIPs per publikation (log-skala)", 
                type="log", 
                tickmode="array",
                tickvals=[1, 2, 5, 10, 20],
                ticktext=["1", "2", "5","10","20"],
                minor=dict(ticks="")),
            xaxis=dict(title = ""),
            height=max(400, 80 * len(_dist_facs)),
            margin=dict(t=60, l=100, r=60),
            legend_title="Mål",
            title=f"KU-VIPs per publikation per fakultet, {year}",
        )
        st.plotly_chart(fig_dist, width="stretch", key=f"dg_vip_dist_{year}_{mode}",
            config={"toImageButtonOptions": {"format": "png", "scale": 3}})

        st.markdown(
            "Y-aksen i figuren ovenfor er logaritmisk, så både små medianer og store max-værdier kan ses i samme graf. "
            "Bemærk, at en max-værdi på 18 betyder, at der findes mindst én publikation i pågældende år med 18 KU-VIP-forfattere - ".format(year)
            + "den genererer alene 153 forfatterpar, hvilket kan give udsving i netværksanalyserne."
        )

        # Tabel
        _dist_rows = []
        for f in _dist_facs:
            d = _dist[f]
            _dist_rows.append({
                "Fakultet":      f,
                "Publikationer": d.get("n_pubs", 0),
                "Median":        round(d.get("median", 0), 1),
                "Gennemsnit":    round(d.get("mean",   0), 1),
                "P95":           round(d.get("p95",    0), 1),
                "Max":           int(d.get("max",      0)),
            })
        _dist_schema = [
            ("Fakultet",      pa.string()),
            ("Publikationer", pa.int64()),
            ("Median",        pa.float64()),
            ("Gennemsnit",    pa.float64()),
            ("P95",           pa.float64()),
            ("Max",           pa.int64()),
        ]
        with st.expander("Se tabel"):
            st.dataframe(build_table(_dist_rows, _dist_schema), hide_index=True, width="stretch")
            st.download_button(
                "Download (.xlsx)",
                data=rows_to_excel_bytes(_dist_rows, [n for n, _ in _dist_schema]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_dg_vip_dist_{year}_{mode}",
            )


    # ── Publikationstyper ─────────────────────────────────────────────
    _pt          = (publikationstyper or {}).get(yr_str, {})
    fac_order_pt  = [f for f in FAC_ORDER if f in _pt]
    all_pub_types = sorted(
        {t for fd in _pt.values() for t in fd},
        key=lambda t: -sum(_pt.get(f, {}).get(t, 0) for f in fac_order_pt)
    )
    _pt_colors    = ku_color_sequence(len(all_pub_types))

    st.markdown("---")

    st.markdown(
f"""#### Publikationstyper per fakultet
Publikationstyperne viser tilsvarende forskelle mellem fakulteterne. For at få et mere
sammenligneligt billede er CURIS-publikationsformaterne lagt sammen i kollapsede kategorier, 
f.eks. samles konferenceformater i én fælles kategori. Tallene angiver antallet af
publikationer, ikke vægtede mål. 

""")
    # Kategoriinddeling
    _tab_abs_pt, _tab_pct_pt = st.tabs(["Antal", "Andel (%)"])

    with _tab_abs_pt:
        fig2 = go.Figure()
        for i, t in enumerate(all_pub_types):
            _vals = [_pt.get(f, {}).get(t, 0) for f in fac_order_pt]
            fig2.add_trace(go.Bar(
                x=fac_order_pt, y=_vals, name=t,
                text=[str(v) if v > 0 else "" for v in _vals],
                textposition="inside",
                marker_color=_pt_colors[i],
            ))
        fig2.update_layout(
            barmode="stack",
            yaxis_title="Antal publikationer", 
            legend_title="Type", 
            height=500, 
            margin=dict(t=50, r=180),
            legend=dict(traceorder="reversed"),
            title=f"Fakulteternes publikationstypeantal, {year}"
            )
        st.plotly_chart(fig2, width="stretch", key=f"dg_pubtype_abs_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
    )

    with _tab_pct_pt:
        fig3 = go.Figure()
        for i, t in enumerate(all_pub_types):
            _vals_pct = []
            for f in fac_order_pt:
                _tot = sum(_pt.get(f, {}).values()) or 1
                _vals_pct.append(round(_pt.get(f, {}).get(t, 0) / _tot * 100, 1))
            fig3.add_trace(go.Bar(
                x=fac_order_pt, y=_vals_pct, name=t,
                text=[f"{v:.1f}%" if v > 0 else "" for v in _vals_pct],
                textposition="inside",
                marker_color=_pt_colors[i],
            ))
        fig3.update_layout(
            barmode="stack",
            yaxis=dict(title="Andel (%)", 
            range=[0, 100]),
            legend_title="Type", 
            height=500, 
            margin=dict(t=50, r=180),
            legend=dict(traceorder="reversed"),
            title = f"Fakulteternes andele af publikationstyper, {year}")
        st.plotly_chart(fig3, width="stretch", key=f"dg_pubtype_pct_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
    )

    if pubtype_map:
        _map_rows = sorted(set((raw, mapped) for raw, mapped in pubtype_map.items()))
        with st.expander("Se kategoriinddeling"):
            _map_schema = [("CURIS-type", pa.string()), ("Kategori", pa.string())]
            _map_data   = [{"CURIS-type": r, "Kategori": m} for r, m in _map_rows]
            st.dataframe(build_table(_map_data, _map_schema), hide_index=True, width="stretch")

# ---------------------------------------------------------------------------


# ===========================================================================
# KØN TAB
# ===========================================================================

def render_tab_køn(year, mode, raw_nodes, raw_edges, node_meta,
                   selected_facs, selected_insts, selected_grps,
                   all_years_data=None, edges_keep=None,
                   forfatterpositioner=None, sex_pairs=None, show_intra=True, filter_caption = None):
    st.subheader("Kønsfordeling i sampublicering")
    st.markdown(
f"""
Fanen viser, hvordan mænd og kvinder bidrager til KU's sampubliceringsaktivitet - og hvordan de indgår i forfatterpar med hinanden. 
Køn er bestemt ud fra CPR-nummeret i HR-data (lige slutciffer = kvinde, ulige slutciffer = mand). 

Fanen opererer med tre forskellige opgørelser af forfatterpar, som bruges til forskellige formål:

- **Netværkskanter** er forfatterpar *mellem* forskellige organisatoriske enheder - f.eks. en lektor på HUM og en professor
på SAMF eller mellem en ph.d. og lektor på JUR. Det er disse par, der ses i netværksvisningen, og de afspejler KU's tværgående
sampubliceringsmønstre. 
- **Alle forfatterpar, inkl. intra-enhed** tæller derudover par inden for samme enhed - f.eks. to mandlige professorer på
Kemisk Institut, der sampublicerer. Denne opgørelse er mere komplet, men kan ikke vises i netværket, da en cirkel ikke kan have
et forfattterpar med sig selv. 
- **Homofili-indekset** bruger altid alle forfatterpar, inkl. intra-enhed, da det ellers ville undervurdere tendensen til at 
publicere med kolleger af samme køn. 

**Fortolkningsnote**: analyserne i denne fane beskriver sampubliceringsmønstre - ikke præferencer eller bevidste valg hos
den enkelte forsker. Mønstre i køn kan i høj gra afspejle faglige miljøer, organisering og rekrutteringshistorik. 

Resultaterne bør derfor tolkes som **strukturelle mønstre**, ikke individuelle adfærdsmæssige forklaringer. 
""")



    # Bestem hvilke niveauer der er tilgængelige i denne mode
    _level_options = []
    if "F" in mode: 
        _level_options.append("Fakulteter")
    if "I" in mode: 
        _level_options.append("Institutter")
    if "G" in mode:
        _level_options.append("Stillingsgrupper") 

    _default_level = (
        ["Fakulteter"] if "F" in mode else
        ["Institutter"] if "I" in mode else
        ["Stillingsgrupper"]
    )
    _selected_levels = st.multiselect(
        "**Vis kønsfordeling for:**",
        options=_level_options,
        default=_default_level,
        key=f"køn_levels_{year}_{mode}",
    )
    if not _selected_levels:
        st.error("Vælg mindst ét organisatorisk niveau ovenfor.")
        return

    _level_map = {
        "Fakulteter":       ("fac",  "Fakultet"),
        "Institutter":      ("inst", "Institut"),
        "Stillingsgrupper": ("grp",  "Stillingsgruppe"),
    }

    # ── Lokale filtre ─────────────────────────────────────────────────────────
    _all_facs_køn = sorted({m.get("fac", "") for m in node_meta.values() if m.get("fac")})
    _all_insts_køn = sorted({m.get("inst", "") for m in node_meta.values() if m.get("inst")})
    _all_grps_køn = sorted({m.get("grp", "") for m in node_meta.values() if m.get("grp")}, key=lambda g: HIERARKI.get(g, 999))

    _show_fac_filter  = bool(_all_facs_køn)  and "Fakulteter"       in _selected_levels
    _show_inst_filter = bool(_all_insts_køn) and "Institutter"      in _selected_levels
    _show_grp_filter  = bool(_all_grps_køn)  and "Stillingsgrupper" in _selected_levels

    _køn_fac_filter = st.multiselect(
        "**Filtrer på fakultet**", options=_all_facs_køn if _show_fac_filter else [],
        default=[], key=f"køn_fac_filter_{year}_{mode}",
        placeholder="Alle fakulteter", disabled=not _show_fac_filter,
    ) if _show_fac_filter else []

    _køn_inst_filter = st.multiselect(
        "**Filtrer på institut**",
        options=[i for i in _all_insts_køn
                 if not _køn_fac_filter or any(
                     node_meta.get(nid, {}).get("fac") in _køn_fac_filter
                     for nid in node_meta if node_meta[nid].get("inst") == i
                 )] if _show_inst_filter else [],
        default=[], key=f"køn_inst_filter_{year}_{mode}",
        placeholder="Alle institutter", disabled=not _show_inst_filter,
    ) if _show_inst_filter else []

    _køn_grp_filter = st.multiselect(
        "**Filtrer på stillingsgruppe**", options=_all_grps_køn if _show_grp_filter else [],
        default=[], key=f"køn_grp_filter_{year}_{mode}",
        placeholder="Alle stillingsgrupper", disabled=not _show_grp_filter,
    ) if _show_grp_filter else []

    _n_filters = len(_køn_fac_filter) + len(_køn_inst_filter) + len(_køn_grp_filter)
    _hide_ku_samlet = len(_selected_levels) > 1

    def _køn_node_ok(m: dict) -> bool:
        if _køn_fac_filter  and m.get("fac", "")  not in _køn_fac_filter:  return False
        if _køn_inst_filter and m.get("inst", "") not in _køn_inst_filter: return False
        if _køn_grp_filter  and m.get("grp", "")  not in _køn_grp_filter:  return False
        return True
    
    def _køn_node_ok_by_org(org: str) -> bool:
            """Filtrerer org-nøgler baseret på lokale filtre."""
            if _køn_fac_filter and "Fakulteter" in _selected_levels:
                fac = org.split(" - ")[0]
                if fac not in _køn_fac_filter:
                    return False
            if _køn_inst_filter and "Institutter" in _selected_levels:
                inst = org.split(" - ")[1] if " - " in org else org
                if inst not in _køn_inst_filter:
                    return False
            if _køn_grp_filter and "Stillingsgrupper" in _selected_levels:
                grp = org.split(" - ")[-1]
                if grp not in _køn_grp_filter:
                    return False
            return True

    # Filtrerede versioner til brug i hele fanen
    _raw_nodes_f   = {nid: m for nid, m in raw_nodes.items()   if _køn_node_ok(m)}
    _node_meta_f   = {nid: m for nid, m in node_meta.items()   if _køn_node_ok(m)}
    # Filtrerede versioner til brug i hele fanen
    _edges_f = [e for e in edges_keep
                if _køn_node_ok(node_meta.get(e[0], {})) or _køn_node_ok(node_meta.get(e[1], {}))]
    _raw_edges_f   = [e for e in raw_edges
                      if _køn_node_ok(node_meta.get(e[0], {})) or _køn_node_ok(node_meta.get(e[1], {}))]

    # ── Bestem kombineret nøgle baseret på valgte niveauer ───────────────────
    sex_display = {"m": "Mænd", "k": "Kvinder"}
    sex_colors  = {"m": "#425570", "k": "#901a1E"}
    combo_display = {"k-k": "Kvinde-Kvinde", "k-m": "Kvinde-Mand", "m-m": "Mand-Mand"}
    combo_colors  = {"k-k": "#901a1E", "k-m": "#bac7d9", "m-m": "#425570"}

    def _org_key(m: dict) -> str:
        parts = []
        if "Fakulteter"       in _selected_levels: parts.append(m.get("fac",  "") or "")
        if "Institutter"      in _selected_levels: parts.append(m.get("inst", "") or "")
        if "Stillingsgrupper" in _selected_levels: parts.append(m.get("grp",  "") or "")
        return " - ".join(p for p in parts if p)

    _merged_to_raw: dict[str, dict] = {}
    for raw_nid, raw_m in _raw_nodes_f.items():
        merged_key = _org_key(raw_m)
        sex = raw_m.get("sex", "")
        if merged_key and sex:
            # Sæt sex på den merged key - bruges til ew-beregning
            _merged_to_raw[raw_nid] = raw_m

    group_label = " - ".join(
        {"Fakulteter": "Fakultet", "Institutter": "Institut", "Stillingsgrupper": "Stillingsgruppe"}[l]
        for l in ["Fakulteter", "Institutter", "Stillingsgrupper"] if l in _selected_levels
    )

    # Byg inst -> fac map fra node_meta så vi altid kan finde fakultetet
    # for et institut, også når kun "Institutter" er valgt som niveau.
    _inst_to_fac_local: dict[str, str] = {}
    for _m in node_meta.values():
        _i = _m.get("inst")
        _f = _m.get("fac")
        if not _i or not _f:
            continue
        # Normalisér til forkortelse hvis vi får fuldt fakultetsnavn
        _f_norm = FAC_ABBRS.get(_f, _f)
        # Foretræk gyldige (i FAC_ORDER) værdier; overskriv ugyldige
        if _i not in _inst_to_fac_local or _inst_to_fac_local[_i] not in FAC_ORDER:
            _inst_to_fac_local[_i] = _f_norm

    def _fac_rank(fac: str) -> int:
        return FAC_ORDER.index(fac) if fac in FAC_ORDER else 999

    def _parts_of(s: str) -> tuple[str, str, str]:
        """Returnér (fac, inst, grp) uanset hvilke niveauer der er valgt -
        slår fakultet op via _inst_to_fac_local hvis 'Fakulteter' ikke er valgt."""
        fac = inst = grp = ""
        rest = s

        if "Fakulteter" in _selected_levels:
            head, sep, tail = rest.partition(" - ")
            fac = head
            rest = tail if sep else ""

        if "Institutter" in _selected_levels:
            # Institutnavne kan selv indeholde " - " (f.eks. Saxo-Instituttet),
            # så vi slår op i _inst_to_fac_local i stedet for at splitte blindt.
            # Find det længste kendte institutnavn der er præfiks for `rest`.
            matched = ""
            for cand in _inst_to_fac_local:
                if (rest == cand or rest.startswith(cand + " - ")) and len(cand) > len(matched):
                    matched = cand
            if matched:
                inst = matched
                rest = rest[len(matched):]
                if rest.startswith(" - "):
                    rest = rest[3:]
            else:
                # Fallback: brug en simpel split (f.eks. når institut ikke
                # findes i mappen - bør ikke ske, men vi crasher ikke).
                head, sep, tail = rest.partition(" - ")
                inst = head
                rest = tail if sep else ""

        if "Stillingsgrupper" in _selected_levels:
            grp = rest

        if not fac and inst:
            fac = _inst_to_fac_local.get(inst, "")
        return fac, inst, grp

    def _sort_key(s: str) -> tuple:
        fac, inst, grp = _parts_of(s)
        return (_fac_rank(fac), fac, inst, HIERARKI.get(grp, 999))

    def _add_separators(keys: list) -> list:
        if not keys:
            return keys

        # Ingen separator hvis kun ét niveau er valgt - så er der ingen gruppering at lave
        _multi_level = sum(l in _selected_levels for l in 
                        ("Fakulteter", "Institutter", "Stillingsgrupper")) > 1
        if not _multi_level:
            return keys

        result = []
        prev_fac = None
        _sep_count = [0]
        def _sep():
            _sep_count[0] += 1
            return " " * _sep_count[0]

        for k in keys:
            fac, _inst, _grp = _parts_of(k)
            # Indsæt separator når fakultet skifter
            if prev_fac is not None and fac != prev_fac:
                result.append(_sep())
            result.append(k)
            prev_fac = fac
        return result

    # ── Fælles bar/figur-styling for alle plots i Køn-fanen ──────────────────
    _n_active_levels = sum(l in _selected_levels for l in 
                       ("Fakulteter", "Institutter", "Stillingsgrupper"))
    _BAR_W = 0.78 if _n_active_levels <= 1 else 0.55
    _SEP_H = 0   if _n_active_levels <= 1 else 6

    def _bar_layout(plot_keys: list, *, extra_top: int = 50, extra_bottom: int = 20,
                grouped: bool = False) -> dict:
        n_data = len(plot_keys)
        row_h = int(28 * (1.75 if grouped else 1.0))
        height = int(row_h * n_data + extra_top + extra_bottom + 30)
        height = max(280, height)
        return dict(
            height=height,
            bargap=0.05,
            bargroupgap=0.0,
            margin=dict(l=20, r=30, t=extra_top, b=extra_bottom),
        )

    st.markdown("---")

    # ── Forfatterbidrag fordelt på køn ───────────────────────────────────────
    st.markdown(f"#### Forfatterantal fordelt på køn, {year}")
    sex_size: dict[str, dict[str, int]] = {}
    for nid, m in _raw_nodes_f.items():
        if m.get("type") != "grp":
            continue
        key = _org_key(m)
        sex = m.get("sex", "ukendt")
        size = m.get("size", 0)
        if not key:
            continue
        sex_size.setdefault(key, {})
        sex_size[key][sex] = sex_size[key].get(sex, 0) + size

    groups_sorted = sorted(sex_size.keys(), key=_sort_key)
    sexes         = sorted({s for d in sex_size.values() for s in d})

    _ku_sex: dict[str, int] = {}
    for nid, m in raw_nodes.items():
        if m.get("type") != "grp":
            continue
        s = m.get("sex", "ukendt")
        v = m.get("size", 0)
        _ku_sex[s] = _ku_sex.get(s, 0) + v
    _ku_label      = "KU samlet"

    _groups_with_sep = _add_separators(groups_sorted)
    _sex_size_plot = {} if _hide_ku_samlet else {_ku_label: _ku_sex}
    for g in _groups_with_sep:
        _sex_size_plot[g] = sex_size.get(g, {})
    _groups_plot = ([] if _hide_ku_samlet else [_ku_label]) + _groups_with_sep

    def _build_y_positions(keys: list) -> tuple[list, list, list]:
        _multi_level = sum(l in _selected_levels for l in 
                        ("Fakulteter", "Institutter", "Stillingsgrupper")) > 1
        
        INTRA_GAP = 0.6
        INTER_GAP = 0.95
        
        _inst_in = "Institutter"      in _selected_levels
        _grp_in  = "Stillingsgrupper" in _selected_levels
        
        if not _multi_level:
            # 1-niveau mode: brug samme INTRA_GAP som fler-niveau, så søjlebredderne 
            # bliver visuelt konsistente på tværs af modes (alle bruger width=0.55).
            positions = [i * INTRA_GAP for i in range(len(keys))]
            return positions, positions, list(keys)

        positions = []
        labels = []
        prev_fac = None
        y = 0.0
        for k in keys:
            fac, inst, grp = _parts_of(k)
            if prev_fac and fac and fac != prev_fac:
                y += INTER_GAP
            else:
                y += INTRA_GAP if positions else 0
            positions.append(y)
            
            _fac_in = "Fakulteter" in _selected_levels
            if _inst_in and _grp_in:
                labels.append(f"{inst} | {grp}" if (inst and grp) else (inst or grp or k))
            elif _fac_in and _grp_in:
                labels.append(f"{fac} | {grp}" if (fac and grp) else (fac or grp or k))
            elif _grp_in:
                labels.append(grp or k)
            elif _inst_in:
                labels.append(inst or k)
            else:
                labels.append(fac or k)
            
            if fac:
                prev_fac = fac
        
        return positions, positions, labels

    _sex_size_plot = {} if _hide_ku_samlet else {_ku_label: _ku_sex}
    for g in groups_sorted:
        _sex_size_plot[g] = sex_size.get(g, {})
    _groups_plot = ([_ku_label] if not _hide_ku_samlet else []) + groups_sorted
    _y_pos, _tick_pos, _tick_labels = _build_y_positions(_groups_plot)

    _tab_abs, _tab_pct = st.tabs(["Antal", "Andel (%)"])

    with _tab_abs:
        fig1 = go.Figure()
        for sex in sexes:
            fig1.add_trace(go.Bar(
                name=sex_display.get(sex, sex),
                y=_y_pos,
                x=[_sex_size_plot[g].get(sex, 0) for g in _groups_plot],
                orientation="h",
                marker_color=sex_colors.get(sex, "#aaaaaa"),
                text=[_sex_size_plot[g].get(sex, 0) for g in _groups_plot],
                textposition="inside",
                width=0.55,
            ))
        fig1.update_layout(
            barmode="stack",
            yaxis=dict(
                tickmode="array",
                tickvals=_tick_pos,
                ticktext=_tick_labels,
                autorange="reversed",
                showgrid=False,
                zeroline=False,
            ),
            xaxis_title="Forfatterantal",
            title=f"Forfatterantal fordelt på køn, {year}",
            legend_title="Køn",
            **_bar_layout(_groups_plot),
        )
        st.plotly_chart(fig1, width='stretch',
            config={"toImageButtonOptions": {"format": "png", "scale": 3}})
        if filter_caption:
            st.caption(filter_caption)

    with _tab_pct:
        fig1p = go.Figure()
        for sex in sexes:
            _pct_vals = [
                round(100 * _sex_size_plot[g].get(sex, 0) / (sum(_sex_size_plot[g].values()) or 1), 1)
                for g in _groups_plot
            ]
            fig1p.add_trace(go.Bar(
                name=sex_display.get(sex, sex),
                y=_y_pos,                                   # ← konverteret
                x=_pct_vals,
                orientation="h",
                marker_color=sex_colors.get(sex, "#aaaaaa"),
                text=[f"{v}%" for v in _pct_vals],
                textposition="inside",
                width=0.55,
            ))
        fig1p.update_layout(
            barmode="stack",
            xaxis_title="Andel (%)",
            xaxis_range=[0, 100],
            legend_title="Køn",
            title=f"Forfatterandele fordelt på køn, {year}",
            yaxis=dict(
                tickmode="array",
                tickvals=_tick_pos,
                ticktext=_tick_labels,
                autorange="reversed",
                showgrid=False,
                zeroline=False,
            ),
            **_bar_layout(_groups_plot),
        )
        st.plotly_chart(fig1p, width='stretch',
            config={"toImageButtonOptions": {"format": "png", "scale": 3}})
        if filter_caption:
            st.caption(filter_caption)

    rows = []
    for g in groups_sorted:
        row = {group_label: g}
        total = sum(sex_size[g].values())
        for sex in sexes:
            n = sex_size[g].get(sex, 0)
            row[sex_display.get(sex, sex)] = n
            row[f"{sex_display.get(sex, sex)} (%)"] = round(100 * n / total, 1) if total else 0.0
        row["Total"] = total
        rows.append(row)
    if rows:
        schema_fields = (
            [(group_label, pa.string())] +
            [(sex_display.get(s, s), pa.int64()) for s in sexes] +
            [(f"{sex_display.get(s, s)} (%)", pa.float64()) for s in sexes] +
            [("Total", pa.int64())]
        )
        with st.expander("Se tabel"):
            st.dataframe(build_table(rows, schema_fields), hide_index=True, width="stretch")
            st.download_button(
                "Download (.xlsx)",
                data=rows_to_excel_bytes(rows, [n for n, _ in schema_fields]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"sex_forfatterantal_{year}_{mode}"
            )

    # ── Forfatterpositioner ───────────────────────────────────────────────────
    if forfatterpositioner and "Stillingsgrupper" in _selected_levels:
        st.markdown(
f"""#### Forfatterpositioner fordelt på køn, {year}
Figuren viser fordelingen af **førsteforfatter**, **mellemforfatter** og **sidsteforfatter**
fordelt på køn og stillingsgruppe. Forfatterposition er bestemt af rækkefølgen i CURIS -
førsteforfatter er den først listede forfatter, sidsteforfatter den sidst listede.
Kun publikationer med mindst to KU-VIP-forfattere indgår, da position ikke er meningsfuld
for soloartikler. Bemærk at normer for forfatterrækkefølge varierer betydeligt på tværs
af fagområder - på naturvidenskabelige og medicinske felter angiver sidstepladsen typisk
den ansvarlige seniorforskere, mens humanistiske og samfundsvidenskabelige traditioner
ofte bruger alfabetisk eller bidragsbaseret rækkefølge.
""")
        yr_data = forfatterpositioner.get(str(year), {})
        pos_labels = [("first", "Førsteforfatter"), ("middle", "Mellemforfatter"), ("last", "Sidsteforfatter")]
        sex_keys   = [("K", "Kvinder"), ("M", "Mænd")]
        pos_sex_colors = {
            ("first",  "K"): "#901a1E", ("middle", "K"): "#c45c61", ("last", "K"): "#e8a8aa",
            ("first",  "M"): "#122947", ("middle", "M"): "#425570", ("last", "M"): "#7a9bbf",
        }
        def _merge_sex_data(lookup_key: str, active_units: list) -> dict:
            merged: dict[str, dict[str, dict[str, int]]] = {}
            for unit in active_units:
                unit_data = yr_data.get(lookup_key, {}).get(unit, {})
                for sk, grp_dict in unit_data.items():
                    for grp, counts in grp_dict.items():
                        merged.setdefault(sk, {}).setdefault(grp, {})
                        for pos, n in counts.items():
                            merged[sk][grp][pos] = merged[sk][grp].get(pos, 0) + n
            return merged

        _facs_for_pos = _køn_fac_filter or _all_facs_køn

        if _køn_inst_filter:
            # Et eksplicit institut-filter er aktivt: aggreger på inst_sex
            # og brug rene stillingsgruppe-labels
            _pos_sex_data = _merge_sex_data("inst_sex", _køn_inst_filter)
            all_grps = sorted(
                {grp for sk, _ in sex_keys
                     for grp in _pos_sex_data.get(sk, {})
                     if not _køn_grp_filter or grp in _køn_grp_filter},
                key=lambda g: HIERARKI.get(g, 999),
            )

        elif "Institutter" in _selected_levels:
            # SFIG eller SIG: aggreger på inst_sex og byg "Fakultet - Institut - Stillingsgruppe"
            # så _build_y_positions kan vise det som "Institut | Stillingsgruppe"
            # og fakultets-grupperingen virker via _parts_of
            _all_insts_in_facs = sorted({
                m.get("inst", "") for m in node_meta.values()
                if m.get("inst") and m.get("fac") in _facs_for_pos
            })
            _inst_to_fac_pos = {
                m.get("inst"): m.get("fac")
                for m in node_meta.values() if m.get("inst") and m.get("fac")
            }

            _pos_sex_data = {}
            for inst in _all_insts_in_facs:
                fac = _inst_to_fac_pos.get(inst, "")
                if not fac:
                    continue
                inst_data = _merge_sex_data("inst_sex", [inst])
                for sk, grp_dict in inst_data.items():
                    _pos_sex_data.setdefault(sk, {})
                    for grp, counts in grp_dict.items():
                        if _køn_grp_filter and grp not in _køn_grp_filter:
                            continue
                        # Format: "Fakultet - Institut - Stillingsgruppe" så _parts_of kan parse det
                        if "Fakulteter" in _selected_levels:
                            combined = f"{fac} - {inst} - {grp}"
                        else:
                            # SIG: ingen fakultet i key, men _parts_of slår op via _inst_to_fac_local
                            combined = f"{inst} - {grp}"
                        _pos_sex_data[sk][combined] = counts

            all_grps = sorted(
                {g for sk, _ in sex_keys for g in _pos_sex_data.get(sk, {})},
                key=_sort_key,
            )

        elif "Fakulteter" in _selected_levels:
            # SFG (uden institut): aggreger på fac_sex og byg "Fakultet - Stillingsgruppe"
            _pos_sex_data = {}
            for fac in sorted(_facs_for_pos):
                fac_data = _merge_sex_data("fac_sex", [fac])
                for sk, grp_dict in fac_data.items():
                    _pos_sex_data.setdefault(sk, {})
                    for grp, counts in grp_dict.items():
                        if _køn_grp_filter and grp not in _køn_grp_filter:
                            continue
                        combined = f"{fac} - {grp}"
                        _pos_sex_data[sk][combined] = counts
            all_grps = sorted(
                {g for sk, _ in sex_keys for g in _pos_sex_data.get(sk, {})},
                key=_sort_key,
            )

        elif _køn_fac_filter:
            _pos_sex_data = _merge_sex_data("fac_sex", _køn_fac_filter)
            all_grps = sorted(
                {grp for sk, _ in sex_keys
                     for grp in _pos_sex_data.get(sk, {})
                     if not _køn_grp_filter or grp in _køn_grp_filter},
                key=lambda g: HIERARKI.get(g, 999),
            )

        else:
            # SG (kun stillingsgrupper): aggreger på "sex" på KU-niveau
            _pos_sex_data = yr_data.get("sex", {})
            all_grps = sorted(
                {grp for sk, _ in sex_keys
                     for grp in _pos_sex_data.get(sk, {})
                     if not _køn_grp_filter or grp in _køn_grp_filter},
                key=lambda g: HIERARKI.get(g, 999),
            )

        if all_grps:
            _tab_pos_abs, _tab_pos_pct = st.tabs(["Antal", "Andel (%)"])
            
            def _make_pos_fig(use_pct: bool) -> go.Figure:
                _pos_y, _pos_tick, _pos_labels = _build_y_positions(all_grps)
                fig = go.Figure()
                for sex_key, sex_label in sex_keys:
                    sex_data = _pos_sex_data.get(sex_key, {})
                    for pos_name, pos_label in pos_labels:
                        if use_pct:
                            vals, texts = [], []
                            for g in all_grps:
                                gd = sex_data.get(g, {})
                                tot = sum(
                                    _pos_sex_data.get(sk, {}).get(g, {}).get(p, 0)
                                    for sk, _ in sex_keys
                                    for p, _ in pos_labels
                                ) or 1
                                pct = round(100 * gd.get(pos_name, 0) / tot, 1)
                                vals.append(pct)
                                texts.append(f"{fmt_ui(pct)}%")
                        else:
                            vals  = [sex_data.get(g, {}).get(pos_name, 0) for g in all_grps]
                            texts = [str(v) for v in vals]
                        
                        fig.add_trace(go.Bar(
                            name=f"{sex_label} - {pos_label}",
                            y=_pos_y,                                # ← konverteret
                            x=vals,
                            orientation="h",
                            marker_color=pos_sex_colors[(pos_name, sex_key)],
                            legendgroup=sex_key,
                            legendgrouptitle_text=sex_label if pos_name == "first" else None,
                            text=texts, textposition="inside",
                            width=0.55,
                            customdata=all_grps,
                            hovertemplate=f"<b>%{{customdata}}</b> - {sex_label}<br>{pos_label}: %{{x}}" +
                                        ("%" if use_pct else "") + "<extra></extra>",
                        ))
                
                fig.update_layout(
                    barmode="stack",
                    xaxis_title="Andel af forfattere (%)" if use_pct else "Forfatterantal",
                    xaxis=dict(range=[0, 100]) if use_pct else {},
                    yaxis=dict(
                        tickmode="array",
                        tickvals=_pos_tick,
                        ticktext=_pos_labels,
                        autorange="reversed",
                        showgrid=False,
                        zeroline=False,
                    ),
                    title=f"Forfatterpositioner fordelt på køn, {year}" + (" - andel (%)" if use_pct else ""),
                    **_bar_layout(all_grps),
                )
                return fig
            

            with _tab_pos_abs:
                st.plotly_chart(_make_pos_fig(False), width="stretch", key=f"pos_sex_abs_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}})
                if filter_caption:
                    st.caption(filter_caption)
            
            with _tab_pos_pct:
                st.markdown("Andelen viser hvor stor en del af hvert køns forfattere, der falder i hver forfatterposition.")
                st.plotly_chart(_make_pos_fig(True), width="stretch", key=f"pos_sex_pct_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}})
                if filter_caption:
                    st.caption(filter_caption)
            
            _pos_rows = []
            
            for g in all_grps:
                row = {"Stillingsgruppe": g}
                for sex_key, sex_label in sex_keys:
                    sex_data = _pos_sex_data.get(sex_key, {})
                    gd = sex_data.get(g, {})
                    tot = sum(gd.get(p, 0) for p, _ in pos_labels) or 1
                    for pos_name, pos_label in pos_labels:
                        v = gd.get(pos_name, 0)
                        row[f"{sex_label} - {pos_label} (n)"]  = v
                        row[f"{sex_label} - {pos_label} (%)"] = round(100 * v / tot, 1)
                _pos_rows.append(row)
            _pos_col_names = ["Stillingsgruppe"] + [
                f"{sex_label} - {pos_label} {suffix}"
                for sex_key, sex_label in sex_keys
                for pos_name, pos_label in pos_labels
                for suffix in ["(n)", "(%)"]
            ]
            _pos_schema = [("Stillingsgruppe", pa.string())] + [
                (c, pa.float64() if "(%)" in c else pa.int64()) for c in _pos_col_names[1:]
            ]
            with st.expander("Se tabel"):
                st.dataframe(build_table(_pos_rows, _pos_schema), hide_index=True, width="stretch")
                st.download_button(
                    "Download (.xlsx)",
                    data=rows_to_excel_bytes(_pos_rows, _pos_col_names),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_pos_sex_{year}_{mode}",
                )
    
    st.markdown("---")

    # ── Sampubliceringsaktivitet per køn ─────────────────────────────────────
    st.markdown(f"#### Sampubliceringsaktivitet fordelt på køn, {year}")
    st.markdown(f"""
Figuren viser den samlede sampubliceringsaktivitet for kvindelige og mandlige forskere, **målt
som forfatterpar**. Et forfatterpar opstår når to forfattere sampublicerer - en publikation med
fire forfattere giver således seks forfatterpar. Hver forfatter tildeles halvdelen af forfatterpar
per par, så tallene kan summeres på tværs af enheder.
""")

    _grp_sex_ew:   dict[str, dict[str, float]] = {}
    _grp_sex_size: dict[str, dict[str, int]]   = {}

    for nid, m in _raw_nodes_f.items():
        if m.get("type") != "grp":
            continue
        key = _org_key(m)
        sex = m.get("sex", "")
        if not key or not sex:
            continue
        _grp_sex_size.setdefault(key, {})
        _grp_sex_size[key][sex] = _grp_sex_size[key].get(sex, 0) + m.get("size", 0)

    for u, v, w, *_ in _edges_f:
        for nid in (u, v):
            m = _node_meta_f.get(nid, {})
            key = _org_key(m)
            sex = m.get("sex", "")
            if key and sex:
                _grp_sex_ew.setdefault(key, {})
                _grp_sex_ew[key][sex] = _grp_sex_ew[key].get(sex, 0.0) + w

    # Alle forfatterpar inkl. intra-node fra sex_pairs
    _grp_sex_ew_all: dict[str, dict[str, float]] = {}
    for raw_key, counts in (sex_pairs or {}).items():
        parts = raw_key.split("|")
        if len(parts) != 4:
            continue
        fac, inst, grp, focal_sex = parts
        if selected_facs  and fac  not in selected_facs:  continue
        if selected_insts and inst not in selected_insts: continue
        if selected_grps  and grp  not in selected_grps:  continue
        if _køn_fac_filter  and fac  not in _køn_fac_filter:  continue
        if _køn_inst_filter and inst not in _køn_inst_filter: continue
        if _køn_grp_filter  and grp  not in _køn_grp_filter:  continue
        org_parts = []
        if "Fakulteter"       in _selected_levels: org_parts.append(fac)
        if "Institutter"      in _selected_levels: org_parts.append(inst)
        if "Stillingsgrupper" in _selected_levels: org_parts.append(grp)
        org_key = " - ".join(org_parts) if org_parts else fac
        total = counts.get("same", 0) + counts.get("cross", 0)
        _grp_sex_ew_all.setdefault(org_key, {})
        _grp_sex_ew_all[org_key][focal_sex] = _grp_sex_ew_all[org_key].get(focal_sex, 0.0) + total

    _grps_sx = groups_sorted

    _ku_sx_ew: dict[str, float] = {}
    for u, v, w, *rest in edges_keep:
        combo = rest[0] if rest and rest[0] else None
        if not combo or "-" not in combo:
            continue
        sex_u, sex_v = combo.split("-", 1)
        _ku_sx_ew[sex_u] = _ku_sx_ew.get(sex_u, 0.0) + w 
        _ku_sx_ew[sex_v] = _ku_sx_ew.get(sex_v, 0.0) + w
    _ku_sx_label      = "KU samlet"
    _grps_sx_with_sep = _add_separators(_grps_sx)
    _grp_sex_ew_plot  = (
        {g: _grp_sex_ew.get(g, {}) for g in _grps_sx_with_sep}
        if _hide_ku_samlet else
        {_ku_sx_label: _ku_sx_ew, **{g: _grp_sex_ew.get(g, {}) for g in _grps_sx_with_sep}}
    )
    _grps_sx_plot = ([] if _hide_ku_samlet else [_ku_sx_label]) + _grps_sx

    def _make_sx_fig(ew_dict: dict, plot_list: list, use_pct: bool, key_suffix: str):
        _sx_y, _sx_tick, _sx_labels = _build_y_positions(plot_list)
        fig = go.Figure()
        for sex, label, color in [("k", "Kvinder", "#901a1E"), ("m", "Mænd", "#425570")]:
            if use_pct:
                vals = [round(100 * ew_dict.get(g, {}).get(sex, 0) / (sum(ew_dict.get(g, {}).values()) or 1), 1) for g in plot_list]
                texts = [f"{fmt_ui(v)}%" for v in vals]
            else:
                vals = [ew_dict.get(g, {}).get(sex, 0) for g in plot_list]
                texts = [f"{fmt_ui(v)}" for v in vals]
            fig.add_trace(go.Bar(
                name=label,
                y=_sx_y,                                     # ← konverteret
                x=vals,
                orientation="h",
                marker_color=color,
                text=texts, textposition="inside",
                width=0.55,
                customdata=plot_list,
                hovertemplate="<b>%{customdata}</b><br>%{x:.1f}" + 
                            ("%" if use_pct else " forfatterpar") + "<extra></extra>",
            ))
        fig.update_layout(
            barmode="stack",
            xaxis=dict(title="Andel (%)", range=[0, 100]) if use_pct else dict(title="Antal forfatterpar"),
            yaxis=dict(
                tickmode="array",
                tickvals=_sx_tick,
                ticktext=_sx_labels,
                autorange="reversed",
                showgrid=False,
                zeroline=False,
            ),
            legend_title="Køn",
            title=f"Forfatterpar fordelt på køn{' (inkl. intra-enhed)' if _use_sx_all else ''}, {year}" + (" - andel (%)" if use_pct else ""),
            **_bar_layout(plot_list),
        )
        return fig

    # Byg plot-lister for begge versioner
    _grps_sx_all = groups_sorted
    _ku_sx_ew_all: dict[str, float] = {}
    for raw_key, counts in (sex_pairs or {}).items():
        parts = raw_key.split("|")
        if len(parts) == 4:
            _ku_sx_ew_all[parts[3]] = _ku_sx_ew_all.get(parts[3], 0.0) + counts.get("same", 0) + counts.get("cross", 0)

    _grp_sex_ew_all_plot = (
        {g: _grp_sex_ew_all.get(g, {}) for g in _grps_sx_all}
        if _hide_ku_samlet else
        {_ku_sx_label: _ku_sx_ew_all, **{g: _grp_sex_ew_all.get(g, {}) for g in _grps_sx_all}}
    )
    _grps_sx_all_plot = ([] if _hide_ku_samlet else [_ku_sx_label]) + _grps_sx_all

    _lost_keys = set(_grp_sex_ew) - set(groups_sorted)
    if _lost_keys:
        st.write(f"DEBUG: Mistede grupper: {_lost_keys}")
        st.write(f"DEBUG: Deres edge-vægte: {[(k, _grp_sex_ew[k]) for k in _lost_keys]}")

    _use_sx_all = st.toggle(
        "Inkludér intra-enhed forfatterpar (f.eks. to professorer på samme institut)",
        value=False,
        key=f"sx_all_{year}_{mode}",
        help="Til: alle forfatterpar inkl. par inden for samme enhed. Fra: kun netværkskanter.",
        disabled=not show_intra
    )

    _active_sx_plot = _grps_sx_all_plot  if _use_sx_all else _grps_sx_plot
    _active_sx_ew   = _grp_sex_ew_all_plot if _use_sx_all else _grp_sex_ew_plot

    if _use_sx_all:
        st.caption("Alle forfatterpar inkl. par inden for samme enhed (f.eks. to mandlige professorer på Kemisk Institut).")
    else:
        st.caption("Kun forfatterpar *mellem* de valgte organisatoriske enheder - som vist i netværket.")

    #st.write(f"DEBUG: _groups_plot len={len(_groups_plot)}, _active_sx_plot len={len(_active_sx_plot)}")

    tab_abs, tab_pct = st.tabs(["Antal", "Andel (%)"])
    with tab_abs:
        st.plotly_chart(_make_sx_fig(_active_sx_ew, _active_sx_plot, False, "abs"),
            width='content', key=f"sx_abs_{year}_{mode}",
            config={"toImageButtonOptions": {"format": "png", "scale": 3}})
        if filter_caption:
            st.caption(filter_caption)
        _tbl_rows_abs = [
            {group_label: g,
             "Kvinder (forfatterpar)": round(_active_sx_ew.get(g, {}).get("k", 0), 1),
             "Mænd (forfatterpar)":    round(_active_sx_ew.get(g, {}).get("m", 0), 1)}
            for g in _active_sx_plot if g != "---"
        ]
        _tbl_schema_abs = [
            (group_label,              pa.string()),
            ("Kvinder (forfatterpar)", pa.float64()),
            ("Mænd (forfatterpar)",    pa.float64()),
        ]
        with st.expander("Se tabel"):
            st.dataframe(build_table(_tbl_rows_abs, _tbl_schema_abs), hide_index=True, width="stretch")
            st.download_button("Download (.xlsx)",
                data=rows_to_excel_bytes(_tbl_rows_abs, [n for n, _ in _tbl_schema_abs]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_sx_abs_{year}_{mode}")

    with tab_pct:
        st.plotly_chart(_make_sx_fig(_active_sx_ew, _active_sx_plot, True, "pct"),
            width='content', key=f"sx_pct_{year}_{mode}",
            config={"toImageButtonOptions": {"format": "png", "scale": 3}})
        if filter_caption:
            st.caption(filter_caption)
        _tbl_rows_pct = [
            {group_label: g,
             "Kvinder (%)": round(100 * _active_sx_ew.get(g, {}).get("k", 0) / (sum(_active_sx_ew.get(g, {}).values()) or 1), 1),
             "Mænd (%)":    round(100 * _active_sx_ew.get(g, {}).get("m", 0) / (sum(_active_sx_ew.get(g, {}).values()) or 1), 1)}
            for g in _active_sx_plot if g != "---"
        ]
        _tbl_schema_pct = [
            (group_label,  pa.string()),
            ("Kvinder (%)", pa.float64()),
            ("Mænd (%)",    pa.float64()),
        ]
        with st.expander("Se tabel"):
            st.dataframe(build_table(_tbl_rows_pct, _tbl_schema_pct), hide_index=True, width="stretch")
            st.download_button("Download (.xlsx)",
                data=rows_to_excel_bytes(_tbl_rows_pct, [n for n, _ in _tbl_schema_pct]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_sx_pct_{year}_{mode}")

    st.markdown("---")

    # ── Forfatterpar per kønskombination ─────────────────────────────────────
    st.markdown(f"""#### Sampubliceringer per kønskombination

En publikation med flere forfattere indgår med flere forfatterpar og kan derfor optræde under
flere kønskombinationer samtidig. Sampubliceringsopgørelserne tæller således **forfatterpar** -
og ikke unikke publikationer.

""")

    _hom_org_combo: dict[str, dict[str, dict[str, int]]] = defaultdict(lambda: {"k": {"same": 0, "cross": 0}, "m": {"same": 0, "cross": 0}})
    for raw_key, counts in (sex_pairs or {}).items():
        parts = raw_key.split("|")
        if len(parts) != 4:
            continue
        fac, inst, grp, focal_sex = parts

        # Globale filtre
        if selected_facs and fac not in selected_facs: continue
        if selected_insts and inst not in selected_insts: continue
        if selected_grps and grp not in selected_grps: continue
        # Lokale filtre
        if _køn_fac_filter and fac not in _køn_fac_filter: continue
        if _køn_inst_filter and inst not in _køn_inst_filter: continue
        if _køn_grp_filter and grp not in _køn_grp_filter: continue

        org_parts = []
        if "Fakulteter"       in _selected_levels: org_parts.append(fac)
        if "Institutter"      in _selected_levels: org_parts.append(inst)
        if "Stillingsgrupper" in _selected_levels: org_parts.append(grp)
        org_key = " - ".join(org_parts) if org_parts else fac

        _hom_org_combo[org_key][focal_sex]["same"]  += counts.get("same", 0)
        _hom_org_combo[org_key][focal_sex]["cross"] += counts.get("cross", 0)

    org_combo: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for u, v, w, *rest in _edges_f:
        combo = rest[0] if rest and rest[0] else "ukendt"
        w = int(round(w))
        org_u = _org_key(_node_meta_f.get(u, {}))
        org_v = _org_key(_node_meta_f.get(v, {}))
        for org in ({org_u, org_v} if org_u != org_v else {org_u}):
            if org:
                org_combo[org][combo] += w

    _ku_combo: dict[str, int] = {}
    for u, v, w, *rest in edges_keep:
        combo = rest[0] if rest and rest[0] else "ukendt"
        _ku_combo[combo] = _ku_combo.get(combo, 0) + int(round(w))
    _ku_combo_label = "KU samlet"

    # Brug samme nøgle-mængde som forfatterantal-plottet for konsistent layout
    _orgs_plot = ([] if _hide_ku_samlet else [_ku_combo_label]) + groups_sorted
    _org_combo_plot = ({_ku_combo_label: _ku_combo} if not _hide_ku_samlet else {})
    _org_combo_plot.update({org: org_combo.get(org, {}) for org in groups_sorted})
    combos = sorted({c for d in org_combo.values() for c in d})

    # Byg _org_combo_all fra _hom_org_combo (sex_pairs) med samme k-k/k-m/m-m format
    _org_combo_all: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for org, sx in _hom_org_combo.items():
        _org_combo_all[org]["k-k"] = sx["k"]["same"] / 2
        _org_combo_all[org]["m-m"] = sx["m"]["same"] / 2
        _org_combo_all[org]["k-m"] = sx["k"]["cross"]  # = m's cross, tæl én gang

    # KU samlet for alle par
    _ku_combo_all: dict[str, float] = {}
    for d in _org_combo_all.values():
        for c, v in d.items():
            _ku_combo_all[c] = _ku_combo_all.get(c, 0.0) + v

    _use_combo_all = st.toggle(
        "Inkludér intra-enhed forfatterpar (f.eks. to professorer på samme institut)",
        value=False,
        key=f"combo_all_{year}_{mode}",
        help="Til: alle forfatterpar inkl. par inden for samme enhed. Fra: kun netværkskanter.",
        disabled=not show_intra
    )

    if _use_combo_all:
        st.caption("Alle forfatterpar inkl. par inden for samme enhed (f.eks. to mandlige professorer på Kemisk Institut).")
    else:
        st.caption("Kun forfatterpar *mellem* de valgte organisatoriske enheder - som vist i netværket.")

    if _use_combo_all:
        _active_orgs_plot = ([] if _hide_ku_samlet else [_ku_combo_label]) + groups_sorted
        _active_combo_plot = ({_ku_combo_label: _ku_combo_all} if not _hide_ku_samlet else {})
        _active_combo_plot.update({org: dict(_org_combo_all.get(org, {})) for org in groups_sorted})
    else:
        _active_orgs_plot  = _orgs_plot
        _active_combo_plot = _org_combo_plot

    _active_combos = sorted({c for d in _active_combo_plot.values() for c in d})

    # Numeriske y-positioner for kompakt fakultets-gruppering
    _combo_y, _combo_tick, _combo_labels = _build_y_positions(_active_orgs_plot)

    tab_act, tab_pct = st.tabs(["Antal", "Andel (%)"])

    with tab_act:
        fig2 = go.Figure()
        for combo in _active_combos:
            fig2.add_trace(go.Bar(
                name=combo_display.get(combo, combo),
                y=_combo_y,
                x=[_active_combo_plot[org].get(combo, 0) for org in _active_orgs_plot],
                orientation="h",
                marker_color=combo_colors.get(combo, "#aaaaaa"),
                text=[fmt_ui(_active_combo_plot[org].get(combo, 0)) for org in _active_orgs_plot],
                textposition="inside",
                width=0.55,
                customdata=_active_orgs_plot,
                hovertemplate="<b>%{customdata}</b><br>%{x}<extra></extra>",
            ))
        fig2.update_layout(
            barmode="stack",
            yaxis=dict(
                tickmode="array",
                tickvals=_combo_tick,
                ticktext=_combo_labels,
                autorange="reversed",
                showgrid=False,
                zeroline=False,
            ),
            xaxis_title="Antal forfatterpar",
            legend_title="Kønskombination",
            title=f"Forfatterpar per kønskombination{' (inkl. intra-enhed)' if _use_combo_all else ''}, {year}",
            **_bar_layout(_active_orgs_plot),
        )
        st.plotly_chart(fig2, width="stretch",
            config={"toImageButtonOptions": {"format": "png", "scale": 3}})
        if filter_caption:
            st.caption(filter_caption)

        combo_rows = []
        for org in _active_orgs_plot:
            row = {group_label: org}
            total = sum(_active_combo_plot[org].values())
            for combo in _active_combos:
                row[combo_display.get(combo, combo)] = round(_active_combo_plot[org].get(combo, 0), 1)
            row["Total"] = round(total, 1)
            combo_rows.append(row)
        combo_schema = (
            [(group_label, pa.string())] +
            [(combo_display.get(c, c), pa.float64()) for c in _active_combos] +
            [("Total", pa.float64())]
        )
        with st.expander("Se tabel"):
            st.dataframe(build_table(combo_rows, combo_schema), hide_index=True, width="stretch")
            st.download_button("Download (.xlsx)",
                data=rows_to_excel_bytes(combo_rows, [n for n, _ in combo_schema]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"sex_combo_{year}_{mode}")

    with tab_pct:
        fig2p = go.Figure()
        for combo in _active_combos:
            _pct_vals = [
                round(100 * _active_combo_plot[org].get(combo, 0) / (sum(_active_combo_plot[org].values()) or 1), 1)
                for org in _active_orgs_plot
            ]
            fig2p.add_trace(go.Bar(
                name=combo_display.get(combo, combo),
                y=_combo_y,
                x=_pct_vals,
                orientation="h",
                marker_color=combo_colors.get(combo, "#aaaaaa"),
                text=[f"{v}%" for v in _pct_vals],
                textposition="inside",
                width=0.55,
                customdata=_active_orgs_plot,
                hovertemplate="<b>%{customdata}</b><br>%{x}%<extra></extra>",
            ))
        fig2p.update_layout(
            barmode="stack",
            yaxis=dict(
                tickmode="array",
                tickvals=_combo_tick,
                ticktext=_combo_labels,
                autorange="reversed",
                showgrid=False,
                zeroline=False,
            ),
            xaxis=dict(title="Andel (%)", range=[0, 100]),
            legend_title="Kønskombination",
            title=f"Forfatterpar per kønskombination{' (inkl. intra-enhed)' if _use_combo_all else ''} - andel (%), {year}",
            **_bar_layout(_active_orgs_plot),
        )
        st.plotly_chart(fig2p, width="stretch",
            config={"toImageButtonOptions": {"format": "png", "scale": 3}})
        if filter_caption:
            st.caption(filter_caption)

        combo_rows_pct = []
        for org in _active_orgs_plot:
            row = {group_label: org}
            total = sum(_active_combo_plot[org].values())
            for combo in _active_combos:
                n = _active_combo_plot[org].get(combo, 0)
                row[combo_display.get(combo, combo)] = round(n, 1)
                row[f"{combo_display.get(combo, combo)} (%)"] = round(100 * n / total, 1) if total else 0.0
            row["Total"] = round(total, 1)
            combo_rows_pct.append(row)
        combo_schema_pct = (
            [(group_label, pa.string())] +
            [(combo_display.get(c, c), pa.float64()) for c in _active_combos] +
            [(f"{combo_display.get(c, c)} (%)", pa.float64()) for c in _active_combos] +
            [("Total", pa.float64())]
        )
        with st.expander("Se tabel"):
            st.dataframe(build_table(combo_rows_pct, combo_schema_pct), hide_index=True, width="stretch")
            st.download_button("Download (.xlsx)",
                data=rows_to_excel_bytes(combo_rows_pct, [n for n, _ in combo_schema_pct]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"combo_sex_{year}_{mode}")
    
    st.markdown("---")

    st.markdown(
f"""
#### Kønshomofili

Fanen undersøger, om forskere på KU publicerer med kolleger af samme køn oftere end 
den lokale kønsfordeling ville forudsige - såkaldt **kønshomofili**. 

Opgørelsen er lavet på **individniveau**: hvert forfatterpar tælles fra begge forfatteres perspektiv.
Det betyder, at et forfatterpar mellem to mandlige professorer på Kemisk Institut tæller som 
ét *same-gender* par fra hver af de to mænds perspektiv - og disse intra-enhed par er **inkluderet**,
i modsætning til øvrige netværksanalyse, der kun viser kanter *mellem* enheder. 

**Beregning for en given enhed:**

1. **Baseline** - kønnets andel af unikke forfattere i enheden (f.eks. 30% kvinder)
2. **Same-gender rate** - andelen af forfatterpar, set fra kønnets eget perspektiv, der er med
en forfatter af samme køn (f.eks. andelen af kvinders par, der er med en anden kvinde)
3. **Homofili-indeks** = same-gender rate / baseline 

Et indeks på 1.0 betyder, at man publicerer med samme køn præcis så ofte som en tilfældig mixing 
ville forudsige. **Klubfaktoren** sætter mænds og kvinders indeks i forhold til hinanden.

Opgørelserne bygger på **alle {sum(sum(v.values()) for v in (sex_pairs or {}).values())} 
forfatterpar-endpoints** i det valgte udsnit - inklusiv par inden for samme enhed.

Homofili-indekset sammenligner den observerede same-gender rate med kønsfordelingen i den valgte organisatoriske
enhed (f.eks. et fakultet eller institut). Men sampublicering foregår primært inden for snævrere grupper - 
laboratorier, centre og forskningsgrupper - som kan have skæve kønsfordelinger. Et højt indeks kan derfor
afspejle, at mænd og kvinder er koncentrerede i forskellige faglige miljøer, snarere end en præference for
at publicere med eget køn. Indekset bør derfor tolkes som et mål for *kønssegregation i sampubliceringsmønstre* - 
ikke nødvendigvis som et udtryk for bevidste valg.

Enheder med færre end fem forfatterpar vises ikke, da indekset er ustabilt ved få observationer.
""")
    # KU samlet
    _ku_sp = {"k": {"same": 0, "cross": 0}, "m": {"same": 0, "cross": 0}}
    for raw_key, counts in (sex_pairs or {}).items():
        parts = raw_key.split("|")
        if len(parts) != 4:
            continue
        focal_sex = parts[3]
        if focal_sex in _ku_sp:
            _ku_sp[focal_sex]["same"]  += counts.get("same", 0)
            _ku_sp[focal_sex]["cross"] += counts.get("cross", 0)

    _min_pairs = 5
    baseline_k_ku = _ku_sex.get("k", 0) / (sum(_ku_sex.values()) or 1)
    baseline_m_ku = _ku_sex.get("m", 0) / (sum(_ku_sex.values()) or 1)
    pairs_k_ku = _ku_sp["k"]["same"] + _ku_sp["k"]["cross"]
    pairs_m_ku = _ku_sp["m"]["same"] + _ku_sp["m"]["cross"]
    rate_k_ku  = _ku_sp["k"]["same"] / (pairs_k_ku or 1)
    rate_m_ku  = _ku_sp["m"]["same"] / (pairs_m_ku or 1)
    index_k_ku = rate_k_ku / baseline_k_ku if (baseline_k_ku > 0 and pairs_k_ku >= _min_pairs) else None
    index_m_ku = rate_m_ku / baseline_m_ku if (baseline_m_ku > 0 and pairs_m_ku >= _min_pairs) else None
    klubfaktor_ku = round(index_m_ku / index_k_ku, 2) if index_k_ku and index_k_ku > 0 else None

    homofili_rows = []
    _hom_baseline: dict[str, dict[str, int]] = {}
    for nid, m in _raw_nodes_f.items():
        if m.get("type") != "grp":
            continue
        key = _org_key(m)
        sex = m.get("sex", "")
        if not key or not sex:
            continue
        _hom_baseline.setdefault(key, {})
        _hom_baseline[key][sex] = _hom_baseline[key].get(sex, 0) + m.get("size", 0)

    for org in sorted(_hom_org_combo.keys(), key=_sort_key):
        fb = _hom_baseline.get(org, {})
        total_fb = sum(fb.values()) or 1
        baseline_k = fb.get("k", 0) / total_fb
        baseline_m = fb.get("m", 0) / total_fb
        k_data = _hom_org_combo[org]["k"]
        m_data = _hom_org_combo[org]["m"]
        _min_pairs = 5
        pairs_k = k_data["same"] + k_data["cross"]
        pairs_m = m_data["same"] + m_data["cross"]
        rate_k  = k_data["same"] / (pairs_k or 1)
        rate_m  = m_data["same"] / (pairs_m or 1)
        index_k = round(rate_k / baseline_k, 2) if (baseline_k > 0 and pairs_k >= _min_pairs) else None
        index_m = round(rate_m / baseline_m, 2) if (baseline_m > 0 and pairs_m >= _min_pairs) else None
        klubfaktor = round(index_m / index_k, 2) if (index_k and index_m and index_k > 0) else None
        homofili_rows.append({
            group_label:          org,
            "Homofili-indeks K":  index_k,
            "Homofili-indeks M":  index_m,
            "Klubfaktor (M/K)":   klubfaktor,
        })

    # ── KU samlet rækken ──
    if not _hide_ku_samlet:
        homofili_rows = [{
            group_label:         "KU samlet",
            "Homofili-indeks K": round(index_k_ku, 2) if index_k_ku else None,
            "Homofili-indeks M": round(index_m_ku, 2) if index_m_ku else None,
            "Klubfaktor (M/K)":  klubfaktor_ku,
        }] + homofili_rows

    _orgs_hom  = [r[group_label] for r in homofili_rows]
    _idx_k     = [r["Homofili-indeks K"] or 0 for r in homofili_rows]
    _idx_m     = [r["Homofili-indeks M"] or 0 for r in homofili_rows]
    _klub      = [r["Klubfaktor (M/K)"] or 0 for r in homofili_rows]

    _hom_y, _hom_tick, _hom_labels = _build_y_positions(_orgs_hom)

    _tab_hom_idx, _tab_hom_klub = st.tabs(["Homofili-indeks", "Klubfaktor"])

    with _tab_hom_idx:
        st.markdown(
f"""
Indekset viser, om hvert køn publicerer med samme køn oftere eller sjældnere end den lokale
kønsfordeling forudsiger:

- **indeks > 1** - publicerer med samme køn **oftere** end forventet -> tendens til klubdannelse
- **indeks = 1** - ingen tendens, svarer til tilfældig mixing
- **indeks < 1** - publicerer med samme køn **sjældnere** end forventet -> krydskøns-præference

**Eksempel:** en enhed har 20% mænd og 80% kvinder. Hvis mænd alligevel indgår i 40% af deres
forfatterpar med andre mænd er indekset 0.40/0.20 = 2.0 -> mænd publicerer med mænd dobbelt så ofte, 
som tilfældig mixing ville forudsige. Hvis kvinder tilsvarende indgår i 60% af deres par med andre
kvinder, er kvindeindekset 0.60/0.80 = 0.75 - kvinder publicerer faktisk sjældnere med andre
kvinder end forventet.
""")
        _fig_hom = go.Figure()
        for vals, color, label in [(_idx_k, "#901a1E", "Kvinder"), (_idx_m, "#425570", "Mænd")]:
            _fig_hom.add_trace(go.Bar(
                name=label,
                y=_hom_y,
                x=vals,
                orientation="h",
                marker_color=color,
                width=0.28,
                text=[f"{v:.2f}" for v in vals],
                textposition="inside",
                customdata=_orgs_hom,
                hovertemplate="<b>%{customdata}</b><br>Homofili-indeks: %{x:.2f}<extra></extra>",
            ))
        _fig_hom.add_vline(x=1.0, line_dash="dash", line_color="#3d3d3d",
            annotation_text="Ingen homofili (= 1)", annotation_position="top right")
        _layout_hom = _bar_layout(_orgs_hom, grouped=True)
        _layout_hom["bargroupgap"] = 0.0  # ingen luft mellem k- og m-søjlen i samme gruppe
        _fig_hom.update_layout(
            barmode="group",
            xaxis_title="Homofili-indeks",
            legend_title="Køn",
            yaxis=dict(
                tickmode="array",
                tickvals=_hom_tick,
                ticktext=_hom_labels,
                autorange="reversed",
                showgrid=False,
                zeroline=False,
            ),
            title=f"Kønshomofili per {group_label.lower()}, {year}",
            **_layout_hom,
        )
        st.plotly_chart(_fig_hom, width='content',
            config={"toImageButtonOptions": {"format": "png", "scale": 3}})
        if filter_caption:
            st.caption(filter_caption)

    with _tab_hom_klub:
        st.markdown(
f"""
Klubfaktoren = mændenes homofili-indeks / kvindernes homofili-indeks. Den viser, om der er 
forskel på de to køns tendens til at publicere inden for samme køn, korrigeres for kønsfordelingen:

- **klubfaktor > 1** - mænd klumper sig relativt mere sammen end kvinder -> *mandeklub-tendens*
- **klubfaktor = 1** - ingen forskel mellem kønnenes homofilimønstre
- **klubfaktor < 1** - kvinder klumper sig relativt mere end mænd -> *kvindeklub-tendens*

Klubfaktoren beregnes kun hvor begge køn har mindst fem forfatterpar i enheden - samme tærskel 
som for homofili-indekset.
""")
        _fig_klub = go.Figure()
        _fig_klub.add_trace(go.Bar(
            name="Klubfaktor",
            y=_hom_y,
            x=_klub,
            orientation="h",
            marker_color=["#122947" if v >= 1 else "#901a1E" for v in _klub],
            width=0.55,
            text=[f"{v:.2f}" for v in _klub],
            textposition="inside",
            customdata=_orgs_hom,
            hovertemplate="<b>%{customdata}</b><br>Klubfaktor: %{x:.2f}<extra></extra>",
        ))
        _fig_klub.add_vline(x=1.0, line_dash="dash", line_color="#3d3d3d",
            annotation_text="Ingen forskel (= 1)", annotation_position="top right")
        _fig_klub.update_layout(
            xaxis_title="Klubfaktor (M/K)",
            yaxis=dict(
                tickmode="array",
                tickvals=_hom_tick,
                ticktext=_hom_labels,
                autorange="reversed",
                showgrid=False,
                zeroline=False,
            ),
            title=f"Klubfaktor per {group_label.lower()}, {year}",
            **_bar_layout(_orgs_hom, grouped=False),
        )
        st.plotly_chart(_fig_klub, width='content',
            config={"toImageButtonOptions": {"format": "png", "scale": 3}})
        if filter_caption:
            st.caption(filter_caption)


    st.markdown("---")


    # ── Udvikling over tid ────────────────────────────────────────────────────
    if all_years_data and len(all_years_data) >= 2:
        st.subheader("Udvikling over tid")
        years_sorted = sorted(all_years_data.keys())
        _tab_bidrag, _tab_combo, _tab_homofili = st.tabs([
            "Forfatterantal per køn", "Forfatterpar per kønskombination", "Kønshomofili"
        ])

        with _tab_bidrag:
            _all_sexes = sorted({s for snap in all_years_data.values() for s in snap.get("sex_bidrag", {})})
            fig_sb = go.Figure()
            for sex in _all_sexes:
                fig_sb.add_trace(go.Scatter(
                    x=years_sorted,
                    y=[all_years_data[y].get("sex_bidrag", {}).get(sex, 0) for y in years_sorted],
                    name=sex_display.get(sex, sex), mode="lines+markers",
                    line=dict(color=sex_colors.get(sex, "#3d3d3d"), width=2), marker=dict(size=8)
                ))
            fig_sb.add_trace(go.Scatter(
                x=years_sorted,
                y=[sum(all_years_data[y].get("sex_bidrag", {}).get(sex, 0) for sex in _all_sexes) for y in years_sorted],
                name="Total", mode="lines+markers",
                line=dict(color="#3d3d3d", width=2, dash="dot"), marker=dict(size=8),
            ))
            fig_sb.update_layout(
                xaxis=dict(tickmode="array", tickvals=years_sorted, dtick=1),
                yaxis_title="Forfatterantal", legend_title="Køn", height=500, margin=dict(t=50),
                title=f"Forfatterantal per køn over tid, {years_sorted[0]}–{years_sorted[-1]}"
            )
            st.plotly_chart(fig_sb, width="stretch",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}})
            if filter_caption:
                st.caption(filter_caption)
            _tbl_bidrag = (
                [{"Køn": sex_display.get(sex, sex),
                  **{str(y): all_years_data[y].get("sex_bidrag", {}).get(sex, 0) for y in years_sorted}}
                 for sex in _all_sexes] +
                [{"Køn": "Total",
                  **{str(y): sum(all_years_data[y].get("sex_bidrag", {}).get(sex, 0) for sex in _all_sexes) for y in years_sorted}}]
            )
            _tbl_bidrag_schema = [("Køn", pa.string())] + [(str(y), pa.int64()) for y in years_sorted]
            with st.expander("Se tabel"):
                st.dataframe(build_table(_tbl_bidrag, _tbl_bidrag_schema), hide_index=True, width="stretch")
                st.download_button("Download (.xlsx)",
                    data=rows_to_excel_bytes(_tbl_bidrag, [n for n, _ in _tbl_bidrag_schema]),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_bidrag_tid_{year}_{mode}")

        with _tab_combo:
            _use_combo_all_tid = st.toggle(
                "Inkludér intra-enhed forfatterpar (f.eks. to professorer på samme institut)",
                value=False,
                key=f"combo_all_tid_{year}_{mode}",
                help="Til: alle forfatterpar inkl. par inden for samme enhed. Fra: kun netværkskanter.",
                disabled=not show_intra
            )
            _combo_key = "combo_pubs_all" if _use_combo_all_tid else "combo_pubs"
            if _use_combo_all_tid:
                st.caption("Alle forfatterpar inkl. par inden for samme enhed.")
            else:
                st.caption("Kun forfatterpar *mellem* de valgte organisatoriske enheder - som vist i netværket.")
            _all_combos_tid = sorted({c for snap in all_years_data.values() for c in snap.get(_combo_key, {})})
            fig_combo_tid = go.Figure()
            for combo in _all_combos_tid:
                fig_combo_tid.add_trace(go.Scatter(
                    x=years_sorted,
                    y=[all_years_data[y].get(_combo_key, {}).get(combo, 0) for y in years_sorted],
                    name=combo_display.get(combo, combo), mode="lines+markers",
                    line=dict(color=combo_colors.get(combo, "#3d3d3d"), width=2), marker=dict(size=8),
                ))
            fig_combo_tid.add_trace(go.Scatter(
                x=years_sorted,
                y=[sum(all_years_data[y].get(_combo_key, {}).get(c, 0) for c in _all_combos_tid) for y in years_sorted],
                name="Total", mode="lines+markers",
                line=dict(color="#3d3d3d", width=2, dash="dot"), marker=dict(size=8),
            ))
            fig_combo_tid.update_layout(
                xaxis=dict(tickmode="array", tickvals=years_sorted, dtick=1),
                yaxis_title="Antal forfatterpar", legend_title="Kønskombination",
                height=320, margin=dict(t=50),
                title=f"Forfatterpar per kønskombination{' (inkl. intra-enhed)' if _use_combo_all_tid else ''} over tid, {years_sorted[0]}–{years_sorted[-1]}"
            )
            st.plotly_chart(fig_combo_tid, width='content', key=f"combo_tid_{year}_{mode}",
                config={"toImageButtonOptions": {"format": "png", "scale": 3}})
            if filter_caption:
                st.caption(filter_caption)
            _tbl_combo = (
                [{"Kønskombination": combo_display.get(combo, combo),
                  **{str(y): all_years_data[y].get(_combo_key, {}).get(combo, 0) for y in years_sorted}}
                 for combo in _all_combos_tid] +
                [{"Kønskombination": "Total",
                  **{str(y): sum(all_years_data[y].get(_combo_key, {}).get(c, 0) for c in _all_combos_tid) for y in years_sorted}}]
            )
            _tbl_combo_schema = [("Kønskombination", pa.string())] + [(str(y), pa.int64()) for y in years_sorted]
            with st.expander("Se tabel"):
                st.dataframe(build_table(_tbl_combo, _tbl_combo_schema), hide_index=True, width="stretch")
                st.download_button("Download (.xlsx)",
                    data=rows_to_excel_bytes(_tbl_combo, [n for n, _ in _tbl_combo_schema]),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_combo_tid_{year}_{mode}")

        with _tab_homofili:
            st.markdown(
                "Homofili-indeks per køn per år - KU samlet. "
                "Et indeks > 1 betyder at kønsgruppen publicerer med samme køn oftere end "
                "den samlede kønsfordeling ville forudsige."
            )
            _hom_years = [y for y in years_sorted if all_years_data[y].get("hom_index_ku")]
            _klub_years = [y for y in _hom_years
                           if all_years_data[y]["hom_index_ku"].get("k")
                           and all_years_data[y]["hom_index_ku"].get("m")]
            _klub_vals = [
                round(all_years_data[y]["hom_index_ku"]["m"] /
                      all_years_data[y]["hom_index_ku"]["k"], 3)
                for y in _klub_years]
            fig_hom_tid = go.Figure()
            for sex, label, color in [("k", "Kvinder", "#901a1E"), ("m", "Mænd", "#425570")]:
                fig_hom_tid.add_trace(go.Scatter(
                    x=_hom_years,
                    y=[all_years_data[y]["hom_index_ku"].get(sex) for y in _hom_years],
                    name=label, mode="lines+markers",
                    line=dict(color=color, width=2), marker=dict(size=8),
                ))
            fig_hom_tid.add_trace(go.Scatter(
                x=_klub_years, y=_klub_vals,
                name="Klubfaktor (M/K)", mode="lines+markers",
                line=dict(color="#122947", width=2), marker=dict(size=8),
            ))
            fig_hom_tid.add_hline(y=1.0, line_dash="dash", line_color="#3d3d3d",
                annotation_text="Ingen homofili (= 1)", annotation_position="bottom right")
            fig_hom_tid.update_layout(
                xaxis=dict(tickmode="array", tickvals=_hom_years, dtick=1),
                yaxis_title="Homofili-indeks", legend_title="Køn",
                height=360, margin=dict(t=50),
                title=f"Homofili og klubfaktor over tid, {years_sorted[0]}–{years_sorted[-1]}"
            )
            st.plotly_chart(fig_hom_tid, width='content',
                config={"toImageButtonOptions": {"format": "png", "scale": 3}})
            if filter_caption:
                st.caption(filter_caption)
            _tbl_hom = (
                [{"Metrik": label,
                  **{str(y): round(all_years_data[y]["hom_index_ku"].get(sex) or 0, 3) if y in _hom_years else None
                     for y in years_sorted}}
                 for sex, label, _ in [("k", "Kvinder (homofili)", None), ("m", "Mænd (homofili)", None)]] +
                [{"Metrik": "Klubfaktor (M/K)",
                  **{str(y): round(all_years_data[y]["hom_index_ku"]["m"] / all_years_data[y]["hom_index_ku"]["k"], 3)
                     if y in _klub_years else None
                     for y in years_sorted}}]
            )
            _tbl_hom_schema = [("Metrik", pa.string())] + [(str(y), pa.float64()) for y in years_sorted]
            with st.expander("Se tabel"):
                st.dataframe(build_table(_tbl_hom, _tbl_hom_schema), hide_index=True, width="stretch")
                st.download_button("Download (.xlsx)",
                    data=rows_to_excel_bytes(_tbl_hom, [n for n, _ in _tbl_hom_schema]),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_hom_tid_{year}_{mode}")





# ===========================================================================
# NATIONALITETER TAB
# ===========================================================================
def render_tab_nationaliteter(year, mode, raw_nodes, raw_edges, node_meta,
                               selected_facs, selected_insts, selected_grps,
                               raw_nodes_unfiltered=None, raw_edges_unfiltered=None,
                               all_years_data=None, cs_pairs=None, cs_nat_pairs=None,
                               raw_nodes_pre_cs=None, selected_citizenships=None, show_intra=True, edges_keep=None, filter_caption = None):

    st.subheader("Statsborgerskab blandt KU-VIP-forfattere")
    st.markdown(
f"""
Fanen viser fordelingen af statsborgerskab blandt de VIP-forfattere, der er ansat på KU og indgår
i sampubliceringer. Det er altså ikke et mål for internationalt samarbejde, men derimod for den
nationale sammensætning af KU's VIP-personale. Statsborgerskab er bestemt ud fra HR-data.

Fanen opererer med to opgørelser af forfatterpar - ligesom Køn-fanen:

- **Netværkskanter** er forfatterpar *mellem* forskellige organisatoriske enheder. Det er disse 
  par der vises i netværksvisningen.
- **Alle forfatterpar, inkl. intra-enhed** tæller derudover par inden for samme enhed - f.eks. 
  to danske professorer på Kemisk Institut. Denne opgørelse bruges i nationalitetshomofili-beregningen.

**Fortolkningsnote**: analyserne i denne fane beskriver sampubliceringsmønstre - ikke præferencer eller bevidste valg hos
den enkelte forsker. Mønstre i nationalitet kan i høj grad afspejle faglige miljøer, organisering og resrutteringshistorik. 

Resultaterne bør derfor tolkes som **strukturelle mønstre**, ikke individuelle adfærdsmæssige forklaringer. 
""")

    _nat_level_options = []
    if fac_in_mode(mode):  _nat_level_options.append("Fakulteter")
    if inst_in_mode(mode): _nat_level_options.append("Institutter")
    if grp_in_mode(mode):  _nat_level_options.append("Stillingsgrupper")

    _nat_default_level = (
        ["Fakulteter"] if fac_in_mode(mode) else
        ["Institutter"] if inst_in_mode(mode) else
        ["Stillingsgrupper"]
    )
    _nat_selected_levels = st.multiselect(
        "**Vis nationalitetsfordeling for:**",
        options=_nat_level_options,
        default=_nat_default_level,
        key=f"nat_levels_{year}_{mode}",
    )
    if not _nat_selected_levels:
        st.error("Vælg mindst ét organisatorisk niveau ovenfor.")
        return

    group_label = " - ".join(
        {"Fakulteter": "Fakultet", "Institutter": "Institut", "Stillingsgrupper": "Stillingsgruppe"}[l]
        for l in ["Fakulteter", "Institutter", "Stillingsgrupper"] if l in _nat_selected_levels
    )
    # Behold group_key til bagudkompatibilitet i kode der kun bruger ét niveau
    group_key = ("fac"  if _nat_selected_levels == ["Fakulteter"] else
                 "inst" if _nat_selected_levels == ["Institutter"] else
                 "grp"  if _nat_selected_levels == ["Stillingsgrupper"] else "fac")

    _edges = edges_keep if edges_keep is not None else raw_edges
    
    # ── Lokale filtre ─────────────────────────────────────────────────────────
    _all_facs_nat  = sorted({m.get("fac","")  for m in node_meta.values() if m.get("fac")})
    _all_insts_nat = sorted({m.get("inst","") for m in node_meta.values() if m.get("inst")})
    _all_grps_nat  = sorted({m.get("grp","")  for m in node_meta.values() if m.get("grp")},
                             key=lambda g: HIERARKI.get(g, 999))

    _show_fac  = bool(_all_facs_nat)  and "Fakulteter"       in _nat_selected_levels
    _show_inst = bool(_all_insts_nat) and "Institutter"      in _nat_selected_levels
    _show_grp  = bool(_all_grps_nat)  and "Stillingsgrupper" in _nat_selected_levels

    _nat_fac_filter = st.multiselect(
        "**Filtrer på fakultet**", options=_all_facs_nat if _show_fac else [],
        default=[], key=f"nat_fac_{year}_{mode}",
        placeholder="Alle fakulteter", disabled=not _show_fac,
    ) if _show_fac else []

    _nat_inst_filter = st.multiselect(
        "**Filtrer på institut**",
        options=[i for i in _all_insts_nat
                 if not _nat_fac_filter or any(
                     node_meta[nid].get("fac") in _nat_fac_filter
                     for nid in node_meta if node_meta[nid].get("inst") == i
                 )] if _show_inst else [],
        default=[], key=f"nat_inst_{year}_{mode}",
        placeholder="Alle institutter", disabled=not _show_inst,
    ) if _show_inst else []

    _nat_grp_filter = st.multiselect(
        "**Filtrer på stillingsgruppe**", options=_all_grps_nat if _show_grp else [],
        default=[], key=f"nat_grp_{year}_{mode}",
        placeholder="Alle stillingsgrupper", disabled=not _show_grp,
    ) if _show_grp else []

    _nat_n_filters = len(_nat_fac_filter) + len(_nat_inst_filter) + len(_nat_grp_filter)
    _hide_nat_ku_samlet = len(_nat_selected_levels) > 1 or _nat_n_filters >= 1

    def _nat_node_ok(m: dict) -> bool:
        if _nat_fac_filter  and m.get("fac","")  not in _nat_fac_filter:  return False
        if _nat_inst_filter and m.get("inst","") not in _nat_inst_filter: return False
        if _nat_grp_filter  and m.get("grp","")  not in _nat_grp_filter:  return False
        return True

    def _nat_nid_ok(nid: str) -> bool:
        parts = nid.split("|")
        if len(parts) < 3: return True
        if _nat_fac_filter  and parts[0] not in _nat_fac_filter:  return False
        if _nat_inst_filter and parts[1] not in _nat_inst_filter: return False
        if _nat_grp_filter  and parts[2] not in _nat_grp_filter:  return False
        return True
    
    def _nat_org_key(m: dict) -> str:
        parts = []
        if "Fakulteter"       in _nat_selected_levels: parts.append(m.get("fac",  "") or "")
        if "Institutter"      in _nat_selected_levels: parts.append(m.get("inst", "") or "")
        if "Stillingsgrupper" in _nat_selected_levels: parts.append(m.get("grp",  "") or "")
        return " - ".join(p for p in parts if p)

     # ── Faculty-gruppering for institutter (samme princip som Køn-fanen) ─────
    _inst_to_fac_local: dict[str, str] = {}
    for _m in node_meta.values():
        _i = _m.get("inst")
        _f = _m.get("fac")
        if not _i or not _f:
            continue
        # Normalisér til forkortelse hvis vi får fuldt fakultetsnavn
        _f_norm = FAC_ABBRS.get(_f, _f)
        # Foretræk gyldige (i FAC_ORDER) værdier; overskriv ugyldige
        if _i not in _inst_to_fac_local or _inst_to_fac_local[_i] not in FAC_ORDER:
            _inst_to_fac_local[_i] = _f_norm

    def _nat_fac_rank(fac: str) -> int:
        return FAC_ORDER.index(fac) if fac in FAC_ORDER else 999

    def _nat_parts_of(s: str) -> tuple[str, str, str]:
        fac = inst = grp = ""
        rest = s

        if "Fakulteter" in _nat_selected_levels:
            head, sep, tail = rest.partition(" - ")
            fac = head
            rest = tail if sep else ""

        if "Institutter" in _nat_selected_levels:
            matched = ""
            for cand in _inst_to_fac_local:
                if (rest == cand or rest.startswith(cand + " - ")) and len(cand) > len(matched):
                    matched = cand
            if matched:
                inst = matched
                rest = rest[len(matched):]
                if rest.startswith(" - "):
                    rest = rest[3:]
            else:
                head, sep, tail = rest.partition(" - ")
                inst = head
                rest = tail if sep else ""

        if "Stillingsgrupper" in _nat_selected_levels:
            grp = rest

        if not fac and inst:
            fac = _inst_to_fac_local.get(inst, "")
        return fac, inst, grp

    def _nat_sort_key(s: str) -> tuple:
        fac, inst, grp = _nat_parts_of(s)
        return (_nat_fac_rank(fac), fac, inst, HIERARKI.get(grp, 999))

    def _nat_add_separators(keys: list) -> list:
        if not keys:
            return keys
        _multi_level = sum(l in _nat_selected_levels for l in
                        ("Fakulteter", "Institutter", "Stillingsgrupper")) > 1
        if not _multi_level:
            return keys
        result = []
        prev_fac = None
        _sep_count = [0]
        def _sep():
            _sep_count[0] += 1
            return " " * _sep_count[0]
        for k in keys:
            fac, _inst, _grp = _nat_parts_of(k)
            if prev_fac is not None and fac != prev_fac:
                result.append(_sep())
            result.append(k)
            prev_fac = fac
        return result

    def _build_nat_y_positions(keys: list) -> tuple[list, list, list]:
        _multi_level = sum(l in _nat_selected_levels for l in
                        ("Fakulteter", "Institutter", "Stillingsgrupper")) > 1
        INTRA_GAP = 0.6
        INTER_GAP = 0.95
        _inst_in = "Institutter"      in _nat_selected_levels
        _grp_in  = "Stillingsgrupper" in _nat_selected_levels
        if not _multi_level:
            positions = [i * INTRA_GAP for i in range(len(keys))]
            return positions, positions, list(keys)
        positions = []
        labels = []
        prev_fac = None
        y = 0.0
        for k in keys:
            fac, inst, grp = _nat_parts_of(k)
            if prev_fac and fac and fac != prev_fac:
                y += INTER_GAP
            else:
                y += INTRA_GAP if positions else 0
            positions.append(y)
            _fac_in = "Fakulteter" in _nat_selected_levels
            if _inst_in and _grp_in:
                labels.append(f"{inst} | {grp}" if (inst and grp) else (inst or grp or k))
            elif _fac_in and _grp_in:
                labels.append(f"{fac} | {grp}" if (fac and grp) else (fac or grp or k))
            elif _grp_in:
                labels.append(grp or k)
            elif _inst_in:
                labels.append(inst or k)
            else:
                labels.append(fac or k)
            if fac:
                prev_fac = fac
        return positions, positions, labels

    _NAT_BAR_W = 0.55

    def _nat_bar_layout(plot_keys: list, *, extra_top: int = 50, extra_bottom: int = 20,
                    grouped: bool = False) -> dict:
        """Returnér ensartet height/margin/bargap for horisontale bar charts.
        Sæt grouped=True for figurer med flere traces side-om-side (barmode='group'),
        så hver række får ekstra lodret plads."""
        n_rows = len(plot_keys)
        n_data = len(plot_keys)
        # Grupperede figurer skal have ~75% mere lodret plads pr. række
        row_h = int(28 * (1.75 if grouped else 1.0))
        height = int(row_h * n_data + extra_top + extra_bottom + 30)
        height = max(280, height)
        return dict(
            height=height,
            bargap=0.05,
            bargroupgap=0.0,
            margin=dict(l=20, r=30, t=extra_top, b=extra_bottom),
        )

    _raw_nodes_all = raw_nodes_pre_cs if raw_nodes_pre_cs else raw_nodes
    _raw_nodes_f = {nid: m for nid, m in _raw_nodes_all.items() if _nat_node_ok(m)}

    def _org_from_node_id(nid: str) -> str:
        parts = nid.split("|")
        keys = []
        if "Fakulteter"       in _nat_selected_levels: keys.append(parts[0] if len(parts) > 0 else "")
        if "Institutter"      in _nat_selected_levels: keys.append(parts[1] if len(parts) > 1 else "")
        if "Stillingsgrupper" in _nat_selected_levels: keys.append(parts[2] if len(parts) > 2 else "")
        return " - ".join(k for k in keys if k) or "ukendt"

    cs_size:  dict[str, dict[str, int]] = {}
    cs_total: dict[str, int] = {}
    for nid, m in _raw_nodes_f.items():
        if m.get("type") != "grp": continue
        org = _nat_org_key(m)
        cs  = m.get("statsborgerskab", "Ukendt") or "Ukendt"
        size = m.get("size", 0)
        if not org: continue
        cs_size.setdefault(org, {})
        cs_size[org][cs] = cs_size[org].get(cs, 0) + size
        cs_total[cs]     = cs_total.get(cs, 0) + size

    groups_sorted = sorted(cs_size.keys(), key=_nat_sort_key)

    # KU-samlet række til brug i plots (ikke tabellen)
    _ku_label      = "KU samlet"


    # ── Top-N nationaliteter (overall) ───────────────────────────────────────
    all_cs_sorted = sorted(cs_total.items(), key=lambda x: -x[1])
    max_n         = len(all_cs_sorted)
    if max_n == 0:
        st.error("Ingen nationalitetsdata tilgængelig for det valgte udsnit.")
        return

    if selected_citizenships:
        top_cs   = [c for c in selected_citizenships if c in cs_total]
        top_vals = [cs_total.get(c, 0) for c in top_cs]
    else:
        top_cs   = [c for c, _ in all_cs_sorted]
        top_vals = [v for _, v in all_cs_sorted]

    other_cs = [c for c, _ in all_cs_sorted if c not in top_cs]
    
    if other_cs:
        # Aggregér ikke-top-N til "Andre" i cs_size
        for org in cs_size:
            andre_sum = sum(cs_size[org].get(c, 0) for c in other_cs)
            if andre_sum > 0:
                cs_size[org]["Andre"] = andre_sum
        # Fjern de individuelle ikke-top-N fra cs_size
        for org in cs_size:
            for c in other_cs:
                cs_size[org].pop(c, None)
        # Opdater cs_total
        cs_total["Andre"] = sum(cs_total.pop(c, 0) for c in other_cs)
        # Tilføj "Andre" sidst i top_cs
        top_cs = top_cs + ["Andre"]
        top_vals = top_vals + [cs_total.get("Andre", 0)]
        # Opdater _ku_cs
    
    _ku_cs: dict[str, int] = {}
    for d in cs_size.values():
        for cs_k, v in d.items():
            _ku_cs[cs_k] = _ku_cs.get(cs_k, 0) + v
    _cs_size_plot  = ({} if _hide_nat_ku_samlet else {_ku_label: _ku_cs})
    for _g in groups_sorted:
        _cs_size_plot[_g] = cs_size.get(_g, {})
    _groups_plot   = ([] if _hide_nat_ku_samlet else [_ku_label]) + groups_sorted

    st.markdown("---")

    st.markdown(f"#### Forfattere fordelt på nationalitet i {year}")
    st.markdown(
f"""
Figuren nedenfor viser fordelingen af statsborgerskab blandt KU's VIP-forfattere i {year}, 
opdelt på den valgte organisatoriske enhed. Hver forfatter tælles én gang uanset antal publikationer. 
Nationaliteter uden for de udvalgte grupperes under **Andre**.

Brug feltet i sidepanelet under **Diversitet** til at justere, hvor mange nationaliteter
der inkluderes i analyserne - de øvrige samles automatisk i Andre-kategorien og vises ikke
enkeltvis i netværket.
""")

    _tab_abs, _tab_pct = st.tabs(["Antal", "Andel (%)"])

    nat_colors = ku_color_sequence(len(top_cs) + 1)

    _nat_y_pos, _nat_tick_pos, _nat_tick_labels = _build_nat_y_positions(_groups_plot)

    def _nat_traces(normalise = False):
        traces = []
        for i, cs in enumerate(top_cs):
            raw_y = [_cs_size_plot[g].get(cs, 0) for g in _groups_plot]
            y = ([round(100 * v / (sum(_cs_size_plot[g].values()) or 1), 1)
                  for g, v in zip(_groups_plot, raw_y)]
                 if normalise else raw_y)
            color = "#cccccc" if cs == "Andre" else nat_colors[i]
            traces.append(go.Bar(
                name=country_name(cs) if cs != "Andre" else "Andre",
                y=_nat_y_pos,
                x=y,
                orientation="h",
                marker_color=color,
                text=[f"{v}%" for v in y] if normalise else y,
                textposition="inside",
                width=_NAT_BAR_W,
            ))
        return traces

    with _tab_abs:
        fig1 = go.Figure(_nat_traces(normalise=False))
        fig1.update_layout(
            barmode="stack",
            xaxis_title="Forfatterantal",
            yaxis=dict(
                tickmode="array",
                tickvals=_nat_tick_pos,
                ticktext=_nat_tick_labels,
                autorange="reversed",
                showgrid=False,
                zeroline=False,
            ),
            legend_title="Statsborgerskab",
            title=f"Forfatterantal fordelt på nationalitet, {year}",
            **_nat_bar_layout(_groups_plot),
        )
        st.plotly_chart(fig1, width="stretch",
            config={"toImageButtonOptions": {"format": "png", "scale": 3}})
        if filter_caption:
            st.caption(filter_caption)
        _rows_abs = []
        for g in groups_sorted:
            row = {group_label: g}
            for cs in top_cs:
                row[country_name(cs) if cs != "Andre" else "Andre"] = cs_size[g].get(cs, 0)
            row["Andre"] = sum(v for k, v in cs_size[g].items() if k not in top_cs)
            row["Total"] = sum(cs_size[g].values())
            _rows_abs.append(row)
        _schema_abs = (
            [(group_label, pa.string())] +
            [(country_name(cs) if cs != "Andre" else "Andre", pa.int64()) for cs in top_cs] +
            [("Andre", pa.int64()), ("Total", pa.int64())]
        )
        with st.expander("Se tabel"):
            st.dataframe(build_table(_rows_abs, _schema_abs), hide_index=True, width="stretch")
            st.download_button("Download (.xlsx)",
                data=rows_to_excel_bytes(_rows_abs, [n for n, _ in _schema_abs]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"nat_bidrag_abs_{year}_{mode}")

    with _tab_pct:
        fig1p = go.Figure(_nat_traces(normalise=True))
        fig1p.update_layout(
            barmode="stack",
            yaxis=dict(
                tickmode="array",
                tickvals=_nat_tick_pos,
                ticktext=_nat_tick_labels,
                autorange="reversed",
                showgrid=False,
                zeroline=False,
            ),
            xaxis_title="Andel (%)",
            xaxis_range=[0, 100],
            legend_title="Statsborgerskab",
            title=f"Forfatterantal fordelt på nationalitet - andel (%), {year}",
            **_nat_bar_layout(_groups_plot),
        )
        st.plotly_chart(fig1p, width="stretch",
            config={"toImageButtonOptions": {"format": "png", "scale": 3}})
        if filter_caption:
            st.caption(filter_caption)
        _rows_pct = []
        for g in groups_sorted:
            total = sum(cs_size[g].values()) or 1
            row = {group_label: g}
            for cs in top_cs:
                row[f"{country_name(cs) if cs != 'Andre' else 'Andre'} (%)"] = round(100 * cs_size[g].get(cs, 0) / total, 1)
            row["Andre (%)"] = round(100 * sum(v for k, v in cs_size[g].items() if k not in top_cs) / total, 1)
            _rows_pct.append(row)
        _schema_pct = (
            [(group_label, pa.string())] +
            [(f"{country_name(cs) if cs != 'Andre' else 'Andre'} (%)", pa.float64()) for cs in top_cs] +
            [("Andre (%)", pa.float64())]
        )
        with st.expander("Se tabel"):
            st.dataframe(build_table(_rows_pct, _schema_pct), hide_index=True, width="stretch")
            st.download_button("Download (.xlsx)",
                data=rows_to_excel_bytes(_rows_pct, [n for n, _ in _schema_pct]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"nat_bidrag_pct_{year}_{mode}")
    
    st.markdown("---")

    # ── Forfatterpar fordelt på nationalitet ─────────────────────────────────
    st.markdown(f"#### Forfatterpar fordelt på nationalitet i {year}")

    st.markdown(
f"""
Figuren viser, hvor mange forfatterpar hvert statsborgerskab bidrager til - dvs. hvor mange
gange en forfatter med et givet statsborgerskab indgår i et forfatterpar med en anden KU-VIP-forfatter.
En forfatter kan bidrage til mange par på én gang, f.eks. hvis en artikel har fem forfattere.

Brug togglen nedenfor til at vælge, om kun netværkskanter (par *mellem* enheder) eller alle
forfatterpar inkl. intra-enhed skal indgå.
""")

    # Beregn forfatterpar per nationalitet fra raw_edges
    # Hvert forfatterpar tæller med hos begge endepunkters nationalitet (÷2 for ikke at dobbelt-tælle)
    cs_ew: dict[str, float] = {}
    for edge in _edges:
        u, v, w = edge[0], edge[1], edge[2]
        m_u = raw_nodes.get(u, {})
        m_v = raw_nodes.get(v, {})
        if not _nat_node_ok(m_u) and not _nat_node_ok(m_v):
            continue
        cs_u = m_u.get("statsborgerskab", "Ukendt") or "Ukendt"
        cs_v = m_v.get("statsborgerskab", "Ukendt") or "Ukendt"
        if _nat_node_ok(m_u):
            cs_ew[cs_u] = cs_ew.get(cs_u, 0.0) + w
        if _nat_node_ok(m_v):
            cs_ew[cs_v] = cs_ew.get(cs_v, 0.0) + w

    _ew_top_sorted = sorted(cs_ew.items(), key=lambda x: -x[1])
    _ew_top_cs     = [c for c, _ in _ew_top_sorted if c in top_cs]  # samme top-N som ovenfor
    _ew_andre      = sum(v for c, v in _ew_top_sorted if c not in top_cs)
    _ew_grand      = (sum(cs_ew.values()) / 2) or 1

    def _ew_traces(normalise=False):
        traces = []
        for i, cs in enumerate(_ew_top_cs):
            v = cs_ew.get(cs, 0.0)
            x = round(100 * v / _ew_grand, 1) if normalise else round(v, 1)
            traces.append(go.Bar(
                name=country_name(cs) if cs != "Andre" else "Andre",
                y=[cs],
                x=[x],
                orientation="h",
                marker_color=nat_colors[_ew_top_cs.index(cs) % len(nat_colors)],
                text=[f"{x}%" if normalise else f"{fmt_ui(x)}"],
                textposition="inside",
            ))
        if _ew_andre > 0:
            x = round(100 * _ew_andre / _ew_grand, 1) if normalise else round(_ew_andre, 1)
            traces.append(go.Bar(
                name="Andre",
                y=["Andre"],
                x=[x],
                orientation="h",
                marker_color="#cccccc",
                text=[f"{x}%" if normalise else f"{fmt_ui(x)}"],
                textposition="inside",
            ))
        return traces

    # Brug samme struktur som forfatterantal-plottet: grouped bars per org-enhed
    cs_ew_by_org: dict[str, dict[str, float]] = {}
    for edge in _edges:
        u, v, w = edge[0], edge[1], edge[2]
        m_u = raw_nodes.get(u, {})
        m_v = raw_nodes.get(v, {})
        cs_u = m_u.get("statsborgerskab", "Ukendt") or "Ukendt"
        cs_v = m_v.get("statsborgerskab", "Ukendt") or "Ukendt"
        org_u = _org_from_node_id(u)
        org_v = _org_from_node_id(v)
        if _nat_node_ok(m_u) and org_u in cs_size:
            cs_ew_by_org.setdefault(org_u, {})
            cs_ew_by_org[org_u][cs_u] = cs_ew_by_org[org_u].get(cs_u, 0.0) + w 
        if _nat_node_ok(m_v) and org_v in cs_size:
            cs_ew_by_org.setdefault(org_v, {})
            cs_ew_by_org[org_v][cs_v] = cs_ew_by_org[org_v].get(cs_v, 0.0) + w 

    # KU-samlet række for forfatterpar
    _ku_ew: dict[str, float] = {}
    for d in cs_ew_by_org.values():
        for cs_k, v in d.items():
            _ku_ew[cs_k] = _ku_ew.get(cs_k, 0.0) + v
    _ku_ew_label       = "KU samlet"
    _cs_ew_by_org_plot = ({} if _hide_nat_ku_samlet else {_ku_ew_label: _ku_ew})
    _cs_ew_by_org_plot.update(cs_ew_by_org)
    _groups_ew_plot    = ([] if _hide_nat_ku_samlet else [_ku_ew_label]) + groups_sorted

    def _ew_org_traces(normalise=False):
        traces = []
        for i, cs in enumerate(top_cs):
            raw_y = [_cs_ew_by_org_plot.get(g, {}).get(cs, 0.0) for g in _groups_ew_plot]
            y = ([round(100 * v / (sum(_cs_ew_by_org_plot.get(g, {}).values()) or 1), 1)
                  for g, v in zip(_groups_ew_plot, raw_y)]
                 if normalise else [round(v, 1) for v in raw_y])
            traces.append(go.Bar(
                name=cs,
                y=_groups_ew_plot,
                x=y,
                orientation="h",
                marker_color=nat_colors[i],
                text=[f"{v}%" for v in y] if normalise else [f"{fmt_ui(v)}" for v in y],
                textposition="inside",
            ))
        andre_raw = [sum(v for k, v in _cs_ew_by_org_plot.get(g, {}).items() if k not in top_cs)
                     for g in _groups_ew_plot]
        if any(v > 0 for v in andre_raw):
            andre_y = ([round(100 * v / (sum(_cs_ew_by_org_plot.get(g, {}).values()) or 1), 1)
                        for g, v in zip(_groups_ew_plot, andre_raw)]
                       if normalise else [round(v, 1) for v in andre_raw])
            traces.append(go.Bar(
                name="Andre",
                y=_groups_ew_plot,
                x=andre_y,
                orientation="h",
                marker_color="#cccccc",
                text=[f"{v}%" for v in andre_y] if normalise else [f"{fmt_ui(v)}" for v in andre_y],
                textposition="inside",
            ))
        return traces

    _use_all_ew = st.toggle(
        "Inkludér intra-enhed forfatterpar (f.eks. to professorer på samme institut)",
        value=False,
        key=f"nat_ew_all_{year}_{mode}",
        help="Til: alle forfatterpar inkl. par inden for samme enhed. Fra: kun netværkskanter.",
        disabled=not show_intra
    )
    if _use_all_ew:
        st.caption("Alle forfatterpar inkl. par inden for samme enhed (f.eks. to mandlige professorer på Kemisk Institut).")
    else:
        st.caption("Kun forfatterpar *mellem* de valgte organisatoriske enheder - som vist i netværket.")


    # Byg cs_ew_by_org_all fra cs_pairs
    cs_ew_by_org_all: dict[str, dict[str, float]] = {}
    for raw_key, counts in (cs_pairs or {}).items():
        parts = raw_key.split("|")
        if len(parts) != 4: continue
        fac, inst, grp, focal_cs = parts
        if selected_facs  and fac  not in selected_facs:  continue
        if selected_insts and inst not in selected_insts: continue
        if selected_grps  and grp  not in selected_grps:  continue
        if _nat_fac_filter  and fac  not in _nat_fac_filter:  continue
        if _nat_inst_filter and inst not in _nat_inst_filter: continue
        if _nat_grp_filter  and grp  not in _nat_grp_filter:  continue
        org_key_all = _nat_org_key({"fac": fac, "inst": inst, "grp": grp})
        cs_ew_by_org_all.setdefault(org_key_all, {})
        total = counts.get("same", 0) + counts.get("cross", 0)
        cs_ew_by_org_all[org_key_all][focal_cs] = cs_ew_by_org_all[org_key_all].get(focal_cs, 0.0) + total

    _ku_ew_all: dict[str, float] = {}
    for d in cs_ew_by_org_all.values():
        for k, v in d.items():
            _ku_ew_all[k] = _ku_ew_all.get(k, 0.0) + v

    # Aktiv datakilde afhænger af toggle
    _active_ew_by_org = cs_ew_by_org_all if _use_all_ew else cs_ew_by_org
    _active_ku_ew     = _ku_ew_all        if _use_all_ew else _ku_ew
    _active_top_cs    = top_cs  # netværkskanter bruger alle nationaliteter; all bruger DK/intl

    _cs_active_plot = ({} if _hide_nat_ku_samlet else {_ku_ew_label: _active_ku_ew})
    _cs_active_plot.update(_active_ew_by_org)
    _active_sorted = sorted(_active_ew_by_org.keys(), key=_nat_sort_key)
    _groups_active_plot = ([] if _hide_nat_ku_samlet else [_ku_ew_label]) + _active_sorted

    _ew_active_y, _ew_active_tick, _ew_active_labels = _build_nat_y_positions(_groups_active_plot)

    def _ew_active_traces(normalise=False):
        traces = []
        for i, cs in enumerate(top_cs):
            raw_y = [_cs_active_plot.get(g, {}).get(cs, 0.0) for g in _groups_active_plot]
            y = ([round(100*v/(sum(_cs_active_plot.get(g,{}).values()) or 1),1)
                  for g,v in zip(_groups_active_plot, raw_y)]
                 if normalise else [round(v,1) for v in raw_y])
            traces.append(go.Bar(
                name=country_name(cs),
                y=_ew_active_y, 
                x=y, 
                orientation="h",
                marker_color=nat_colors[i], 
                width=_NAT_BAR_W,
                text=[f"{v}%" for v in y] if normalise else [f"{fmt_ui(v)}" for v in y],
                textposition="inside"))
        
        andre_raw = [sum(v for k, v in _cs_active_plot.get(g,{}).items() if k not in top_cs)
                     for g in _groups_active_plot]
        if any(v > 0 for v in andre_raw):
            andre_y = ([round(100*v/(sum(_cs_active_plot.get(g,{}).values()) or 1),1)
                        for g,v in zip(_groups_active_plot, andre_raw)]
                       if normalise else [round(v,1) for v in andre_raw])
            traces.append(go.Bar(name="Andre", y=_ew_active_y, x=andre_y, orientation="h",
                marker_color="#cccccc", width=_NAT_BAR_W,
                text=[f"{v}%" for v in andre_y] if normalise else [f"{fmt_ui(v)}" for v in andre_y],
                textposition="inside"))
        return traces

    _sub_abs, _sub_pct = st.tabs(["Antal", "Andel (%)"])
    with _sub_abs:
        fig_ew = go.Figure(_ew_active_traces(normalise=False))
        fig_ew.update_layout(barmode="stack", xaxis_title="Antal forfatterpar",
            yaxis=dict(
                tickmode="array",
                tickvals=_ew_active_tick,
                ticktext=_ew_active_labels,
                autorange="reversed",
                showgrid=False,
                zeroline=False,
            ),
            legend_title="Statsborgerskab",
            title=f"Forfatterpar fordelt på nationalitet{' (inkl. intra-enhed)' if _use_all_ew else ''}, {year}",
            **_nat_bar_layout(_groups_active_plot))
        st.plotly_chart(fig_ew, width='content', key=f"nat_ew_abs_{year}_{mode}",
            config={"toImageButtonOptions": {"format": "png", "scale": 3}})
        if filter_caption:
            st.caption(filter_caption)
        _ew_rows_abs = []
        for g in _groups_active_plot:
            row = {group_label: g}
            for cs in top_cs:
                row[country_name(cs)] = round(_cs_active_plot.get(g, {}).get(cs, 0.0), 1)
            row["Andre"] = round(sum(v for k, v in _cs_active_plot.get(g, {}).items() if k not in top_cs), 1)
            row["Total"] = round(sum(_cs_active_plot.get(g, {}).values()), 1)
            _ew_rows_abs.append(row)
        _ew_schema_abs = (
            [(group_label, pa.string())] +
            [(country_name(cs) if cs != "Andre" else "Andre", pa.float64()) for cs in top_cs] +
            [("Andre", pa.float64()), ("Total", pa.float64())]
        )
        with st.expander("Se tabel"):
            st.dataframe(build_table(_ew_rows_abs, _ew_schema_abs), hide_index=True, width="stretch")
            st.download_button("Download (.xlsx)",
                data=rows_to_excel_bytes(_ew_rows_abs, [n for n, _ in _ew_schema_abs]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_nat_ew_abs_{year}_{mode}")

    with _sub_pct:
        fig_ewp = go.Figure(_ew_active_traces(normalise=True))
        fig_ewp.update_layout(barmode="stack",
            yaxis=dict(
                tickmode="array",
                tickvals=_ew_active_tick,
                ticktext=_ew_active_labels,
                autorange="reversed",
                showgrid=False,
                zeroline=False,
            ),
            xaxis_title="Andel (%)", xaxis_range=[0,100], legend_title="Statsborgerskab",
            title=f"Forfatterpar fordelt på nationalitet{' (inkl. intra-enhed)' if _use_all_ew else ''} - andel (%), {year}",
            **_nat_bar_layout(_groups_active_plot))
        st.plotly_chart(fig_ewp, width='content', key=f"nat_ew_pct_{year}_{mode}",
            config={"toImageButtonOptions": {"format": "png", "scale": 3}})
        if filter_caption:
            st.caption(filter_caption)
        _ew_rows_pct = []
        for g in _groups_active_plot:
            total = sum(_cs_active_plot.get(g, {}).values()) or 1
            row = {group_label: g}
            for cs in top_cs:
                row[f"{country_name(cs) if cs != 'Andre' else 'Andre'} (%)"] = round(100 * _cs_active_plot.get(g, {}).get(cs, 0.0) / total, 1)
            row["Andre (%)"] = round(100 * sum(v for k, v in _cs_active_plot.get(g, {}).items() if k not in top_cs) / total, 1)
            _ew_rows_pct.append(row)
        _ew_schema_pct = (
            [(group_label, pa.string())] +
            [(f"{country_name(cs) if cs != 'Andre' else 'Andre'} (%)", pa.float64()) for cs in top_cs] +
            [("Andre (%)", pa.float64())]
        )
        with st.expander("Se tabel"):
            st.dataframe(build_table(_ew_rows_pct, _ew_schema_pct), hide_index=True, width="stretch")
            st.download_button("Download (.xlsx)",
                data=rows_to_excel_bytes(_ew_rows_pct, [n for n, _ in _ew_schema_pct]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_nat_ew_pct_{year}_{mode}")

    st.markdown("---")

    # ── Heatmap ───────────────────────────────────────────────────────────────
    st.markdown("#### Sampubliceringer mellem nationaliteter")
    st.markdown(
f"""
Heatmap over sampubliceringsvægt mellem de mest repræsenterede nationaliteter - inklusiv 
par inden for samme enhed (f.eks. to danske professorer på Kemisk Institut). Kun forfatterpar, 
hvor begge forfattere har en kendt nationalitet, medtages."""
    )

    _nid_to_cs: dict[str, str] = {
        nid: (m.get("statsborgerskab", "") or "Ukendt")
        for nid, m in (raw_nodes_pre_cs or raw_nodes).items()
        if m.get("type") == "grp"
    }

    # Tæl totaler per nationalitetskombination
    total_cs_combo: dict[str, int] = {}
    for edge in _edges:
        if not _nat_nid_ok(edge[0]) and not _nat_nid_ok(edge[1]):
            continue
        cs_u = _nid_to_cs.get(edge[0], "Ukendt")
        cs_v = _nid_to_cs.get(edge[1], "Ukendt")
        w    = int(round(edge[2]))
        combo = f"{min(cs_u,cs_v)}-{max(cs_u,cs_v)}"
        total_cs_combo[combo] = total_cs_combo.get(combo, 0) + w

    _use_all_heat = st.toggle(
        "Inkludér intra-enhed forfatterpar (f.eks. to professorer på samme institut)",
        value=False,
        key=f"nat_heat_all_{year}_{mode}",
        help="Til: alle forfatterpar inkl. par inden for samme enhed. Fra: kun netværkskanter.",
        disabled=not show_intra
    )
    if _use_all_heat:
        st.caption("Alle forfatterpar inkl. par inden for samme enhed (f.eks. to mandlige professorer på Kemisk Institut).")
    else:
        st.caption("Kun forfatterpar *mellem* de valgte organisatoriske enheder - som vist i netværket.")


    _cs_matrix: dict[tuple, float] = {}

    # Inter-enhed kanter fra raw_edges (altid med)
    for edge in _edges:
        u, v, w = edge[0], edge[1], edge[2]
        if not _nat_nid_ok(u) and not _nat_nid_ok(v):
            continue
        cs_u = _nid_to_cs.get(u, "")
        cs_v = _nid_to_cs.get(v, "")
        if not cs_u or not cs_v:
            continue
        cs_u_bin = cs_u if cs_u in top_cs else "Andre"
        cs_v_bin = cs_v if cs_v in top_cs else "Andre"
        for a, b in [(cs_u_bin, cs_v_bin), (cs_v_bin, cs_u_bin)]:
            _cs_matrix[(a, b)] = _cs_matrix.get((a, b), 0.0) + w

    # Intra-enhed kanter fra cs_nat_pairs (kun når toggle er slået til)
    if _use_all_heat:
        for raw_key, partner_counts in (cs_nat_pairs or {}).items():
            parts = raw_key.split("|")
            if len(parts) != 4: continue
            fac, inst, grp, focal_cs = parts
            if selected_facs  and fac  not in selected_facs:  continue
            if selected_insts and inst not in selected_insts: continue
            if selected_grps  and grp  not in selected_grps:  continue
            if _nat_fac_filter  and fac  not in _nat_fac_filter:  continue
            if _nat_inst_filter and inst not in _nat_inst_filter: continue
            if _nat_grp_filter  and grp  not in _nat_grp_filter:  continue
            focal_bin = focal_cs if focal_cs in top_cs else "Andre"
            for partner_cs, count in partner_counts.items():
                partner_bin = partner_cs if partner_cs in top_cs else "Andre"
                _cs_matrix[(focal_bin, partner_bin)] = _cs_matrix.get((focal_bin, partner_bin), 0.0) + count / 2

    _heat_labels = [cs for cs in top_cs if cs != "Andre"]
    if any(k[0] == "Andre" or k[1] == "Andre" for k in _cs_matrix):
        _heat_labels = _heat_labels + ["Andre"]

    if _cs_matrix:
        _z    = [[_cs_matrix.get((a, b), 0.0) for b in _heat_labels] for a in _heat_labels]
        _text = [[f"{fmt_ui(_cs_matrix.get((a,b),0),0)}" for b in _heat_labels] for a in _heat_labels]
        _fig_heat = go.Figure(go.Heatmap(
            z=_z, x=_heat_labels, y=_heat_labels,
            colorscale=[[0, "#f0f4f8"], [1, "#122947"]],
            text=_text, texttemplate="%{text}",
            hovertemplate="%{y} → %{x}: %{z:.0f} forfatterpar<extra></extra>",
            colorbar=dict(title="Antal forfatterpar"),
        ))
        _fig_heat.update_layout(
            xaxis_title="Statsborgerskab", yaxis_title="Statsborgerskab",
            height=max(400, 45 * len(_heat_labels)),
            margin=dict(l=80, b=100, t=50, r=20),
            title=f"Sampubliceringer mellem nationaliteter{' (inkl. intra-enhed)' if _use_all_heat else ''}, {year}"
        )
        st.plotly_chart(_fig_heat, width="content",
            config={"toImageButtonOptions": {"format": "png", "scale": 3}})
        if filter_caption:
            st.caption(filter_caption)

        _pair_rows = []
        seen_pairs = set()
        for (a, b), w in sorted(_cs_matrix.items(), key=lambda x: -x[1]):
            key = tuple(sorted((a, b)))
            if key in seen_pairs:
                continue
            seen_pairs.add(key)
            _pair_rows.append({
                "Nationalitet A": country_name(a), "Nationalitet B": country_name(b),
                "Forfatterpar": round(w, 1),
                "Type": "Intra" if a == b else "Inter",
            })
        _pair_schema = [
            ("Nationalitet A", pa.string()), ("Nationalitet B", pa.string()),
            ("Forfatterpar", pa.float64()), ("Type", pa.string()),
        ]
        with st.expander("Se tabel"):
            st.dataframe(build_table(_pair_rows[:50], _pair_schema), hide_index=True, width="stretch")
            st.download_button(
                "Download (.xlsx)",
                data=rows_to_excel_bytes(_pair_rows, [n for n, _ in _pair_schema]),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"nat_par_{year}_{mode}",
            )
    else:
        st.error("Ingen forfatterpar med kendte nationaliteter for begge endepunkter i det valgte udsnit.")

    # Perspektivtekst fra cs_nat_pairs (inkl. intra-enhed)
    _dk_same = _dk_cross = _intl_same = _intl_cross = 0
    for raw_key, partner_counts in (cs_nat_pairs or {}).items():
        parts = raw_key.split("|")
        if len(parts) != 4: continue
        fac, inst, grp, focal_cs = parts
        if selected_facs  and fac  not in selected_facs:  continue
        if selected_insts and inst not in selected_insts: continue
        if selected_grps  and grp  not in selected_grps:  continue
        if _nat_fac_filter  and fac  not in _nat_fac_filter:  continue
        if _nat_inst_filter and inst not in _nat_inst_filter: continue
        if _nat_grp_filter  and grp  not in _nat_grp_filter:  continue
        for partner_cs, count in partner_counts.items():
            if focal_cs == "DK":
                if partner_cs == "DK": _dk_same  += count
                else:                  _dk_cross += count
            else:
                if partner_cs != "DK": _intl_same  += count
                else:                  _intl_cross += count

    _dk_total   = (_dk_same   + _dk_cross)   or 1
    _intl_total = (_intl_same + _intl_cross) or 1


    dk_tot: dict[str, int] = {"DK": 0, "International": 0}
    for nid, m in _raw_nodes_f.items():
        if m.get("type") != "grp":
            continue
        cs   = m.get("statsborgerskab", "")
        size = m.get("size", 0)
        key  = "DK" if cs == "DK" else "International"
        dk_tot[key] += size

    dk_ew: dict[str, float] = {"DK": 0.0, "International": 0.0}
    for edge in _edges:
        u, v, w = edge[0], edge[1], edge[2]
        if not _nat_nid_ok(u) and not _nat_nid_ok(v):
            continue
        cs_u = _nid_to_cs.get(u, "")
        cs_v = _nid_to_cs.get(v, "")
        if cs_u:
            dk_ew["DK" if cs_u == "DK" else "International"] += w
        if cs_v:
            dk_ew["DK" if cs_v == "DK" else "International"] += w

    orgs_for_comparison = sorted(cs_size.keys())

    org_cs_ew: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for edge in _edges:
        u, v, w = edge[0], edge[1], edge[2]
        if not _nat_nid_ok(u) and not _nat_nid_ok(v):
            continue
        cs_u = _nid_to_cs.get(u, "")
        cs_v = _nid_to_cs.get(v, "")
        org_u = _org_from_node_id(u)
        org_v = _org_from_node_id(v)
        key_u = "DK" if cs_u == "DK" else "International"
        key_v = "DK" if cs_v == "DK" else "International"
        if _nat_nid_ok(u) and org_u in cs_size:
            org_cs_ew[org_u][key_u] += w
        if _nat_nid_ok(v) and org_v in cs_size:
            org_cs_ew[org_v][key_v] += w

    share_rows = []
    for org in orgs_for_comparison:
        _org_tot = {}
        for nid, m in raw_nodes.items():
            if m.get("type") != "grp" or _org_from_node_id(nid) != org:
                continue
            cs   = m.get("statsborgerskab", "")
            size = m.get("size", 0)
            key  = "DK" if cs == "DK" else "International"
            _org_tot[key] = _org_tot.get(key, 0) + size
        ew     = org_cs_ew.get(org, {})
        fb_tot = sum(_org_tot.values()) or 1
        ew_tot = sum(ew.values()) or 1
        for key in ["DK", "International"]:
            pct_fb = round(100 * _org_tot.get(key, 0) / fb_tot, 1)
            pct_ew = round(100 * ew.get(key, 0.0) / ew_tot, 1)
            share_rows.append({
                group_label:                  org,
                "Nationalitet":               key,
                "Andel forfattere (%)":  pct_fb,
                "Andel sampubliceringer (%)": pct_ew,
                "Forskel (pp)":               round(pct_ew - pct_fb, 1),
            })

# MED:
    # Intra-enhed forfatterpar fra cs_nat_pairs
    org_cs_ew_all: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for raw_key, partner_counts in (cs_nat_pairs or {}).items():
        parts = raw_key.split("|")
        if len(parts) != 4: continue
        fac, inst, grp, focal_cs = parts
        if selected_facs  and fac  not in selected_facs:  continue
        if selected_insts and inst not in selected_insts: continue
        if selected_grps  and grp  not in selected_grps:  continue
        if _nat_fac_filter  and fac  not in _nat_fac_filter:  continue
        if _nat_inst_filter and inst not in _nat_inst_filter: continue
        if _nat_grp_filter  and grp  not in _nat_grp_filter:  continue
        org = _nat_org_key({"fac": fac, "inst": inst, "grp": grp})
        cs_label = "DK" if focal_cs == "DK" else "International"
        for count in partner_counts.values():
            org_cs_ew_all[org][cs_label] += count / 2

    share_rows = []
    for org in orgs_for_comparison:
        _org_tot = {}
        for nid, m in raw_nodes.items():
            if m.get("type") != "grp" or _org_from_node_id(nid) != org:
                continue
            cs   = m.get("statsborgerskab", "")
            size = m.get("size", 0)
            key  = "DK" if cs == "DK" else "International"
            _org_tot[key] = _org_tot.get(key, 0) + size
        ew_net = org_cs_ew.get(org, {})
        ew_all = org_cs_ew_all.get(org, {})
        fb_tot      = sum(_org_tot.values()) or 1
        ew_net_tot  = sum(ew_net.values()) or 1
        ew_all_tot  = sum(ew_all.values()) or 1
        for key in ["DK", "International"]:
            pct_fb      = round(100 * _org_tot.get(key, 0)     / fb_tot,     1)
            pct_ew_net  = round(100 * ew_net.get(key, 0.0)     / ew_net_tot, 1)
            pct_ew_all  = round(100 * ew_all.get(key, 0.0)     / ew_all_tot, 1)
            share_rows.append({
                group_label:                               org,
                "Nationalitet":                            key,
                "Andel forfattere (%)":               pct_fb,
                "Andel sampubliceringer - netværk (%)":    pct_ew_net,
                "Andel sampubliceringer - alle par (%)":   pct_ew_all,
                "Forskel netværk (pp)":                    round(pct_ew_net - pct_fb, 1),
                "Forskel alle par (pp)":                   round(pct_ew_all - pct_fb, 1),
            })

    st.markdown("---")
    
    # ── Nationalitetshomofili ─────────────────────────────────────────────────
    st.markdown("#### Nationalitetshomofili - er der en 'danskerklub'?")
    st.markdown(f"""
Indekset viser om danske og internationale forskere publicerer med kolleger af samme nationalitetsgruppe 
oftere end den lokale sammensætning ville forudsige. Beregningen er identisk med kønshomofili-indekset:

1. **Baseline** - andelen af DK/internationale forfattere i enheden
2. **Same-group rate** - andelen af forfatterpar, set fra gruppens eget perspektiv, der er med en 
   kollega fra samme gruppe
3. **Homofili-indeks** = same-group rate / baseline

Opgørelsen inkluderer intra-enhed par (f.eks. to danske professorer på Kemisk Institut) og alle nationaliteter - også
dem, der ikke er blevet valgt i sidepanelet. 

**Metodisk forbehold:** Samme forbehold som ved kønshomofili gælder - forskere publicerer primært 
inden for snævrere grupper der kan have skæve sammensætninger ift. den valgte enhed. Desuden
kan udgivelsessprog også have en effekt på nationalitetshomofili. 

Enheder med færre end fem forfatterpar vises ikke, da indekset er ustabilt ved få observationer.
""")

    _cs_hom_combo: dict[str, dict[str, dict[str, int]]] = defaultdict(
        lambda: {"DK": {"same": 0, "cross": 0}, "intl": {"same": 0, "cross": 0}})
    for raw_key, counts in (cs_pairs or {}).items():
        parts = raw_key.split("|")
        if len(parts) != 4:
            continue
        fac, inst, grp, focal_cs = parts
        if selected_facs  and fac  not in selected_facs:  continue
        if selected_insts and inst not in selected_insts: continue
        if selected_grps  and grp  not in selected_grps:  continue
        if _nat_fac_filter  and fac  not in _nat_fac_filter:  continue
        if _nat_inst_filter and inst not in _nat_inst_filter: continue
        if _nat_grp_filter  and grp  not in _nat_grp_filter:  continue
        org_key = _nat_org_key({"fac": fac, "inst": inst, "grp": grp})
        focal_cs_dk = "DK" if focal_cs == "DK" else "intl"
        _cs_hom_combo[org_key][focal_cs_dk]["same"]  += counts.get("same", 0)
        _cs_hom_combo[org_key][focal_cs_dk]["cross"] += counts.get("cross", 0)

    # Baseline fra cs_size (allerede beregnet)
    _hom_source = raw_nodes_pre_cs if raw_nodes_pre_cs is not None else raw_nodes

    # Per-org baseline: forfatterbidrag pr. enhed, opdelt på DK/intl
    _cs_hom_baseline: dict[str, dict[str, int]] = {}
    for nid, m in _hom_source.items():
        if m.get("type") != "grp":
            continue
        # Anvend de samme org-filtre som ellers i fanen
        if not _nat_node_ok(m):
            continue
        if selected_facs  and m.get("fac","")  not in selected_facs:  continue
        if selected_insts and m.get("inst","") not in selected_insts: continue
        if selected_grps  and m.get("grp","")  not in selected_grps:  continue
        org = _nat_org_key(m)
        if not org:
            continue
        cs_dk = "DK" if m.get("statsborgerskab", "") == "DK" else "intl"
        size = m.get("size", 0)
        _cs_hom_baseline.setdefault(org, {"DK": 0, "intl": 0})
        _cs_hom_baseline[org][cs_dk] += size

    _cs_hom_rows = []
    for org in sorted(_cs_hom_combo.keys()):
        fb = _cs_hom_baseline.get(org, {"DK": 0, "intl": 0})
        total_fb = (fb["DK"] + fb["intl"]) or 1
        baseline_dk   = fb["DK"]   / total_fb
        baseline_intl = fb["intl"] / total_fb

        for cs_grp, label, baseline in [("DK", "DK", baseline_dk), ("intl", "International", baseline_intl)]:
            d = _cs_hom_combo[org][cs_grp]
            pairs = d["same"] + d["cross"] or 1
            rate  = d["same"] / pairs
            min_pairs = 5  # kræv mindst 5 forfatterpar for at vise indeks
            index = round(rate / baseline, 2) if (baseline > 0 and (d["same"] + d["cross"]) >= min_pairs) else None
            _cs_hom_rows.append({
                group_label: org,
                "Gruppe":            label,
                "Homofili-indeks":   index,
            })

    # KU samlet - samme princip: baseline mod ALLE nationaliteter
    _ku_cs_sp = {"DK": {"same": 0, "cross": 0}, "intl": {"same": 0, "cross": 0}}
    for raw_key, counts in (cs_pairs or {}).items():
        parts = raw_key.split("|")
        if len(parts) != 4:
            continue
        focal_cs = parts[3]
        focal_cs_dk = "DK" if focal_cs == "DK" else "intl"
        _ku_cs_sp[focal_cs_dk]["same"]  += counts.get("same", 0)
        _ku_cs_sp[focal_cs_dk]["cross"] += counts.get("cross", 0)

    _ku_baseline_counts = {"DK": 0, "intl": 0}
    for nid, m in _hom_source.items():
        if m.get("type") != "grp":
            continue
        cs_dk = "DK" if m.get("statsborgerskab", "") == "DK" else "intl"
        _ku_baseline_counts[cs_dk] += m.get("size", 0)
    _ku_total_baseline = (_ku_baseline_counts["DK"] + _ku_baseline_counts["intl"]) or 1
    _ku_cs_baseline_dk   = _ku_baseline_counts["DK"]   / _ku_total_baseline
    _ku_cs_baseline_intl = _ku_baseline_counts["intl"] / _ku_total_baseline

    _cs_hom_ku = {}
    for cs_grp, label, baseline in [("DK", "DK", _ku_cs_baseline_dk), ("intl", "International", _ku_cs_baseline_intl)]:
        d = _ku_cs_sp[cs_grp]
        pairs = d["same"] + d["cross"] or 1
        rate  = d["same"] / pairs
        _cs_hom_ku[label] = round(rate / baseline, 2) if (baseline > 0 and (d["same"] + d["cross"]) >= min_pairs) else None

    # Plot
    if not _hide_nat_ku_samlet:
        ku_klub = round(_cs_hom_ku.get("DK") / _cs_hom_ku.get("International"), 2) if _cs_hom_ku.get("DK") and _cs_hom_ku.get("International") else None
        _cs_hom_rows = [
            {group_label: "KU samlet", "Gruppe": "DK",            "Homofili-indeks": _cs_hom_ku.get("DK")},
            {group_label: "KU samlet", "Gruppe": "International", "Homofili-indeks": _cs_hom_ku.get("International")},
        ] + _cs_hom_rows

    _orgs_cs_hom = (
        (["KU samlet"] if not _hide_nat_ku_samlet else []) +
        sorted(
            {r[group_label] for r in _cs_hom_rows if r[group_label] != "KU samlet"},
            key=_nat_sort_key
        )
    )

    _cs_klub_rows = []
    for org in _orgs_cs_hom:
        if org == "KU samlet":
            klub = round(_cs_hom_ku.get("DK") / _cs_hom_ku.get("International"), 2) if _cs_hom_ku.get("DK") and _cs_hom_ku.get("International") else None
        else:
            idx_dk   = next((r["Homofili-indeks"] for r in _cs_hom_rows if r[group_label] == org and r["Gruppe"] == "DK"), None)
            idx_intl = next((r["Homofili-indeks"] for r in _cs_hom_rows if r[group_label] == org and r["Gruppe"] == "International"), None)
            klub = round(idx_dk / idx_intl, 2) if idx_dk and idx_intl and idx_intl > 0 else None
        _cs_klub_rows.append({group_label: org, "Klubfaktor (DK/Intl)": klub})

    _hom_cs_y, _hom_cs_tick, _hom_cs_labels = _build_nat_y_positions(_orgs_cs_hom)

    _tab_cs_hom_idx, _tab_cs_hom_klub = st.tabs(["Homofili-indeks", "Klubfaktor"])

    with _tab_cs_hom_idx:
        st.markdown(
"""
- **> 1** - publicerer med samme gruppe oftere end forventet → tendens til clustering
- **= 1** - ingen tendens
- **< 1** - publicerer med samme gruppe sjældnere end forventet
""")
        _fig_cs_hom = go.Figure()
        for grp_label, color in [("DK", "#425570"), ("International", "#901a1E")]:
            vals = [next((r["Homofili-indeks"] or 0 for r in _cs_hom_rows
                          if r[group_label] == org and r["Gruppe"] == grp_label), 0)
                    if org and org.strip() else 0
                    for org in _orgs_cs_hom]
            _fig_cs_hom.add_trace(go.Bar(
                name=grp_label, y=_hom_cs_y, x=vals, orientation="h",
                marker_color=color, width=0.28,
                text=[f"{v:.2f}" for v in vals], textposition="inside",
                customdata=_orgs_cs_hom,
                hovertemplate="<b>%{customdata}</b><br>Homofili-indeks: %{x:.2f}<extra></extra>",
            ))
        _fig_cs_hom.add_vline(x=1.0, line_dash="dash", line_color="#3d3d3d",
            annotation_text="Ingen homofili (= 1)", annotation_position="top right")
        _layout_cs_hom = _nat_bar_layout(_orgs_cs_hom, grouped=True)
        _layout_cs_hom["bargroupgap"] = 0.0
        _fig_cs_hom.update_layout(
            barmode="group", xaxis_title="Homofili-indeks", legend_title="Gruppe",
            yaxis=dict(
                tickmode="array",
                tickvals=_hom_cs_tick,
                ticktext=_hom_cs_labels,
                autorange="reversed",
                showgrid=False,
                zeroline=False,
            ),
            title=f"Nationalitetshomofili per {group_label.lower()}, {year}",
            **_layout_cs_hom,
        )
        st.plotly_chart(_fig_cs_hom, width='content',
            config={"toImageButtonOptions": {"format": "png", "scale": 3}})
        if filter_caption:
            st.caption(filter_caption)
        
    with _tab_cs_hom_klub:
        st.markdown(
"""
Klubfaktoren = DK-homofili-indeks / international-homofili-indeks.

- **> 1** - danske forskere klumper sig relativt mere sammen end internationale → *danskerklub*-tendens
- **= 1** - ingen forskel
- **< 1** - internationale forskere klumper sig relativt mere
""")
        _klub_vals = [
            (next((r["Klubfaktor (DK/Intl)"] for r in _cs_klub_rows
                   if r[group_label] == org), 0) or 0)
            if org and org.strip() else 0
            for org in _orgs_cs_hom
        ]
        _fig_cs_klub = go.Figure()
        _fig_cs_klub.add_trace(go.Bar(
            name="Klubfaktor", y=_hom_cs_y, x=_klub_vals, orientation="h",
            marker_color=["#425570" if (v or 0) >= 1 else "#901a1E" for v in _klub_vals],
            width=0.55,
            text=[f"{v:.2f}" for v in _klub_vals], textposition="inside",
            customdata=_orgs_cs_hom,
            hovertemplate="<b>%{customdata}</b><br>Klubfaktor: %{x:.2f}<extra></extra>",
        ))
        _fig_cs_klub.add_vline(x=1.0, line_dash="dash", line_color="#3d3d3d",
            annotation_text="Ingen forskel (= 1)", annotation_position="top right")
        _fig_cs_klub.update_layout(
            xaxis_title="Klubfaktor (DK/Intl)",
            yaxis=dict(
                tickmode="array",
                tickvals=_hom_cs_tick,
                ticktext=_hom_cs_labels,
                autorange="reversed",
                showgrid=False,
                zeroline=False,
            ),
            title=f"Nationalitetsklubfaktor (DK/International) per {group_label.lower()}, {year}",
            **_nat_bar_layout(_orgs_cs_hom, grouped=False),
        )
        ku_klub = round(_cs_hom_ku.get("DK") / _cs_hom_ku.get("International"), 2) if _cs_hom_ku.get("DK") and _cs_hom_ku.get("International") else None
        st.plotly_chart(_fig_cs_klub, width='content',
            config={"toImageButtonOptions": {"format": "png", "scale": 3}})
        if filter_caption:
            st.caption(filter_caption)

    _cs_hom_tabel = []
    for org in _orgs_cs_hom:
        idx_dk   = next((r["Homofili-indeks"] for r in _cs_hom_rows if r[group_label] == org and r["Gruppe"] == "DK"), None)
        idx_intl = next((r["Homofili-indeks"] for r in _cs_hom_rows if r[group_label] == org and r["Gruppe"] == "International"), None)
        klub = next((r["Klubfaktor (DK/Intl)"] for r in _cs_klub_rows if r[group_label] == org), None)
        _cs_hom_tabel.append({
            group_label:              org,
            "Homofili-indeks DK":     idx_dk,
            "Homofili-indeks Intl":   idx_intl,
            "Klubfaktor (DK/Intl)":   klub,
        })
    _cs_hom_schema = [
        (group_label,               pa.string()),
        ("Homofili-indeks DK",      pa.float64()),
        ("Homofili-indeks Intl",    pa.float64()),
        ("Klubfaktor (DK/Intl)",    pa.float64()),
    ]
    with st.expander("Se tabel"):
        st.dataframe(build_table(_cs_hom_tabel, _cs_hom_schema), hide_index=True, width="stretch")
        st.download_button(
            "Download (.xlsx)",
            data=rows_to_excel_bytes(_cs_hom_tabel, [n for n, _ in _cs_hom_schema]),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_cs_homofili_{year}_{mode}",
        )
    

    st.markdown("---")

    # _______Udvikling over tid
    if all_years_data and len(all_years_data) >= 2:
        st.subheader("Udvikling over tid")
        years_sorted = sorted(all_years_data.keys())
        _all_cs = top_cs

        nat_colors_time = ku_color_sequence(len(_all_cs) + 1)
        _tab_nat_tid, _tab_intl_tid, _tab_hom_tid = st.tabs([
            "Forfatterantal per nationalitet",
            "Andel internationale forfatterpar",
            "Nationalitetshomofili over tid",
        ])

        with _tab_nat_tid:
            fig_nat = go.Figure()
            for i, cs in enumerate(_all_cs):
                fig_nat.add_trace(go.Scatter(
                    name=country_name(cs) if cs != "Andre" else "Andre",
                    x=years_sorted,
                    y=[all_years_data[y].get("nat_bidrag", {}).get(cs, 0) for y in years_sorted],
                    mode="lines+markers",
                    line=dict(color=nat_colors_time[i], width=2),
                    marker=dict(size=8),
                ))

            fig_nat.add_trace(go.Scatter(
                x=years_sorted,
                y=[sum(all_years_data[y].get("nat_bidrag", {}).get(cs, 0) for cs in _all_cs)
                for y in years_sorted],
                name="Total (top-N)",
                mode="lines+markers",
                line=dict(color="#3d3d3d", width=2, dash="dot"),
                marker=dict(size=8),
            ))
            fig_nat.add_trace(go.Scatter(
                x=years_sorted,
                y=[sum(v for k, v in all_years_data[y].get("nat_bidrag", {}).items() if k != "DK")
                for y in years_sorted],
                name="International (alle)",
                mode="lines+markers",
                line=dict(color="#901a1E", width=2, dash="dash"),
                marker=dict(size=8),
            ))

            fig_nat.update_layout(
                xaxis=dict(tickmode="array", tickvals=years_sorted, dtick=1),
                yaxis_title="Forfatterantal",
                legend_title="Statsborgerskab",
                height=500,
                margin=dict(t=50),
                title=f"Forfatterantal fordelt på statsborgerskab over tid, {years_sorted[0]}–{years_sorted[-1]}"
            )
            st.plotly_chart(fig_nat, width="stretch",
            config={
                "toImageButtonOptions": {
                    "format": "png",
                    "scale": 3,
                }
            }
            )
            if filter_caption:
                st.caption(filter_caption)

            # Tabel
            _nat_pivot = (
                [{"Nationalitet": country_name(cs),
                  **{str(y): all_years_data[y].get("nat_bidrag", {}).get(cs, 0) for y in years_sorted}}
                 for cs in _all_cs] +
                [{"Nationalitet": "Total",
                  **{str(y): sum(all_years_data[y].get("nat_bidrag", {}).get(cs, 0) for cs in _all_cs) for y in years_sorted}}]
            )
            _nat_pivot_schema = [("Nationalitet", pa.string())] + [(str(y), pa.int64()) for y in years_sorted]
            with st.expander("Se tabel"):
                st.dataframe(build_table(_nat_pivot, _nat_pivot_schema), hide_index=True, width="stretch")
                st.download_button("Download (.xlsx)",
                    data=rows_to_excel_bytes(_nat_pivot, [n for n, _ in _nat_pivot_schema]),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_nat_tid_{mode}")

        with _tab_intl_tid:
            st.markdown(
                "Andelen af forfatterpar der involverer mindst én international forfatter, "
                "beregnet per år. En stigende kurve indikerer øget internationalt samarbejde."
            )
            _use_all_intl = st.toggle(
                "Inkludér intra-enhed forfatterpar (f.eks. to professorer på samme institut)",
                value=False,
                key=f"intl_tid_all_{year}_{mode}",
                help="Til: alle forfatterpar inkl. par inden for samme enhed. Fra: kun netværkskanter.",
                disabled=not show_intra
            )
            if _use_all_intl:
                st.caption("Alle forfatterpar inkl. par inden for samme enhed (f.eks. to mandlige professorer på Kemisk Institut).")
            else:
                st.caption("Kun forfatterpar *mellem* de valgte organisatoriske enheder - som vist i netværket.")


            _nat_ew_key = "nat_ew_all" if _use_all_intl else "nat_ew"
            _intl_rows = []
            _intl_y_pct = []
            for yr in years_sorted:
                _nat_ew  = all_years_data[yr].get(_nat_ew_key, {})
                _dk_dk   = _nat_ew.get("dk_dk",  0.0)
                _dk_int  = _nat_ew.get("dk_int",  0.0)
                _int_int = _nat_ew.get("int_int", 0.0)
                _tot_ew  = (_dk_dk + _dk_int + _int_int) or 1
                _pct_dk_dk   = round(100 * _dk_dk   / _tot_ew, 1)
                _pct_dk_int  = round(100 * _dk_int  / _tot_ew, 1)
                _pct_int_int = round(100 * _int_int / _tot_ew, 1)
                _intl_y_pct.append(_pct_int_int)
                _intl_rows.append({"År": int(yr),
                    "DK-DK (%)":          _pct_dk_dk,
                    "DK-International (%)": _pct_dk_int,
                    "Int-Int (%)":         _pct_int_int,
                    "DK-DK (n)":          round(_dk_dk, 1),
                    "DK-International (n)": round(_dk_int, 1),
                    "Int-Int (n)":         round(_int_int, 1),
                })
            fig_intl = go.Figure()
            for _label, _key, _color in [
                ("DK-DK (%)",           "DK-DK (%)",           "#425570"),
                ("DK-International (%)", "DK-International (%)", "#bac7d9"),
                ("Int-Int (%)",          "Int-Int (%)",          "#901a1E"),
            ]:
                fig_intl.add_trace(go.Scatter(
                    x=years_sorted,
                    y=[r[_key] for r in _intl_rows],
                    name=_label, mode="lines+markers",
                    line=dict(color=_color, width=2), marker=dict(size=8),
                ))
            fig_intl.update_layout(
                xaxis=dict(tickmode="array", tickvals=years_sorted, dtick=1),
                yaxis_title="Andel af forfatterpar (%)",
                height=400, margin=dict(t=50),
                title=f"Andel internationale forfatterpar{' (inkl. intra-enhed)' if _use_all_intl else ''} over tid, {years_sorted[0]}–{years_sorted[-1]}"
            )
            st.plotly_chart(fig_intl, width="stretch", key=f"intl_tid_{year}_{mode}",
        config={
            "toImageButtonOptions": {
                "format": "png",
                "scale": 3,
            }
        }
            )
            if filter_caption:
                st.caption(filter_caption)

            _intl_metrics = [
                "DK-DK (%)", "DK-International (%)", "Int-Int (%)",
                "DK-DK (n)", "DK-International (n)", "Int-Int (n)",
            ]
            _intl_pivot = [
                {"Metrik": m,
                 **{str(r["År"]): r[m] for r in _intl_rows}}
                for m in _intl_metrics
            ]
            _intl_pivot_schema = [("Metrik", pa.string())] + [(str(y), pa.float64()) for y in years_sorted]
            with st.expander("Se tabel"):
                st.dataframe(build_table(_intl_pivot, _intl_pivot_schema), hide_index=True, width="stretch")
                st.download_button("Download (.xlsx)",
                    data=rows_to_excel_bytes(_intl_pivot, [n for n, _ in _intl_pivot_schema]),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_intl_tid_{mode}")

        with _tab_hom_tid:
                st.markdown(
"""
Homofili-indeks for DK og internationale forfattere per år - KU samlet. Et indeks > 1 
betyder at gruppen publicerer med samme gruppe oftere end den samlede sammensætning 
ville forudsige. 

**Klubfaktoren** (DK/International) viser det relative forhold mellem de to gruppers homofili: 
en klubfaktor > 1 betyder at danske forskere har en stærkere tendens til at publicere med 
andre danskere end internationale forskere har til at publicere med andre internationale - 
altså en *danskerklub*-tendens. En klubfaktor < 1 indikerer det modsatte.
"""
                )
                _hom_years_cs = [y for y in years_sorted if all_years_data[y].get("cs_hom_index_ku")]
                _klub_years_cs = [y for y in _hom_years_cs
                                  if all_years_data[y]["cs_hom_index_ku"].get("DK")
                                  and all_years_data[y]["cs_hom_index_ku"].get("International")]
                _klub_vals_cs = [
                    round(all_years_data[y]["cs_hom_index_ku"]["DK"] /
                          all_years_data[y]["cs_hom_index_ku"]["International"], 3)
                    for y in _klub_years_cs]
                fig_hom_cs = go.Figure()
                for label, color in [("DK", "#425570"), ("International", "#901a1E")]:
                    fig_hom_cs.add_trace(go.Scatter(
                        x=_hom_years_cs,
                        y=[all_years_data[y]["cs_hom_index_ku"].get(label) for y in _hom_years_cs],
                        name=label, mode="lines+markers",
                        line=dict(color=color, width=2), marker=dict(size=8),
                    ))
                fig_hom_cs.add_trace(go.Scatter(
                    x=_klub_years_cs, y=_klub_vals_cs,
                    name="Klubfaktor (DK/Intl)", mode="lines+markers",
                    line=dict(color="#122947", width=2), marker=dict(size=8),
                ))
                fig_hom_cs.add_hline(y=1.0, line_dash="dash", line_color="#3d3d3d",
                    annotation_text="Ingen homofili (= 1)", annotation_position="bottom right")
                fig_hom_cs.update_layout(
                    xaxis=dict(tickmode="array", tickvals=_hom_years_cs, dtick=1),
                    yaxis_title="Homofili-indeks", legend_title="Gruppe",
                    height=360, margin=dict(t=50),
                    title=f"Nationalitetshomofili og klubfaktor over tid, {years_sorted[0]}–{years_sorted[-1]}"
                )
                st.plotly_chart(fig_hom_cs, width='content',
                    config={"toImageButtonOptions": {"format": "png", "scale": 3}})
                if filter_caption:
                    st.caption(filter_caption)
                
                _hom_cs_pivot = (
                    [{"Gruppe": label,
                      **{str(y): round(all_years_data[y]["cs_hom_index_ku"].get(label) or 0, 3)
                         if y in _hom_years_cs else None
                         for y in years_sorted}}
                     for label in ["DK", "International"]] +
                    [{"Gruppe": "Klubfaktor (DK/Intl)",
                      **{str(y): round(all_years_data[y]["cs_hom_index_ku"]["DK"] /
                                       all_years_data[y]["cs_hom_index_ku"]["International"], 3)
                         if y in _klub_years_cs else None
                         for y in years_sorted}}]
                )
                _hom_cs_schema = [("Gruppe", pa.string())] + [(str(y), pa.float64()) for y in years_sorted]
                with st.expander("Se tabel"):
                    st.dataframe(build_table(_hom_cs_pivot, _hom_cs_schema), hide_index=True, width="stretch")
                    st.download_button("Download (.xlsx)",
                        data=rows_to_excel_bytes(_hom_cs_pivot, [n for n, _ in _hom_cs_schema]),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_hom_cs_tid_{mode}")

if __name__ == "__main__":
    main()